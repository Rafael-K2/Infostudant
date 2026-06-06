"""
SERVIDOR MARWIN - ServidorV15.py
======================================
Rode ESTE arquivo no PC da escola.

O que ele faz ao iniciar:
  1. Prepara o servidor Flask na porta 5000
  2. Conecta-se ao PostgreSQL Neon como fonte de dados
  3. Abre o painel admin Tkinter com IP local e botão copiar URL

Instalação (uma vez só):
  pip install flask flask-cors fpdf2 qrcode pillow psycopg2-binary

O index.html pode usar a API na nuvem (ApiNuvem.py) — configure dados/cloud_config.json.
O painel admin sincroniza cardápio/eventos/config com a nuvem automaticamente.

Organização dos QR Codes: qrcodes_marwin/
    1 Ano - Desenvolvimento de Sistemas/
      2026001_Joao_Silva_1_Ano_Desenvolvimento_de_Sistemas.png
      ...
    2 Ano - Redes de Computadores/
      ...
    3 Ano - Informatica/
      ...
"""

import sys, subprocess

for pkg in ["flask", "flask-cors", "fpdf2", "qrcode", "pillow", "psycopg2-binary"]:
    try:
        __import__(pkg.replace("-", "_").replace("pillow","PIL"))
    except ImportError:
        print(f"Instalando {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

import os, json, csv, datetime, threading, calendar, time, io, base64, re, logging, smtplib, shutil, zipfile, unicodedata, socket
from collections import Counter
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import tkinter as tk
from tkinter import ttk, messagebox

import qrcode
from PIL import Image, ImageTk
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from logging.handlers import TimedRotatingFileHandler

app = Flask(__name__)
CORS(app)

DADOS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "dados"))
os.makedirs(DADOS_DIR, exist_ok=True)

# ── Configuração de Logging ───────────────────────────────────────────────────
LOG_FILE = os.path.join(DADOS_DIR, "marwin.log")
logger = logging.getLogger("marwin")
logger.setLevel(logging.DEBUG)
try:
    log_handler = TimedRotatingFileHandler(
        LOG_FILE, when="midnight", interval=1, backupCount=7, encoding="utf-8"
    )
    log_handler.setLevel(logging.DEBUG)
    log_format = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s",
                                     datefmt="%d/%m/%Y %H:%M:%S")
    log_handler.setFormatter(log_format)
    logger.addHandler(log_handler)
except Exception as e:
    print(f"[LOG] Erro ao configurar logging: {e}")

def _buscar_logo_png():
    if os.path.isdir(DADOS_DIR):
        logos = [n for n in os.listdir(DADOS_DIR) if n.lower().endswith(".png")]
        if logos:
            for nome in logos:
                if "logo" in nome.lower():
                    return os.path.join(DADOS_DIR, nome)
            return os.path.join(DADOS_DIR, logos[0])
    return None

CARDAPIO_FILE    = os.path.join(DADOS_DIR, "cardapio.json")
EVENTOS_FILE     = os.path.join(DADOS_DIR, "eventos.json")
CONFIG_FILE      = os.path.join(DADOS_DIR, "config_sistema.json")
TEMA_FILE        = os.path.join(DADOS_DIR, "tema_config.json")
LISTA_ALUNOS_FILE = os.path.join(DADOS_DIR, "lista_alunos.json")
DB_CONFIG_FILE   = os.path.join(DADOS_DIR, "db_config.json")
CLOUD_CONFIG_FILE = os.path.join(DADOS_DIR, "cloud_config.json")
MESES_PT = (
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
)
DEFAULT_ADMIN_PLAIN = "Marwin2026"
# Prefer env var MARWIN_ADMIN_PASS (pode conter hash bcrypt). Fallback para DEFAULT_ADMIN_PLAIN.
ADMIN_PASSWORD = os.getenv("MARWIN_ADMIN_PASS", DEFAULT_ADMIN_PLAIN)

# Try to import bcrypt if available (optional). If ADMIN_PASSWORD is a bcrypt hash, we'll use it.
try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except Exception:
    BCRYPT_AVAILABLE = False

def ler_json(path, padrao):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return padrao

def salvar_json(path, dados):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def _ler_cloud_config():
    return ler_json(CLOUD_CONFIG_FILE, {"api_url": "", "sincronizar_automatico": True})

def _cloud_api_url():
    return _ler_cloud_config().get("api_url", "").strip().rstrip("/")

def _sync_nuvem(rota, metodo, dados=None):
    base = _cloud_api_url()
    if not base or not _ler_cloud_config().get("sincronizar_automatico", True):
        return False
    import urllib.request
    url = base + rota
    body = json.dumps(dados, ensure_ascii=False).encode("utf-8") if dados is not None else None
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json", "X-Senha": ADMIN_PASSWORD},
        method=metodo,
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        logger.warning(f"Sync nuvem falhou ({rota}): {e}")
        return False

def _sincronizar_tudo_nuvem():
    ok = True
    if not _sync_nuvem("/admin/cardapio", "PUT", ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO)):
        ok = False
    if not _sync_nuvem("/admin/eventos", "PUT", ler_json(EVENTOS_FILE, EVENTOS_PADRAO)):
        ok = False
    if not _sync_nuvem("/admin/config", "PUT", ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"})):
        ok = False
    if ok and _cloud_api_url():
        logger.info("Configurações sincronizadas com a API na nuvem")
    return ok

DB_DEFAULT_CONNECTION = "postgresql://neondb_owner:npg_ydP7rqBR0ZoQ@ep-broad-mountain-apatc77i-pooler.c-7.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"
PG_POOL = None

def _carregar_db_config():
    cfg = ler_json(DB_CONFIG_FILE, {})
    if not cfg.get("connection_string"):
        cfg["connection_string"] = DB_DEFAULT_CONNECTION
        salvar_json(DB_CONFIG_FILE, cfg)
    return cfg

def _iniciar_pool_pg():
    global PG_POOL
    if PG_POOL is not None:
        return
    cfg = _carregar_db_config()
    try:
        import psycopg2
        from psycopg2 import pool
        PG_POOL = pool.ThreadedConnectionPool(1, 10, cfg["connection_string"])
        logger.info("Pool PostgreSQL inicializado")
    except Exception as e:
        PG_POOL = None
        logger.warning("PostgreSQL indisponível")

def get_pg_conn():
    global PG_POOL
    if PG_POOL is None:
        _iniciar_pool_pg()
    if PG_POOL is None:
        return None
    try:
        return PG_POOL.getconn()
    except Exception as e:
        logger.warning("Falha ao obter conexão PostgreSQL")
        return None

def _release_pg_conn(conn):
    global PG_POOL
    if not conn:
        return
    try:
        if PG_POOL:
            PG_POOL.putconn(conn)
        else:
            conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

def _executar_pg(sql, params=None, fetch=False):
    conn = get_pg_conn()
    if conn is None:
        raise RuntimeError("PostgreSQL indisponível")
    cur = None
    try:
        cur = conn.cursor()
        cur.execute(sql, params or ())
        if fetch:
            cols = [d[0] for d in cur.description] if cur.description else []
            rows = cur.fetchall()
            return [dict(zip(cols, row)) for row in rows]
        conn.commit()
        return []
    finally:
        if cur is not None:
            try:
                cur.close()
            except Exception:
                pass
        _release_pg_conn(conn)

def _get_local_ip():
    try:
        ip = socket.gethostbyname(socket.gethostname())
        if ip.startswith("127.") or ip == "0.0.0.0":
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect(("8.8.8.8", 80))
                ip = s.getsockname()[0]
            finally:
                s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _criar_tabelas_neon():
    try:
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS refeitorio (
                data TEXT NOT NULL,
                horaentrada TEXT NOT NULL,
                matricula TEXT NOT NULL,
                nome TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                refeicao TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS frequencia (
                data TEXT NOT NULL,
                horaentrada TEXT NOT NULL,
                matricula TEXT NOT NULL,
                nome TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                aula TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS avaliacoes (
                data TEXT NOT NULL,
                aluno TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                estagio TEXT,
                item TEXT,
                nota TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS sistema_config (
                chave TEXT PRIMARY KEY,
                valor JSONB NOT NULL
            );
            """
        )
        logger.info("Tabelas Neon verificadas/criadas")
    except Exception as e:
        logger.warning(f"Nao foi possivel criar tabelas Neon: {e}")


def _ler_refeitorio_hoje_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, refeicao FROM refeitorio WHERE data = %s",
        (_hoje(),),
        fetch=True
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["refeicao"]]
        for row in rows
    ]


def _inserir_refeitorio_db(registro):
    _executar_pg(
        "INSERT INTO refeitorio (data, horaentrada, matricula, nome, serie, curso, refeicao) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _apagar_refeitorio_data_db(data_alvo):
    _executar_pg("DELETE FROM refeitorio WHERE data = %s", (data_alvo,))


def _refeitorio_duplicado_db(matricula, refeicao):
    try:
        rows = _executar_pg(
            "SELECT horaentrada, nome FROM refeitorio WHERE data = %s AND matricula = %s AND refeicao = %s LIMIT 1",
            (_hoje(), matricula, refeicao),
            fetch=True
        )
        if not rows:
            return None

        total_hoje = _executar_pg(
            "SELECT COUNT(*) AS total FROM refeitorio WHERE data = %s",
            (_hoje(),),
            fetch=True
        )
        total_refeicao = _executar_pg(
            "SELECT COUNT(*) AS total FROM refeitorio WHERE data = %s AND refeicao = %s",
            (_hoje(), refeicao),
            fetch=True
        )
        return {
            "hora": rows[0]["horaentrada"],
            "nome": rows[0]["nome"],
            "total_hoje": total_hoje[0].get("total", 0) if total_hoje else 0,
            "total_refeicao": total_refeicao[0].get("total", 0) if total_refeicao else 0,
        }
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade de refeitorio: {e}")
        return None


def _ler_frequencia_hoje_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, aula FROM frequencia WHERE data = %s",
        (_hoje(),),
        fetch=True
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["aula"]]
        for row in rows
    ]


def _inserir_frequencia_db(registro):
    _executar_pg(
        "INSERT INTO frequencia (data, horaentrada, matricula, nome, serie, curso, aula) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _apagar_frequencia_data_db(data_alvo):
    _executar_pg("DELETE FROM frequencia WHERE data = %s", (data_alvo,))


def _frequencia_duplicado_db(matricula):
    try:
        rows = _executar_pg(
            "SELECT horaentrada, nome FROM frequencia WHERE data = %s AND matricula = %s LIMIT 1",
            (_hoje(), matricula),
            fetch=True
        )
        if not rows:
            return None

        total_hoje = _executar_pg(
            "SELECT COUNT(*) AS total FROM frequencia WHERE data = %s",
            (_hoje(),),
            fetch=True
        )
        return {
            "hora": rows[0]["horaentrada"],
            "nome": rows[0]["nome"],
            "total_hoje": total_hoje[0].get("total", 0) if total_hoje else 0,
        }
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade de frequencia: {e}")
        return None


def _inserir_avaliacao_db(registro):
    _executar_pg(
        "INSERT INTO avaliacoes (data, aluno, serie, curso, estagio, item, nota) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _ler_avaliacoes_db():
    rows = _executar_pg(
        "SELECT data, aluno, serie, curso, estagio, item, nota FROM avaliacoes ORDER BY data DESC",
        (),
        fetch=True
    )
    if not rows:
        return []
    return [
        {
            "Data": row["data"],
            "Aluno": row["aluno"],
            "Serie": row["serie"],
            "Curso": row["curso"],
            "Estagio": row["estagio"],
            "Item": row["item"],
            "Nota": row["nota"],
        }
        for row in rows
    ]


def _ler_refeitorio_todos_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, refeicao FROM refeitorio ORDER BY data DESC, horaentrada DESC",
        (),
        fetch=True,
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["refeicao"]]
        for row in rows
    ]


def _ler_frequencia_todos_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, aula FROM frequencia ORDER BY data DESC, horaentrada DESC",
        (),
        fetch=True,
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["aula"]]
        for row in rows
    ]


def _avaliacoes_para_linhas():
    return [
        [r["Data"], r["Aluno"], r["Serie"], r["Curso"], r["Estagio"], r["Item"], r["Nota"]]
        for r in _ler_avaliacoes_db()
    ]


def _escrever_csv(caminho, header, linhas):
    with open(caminho, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(linhas)


def _csv_bytes(header, linhas):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    w.writerows(linhas)
    out = io.BytesIO(buf.getvalue().encode("utf-8"))
    out.seek(0)
    return out


def _nome_backup_mes_sugerido():
    hoje = datetime.date.today()
    return f"backup_{MESES_PT[hoje.month - 1]}_{hoje.year}"


def _exportar_backup_mensal_csv(nome_base):
    """Exporta todas as tabelas do banco para CSVs em dados/backups/YYYY_MM/."""
    nome_base = re.sub(r'[<>:"/\\|?*]', "_", (nome_base or "").strip())
    if not nome_base:
        raise ValueError("Nome do arquivo inválido")
    pasta = os.path.join(DADOS_DIR, "backups", datetime.date.today().strftime("%Y_%m"))
    os.makedirs(pasta, exist_ok=True)
    arquivos = []
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_avaliacoes_escola.csv"),
        ["Data", "Aluno", "Serie", "Curso", "Estagio", "Item", "Nota"],
        _avaliacoes_para_linhas(),
    )
    arquivos.append(f"{nome_base}_avaliacoes_escola.csv")
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_refeitorio.csv"),
        ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
        _ler_refeitorio_todos_db(),
    )
    arquivos.append(f"{nome_base}_refeitorio.csv")
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_frequencia.csv"),
        ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
        _ler_frequencia_todos_db(),
    )
    arquivos.append(f"{nome_base}_frequencia.csv")
    return pasta, arquivos


def _avaliacao_ja_existe_db(nome, semana_iso, ano_iso):
    if not nome:
        return False
    try:
        rows = _executar_pg(    
            "SELECT data FROM avaliacoes WHERE lower(aluno) = lower(%s)",
            (nome,),
            fetch=True
        )
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade: {e}")
        return False

    if not rows:
        return False
    for row in rows:
        try:
            data_str = row["data"].split(" ")[0]
            data_obj = datetime.datetime.strptime(data_str, "%d/%m/%Y").date()
            if data_obj.isocalendar()[1] == semana_iso and data_obj.isocalendar()[0] == ano_iso:
                return True
        except Exception:
            continue
    return False


def _apagar_avaliacoes_db():
    _executar_pg("DELETE FROM avaliacoes", ())


def _limpar_tabelas_neon():
    _executar_pg("DELETE FROM refeitorio", ())
    _executar_pg("DELETE FROM frequencia", ())
    _executar_pg("DELETE FROM avaliacoes", ())
    logger.info("Tabelas Neon refeitorio, frequencia e avaliacoes limpas")


CARDAPIO_PADRAO = {
    "SEGUNDA": ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "TERCA":   ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "QUARTA":  ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "QUINTA":  ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "SEXTA":   ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
}
EVENTOS_PADRAO = [
    {"data": "03/02", "evento": "Inicio das Aulas"},
    {"data": "01/05", "evento": "Feriado: Dia do Trabalhador"},
    {"data": "24/06", "evento": "Arraia do Marwin"},
    {"data": "07/09", "evento": "Feriado: Independencia do Brasil"},
    {"data": "15/10", "evento": "Dia do Professor"},
    {"data": "25/12", "evento": "Natal"},
]

# ==============================================================================
# ROTAS FLASK
# ==============================================================================
import secrets

def checar_senha(req):
    hdr = req.headers.get("X-Senha", "")
    if not hdr:
        return False
    # If ADMIN_PASSWORD looks like a bcrypt hash and bcrypt is available, verify accordingly
    if BCRYPT_AVAILABLE and isinstance(ADMIN_PASSWORD, str) and ADMIN_PASSWORD.startswith("$2"):
        try:
            return bcrypt.checkpw(hdr.encode("utf-8"), ADMIN_PASSWORD.encode("utf-8"))
        except Exception:
            return False
    # Fallback: constant-time compare with configured plaintext
    return secrets.compare_digest(hdr, ADMIN_PASSWORD)

# Rota para servir o cliente web (index.html)
@app.route('/')
def serve_cliente():
    pasta_raiz = os.path.dirname(os.path.abspath(__file__))
    for nome in ('index.html', 'Index.html'):
        arquivo_index = os.path.join(pasta_raiz, nome)
        if os.path.isfile(arquivo_index):
            return send_from_directory(pasta_raiz, nome)
    return jsonify({'erro': 'index.html não encontrado'}), 404

# Rota para servir arquivos estáticos (CSS, JS, imagens, etc) da pasta raiz
@app.route('/<path:path>')
def serve_static(path):
    try:
        pasta_raiz = os.path.dirname(os.path.abspath(__file__))
        arquivo = os.path.join(pasta_raiz, path)
        # Segurança: evitar acesso fora da pasta raiz
        caminho_absoluto = os.path.abspath(arquivo)
        if not caminho_absoluto.startswith(os.path.abspath(pasta_raiz)):
            return jsonify({'erro': 'Acesso negado'}), 403
        if os.path.isfile(caminho_absoluto):
            return send_from_directory(pasta_raiz, path)
        return jsonify({'erro': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route("/cardapio",  methods=["GET"])
def get_cardapio(): return jsonify(ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO))

@app.route("/eventos",   methods=["GET"])
def get_eventos():  return jsonify(ler_json(EVENTOS_FILE, EVENTOS_PADRAO))

@app.route("/config",    methods=["GET"])
def get_config():   return jsonify(ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"}))

@app.route("/avaliacoes", methods=["GET"])
def get_avaliacoes_publicas():
    """Retorna todos os registros de avaliações em formato JSON"""
    try:
        registros = _ler_avaliacoes_db()
        return jsonify({"avaliacoes": registros, "total": len(registros)})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao carregar avaliacoes: {e}")
        return jsonify({"erro": "Erro ao carregar avaliações"}), 500

@app.route("/avaliacao", methods=["POST"])
def post_avaliacao():
    dados = request.get_json()
    if not dados: return jsonify({"erro": "JSON invalido"}), 400
    nome, serie, curso = dados.get("nome","Anonimo"), dados.get("serie","N/A"), dados.get("curso","N/A")
    respostas = dados.get("respostas", {})
    
    # Verificar se já avaliou esta semana (apenas se não for anônimo)
    if nome and nome.strip().lower() not in {"anonimo", "anônimo"}:
        hoje = datetime.date.today()
        semana_iso = hoje.isocalendar()[1]
        ano_iso = hoje.isocalendar()[0]
        
        try:
            if _avaliacao_ja_existe_db(nome, semana_iso, ano_iso):
                logger.warning(f"Avaliação duplicada detectada: {nome} - Semana {semana_iso}/{ano_iso}")
                return jsonify({"status": "ja_avaliou", "mensagem": "Você já avaliou esta semana"}), 200
        except Exception as e:
            logger.error(f"Erro ao verificar duplicidade de avaliação: {e}")
            return jsonify({"erro": "Banco de dados indisponível"}), 503
    
    data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    registros = []
    for chave, nota in respostas.items():
        estagio, item = chave.split("|", 1)
        registros.append([data_hora, nome, serie, curso, estagio, item, nota])

    try:
        for registro in registros:
            _inserir_avaliacao_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao salvar avaliação: {e}")
        return jsonify({"erro": "Erro ao salvar avaliação"}), 500
    logger.info(f"Avaliação recebida de {nome} ({len(respostas)} itens)")
    return jsonify({"status": "ok"})

@app.route("/avaliacao/verificar", methods=["GET"])
def verificar_avaliacao():
    nome = request.args.get("nome", "").strip()
    if not nome or nome.lower() in {"anonimo", "anônimo"}:
        return jsonify({"ja_avaliou": False}), 200
    
    hoje = datetime.date.today()
    semana_iso = hoje.isocalendar()[1]
    ano_iso = hoje.isocalendar()[0]
    
    try:
        ja_avaliou = _avaliacao_ja_existe_db(nome, semana_iso, ano_iso)
        return jsonify({"ja_avaliou": ja_avaliou}), 200
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao verificar avaliação: {e}")
        return jsonify({"erro": "Erro ao verificar avaliação"}), 500

@app.route("/admin/avaliacoes", methods=["GET"])
def get_avaliacoes():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_avaliacoes_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao carregar avaliacoes: {e}")
        return jsonify({"erro": "Erro ao carregar avaliações"}), 500

@app.route("/admin/avaliacoes", methods=["DELETE"])
def del_avaliacoes():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        _apagar_avaliacoes_db()
        return jsonify({"status": "apagado"})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar avaliações: {e}")
        return jsonify({"erro": "Erro ao apagar avaliações"}), 500

@app.route("/admin/cardapio", methods=["PUT"])
def put_cardapio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    dados = request.get_json()
    salvar_json(CARDAPIO_FILE, dados)
    _sync_nuvem("/admin/cardapio", "PUT", dados)
    return jsonify({"status": "ok"})

@app.route("/admin/eventos", methods=["POST"])
def add_evento():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    ev = request.get_json()
    eventos = ler_json(EVENTOS_FILE, EVENTOS_PADRAO); eventos.append(ev)
    salvar_json(EVENTOS_FILE, eventos)
    _sync_nuvem("/admin/eventos", "PUT", eventos)
    return jsonify({"status": "ok"})

@app.route("/admin/eventos", methods=["DELETE"])
def del_evento():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    ev = request.get_json()
    eventos = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
    eventos = [e for e in eventos if not (e["data"]==ev["data"] and e["evento"]==ev["evento"])]
    salvar_json(EVENTOS_FILE, eventos)
    _sync_nuvem("/admin/eventos", "PUT", eventos)
    return jsonify({"status": "ok"})

@app.route("/admin/config", methods=["PUT"])
def put_config():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    dados = request.get_json()
    salvar_json(CONFIG_FILE, dados)
    _sync_nuvem("/admin/config", "PUT", dados)
    return jsonify({"status": "ok"})

# ==============================================================================
# ROTAS FLASK — REFEITÓRIO
# ==============================================================================
def _hoje():
    return datetime.date.today().strftime("%d/%m/%Y")

def _aula_por_hora(hora):
    """Retorna a aula ou intervalo correspondente ao horário de entrada."""
    try:
        hora_limpa = hora.strip().split(" ")[0]
        if len(hora_limpa) == 5:
            hora_limpa += ":00"
        t = datetime.datetime.strptime(hora_limpa, "%H:%M:%S").time()
    except Exception:
        return "Fora do horário"

    periodos = [
        ("07:10:00", "08:00:00", "Aula 1"),
        ("08:00:00", "08:50:00", "Aula 2"),
        ("08:50:00", "09:10:00", "Intervalo"),
        ("09:10:00", "10:00:00", "Aula 3"),
        ("10:00:00", "10:50:00", "Aula 4"),
        ("10:50:00", "11:40:00", "Aula 5"),
        ("11:40:00", "13:00:00", "Intervalo"),
        ("13:00:00", "13:50:00", "Aula 6"),
        ("13:50:00", "14:40:00", "Aula 7"),
        ("14:40:00", "15:00:00", "Intervalo"),
        ("15:00:00", "15:50:00", "Aula 8"),
        ("15:50:00", "16:40:00", "Aula 9"),
    ]
    for inicio, fim, nome_aula in periodos:
        if datetime.time.fromisoformat(inicio) <= t < datetime.time.fromisoformat(fim):
            return nome_aula
    return "Fora do horário"

def _registros_hoje():
    try:
        return _ler_refeitorio_hoje_db()
    except Exception as e:
        logger.error(f"Erro ao ler refeitorio do banco: {e}")
        return []

@app.route("/refeitorio/registrar", methods=["POST"])
def registrar_refeicao():
    dados = request.get_json()
    if not dados:
        return jsonify({"erro": "JSON invalido"}), 400
    matricula = dados.get("matricula", "").strip()
    refeicao  = dados.get("refeicao", "almoco").strip().lower()
    nome      = dados.get("nome", "Desconhecido").strip()
    serie     = dados.get("serie", "N/A").strip()
    curso     = dados.get("curso", "N/A").strip()
    
    # Limpa dados malformados do leitor USB (não use locals())
    nome = nome or "Desconhecido"
    serie = serie or "N/A"
    curso = curso or "N/A"
    matricula = matricula or ""

    def _limpar_campo(valor: str) -> str:
        if not valor:
            return valor
        v = valor
        if "^" in v or "Ç" in v:
            v = v.replace("^Ç", ":").replace("^", "").replace("Ç", "")
        v = re.sub(r"[^a-zA-Z0-9À-ÿ\s.'-]", "", v).strip()
        return v

    nome = _limpar_campo(nome)
    serie = _limpar_campo(serie)
    curso = _limpar_campo(curso)
    matricula = _limpar_campo(matricula)
    
    today      = _hoje()
    if not matricula:
        return jsonify({"erro": "Matricula nao informada"}), 400

    db_duplicado = _refeitorio_duplicado_db(matricula, refeicao)
    if db_duplicado:
        return jsonify({
            "status": "ja_registrado",
            "nome": db_duplicado["nome"],
            "hora": db_duplicado["hora"],
            "total_hoje": db_duplicado["total_hoje"],
            "total_refeicao": db_duplicado["total_refeicao"],
        }), 200
    
    hora = datetime.datetime.now().strftime("%H:%M:%S")
    periodo = _aula_por_hora(hora)
    registro = [today, hora, matricula, nome, serie, curso, refeicao]

    try:
        _inserir_refeitorio_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao registrar refeição: {e}")
        return jsonify({"erro": "Erro ao registrar refeição"}), 500
    logger.info(f"Refeição registrada: {nome} ({matricula}) - {refeicao}")
    # Contagens após inserir
    registros = _ler_refeitorio_hoje_db()
    total_hoje = len(registros)
    total_refeicao = sum(1 for r in registros if r[6] == refeicao)
    return jsonify({"status": "ok", "nome": nome, "hora": hora, "aula": periodo,
                    "total_hoje": total_hoje, "total_refeicao": total_refeicao}), 200

@app.route("/refeitorio/hoje", methods=["GET"])
def get_refeitorio_hoje():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_refeitorio_hoje_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao ler refeitorio: {e}")
        return jsonify({"erro": "Erro ao ler dados de refeitório"}), 500

@app.route("/refeitorio/exportar", methods=["GET"])
def exportar_refeitorio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        linhas = _ler_refeitorio_todos_db()
        buf = _csv_bytes(
            ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
            linhas,
        )
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"refeitorio_{_hoje().replace('/', '_')}.csv",
            mimetype="text/csv",
        )
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao exportar refeitorio: {e}")
        return jsonify({"erro": "Erro ao exportar refeitório"}), 500

@app.route("/refeitorio/apagar", methods=["DELETE"])
def apagar_refeitorio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    data_alvo = request.args.get("data", _hoje())
    try:
        _apagar_refeitorio_data_db(data_alvo)
        return jsonify({"status": "apagado", "data": data_alvo})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar refeitorio: {e}")
        return jsonify({"erro": "Erro ao apagar refeitório"}), 500

@app.route("/refeitorio/qrcode/<matricula>", methods=["GET"])
def gerar_qrcode_img(matricula):
    nome  = request.args.get("nome", "")
    serie = request.args.get("serie", "")
    curso = request.args.get("curso", "")
    payload = json.dumps({"matricula": matricula, "nome": nome, "serie": serie, "curso": curso}, ensure_ascii=False)
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
    qr.add_data(payload); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
    return send_file(buf, mimetype="image/png", download_name=f"qrcode_{matricula}.png")

# ==============================================================================
# ROTAS FLASK — FREQUÊNCIA
# ==============================================================================

def _registros_freq_hoje():
    try:
        return _ler_frequencia_hoje_db()
    except Exception as e:
        logger.error(f"Erro ao ler frequencia do banco: {e}")
        return []

@app.route("/frequencia/registrar", methods=["POST"])
def registrar_frequencia():
    dados = request.get_json()
    if not dados:
        return jsonify({"erro": "JSON invalido"}), 400
    matricula = dados.get("matricula", "").strip()
    nome      = dados.get("nome", "Desconhecido").strip()
    serie     = dados.get("serie", "N/A").strip()
    curso     = dados.get("curso", "N/A").strip()
    
    # Limpa dados malformados do leitor USB
    def _limpar_campo(valor: str) -> str:
        if not valor:
            return valor
        v = valor
        if "^" in v or "Ç" in v:
            v = v.replace("^Ç", ":").replace("^", "").replace("Ç", "")
        v = re.sub(r"[^a-zA-Z0-9À-ÿ\s.'-]", "", v).strip()
        return v

    nome = _limpar_campo(nome)
    serie = _limpar_campo(serie)
    curso = _limpar_campo(curso)
    matricula = _limpar_campo(matricula)
    
    nome = nome or "Desconhecido"
    serie = serie or "N/A"
    curso = curso or "N/A"
    
    hoje = _hoje()
    if not matricula:
        return jsonify({"erro": "Matricula nao informada"}), 400

    db_duplicado = _frequencia_duplicado_db(matricula)
    if db_duplicado:
        return jsonify({"status": "ja_registrado", "nome": db_duplicado["nome"], "hora": db_duplicado["hora"], "total_hoje": db_duplicado["total_hoje"]}), 200
    
    hora = datetime.datetime.now().strftime("%H:%M:%S")
    aula = _aula_por_hora(hora)
    registro = [hoje, hora, matricula, nome, serie, curso, aula]

    try:
        _inserir_frequencia_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao registrar frequência: {e}")
        return jsonify({"erro": "Erro ao registrar frequência"}), 500

    logger.info(f"Frequência registrada: {nome} ({matricula})")
    
    registros = _ler_frequencia_hoje_db()
    total_hoje = len(registros)
    return jsonify({"status": "ok", "nome": nome, "hora": hora, "aula": aula, "total_hoje": total_hoje}), 200

@app.route("/frequencia/hoje", methods=["GET"])
def get_frequencia_hoje():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_frequencia_hoje_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao ler frequencia: {e}")
        return jsonify({"erro": "Erro ao ler dados de frequência"}), 500

@app.route("/frequencia/exportar", methods=["GET"])
def exportar_frequencia():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        linhas = _ler_frequencia_todos_db()
        buf = _csv_bytes(
            ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
            linhas,
        )
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"frequencia_{_hoje().replace('/', '_')}.csv",
            mimetype="text/csv",
        )
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao exportar frequencia: {e}")
        return jsonify({"erro": "Erro ao exportar frequência"}), 500

@app.route("/frequencia/apagar", methods=["DELETE"])
def apagar_frequencia():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    data_alvo = request.args.get("data", _hoje())
    try:
        _apagar_frequencia_data_db(data_alvo)
        return jsonify({"status": "apagado", "data": data_alvo})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar frequencia: {e}")
        return jsonify({"erro": "Erro ao apagar frequência"}), 500

@app.route("/admin/limpar_neon", methods=["DELETE"])
def admin_limpar_neon():
    if not checar_senha(request):
        return jsonify({"erro": "Acesso negado"}), 403
    try:
        _limpar_tabelas_neon()
        return jsonify({"status": "neon_limpo"}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

# ==============================================================================
# BACKUP POR E-MAIL (opcional)
# ==============================================================================
def _enviar_backup_email():
    """Envia backup compactado por e-mail."""
    try:
        cfg = ler_json(CONFIG_FILE, {})
        if not cfg.get("email_backup_ativo", False):
            return
        
        email_remetente = cfg.get("email_remetente", "")
        email_senha = cfg.get("email_senha_app", "")
        email_destino = cfg.get("email_destino", "")
        email_smtp = cfg.get("email_smtp", "smtp.gmail.com")
        email_porta = cfg.get("email_porta", 587)
        
        if not all([email_remetente, email_senha, email_destino]):
            logger.warning("Backup por e-mail desativado: credenciais incompletas")
            return
        
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_name = f"backup_marwin_{ts}.zip"
        pasta_exp, arquivos_csv = _exportar_backup_mensal_csv(f"email_{ts}")
        with zipfile.ZipFile(zip_name, "w") as zf:
            for nome_arq in arquivos_csv:
                zf.write(os.path.join(pasta_exp, nome_arq), arcname=nome_arq)
            if os.path.exists(LISTA_ALUNOS_FILE):
                zf.write(LISTA_ALUNOS_FILE, arcname="lista_alunos.json")
        
        # Enviar por SMTP
        msg = MIMEMultipart()
        msg["From"] = email_remetente
        msg["To"] = email_destino
        msg["Subject"] = f"Backup MARWIN - {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        body = f"Backup automático do sistema MARWIN - {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        msg.attach(MIMEText(body, "plain", "utf-8"))
        
        with open(zip_name, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename= {zip_name}")
            msg.attach(part)
        
        server = smtplib.SMTP(email_smtp, email_porta)
        server.starttls()
        server.login(email_remetente, email_senha)
        server.send_message(msg)
        server.quit()
        
        os.remove(zip_name)
        logger.info(f"Backup enviado por e-mail para {email_destino}")
    except Exception as e:
        logger.error(f"Erro ao enviar backup por e-mail: {e}")

def _agendar_backup_email():
    """Agenda backup automático para 23h30 todos os dias."""
    def _loop_backup():
        while True:
            agora = datetime.datetime.now()
            proxima_exec = agora.replace(hour=23, minute=30, second=0, microsecond=0)
            if agora >= proxima_exec:
                proxima_exec += datetime.timedelta(days=1)
            tempo_espera = (proxima_exec - agora).total_seconds()
            time.sleep(tempo_espera)
            _enviar_backup_email()
    
    thread = threading.Thread(target=_loop_backup, daemon=True)
    thread.start()

def _iniciar_fluxo_backup_mes(btn_backup=None):
    """Fluxo guiado: nome do arquivo → exportar CSV → confirmação dupla → limpar banco."""
    def _habilitar_btn(estado):
        if btn_backup is not None:
            try:
                btn_backup.config(state=estado)
            except Exception:
                pass

    _habilitar_btn("disabled")

    dlg_nome = tk.Toplevel(janela)
    dlg_nome.title("Backup do Mês")
    dlg_nome.configure(bg=T["BG_CARD"])
    dlg_nome.transient(janela)
    dlg_nome.grab_set()
    dlg_nome.resizable(False, False)

    tk.Label(dlg_nome, text="Nome do arquivo de backup",
             font=("Segoe UI", 12, "bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(padx=20, pady=(16, 6))
    tk.Label(dlg_nome, text="Os CSVs serão salvos em dados/backups/AAAA_MM/",
             font=("Segoe UI", 9), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(padx=20, pady=(0, 10))

    nome_var = tk.StringVar(value=_nome_backup_mes_sugerido())
    entry_nome = tk.Entry(dlg_nome, textvariable=nome_var, font=("Segoe UI", 11),
                          bg=T["ENTRY_BG"], fg=T["FG_TEXT"], width=36)
    entry_nome.pack(padx=20, pady=(0, 16))
    entry_nome.focus_set()
    entry_nome.select_range(0, "end")

    resultado = {"cancelado": True, "nome": ""}

    def _cancelar_inicio():
        resultado["cancelado"] = True
        dlg_nome.destroy()
        _habilitar_btn("normal")

    def _confirmar_nome():
        nome = nome_var.get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "Informe um nome para o backup.", parent=dlg_nome)
            return
        resultado["cancelado"] = False
        resultado["nome"] = nome
        dlg_nome.destroy()

    btn_row = tk.Frame(dlg_nome, bg=T["BG_CARD"])
    btn_row.pack(pady=(0, 16))
    tk.Button(btn_row, text="Cancelar", command=_cancelar_inicio,
              bg=T["FRASE_FG"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)
    tk.Button(btn_row, text="Continuar", command=_confirmar_nome,
              bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)

    dlg_nome.wait_window()
    if resultado["cancelado"]:
        return

    try:
        pasta, arquivos = _exportar_backup_mensal_csv(resultado["nome"])
        lista = "\n".join(f"• {a}" for a in arquivos)
        messagebox.showinfo(
            "Backup exportado",
            f"CSVs gerados com sucesso em:\n{pasta}\n\n{lista}",
        )
        logger.info(f"Backup do mês exportado: {pasta}")
    except Exception as e:
        logger.error(f"Falha ao exportar backup do mês: {e}")
        messagebox.showerror("Erro", f"Falha ao exportar backup:\n{e}")
        _habilitar_btn("normal")
        return

    if not messagebox.askyesno(
        "Confirmar exclusão",
        "Tem certeza que deseja apagar todos os dados do banco?\n"
        "Esta ação não pode ser desfeita.",
        icon="warning",
    ):
        messagebox.showinfo("Cancelado", "Nenhum dado foi removido do banco. O backup CSV foi mantido.")
        _habilitar_btn("normal")
        return

    dlg_final = tk.Toplevel(janela)
    dlg_final.title("Confirmação final")
    dlg_final.configure(bg=T["BG_CARD"])
    dlg_final.transient(janela)
    dlg_final.grab_set()
    dlg_final.resizable(False, False)

    tk.Label(
        dlg_final,
        text="Confirmação final: os dados serão permanentemente removidos do banco.\n"
             "Digite CONFIRMAR para prosseguir.",
        font=("Segoe UI", 10),
        bg=T["BG_CARD"],
        fg=T["FG_TEXT"],
        justify="left",
    ).pack(padx=20, pady=(16, 10))

    conf_var = tk.StringVar()
    entry_conf = tk.Entry(dlg_final, textvariable=conf_var, font=("Segoe UI", 11),
                          bg=T["ENTRY_BG"], fg=T["FG_TEXT"], width=28)
    entry_conf.pack(padx=20, pady=(0, 12))

    btn_apagar = tk.Button(
        dlg_final, text="Apagar dados do banco", state="disabled",
        bg="#c62828", fg="white", font=("Segoe UI", 10, "bold"),
        padx=14, pady=6, bd=0,
    )
    confirmado = {"ok": False}

    def _atualizar_btn(*_):
        if conf_var.get() == "CONFIRMAR":
            btn_apagar.config(state="normal", cursor="hand2")
        else:
            btn_apagar.config(state="disabled", cursor="")

    def _executar_apagar():
        confirmado["ok"] = True
        dlg_final.destroy()

    def _cancelar_final():
        confirmado["ok"] = False
        dlg_final.destroy()

    conf_var.trace_add("write", _atualizar_btn)
    btn_apagar.config(command=_executar_apagar)

    btn_row2 = tk.Frame(dlg_final, bg=T["BG_CARD"])
    btn_row2.pack(pady=(0, 16))
    tk.Button(btn_row2, text="Cancelar", command=_cancelar_final,
              bg=T["FRASE_FG"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)
    btn_apagar.pack(in_=btn_row2, side="left", padx=6)

    dlg_final.wait_window()

    if not confirmado["ok"]:
        messagebox.showinfo("Cancelado", "Nenhum dado foi removido do banco. O backup CSV foi mantido.")
        _habilitar_btn("normal")
        return

    try:
        _limpar_tabelas_neon()
        messagebox.showinfo("Concluído", "Dados do banco removidos com sucesso. O backup CSV foi mantido.")
        logger.info("Backup do mês concluído: banco limpo após exportação")
    except Exception as e:
        logger.error(f"Falha ao limpar banco após backup: {e}")
        messagebox.showerror(
            "Erro",
            f"O backup CSV foi salvo, mas falhou ao apagar o banco:\n{e}",
        )
    finally:
        _habilitar_btn("normal")

# ==============================================================================
# PAINEL ADMIN TKINTER
# ==============================================================================
TEMAS = {
    "claro": {
        "BG_MAIN":"#f2f8f2","BG_CARD":"#ffffff","ACCENT_VIBRANT":"#1b5e20",
        "ACCENT_SOFT":"#388e3c","FG_TEXT":"#1a2a1a","BORDER_GRID":"#c8e6c9",
        "HIGHLIGHT_YELLOW":"#f9a825","OBS_BG":"#e8f5e9","BTN_CARDAPIO":"#2e7d32",
        "FRASE_FG":"#4a7a4a","ENTRY_BG":"#f0f8f0","BTN_VOLTAR":"#1b5e20"
    },
    "escuro": {
        "BG_MAIN":"#0a1a0a","BG_CARD":"#142014","ACCENT_VIBRANT":"#4caf50",
        "ACCENT_SOFT":"#388e3c","FG_TEXT":"#c8e6c9","BORDER_GRID":"#2e5e2e",
        "HIGHLIGHT_YELLOW":"#f9a825","OBS_BG":"#1a3a1a","BTN_CARDAPIO":"#2e7d32",
        "FRASE_FG":"#81c784","ENTRY_BG":"#1a2e1a","BTN_VOLTAR":"#1b5e20"
    }
}

janela = None
T = None
tema_atual = None
url_ngrok_global = ""
btn_tema = None

def carregar_tema():
    if os.path.exists(TEMA_FILE):
        with open(TEMA_FILE) as f: return json.load(f).get("tema","claro")
    return "claro"

def salvar_tema_tk(tema):
    with open(TEMA_FILE,"w") as f: json.dump({"tema":tema},f)

def configurar_janela(win, titulo):
    win.title(titulo); win.state("zoomed"); win.configure(bg=T["BG_MAIN"])
    win.bind("<F11>", lambda e: win.attributes("-fullscreen", not win.attributes("-fullscreen")))
    win.bind("<Escape>", lambda e: win.destroy())

def _configurar_layout_portrait(corpo, painel_lateral, painel_principal, lado="left", largura_limite=980):
    def reajustar(event=None):
        corpo.update_idletasks()
        if corpo.winfo_width() <= largura_limite:
            painel_lateral.pack_forget(); painel_principal.pack_forget()
            painel_lateral.pack(fill="x", padx=0, pady=(0, 12))
            painel_principal.pack(fill="both", expand=True)
        else:
            painel_lateral.pack_forget(); painel_principal.pack_forget()
            if lado == "right":
                painel_principal.pack(side="left", fill="both", expand=True)
                painel_lateral.pack(side="right", fill="y")
            else:
                painel_lateral.pack(side="left", fill="y")
                painel_principal.pack(side="left", fill="both", expand=True)
    corpo.bind("<Configure>", reajustar)
    reajustar()

def criar_scrollable_frame(parent):
    canvas = tk.Canvas(parent, bg=T["BG_MAIN"], highlightthickness=0)
    sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
    sf = tk.Frame(canvas, bg=T["BG_MAIN"])
    wid = canvas.create_window((0,0), window=sf, anchor="nw")
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width))
    sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.configure(yscrollcommand=sb.set)
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
    canvas.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
    return sf

def criar_barra_topo(win, titulo, cmd_voltar=None, subtitulo=None):
    """Barra de topo institucional verde + linha amarela — padrão EEEP Marwin."""
    barra = tk.Frame(win, bg=T["ACCENT_VIBRANT"], height=56)
    barra.pack(fill="x"); barra.pack_propagate(False)
    if cmd_voltar:
        tk.Button(barra, text="← Voltar", font=("Segoe UI", 11, "bold"),
                  bg=T["ACCENT_VIBRANT"], fg="white", bd=0, padx=14, cursor="hand2",
                  activebackground=T["ACCENT_SOFT"],
                  command=cmd_voltar).pack(side="left", padx=16, pady=14)
        tk.Frame(barra, bg="#ffffff", width=1).pack(side="left", fill="y", pady=12)
    tk.Label(barra, text=titulo, font=("Segoe UI", 14, "bold"),
             bg=T["ACCENT_VIBRANT"], fg="white").pack(side="left", padx=16, pady=16)
    if subtitulo:
        tk.Label(barra, text=subtitulo, font=("Segoe UI", 10),
                 bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="right", padx=20)
    else:
        tk.Label(barra, text="EEEP MARWIN  —  Servidor Admin", font=("Segoe UI", 10),
                 bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="right", padx=20)
    tk.Frame(win, bg=T["HIGHLIGHT_YELLOW"], height=4).pack(fill="x")
    return barra

def pedir_senha_tk():
    win = tk.Toplevel(janela); win.title("Acesso Administrativo"); win.resizable(False, False)
    win.configure(bg=T["BG_MAIN"]); win.transient(janela); win.grab_set()
    win.update_idletasks()
    w, h = 420, 300
    sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # Barra verde no topo do diálogo
    barra = tk.Frame(win, bg=T["ACCENT_VIBRANT"], height=44)
    barra.pack(fill="x"); barra.pack_propagate(False)
    tk.Label(barra, text="🔒  Acesso Administrativo", font=("Segoe UI", 12, "bold"),
             bg=T["ACCENT_VIBRANT"], fg="white").pack(side="left", padx=16, pady=12)
    tk.Frame(win, bg=T["HIGHLIGHT_YELLOW"], height=3).pack(fill="x")

    frame = tk.Frame(win, bg=T["BG_CARD"], padx=36, pady=28)
    frame.pack(expand=True, fill="both", padx=24, pady=20)
    tk.Label(frame, text="Digite a senha para continuar",
             font=("Segoe UI", 11), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(pady=(0, 16))
    senha_var = tk.StringVar()
    entry = tk.Entry(frame, textvariable=senha_var, show="*", font=("Segoe UI", 13),
                     bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1,
                     insertbackground=T["ACCENT_VIBRANT"])
    entry.pack(fill="x", ipady=8, pady=(0, 20)); entry.focus()
    resultado = {"senha": None}
    def confirmar(): resultado["senha"] = senha_var.get(); win.destroy()
    def cancelar(): win.destroy()
    bf = tk.Frame(frame, bg=T["BG_CARD"]); bf.pack(fill="x")
    tk.Button(bf, text="Cancelar", bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"],
              font=("Segoe UI", 11, "bold"), bd=0, padx=15, pady=10,
              cursor="hand2", command=cancelar).pack(side="left", expand=True, fill="x", padx=(0, 6))
    tk.Button(bf, text="Entrar", bg=T["ACCENT_VIBRANT"], fg="white",
              font=("Segoe UI", 11, "bold"), bd=0, padx=15, pady=10,
              cursor="hand2", activebackground=T["ACCENT_SOFT"],
              command=confirmar).pack(side="right", expand=True, fill="x", padx=(6, 0))
    win.bind("<Return>", lambda e: confirmar()); win.wait_window(); return resultado["senha"]

def abrir_painel_admin(event=None):
    senha = pedir_senha_tk()
    def _local_check(pw: str) -> bool:
        if not pw:
            return False
        if BCRYPT_AVAILABLE and isinstance(ADMIN_PASSWORD, str) and ADMIN_PASSWORD.startswith("$2"):
            try:
                return bcrypt.checkpw(pw.encode("utf-8"), ADMIN_PASSWORD.encode("utf-8"))
            except Exception:
                return False
        return secrets.compare_digest(pw, ADMIN_PASSWORD)

    if not _local_check(senha):
        if senha is not None:
            logger.warning(f"Tentativa de acesso admin com senha incorreta")
            messagebox.showerror("Acesso negado","Senha incorreta!")
        return
        return

    logger.info("Painel admin aberto com sucesso")
    
    jd = tk.Toplevel(janela); configurar_janela(jd,"Painel Administrativo — EEEP MARWIN")
    jd.attributes("-topmost", True)
    jd.lift()
    jd.focus_force()
    jd.after(100, lambda: jd.attributes("-topmost", False))
    criar_barra_topo(jd, "⚙️  Painel Administrativo",
                     subtitulo=f"Servidor: {url_ngrok_global}" if url_ngrok_global else None)

    nb = ttk.Notebook(jd); nb.pack(fill="both", expand=True, padx=20, pady=15)
    style = ttk.Style(); style.theme_use("default")
    style.configure("TNotebook", background=T["BG_MAIN"], borderwidth=0)
    style.configure("TNotebook.Tab", background=T["OBS_BG"], foreground=T["ACCENT_VIBRANT"],
                    font=("Segoe UI",11,"bold"), padding=[20,10])
    style.map("TNotebook.Tab", background=[("selected",T["BG_CARD"])], foreground=[("selected",T["ACCENT_VIBRANT"])])
    style.configure("Treeview", background=T["BG_CARD"], foreground=T["FG_TEXT"], rowheight=38,
                    fieldbackground=T["BG_CARD"], font=("Segoe UI",11))
    style.map("Treeview", background=[("selected",T["ACCENT_SOFT"])], foreground=[("selected","white")])
    style.configure("Treeview.Heading", background=T["ACCENT_VIBRANT"], foreground="white",
                    font=("Segoe UI",11,"bold"), relief="flat", padding=[0,8])

    aba_dados   = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_dados,   text="  Avaliacoes  ")
    aba_rel     = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_rel,     text="  Relatorio Semanal  ")
    aba_card_ed = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_card_ed, text="  Editar Cardapio  ")
    aba_ev_ed   = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_ev_ed,   text="  Editar Eventos  ")
    aba_ref     = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_ref,     text="  Refeitorio  ")
    aba_freq    = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_freq,    text="  Frequencia  ")
    aba_hist    = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_hist,    text="  Historico  ")
    aba_qr      = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_qr,      text="  QR Codes  ")
    aba_logs    = tk.Frame(nb, bg=T["BG_CARD"]); nb.add(aba_logs,    text="  Logs  ")

    # ── ABA AVALIACOES ────────────────────────────────────────────────────────
    f_nome,f_data,f_item,f_est = tk.StringVar(),tk.StringVar(),tk.StringVar(),tk.StringVar()

    def _eh_almoco_favorito(item: str):
        return "almoco favorito" in item.lower() or "almoço favorito" in item.lower()

    def carregar():
        for i in tabela.get_children(): tabela.delete(i)
        try:
            linhas = _avaliacoes_para_linhas()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar avaliações do banco:\n{e}")
            lbl_cnt.config(text="0 registro(s)")
            return
        for idx, r in enumerate(linhas):
            if len(r) < 7:
                continue
            ok_nome = f_nome.get().lower() in r[1].lower() if f_nome.get() else True
            ok_data = f_data.get() in r[0] if f_data.get() else True
            mapa_e  = {"Comida":"1","Limpeza":"2","Ensino":"3","Semana":"4"}
            ok_est  = (r[4]==mapa_e.get(f_est.get())) if f_est.get() else True
            ok_item = f_item.get().lower() in r[5].lower() if f_item.get() else True
            if ok_nome and ok_data and ok_est and ok_item:
                categoria = "Almoço favorito" if _eh_almoco_favorito(r[5]) else "Avaliação"
                exib = [r[0], r[1], r[2], r[3], r[4], categoria, r[5], r[6]]
                if r[4]=="4" and not _eh_almoco_favorito(r[5]):
                    try:
                        n=float(r[6])
                        exib[7]="Ruim" if n<=1 else ("Medio" if n<=3 else "Bom")
                    except: pass
                tabela.insert("","end",values=exib,tags=("even" if idx%2==0 else "odd",))
        lbl_cnt.config(text=f"{len(tabela.get_children())} registro(s)")

    def apagar():
        if messagebox.askyesno("Aviso","Apagar todos os dados?"):
            try:
                _apagar_avaliacoes_db()
                carregar()
                messagebox.showinfo("Sucesso", "Avaliações apagadas do banco.")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao apagar avaliações:\n{e}")

    def exportar_avaliacoes_pdf():
        try:
            reader = [["Data","Aluno","Serie","Curso","Estagio","Item","Nota"]] + _avaliacoes_para_linhas()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao ler avaliações do banco:\n{e}")
            return
        if len(reader) <= 1:
            messagebox.showwarning("Aviso", "Não há avaliações para exportar.")
            return
        try:
            from fpdf import FPDF
            pdf = FPDF()
            pdf.set_auto_page_break(True, margin=15)
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "AVALIACOES ESCOLARES - EEEP MARWIN", ln=True, align="C")
            pdf.set_font("Arial", "", 12)
            pdf.cell(0, 8, f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
            pdf.ln(6)
            pdf.set_font("Arial", "B", 11)
            pdf.cell(0, 8, "Registros de avaliacao:", ln=True)
            pdf.ln(3)
            pdf.set_font("Arial", "", 10)
            for row in reader[1:]:
                if len(row) < 7: continue
                pdf.multi_cell(0, 6, f"Data: {row[0]} | Aluno: {row[1]} | Serie: {row[2]} | Curso: {row[3]}")
                pdf.multi_cell(0, 6, f"Estagio: {row[4]} | Item: {row[5]} | Nota: {row[6]}")
                pdf.ln(2)
            nome_arq = f"avaliacoes_marwin_{datetime.datetime.now().strftime('%d_%m_%Y')}.pdf"
            pdf.output(nome_arq)
            messagebox.showinfo("Sucesso", f"PDF gerado com sucesso:\n{nome_arq}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", f"Falha ao gerar PDF:\n{e}")

    fcard=tk.Frame(aba_dados,bg=T["BG_CARD"],padx=20,pady=15); fcard.pack(fill="x",padx=20,pady=(20,10))
    tk.Label(fcard,text="FILTROS",font=("Segoe UI",13,"bold"),bg=T["BG_CARD"],fg=T["ACCENT_VIBRANT"]).pack(anchor="w",pady=(0,10))
    tk.Frame(fcard,bg=T["BORDER_GRID"],height=1).pack(fill="x",pady=(0,12))
    
    # LINHA 1 — Filtros de busca (grid)
    linha1=tk.Frame(fcard,bg=T["BG_CARD"]); linha1.pack(fill="x",pady=4)
    linha1.grid_columnconfigure(0, weight=1)
    linha1.grid_columnconfigure(1, weight=1)
    linha1.grid_columnconfigure(2, weight=1)
    linha1.grid_columnconfigure(3, weight=0)
    linha1.grid_columnconfigure(4, weight=0)

    fr_nome=tk.Frame(linha1,bg=T["BG_CARD"]); fr_nome.grid(row=0,column=0,sticky="ew",padx=(0,20))
    tk.Label(fr_nome,text="Nome",font=("Segoe UI",9),bg=T["BG_CARD"],fg=T["FRASE_FG"]).pack(anchor="w")
    tk.Entry(fr_nome,textvariable=f_nome,width=25,font=("Segoe UI",11),bg=T["ENTRY_BG"],fg=T["FG_TEXT"],relief="solid",bd=1).pack(ipady=4)

    fr_data=tk.Frame(linha1,bg=T["BG_CARD"]); fr_data.grid(row=0,column=1,sticky="ew",padx=(0,20))
    tk.Label(fr_data,text="Data",font=("Segoe UI",9),bg=T["BG_CARD"],fg=T["FRASE_FG"]).pack(anchor="w")
    tk.Entry(fr_data,textvariable=f_data,width=12,font=("Segoe UI",11),bg=T["ENTRY_BG"],fg=T["FG_TEXT"],relief="solid",bd=1).pack(ipady=4)

    fr3=tk.Frame(linha1,bg=T["BG_CARD"]); fr3.grid(row=0,column=2,sticky="ew",padx=(0,20))
    tk.Label(fr3,text="Estagio",font=("Segoe UI",9),bg=T["BG_CARD"],fg=T["FRASE_FG"]).pack(anchor="w")
    ttk.Combobox(fr3,textvariable=f_est,values=["","Comida","Limpeza","Ensino","Semana"],
                 state="readonly",width=12,font=("Segoe UI",11)).pack(ipady=2)

    btn_buscar=tk.Button(linha1,text="Buscar",command=carregar,bg=T["ACCENT_SOFT"],fg="white",
              font=("Segoe UI",10,"bold"),padx=18,pady=8,bd=0,cursor="hand2")
    btn_buscar.grid(row=0,column=3,padx=15)
    btn_export=tk.Button(linha1,text="Exportar PDF",command=exportar_avaliacoes_pdf,bg=T["ACCENT_VIBRANT"],fg="white",
              font=("Segoe UI",10,"bold"),padx=18,pady=8,bd=0,cursor="hand2")
    btn_export.grid(row=0,column=4,padx=5)

    cfg_sys = ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"})
    status_var = tk.BooleanVar(value=cfg_sys.get("avaliacoes_ativas", True))
    modo_var = tk.StringVar(value=cfg_sys.get("modo_leitura", "camera"))

    def salvar_config():
        cfg_sys["avaliacoes_ativas"] = status_var.get()
        cfg_sys["modo_leitura"] = modo_var.get()
        salvar_json(CONFIG_FILE, cfg_sys)
        _sync_nuvem("/admin/config", "PUT", cfg_sys)

    def alternar_status():
        salvar_config()

    def alterar_modo():
        salvar_config()

    # LINHA 2 — Controles de configuração (pack)
    linha2=tk.Frame(fcard,bg=T["BG_CARD"]); linha2.pack(fill="x",pady=4)
    
    chk_avaliacoes = tk.Checkbutton(linha2, text="Avaliacoes Ativas", variable=status_var, command=alternar_status,
                   font=("Segoe UI",10,"bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"],
                   activebackground=T["BG_CARD"], selectcolor=T["ENTRY_BG"], cursor="hand2")
    chk_avaliacoes.pack(side="left", padx=(0, 16))
    
    sep1 = tk.Frame(linha2, bg=T["BORDER_GRID"], width=1)
    sep1.pack(side="left", fill="y", padx=(0, 16))
    
    fr4 = tk.Frame(linha2, bg=T["BG_CARD"])
    fr4.pack(side="left", padx=(0, 16))
    tk.Label(fr4, text="Modo de leitura:", font=("Segoe UI",9), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(anchor="w")
    rb_cam = tk.Radiobutton(fr4, text="Webcam", variable=modo_var, value="camera",
                             bg=T["BG_CARD"], fg=T["FG_TEXT"], selectcolor=T["ACCENT_SOFT"],
                             activebackground=T["BG_CARD"], font=("Segoe UI",10), command=alterar_modo)
    rb_cam.pack(side="left", padx=(0, 8))
    rb_usb = tk.Radiobutton(fr4, text="Leitor USB", variable=modo_var, value="usb",
                             bg=T["BG_CARD"], fg=T["FG_TEXT"], selectcolor=T["ACCENT_SOFT"],
                             activebackground=T["BG_CARD"], font=("Segoe UI",10), command=alterar_modo)
    rb_usb.pack(side="left")
    
    sep2 = tk.Frame(linha2, bg=T["BORDER_GRID"], width=1)
    sep2.pack(side="left", fill="y", padx=(0, 16))
    
    btn_apagar = tk.Button(linha2, text="Apagar tudo", command=apagar, bg="#c62828", fg="white",
              font=("Segoe UI",10,"bold"), padx=18, pady=8, bd=0, cursor="hand2")
    btn_apagar.pack(side="right")

    def _ajustar_layout_filtros(event=None):
        largura = fcard.winfo_width()
        if largura < 780:
            for widget in linha1.winfo_children():
                widget.grid_forget()
            row = 0
            for widget in (fr_nome, fr_data, fr3, btn_buscar, btn_export):
                widget.grid(row=row, column=0, sticky="ew", padx=0, pady=(0,6))
                row += 1
            linha1.grid_columnconfigure(0, weight=1)
        else:
            fr_nome.grid(row=0,column=0,sticky="ew",padx=(0,20),pady=0)
            fr_data.grid(row=0,column=1,sticky="ew",padx=(0,20),pady=0)
            fr3.grid(row=0,column=2,sticky="ew",padx=(0,20),pady=0)
            btn_buscar.grid(row=0,column=3,padx=15,pady=0)
            btn_export.grid(row=0,column=4,padx=5,pady=0)
            for col in range(5):
                linha1.grid_columnconfigure(col, weight=1 if col in (0,1,2) else 0)
    linha1.bind("<Configure>", _ajustar_layout_filtros)
    _ajustar_layout_filtros()

    lbl_cnt=tk.Label(aba_dados,text="0 registro(s)",font=("Segoe UI",10),bg=T["BG_CARD"],fg=T["FRASE_FG"])
    lbl_cnt.pack(anchor="w",padx=25,pady=(5,0))
    ft=tk.Frame(aba_dados,bg=T["BORDER_GRID"],padx=1,pady=1); ft.pack(expand=True,fill="both",padx=20,pady=(5,20))
    tabela=ttk.Treeview(ft,columns=("D","A","S","Cr","E","C","I","N"),show="headings")
    for col,txt,w in zip(("D","A","S","Cr","E","C","I","N"),("DATA","ALUNO","SERIE","CURSO","ESTAGIO","CATEGORIA","ITEM","RESPOSTA"),(120,140,80,140,85,130,220,95)):
        tabela.heading(col,text=txt); tabela.column(col,width=w,anchor="center")
    sbh=ttk.Scrollbar(ft, orient="horizontal", command=tabela.xview)
    tabela.configure(xscrollcommand=sbh.set)
    tabela.pack(expand=True,fill="both")
    sbh.pack(fill="x")
    tabela.tag_configure("even",background=T["ENTRY_BG"]); tabela.tag_configure("odd",background=T["BG_CARD"])
    carregar()
    
    # ── Backup do Mês ───────────────────────────────────────────────────────
    cfg_backup_frame = tk.Frame(aba_dados, bg=T["OBS_BG"])
    cfg_backup_frame.pack(fill="x", padx=20, pady=(15, 10))

    tk.Label(cfg_backup_frame, text="BACKUP DO MÊS",
            font=("Segoe UI", 11, "bold"), bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=10, pady=(8, 0))
    tk.Frame(cfg_backup_frame, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=10, pady=(4, 8))

    tk.Label(cfg_backup_frame,
            text="Exporta todos os dados do banco para CSV e, após confirmação em duas etapas, limpa as tabelas.",
            font=("Segoe UI", 9), bg=T["OBS_BG"], fg=T["FG_TEXT"], wraplength=760, justify="left").pack(anchor="w", padx=10, pady=(0, 10))

    btn_backup_frame = tk.Frame(cfg_backup_frame, bg=T["OBS_BG"])
    btn_backup_frame.pack(fill="x", padx=10, pady=(0, 10))
    btn_backup_mes = tk.Button(
        btn_backup_frame, text="Backup do Mês",
        command=lambda: _iniciar_fluxo_backup_mes(btn_backup_mes),
        bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
        padx=14, pady=6, bd=0, cursor="hand2",
    )
    btn_backup_mes.pack(side="left", padx=(0, 8))

    # ── API na nuvem (Opção B) ───────────────────────────────────────────────
    cloud_frame = tk.Frame(aba_dados, bg=T["OBS_BG"])
    cloud_frame.pack(fill="x", padx=20, pady=(0, 10))

    tk.Label(cloud_frame, text="API NA NUVEM (CLIENTE HTML)",
            font=("Segoe UI", 11, "bold"), bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=10, pady=(8, 0))
    tk.Frame(cloud_frame, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=10, pady=(4, 8))
    tk.Label(cloud_frame,
            text="O index.html usa esta URL para gravar no Neon. Cardápio, eventos e config são enviados automaticamente.",
            font=("Segoe UI", 9), bg=T["OBS_BG"], fg=T["FG_TEXT"], wraplength=760, justify="left").pack(anchor="w", padx=10, pady=(0, 8))

    cloud_cfg = _ler_cloud_config()
    cloud_url_var = tk.StringVar(value=cloud_cfg.get("api_url", ""))
    cloud_sync_var = tk.BooleanVar(value=cloud_cfg.get("sincronizar_automatico", True))

    row_cloud = tk.Frame(cloud_frame, bg=T["OBS_BG"])
    row_cloud.pack(fill="x", padx=10, pady=(0, 8))
    tk.Label(row_cloud, text="URL da API:", font=("Segoe UI", 9), bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
    tk.Entry(row_cloud, textvariable=cloud_url_var, font=("Segoe UI", 10), width=48,
             bg=T["ENTRY_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(8, 12))
    tk.Checkbutton(row_cloud, text="Sync automático", variable=cloud_sync_var,
                  font=("Segoe UI", 9), bg=T["OBS_BG"], fg=T["FG_TEXT"],
                  activebackground=T["OBS_BG"], selectcolor=T["ENTRY_BG"]).pack(side="left")

    def _salvar_cloud_config():
        salvar_json(CLOUD_CONFIG_FILE, {
            "api_url": cloud_url_var.get().strip().rstrip("/"),
            "sincronizar_automatico": cloud_sync_var.get(),
        })
        messagebox.showinfo("Salvo", "Configuração da API na nuvem salva.")

    def _sync_cloud_agora():
        _salvar_cloud_config()
        if _sincronizar_tudo_nuvem():
            messagebox.showinfo("Sucesso", "Cardápio, eventos e config enviados para a nuvem.")
        else:
            messagebox.showwarning("Aviso", "Falha ao sincronizar. Verifique a URL e se ApiNuvem.py está no ar.")

    btn_cloud_row = tk.Frame(cloud_frame, bg=T["OBS_BG"])
    btn_cloud_row.pack(fill="x", padx=10, pady=(0, 10))
    tk.Button(btn_cloud_row, text="Salvar URL", command=_salvar_cloud_config,
             bg=T["FRASE_FG"], fg="white", font=("Segoe UI", 10, "bold"),
             padx=12, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
    tk.Button(btn_cloud_row, text="Sincronizar agora", command=_sync_cloud_agora,
             bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI", 10, "bold"),
             padx=12, pady=6, bd=0, cursor="hand2").pack(side="left")

    # ── ABA RELATORIO ─────────────────────────────────────────────────────────
    def gerar_relatorio():
        for w in aba_rel.winfo_children(): w.destroy()

        hoje = datetime.date.today()
        semana_iso = hoje.isocalendar()[1]
        ano_iso    = hoje.isocalendar()[0]
        # Usar janela móvel: últimas 7 dias (hoje e 6 dias anteriores)
        inicio_sem = hoje - datetime.timedelta(days=6)
        fim_sem = hoje

        ds = {"Bom": 0, "Medio": 0, "Ruim": 0}
        ne = {"Comida": {}, "Limpeza": {}, "Ensino": {}, "Semana": {}}
        va = []; total_resp = 0
        alunos_semana = set(); cursos_counter = Counter(); series_counter = Counter()
        def _norm_nome(n: str) -> str:
            if not n:
                return "anonimo"
            s = " ".join(n.split())
            s = s.strip().lower()
            if not s:
                return "anonimo"
            # Normaliza unicode (corrige acentos inconsistentes)
            s = unicodedata.normalize('NFC', s)
            # Tentativa robusta: remover acentos e caracteres não-ASCII
            try:
                s_ascii = unicodedata.normalize('NFD', s)
                s_ascii = ''.join(c for c in s_ascii if unicodedata.category(c) != 'Mn')
                s_alpha = re.sub(r'[^a-z]', '', s_ascii.lower())
            except Exception:
                s_alpha = re.sub(r'[^a-z]', '', s)
            # Se a forma ascii reduzida contém 'anon' ou for 'animo' (corrompido), tratar como anônimo
            if 'anon' in s_alpha or s_alpha.startswith('animo'):
                return 'anonimo'
            return s
        def _aluno_id(n: str, data_str: str) -> str:
            """Retorna um identificador único por avaliador.
            Para nomes válidos, retorna o nome normalizado; para anônimos,
            utiliza o timestamp do registro para diferenciar envios distintos.
            """
            nome_norm = _norm_nome(n)
            if nome_norm == "anonimo":
                # data_str inclui data e hora; usar como separador garante que
                # múltiplas linhas da mesma submissão sejam contadas como uma
                return f"anonimo::{data_str}"
            return nome_norm
        mapa_estagio = {"1": "Comida", "2": "Limpeza", "3": "Ensino", "4": "Semana"}

        try:
            linhas_rel = _avaliacoes_para_linhas()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar avaliações do banco:\n{e}")
            return
        for r in linhas_rel:
            if len(r) < 7:
                continue
            data_str, nome_al, serie_al, curso_al, est, item, ns = (r[0], r[1], r[2], r[3], r[4], r[5], r[6])
            try:
                d_obj = datetime.datetime.strptime(data_str.split(" ")[0], "%d/%m/%Y").date()
                if d_obj < inicio_sem or d_obj > fim_sem:
                    continue
            except Exception:
                continue
            if _eh_almoco_favorito(item):
                va.append(ns)
                alunos_semana.add(_aluno_id(nome_al, data_str))
                if curso_al and curso_al.strip() not in ("N/A", ""):
                    cursos_counter[curso_al.strip()] += 1
                if serie_al and serie_al.strip() not in ("N/A", ""):
                    series_counter[serie_al.strip()] += 1
                continue
            try:
                nota = float(ns)
            except Exception:
                continue
            est_nome = mapa_estagio.get(est, "")
            if not est_nome:
                continue
            if item not in ne[est_nome]:
                ne[est_nome][item] = []
            ne[est_nome][item].append(nota)
            if est == "4":
                if nota <= 1:
                    ds["Ruim"] += 1
                elif nota <= 3:
                    ds["Medio"] += 1
                else:
                    ds["Bom"] += 1
                total_resp += 1
            alunos_semana.add(_aluno_id(nome_al, data_str))
            if curso_al and curso_al.strip() not in ("N/A", ""):
                cursos_counter[curso_al.strip()] += 1
            if serie_al and serie_al.strip() not in ("N/A", ""):
                series_counter[serie_al.strip()] += 1

        cnt = Counter(va)
        if va:
            max_votes = max(cnt.values())
            top_choices = [it for it, vt in cnt.items() if vt == max_votes]
            dv = ", ".join(top_choices); tv = max_votes
        else:
            dv = "Nenhum voto"; tv = 0
        total_favorites = sum(cnt.values())
        total_alunos_unicos = len(alunos_semana)

        cs = tk.Canvas(aba_rel, bg=T["BG_CARD"], highlightthickness=0)
        sbs = ttk.Scrollbar(aba_rel, orient="vertical", command=cs.yview)
        sfs = tk.Frame(cs, bg=T["BG_CARD"])
        sfs.bind("<Configure>", lambda e: cs.configure(scrollregion=cs.bbox("all")))
        window_id = cs.create_window((0, 0), window=sfs, anchor="nw", width=jd.winfo_width() - 60)
        cs.bind("<Configure>", lambda e: cs.itemconfig(window_id, width=e.width))
        cs.configure(yscrollcommand=sbs.set)
        cs.pack(side="left", fill="both", expand=True); sbs.pack(side="right", fill="y")
        cs.bind_all("<MouseWheel>", lambda e: cs.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Cabeçalho
        hdr = tk.Frame(sfs, bg=T["ACCENT_VIBRANT"]); hdr.pack(fill="x")
        tk.Label(hdr, text="RELATÓRIO SEMANAL DE DESEMPENHO",
                 font=("Segoe UI", 16, "bold"), bg=T["ACCENT_VIBRANT"], fg="white",
                 pady=16).pack(side="left", padx=24)
        data_label = f"Semana {semana_iso}/{ano_iso}  —  {hoje.strftime('%d/%m/%Y')}"
        tk.Label(hdr, text=data_label, font=("Segoe UI", 10),
                 bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="right", padx=16)
        tk.Frame(sfs, bg=T["HIGHLIGHT_YELLOW"], height=4).pack(fill="x")

        if total_resp == 0 and not va:
            tk.Label(sfs, text="Aguardando avaliações para gerar o relatório.",
                     font=("Segoe UI", 16), bg=T["BG_CARD"], fg=T["FRASE_FG"], pady=120).pack()
            return

        p_bom = (ds["Bom"] / total_resp) * 100 if total_resp else 0
        if total_resp:
            status = "POSITIVO" if p_bom > 60 else ("REGULAR" if p_bom > 40 else "CRITICO")
            cor_status = {"POSITIVO": "#4caf50", "REGULAR": "#fdd835", "CRITICO": "#f44336"}[status]
        else:
            status = "SEM DADOS"; cor_status = "#9e9e9e"

        # Cards resumo
        summary_frame = tk.Frame(sfs, bg=T["BG_CARD"])
        summary_frame.pack(fill="x", padx=24, pady=(16, 0))
        for c in range(4): summary_frame.columnconfigure(c, weight=1)

        def _resumo_card(parent, titulo, valor, subtitulo, cor, col):
            card = tk.Frame(parent, bg=T["ENTRY_BG"], bd=0,
                            highlightbackground=cor, highlightthickness=2, padx=14, pady=12)
            card.grid(row=0, column=col, sticky="nsew", padx=5, pady=(0, 0))
            tk.Frame(card, bg=cor, height=3).pack(fill="x", pady=(0, 6))
            tk.Label(card, text=titulo, font=("Segoe UI", 8, "bold"),
                     bg=T["ENTRY_BG"], fg=T["FRASE_FG"]).pack(anchor="w")
            tk.Label(card, text=valor, font=("Segoe UI", 18, "bold"),
                     bg=T["ENTRY_BG"], fg=cor).pack(anchor="w", pady=(4, 0))
            if subtitulo:
                tk.Label(card, text=subtitulo, font=("Segoe UI", 9),
                         bg=T["ENTRY_BG"], fg=T["FG_TEXT"]).pack(anchor="w", pady=(3, 0))

        _resumo_card(summary_frame, "ALUNOS ÚNICOS", str(total_alunos_unicos), "Participantes esta semana", T["ACCENT_VIBRANT"], 0)
        _resumo_card(summary_frame, "RESPOSTAS (EXP.)", str(total_resp), "Estágio 4 — Experiência", T["ACCENT_SOFT"], 1)
        _resumo_card(summary_frame, "SATISFAÇÃO", f"{p_bom:.1f}%", status, cor_status, 2)
        _resumo_card(summary_frame, "ALMOÇO FAVORITO", dv[:20] + ("…" if len(dv) > 20 else ""), f"{tv} voto(s) de {total_favorites}", T["HIGHLIGHT_YELLOW"], 3)

        # Badge clima
        badge = tk.Frame(sfs, bg=cor_status, padx=16, pady=10)
        badge.pack(fill="x", padx=24, pady=(10, 0))
        msgs_clima = {
            "POSITIVO": "✅  Clima escolar POSITIVO — maioria dos alunos satisfeita",
            "REGULAR": "⚠️  Clima escolar REGULAR — atenção necessária",
            "CRITICO": "🚨  Clima escolar CRÍTICO — ação imediata recomendada",
            "SEM DADOS": "ℹ️  Sem dados de sentimento registrados esta semana",
        }
        tk.Label(badge, text=msgs_clima.get(status, ""),
                 font=("Segoe UI", 11, "bold"), bg=cor_status, fg="white").pack(side="left")
        tk.Label(badge, text=f"Bom: {ds['Bom']}   Regular: {ds['Medio']}   Ruim: {ds['Ruim']}",
                 font=("Segoe UI", 10), bg=cor_status, fg="white").pack(side="right")

        # Gráfico sentimento
        sentimento_frame = tk.Frame(sfs, bg=T["ENTRY_BG"], bd=0,
                                    highlightbackground=T["BORDER_GRID"], highlightthickness=1,
                                    padx=18, pady=16)
        sentimento_frame.pack(fill="x", padx=24, pady=(14, 0))
        tk.Label(sentimento_frame, text="SENTIMENTO — Estágio 4 (Experiência da Semana)",
                 font=("Segoe UI", 12, "bold"), bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w")
        tk.Label(sentimento_frame, text=f"Baseado em {total_resp} resposta(s) desta semana",
                 font=("Segoe UI", 9), bg=T["ENTRY_BG"], fg=T["FRASE_FG"]).pack(anchor="w", pady=(2, 8))
        cv = tk.Canvas(sentimento_frame, bg=T["ENTRY_BG"], highlightthickness=0, height=100)
        cv.pack(fill="x", pady=2)
        itens_sent = [("🟢  Bom", ds["Bom"], "#4caf50"), ("🟡  Regular", ds["Medio"], "#fdd835"), ("🔴  Ruim", ds["Ruim"], "#f44336")]
        def _desenhar_barras(event=None):
            largura_canvas = max(320, cv.winfo_width())
            cv.delete("all")
            bar_w = max(200, largura_canvas - 200)
            x_label = 100; x_start = x_label + 16
            max_ds = max(ds.values()) if max(ds.values()) > 0 else 1
            for i, (label, val, cor) in enumerate(itens_sent):
                y = 14 + i * 30
                cv.create_rectangle(x_start, y, x_start + bar_w, y + 20, fill=T["BG_CARD"], outline="")
                filled = int((val / max_ds) * bar_w) if bar_w > 0 else 0
                if filled > 0:
                    cv.create_rectangle(x_start, y, x_start + filled, y + 20, fill=cor, outline="")
                cv.create_text(x_label, y + 10, text=label, font=("Segoe UI", 9, "bold"), anchor="e", fill=T["FG_TEXT"])
                pct = (val / total_resp * 100) if total_resp else 0
                cv.create_text(x_start + filled + 8, y + 10,
                               text=f"{val} ({pct:.0f}%)", font=("Segoe UI", 9, "bold"),
                               anchor="w", fill=T["FG_TEXT"])
        cv.bind("<Configure>", _desenhar_barras)
        cv.after(80, _desenhar_barras)

        # Breakdown detalhado por estágio
        tk.Frame(sfs, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=24, pady=(16, 0))
        tk.Label(sfs, text="DETALHAMENTO POR CATEGORIA E ITEM",
                 font=("Segoe UI", 13, "bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(
                 anchor="w", padx=24, pady=(12, 6))

        ICONES_EST = {"Comida": "🍽️", "Limpeza": "🧹", "Ensino": "📚", "Semana": "✨"}
        COR_EST = {"Comida": "#e65100", "Limpeza": "#1565c0", "Ensino": "#6a1b9a", "Semana": "#2e7d32"}

        for est_nome, itens_dict in ne.items():
            if not itens_dict: continue
            bloco = tk.Frame(sfs, bg=T["BG_CARD"], highlightbackground=T["BORDER_GRID"], highlightthickness=1)
            bloco.pack(fill="x", padx=24, pady=(0, 10))
            bloco_hd = tk.Frame(bloco, bg=COR_EST.get(est_nome, T["ACCENT_VIBRANT"]))
            bloco_hd.pack(fill="x")
            media_geral_est = [n for notas in itens_dict.values() for n in notas]
            media_est = sum(media_geral_est) / len(media_geral_est) if media_geral_est else 0
            tk.Label(bloco_hd, text=f"  {ICONES_EST.get(est_nome,'')}  {est_nome.upper()}",
                     font=("Segoe UI", 11, "bold"),
                     bg=COR_EST.get(est_nome, T["ACCENT_VIBRANT"]), fg="white", pady=8).pack(side="left", padx=6)
            tk.Label(bloco_hd, text=f"Média geral: {media_est:.2f}/5.00  |  {len(media_geral_est)} resp.",
                     font=("Segoe UI", 9), bg=COR_EST.get(est_nome, T["ACCENT_VIBRANT"]), fg="white").pack(side="right", padx=12)

            tbl = tk.Frame(bloco, bg=T["BG_CARD"]); tbl.pack(fill="x")
            for ci, txt in enumerate(["ITEM / CRITÉRIO", "MÉDIA", "MIN", "MÁX", "DISTRIBUIÇÃO 1→5"]):
                tk.Label(tbl, text=txt, font=("Segoe UI", 8, "bold"),
                         bg=T["ENTRY_BG"], fg=T["FRASE_FG"], padx=8, pady=5,
                         anchor="w" if ci == 0 else "center").grid(
                             row=0, column=ci, sticky="nsew", padx=1, pady=1)
            tbl.grid_columnconfigure(0, weight=3)
            for ci in range(1, 5): tbl.grid_columnconfigure(ci, weight=1)

            for idx_item, (item_nome, notas) in enumerate(sorted(itens_dict.items())):
                bg_row = T["BG_CARD"] if idx_item % 2 == 0 else T["ENTRY_BG"]
                media_item = sum(notas) / len(notas) if notas else 0
                min_item = min(notas) if notas else 0; max_item = max(notas) if notas else 0
                cor_nota = "#4caf50" if media_item >= 4 else ("#fdd835" if media_item >= 2.5 else "#f44336")
                estrelas = "★" * round(media_item) + "☆" * (5 - round(media_item))
                tk.Label(tbl, text=f"  {item_nome}", font=("Segoe UI", 10),
                         bg=bg_row, fg=T["FG_TEXT"], anchor="w", padx=8, pady=5,
                         wraplength=280, justify="left").grid(row=idx_item+1, column=0, sticky="nsew", padx=1, pady=1)
                frm_nota = tk.Frame(tbl, bg=bg_row); frm_nota.grid(row=idx_item+1, column=1, sticky="nsew", padx=1, pady=1)
                tk.Label(frm_nota, text=f"{media_item:.2f}", font=("Segoe UI", 10, "bold"), bg=bg_row, fg=cor_nota).pack()
                tk.Label(frm_nota, text=estrelas, font=("Segoe UI", 7), bg=bg_row, fg=T["HIGHLIGHT_YELLOW"]).pack()
                tk.Label(tbl, text=f"{min_item:.0f}", font=("Segoe UI", 10),
                         bg=bg_row, fg="#c62828", pady=5).grid(row=idx_item+1, column=2, sticky="nsew", padx=1, pady=1)
                tk.Label(tbl, text=f"{max_item:.0f}", font=("Segoe UI", 10),
                         bg=bg_row, fg="#2e7d32", pady=5).grid(row=idx_item+1, column=3, sticky="nsew", padx=1, pady=1)
                dist_txt = "  ".join([f"{int(n)}★={notas.count(n)}" for n in [1,2,3,4,5] if notas.count(n) > 0])
                tk.Label(tbl, text=dist_txt or "—", font=("Segoe UI", 8),
                         bg=bg_row, fg=T["FRASE_FG"], padx=6, pady=5).grid(row=idx_item+1, column=4, sticky="nsew", padx=1, pady=1)

        # Almoço favorito
        if va:
            tk.Frame(sfs, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=24, pady=(6, 0))
            fav_frame = tk.Frame(sfs, bg=T["ENTRY_BG"], bd=0,
                                 highlightbackground=T["BORDER_GRID"], highlightthickness=1,
                                 padx=18, pady=16)
            fav_frame.pack(fill="x", padx=24, pady=(10, 0))
            tk.Label(fav_frame, text="🍛  RANKING — ALMOÇO FAVORITO DA SEMANA",
                     font=("Segoe UI", 12, "bold"), bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w")
            tk.Label(fav_frame, text=f"Total de votos: {total_favorites}",
                     font=("Segoe UI", 9), bg=T["ENTRY_BG"], fg=T["FRASE_FG"]).pack(anchor="w", pady=(2, 8))
            for pos, (opc, cont) in enumerate(cnt.most_common(), start=1):
                pct_fav = (cont / total_favorites * 100) if total_favorites else 0
                item_f = tk.Frame(fav_frame, bg=T["ENTRY_BG"]); item_f.pack(fill="x", pady=2)
                medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, f"  {pos}.")
                tk.Label(item_f, text=f"{medal}  {opc}", font=("Segoe UI", 11, "bold"),
                         bg=T["ENTRY_BG"], fg=T["FG_TEXT"]).pack(side="left")
                tk.Label(item_f, text=f"{cont} voto(s)  ({pct_fav:.1f}%)",
                         font=("Segoe UI", 10), bg=T["ENTRY_BG"], fg=T["FRASE_FG"]).pack(side="right")

        # Participação por curso/série
        if cursos_counter or series_counter:
            tk.Frame(sfs, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=24, pady=(14, 0))
            tk.Label(sfs, text="PARTICIPAÇÃO POR TURMA",
                     font=("Segoe UI", 12, "bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=24, pady=(10, 6))
            part_frame = tk.Frame(sfs, bg=T["BG_CARD"])
            part_frame.pack(fill="x", padx=24, pady=(0, 10))
            part_frame.grid_columnconfigure(0, weight=1); part_frame.grid_columnconfigure(1, weight=1)
            def _bloco_part(parent, titulo, counter, row, col):
                bf = tk.Frame(parent, bg=T["ENTRY_BG"], bd=0,
                              highlightbackground=T["BORDER_GRID"], highlightthickness=1,
                              padx=12, pady=10)
                bf.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)
                tk.Label(bf, text=titulo, font=("Segoe UI", 10, "bold"),
                         bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", pady=(0, 5))
                total_part = sum(counter.values())
                for item_p, cnt_p in counter.most_common():
                    pf = tk.Frame(bf, bg=T["ENTRY_BG"]); pf.pack(fill="x", pady=1)
                    tk.Label(pf, text=item_p, font=("Segoe UI", 10), bg=T["ENTRY_BG"], fg=T["FG_TEXT"]).pack(side="left")
                    pct_p = (cnt_p / total_part * 100) if total_part else 0
                    tk.Label(pf, text=f"{cnt_p} ({pct_p:.0f}%)", font=("Segoe UI", 9), bg=T["ENTRY_BG"], fg=T["FRASE_FG"]).pack(side="right")
            _bloco_part(part_frame, "Por Curso", cursos_counter, 0, 0)
            _bloco_part(part_frame, "Por Série", series_counter, 0, 1)

        # Exportar PDF
        def exportar_pdf():
            try:
                from fpdf import FPDF
                pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial", "B", 16)
                pdf.cell(200, 10, txt=f"RELATORIO SEMANAL - EEEP MARWIN - Semana {semana_iso}/{ano_iso}", ln=True, align="C")
                pdf.set_font("Arial", "", 12)
                pdf.cell(200, 10, txt=f"Data: {hoje.strftime('%d/%m/%Y')}  |  Alunos: {total_alunos_unicos}  |  Satisfacao: {p_bom:.1f}%", ln=True, align="C")
                pdf.ln(5); pdf.set_font("Arial", "B", 13); pdf.cell(0, 8, txt="Sentimento:", ln=True)
                pdf.set_font("Arial", "", 11)
                pdf.cell(0, 7, txt=f"Bom: {ds['Bom']}  Regular: {ds['Medio']}  Ruim: {ds['Ruim']}  |  Status: {status}", ln=True); pdf.ln(4)
                for est_nome, itens_dict in ne.items():
                    if not itens_dict: continue
                    media_geral_est = [n for notas in itens_dict.values() for n in notas]
                    media_est = sum(media_geral_est)/len(media_geral_est) if media_geral_est else 0
                    pdf.set_font("Arial", "B", 12); pdf.cell(0, 8, txt=f"{est_nome.upper()} (media: {media_est:.2f}/5.00)", ln=True)
                    pdf.set_font("Arial", "", 10)
                    for item_nome, notas in sorted(itens_dict.items()):
                        media_i = sum(notas)/len(notas) if notas else 0
                        dist = "  ".join([f"{int(n)}x{notas.count(n)}" for n in [1,2,3,4,5] if notas.count(n) > 0])
                        pdf.cell(0, 6, txt=f"  - {item_nome}: {media_i:.2f}/5 ({len(notas)} resp.) | {dist}", ln=True)
                    pdf.ln(2)
                if va:
                    pdf.set_font("Arial", "B", 12); pdf.cell(0, 8, txt="Almoco Favorito:", ln=True)
                    pdf.set_font("Arial", "", 10)
                    for opc, cont in cnt.most_common():
                        pdf.cell(0, 6, txt=f"  - {opc}: {cont} voto(s)", ln=True)
                nome_arq = f"relatorio_marwin_semana{semana_iso}_{ano_iso}.pdf"
                pdf.output(nome_arq); messagebox.showinfo("Sucesso", f"PDF salvo como:\n{nome_arq}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha: {e}")

        btn_bar = tk.Frame(sfs, bg=T["BG_CARD"]); btn_bar.pack(fill="x", padx=24, pady=(18, 36))
        tk.Button(btn_bar, text="  EXPORTAR PDF  ", font=("Segoe UI", 11, "bold"),
                  bg=T["ACCENT_VIBRANT"], fg="white", padx=30, pady=12, bd=0,
                  cursor="hand2", command=exportar_pdf).pack(side="left")
        tk.Label(btn_bar, text=f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
                 font=("Segoe UI", 9), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left", padx=14)


    # ── ABA EDITAR CARDAPIO ───────────────────────────────────────────────────
    def setup_cardapio():
        for w in aba_card_ed.winfo_children(): w.destroy()
        hd = tk.Frame(aba_card_ed, bg=T["BG_CARD"]); hd.pack(fill="x", padx=20, pady=(15,5))
        tk.Label(hd, text="Editar Cardapio da Semana", font=("Segoe UI",16,"bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        tk.Label(hd, text="Edite os campos e clique em Salvar", font=("Segoe UI",10), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left", padx=15)
        tk.Frame(aba_card_ed, bg=T["BORDER_GRID"], height=2).pack(fill="x", padx=20, pady=(5,10))
        ca = ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO)
        dias = ["SEGUNDA", "TERCA", "QUARTA", "QUINTA", "SEXTA"]
        cores_dia  = {"SEGUNDA":"#2e7d32","TERCA":"#1565c0","QUARTA":"#6a1b9a","QUINTA":"#e65100","SEXTA":"#c62828"}
        refeicoes  = [("Merenda Manha"),("Almoco"),("Merenda Tarde")]
        ents = {}
        def salvar_card():
            novo = {d:[e.get() for e in ents[d]] for d in dias}
            salvar_json(CARDAPIO_FILE, novo)
            _sync_nuvem("/admin/cardapio", "PUT", novo)
            messagebox.showinfo("Sucesso", "Cardapio atualizado!")
        tk.Button(aba_card_ed, text="SALVAR ALTERACOES", command=salvar_card, bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI",11,"bold"), pady=9, padx=25, bd=0, cursor="hand2").pack(fill="x", padx=20, pady=(0,8))
        tk.Frame(aba_card_ed, bg=T["BORDER_GRID"], height=2).pack(fill="x", padx=20, pady=(0,4))
        cs = tk.Canvas(aba_card_ed, bg=T["BG_CARD"], highlightthickness=0)
        sb = ttk.Scrollbar(aba_card_ed, orient="vertical", command=cs.yview)
        sf = tk.Frame(cs, bg=T["BG_CARD"])
        sf.bind("<Configure>", lambda e: cs.configure(scrollregion=cs.bbox("all")))
        wid = cs.create_window((0,0), window=sf, anchor="nw")
        cs.bind("<Configure>", lambda e: cs.itemconfig(wid, width=e.width))
        cs.configure(yscrollcommand=sb.set); cs.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
        cs.bind_all("<MouseWheel>", lambda e: cs.yview_scroll(int(-1*(e.delta/120)), "units"))
        grade = tk.Frame(sf, bg=T["BG_CARD"]); grade.pack(fill="both", expand=True, padx=15, pady=10)
        grade.grid_columnconfigure(0, weight=1, uniform="col"); grade.grid_columnconfigure(1, weight=1, uniform="col")
        posicoes = [(0,0),(0,1),(1,0),(1,1),(2,0)]
        for idx_dia, dia in enumerate(dias):
            gr, gc = posicoes[idx_dia]; cor = cores_dia[dia]
            if idx_dia == 4:
                card = tk.Frame(grade, bg=T["BG_CARD"], highlightbackground=T["BORDER_GRID"], highlightthickness=1)
                card.grid(row=gr, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
            else:
                card = tk.Frame(grade, bg=T["BG_CARD"], highlightbackground=T["BORDER_GRID"], highlightthickness=1)
                card.grid(row=gr, column=gc, sticky="nsew", padx=5, pady=5)
            dia_hd = tk.Frame(card, bg=cor); dia_hd.pack(fill="x")
            tk.Label(dia_hd, text=f"  {dia}", font=("Segoe UI",13,"bold"), bg=cor, fg="white", pady=9).pack(side="left")
            ents[dia] = []
            if idx_dia == 4:
                inner = tk.Frame(card, bg=T["BG_CARD"], padx=14, pady=10); inner.pack(fill="both", expand=True)
                for j, nome_ref in enumerate(refeicoes):
                    col2 = tk.Frame(inner, bg=T["BG_CARD"]); col2.pack(side="left", expand=True, fill="both", padx=(0,12 if j<2 else 0))
                    tk.Label(col2, text=nome_ref, font=("Segoe UI",9,"bold"), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(anchor="w", pady=(0,4))
                    e = tk.Entry(col2, font=("Segoe UI",11), bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1, insertbackground=cor)
                    e.insert(0, ca.get(dia, ["","",""])[j]); e.pack(fill="x", ipady=8); ents[dia].append(e)
            else:
                inner = tk.Frame(card, bg=T["BG_CARD"], padx=14, pady=10); inner.pack(fill="both", expand=True)
                for j, nome_ref in enumerate(refeicoes):
                    tk.Label(inner, text=nome_ref, font=("Segoe UI",9,"bold"), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(anchor="w", pady=(4,2))
                    e = tk.Entry(inner, font=("Segoe UI",11), bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1, insertbackground=cor)
                    e.insert(0, ca.get(dia, ["","",""])[j]); e.pack(fill="x", ipady=8); ents[dia].append(e)

    # ── ABA EDITAR EVENTOS ────────────────────────────────────────────────────
    def setup_eventos():
        for w in aba_ev_ed.winfo_children(): w.destroy()
        corpo = tk.Frame(aba_ev_ed, bg=T["BG_CARD"]); corpo.pack(fill="both", expand=True)
        esq = tk.Frame(corpo, bg=T["OBS_BG"], width=280); esq.pack(side="left", fill="y"); esq.pack_propagate(False)
        logo_path = _buscar_logo_png()
        if logo_path:
            try:
                from PIL import Image as _PilImg, ImageTk as _PilImgTk
                _img_logo = _PilImg.open(logo_path).convert("RGBA")
                _img_logo.thumbnail((140, 140), _PilImg.LANCZOS)
                _logo_evento_tk = _PilImgTk.PhotoImage(_img_logo, master=esq)
                tk.Label(esq, image=_logo_evento_tk, bg=T["OBS_BG"]).pack(pady=(24, 8))
                esq._logo_evento_ref = _logo_evento_tk
            except Exception:
                pass
        tk.Label(esq, text="Novo Evento", font=("Segoe UI",15,"bold"), bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(pady=(12,4))
        tk.Label(esq, text="Adicione datas e descricoes ao calendario", font=("Segoe UI",9), bg=T["OBS_BG"], fg=T["FRASE_FG"], justify="center").pack(pady=(0,20))
        tk.Frame(esq, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=20, pady=(0,20))
        tk.Label(esq, text="Data", font=("Segoe UI",10,"bold"), bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=20)
        tk.Label(esq, text="Formato: DD/MM  (ex: 15/07)", font=("Segoe UI",8), bg=T["OBS_BG"], fg=T["FRASE_FG"]).pack(anchor="w", padx=20, pady=(2,4))
        ed = tk.Entry(esq, font=("Segoe UI",13), relief="solid", bd=1, bg=T["BG_CARD"], fg=T["FG_TEXT"], insertbackground=T["ACCENT_VIBRANT"])
        ed.pack(fill="x", padx=20, ipady=7, pady=(0,16)); ed.insert(0,"01/01")
        tk.Label(esq, text="Descricao", font=("Segoe UI",10,"bold"), bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=20)
        tk.Label(esq, text="Nome do evento ou feriado", font=("Segoe UI",8), bg=T["OBS_BG"], fg=T["FRASE_FG"]).pack(anchor="w", padx=20, pady=(2,4))
        edesc = tk.Entry(esq, font=("Segoe UI",13), relief="solid", bd=1, bg=T["BG_CARD"], fg=T["FG_TEXT"], insertbackground=T["ACCENT_VIBRANT"])
        edesc.pack(fill="x", padx=20, ipady=7, pady=(0,24))
        dir2 = tk.Frame(corpo, bg=T["BG_CARD"]); dir2.pack(side="left", fill="both", expand=True)
        _configurar_layout_portrait(corpo, esq, dir2)
        dir_hd = tk.Frame(dir2, bg=T["BG_CARD"]); dir_hd.pack(fill="x", padx=25, pady=(20,5))
        tk.Label(dir_hd, text="Eventos Cadastrados", font=("Segoe UI",15,"bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        tk.Frame(dir2, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=25, pady=(0,10))
        fte = tk.Frame(dir2, bg=T["BORDER_GRID"], padx=1, pady=1); fte.pack(fill="both", expand=True, padx=25)
        tev = ttk.Treeview(fte, columns=("D","E"), show="headings")
        tev.heading("D", text="Data"); tev.column("D", width=90, anchor="center")
        tev.heading("E", text="Evento"); tev.column("E", width=400, anchor="w")
        tev.pack(fill="both", expand=True)
        tev.tag_configure("even", background=T["ENTRY_BG"]); tev.tag_configure("odd", background=T["BG_CARD"])
        def carregar_ev():
            for i in tev.get_children(): tev.delete(i)
            evs = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
            try: evs.sort(key=lambda x:(int(x["data"].split("/")[1]),int(x["data"].split("/")[0])))
            except: pass
            for idx,ev in enumerate(evs):
                tev.insert("","end",values=(ev["data"],ev["evento"]),tags=("even" if idx%2==0 else "odd",))
            lbl_total.config(text=f"{len(evs)} evento(s) cadastrado(s)")
        lbl_total = tk.Label(dir2, text="", font=("Segoe UI",9), bg=T["BG_CARD"], fg=T["FRASE_FG"]); lbl_total.pack(anchor="e", padx=25, pady=(4,0))
        btns = tk.Frame(dir2, bg=T["BG_CARD"]); btns.pack(fill="x", padx=25, pady=12)
        def rem_ev():
            sel = tev.selection()
            if not sel: messagebox.showwarning("Aviso","Selecione um evento na lista primeiro."); return
            if messagebox.askyesno("Confirmar","Remover o evento selecionado?"):
                val = tev.item(sel[0],"values"); evs = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
                evs = [e for e in evs if not (e["data"]==val[0] and e["evento"]==val[1])]
                salvar_json(EVENTOS_FILE, evs)
                _sync_nuvem("/admin/eventos", "PUT", evs)
                carregar_ev()
        tk.Button(btns, text="Remover Selecionado", command=rem_ev, bg="#c62828", fg="white", font=("Segoe UI",10,"bold"), pady=9, padx=20, bd=0, cursor="hand2").pack(side="right")
        def add_ev():
            data = ed.get().strip(); desc = edesc.get().strip()
            if not data or not desc: messagebox.showwarning("Campos vazios","Preencha a data e a descricao."); return
            if "/" not in data or len(data)!=5: messagebox.showwarning("Formato invalido","Use o formato DD/MM  ex: 15/07"); return
            evs = ler_json(EVENTOS_FILE, EVENTOS_PADRAO); evs.append({"data":data,"evento":desc})
            salvar_json(EVENTOS_FILE, evs)
            _sync_nuvem("/admin/eventos", "PUT", evs)
            carregar_ev()
            ed.delete(0,"end"); ed.insert(0,"01/01"); edesc.delete(0,"end"); edesc.focus()
        tk.Button(esq, text="Adicionar Evento", command=add_ev, bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI",12,"bold"), pady=13, bd=0, cursor="hand2").pack(fill="x", padx=20)
        esq.bind_all("<Return>", lambda e: add_ev()); carregar_ev()

    # ── ABA REFEITÓRIO ────────────────────────────────────────────────────────
    def setup_refeitorio():
        for w in aba_ref.winfo_children(): w.destroy()

        # Mapeamento de siglas de sala para série/curso
        SIGLAS = {
            "1DS":  {"serie": "1 Ano", "curso": "Desenvolvimento de Sistemas"},
            "1HOS": {"serie": "1 Ano", "curso": "Hospedagem"},
            "1ENF": {"serie": "1 Ano", "curso": "Enfermagem"},
            "1MOD": {"serie": "1 Ano", "curso": "Modelagem do Vestuario"},
            "2DS":  {"serie": "2 Ano", "curso": "Desenvolvimento de Sistemas"},
            "2HOS": {"serie": "2 Ano", "curso": "Hospedagem"},
            "2ENF": {"serie": "2 Ano", "curso": "Enfermagem"},
            "2MOD": {"serie": "2 Ano", "curso": "Modelagem do Vestuario"},
            "3DS":  {"serie": "3 Ano", "curso": "Desenvolvimento de Sistemas"},
            "3HOS": {"serie": "3 Ano", "curso": "Hospedagem"},
            "3ENF": {"serie": "3 Ano", "curso": "Enfermagem"},
            "3MOD": {"serie": "3 Ano", "curso": "Modelagem do Vestuario"},
        }
        SIGLAS_LABEL = {
            "1DS": "1º DS",  "1HOS": "1º HOS", "1ENF": "1º ENF", "1MOD": "1º MOD",
            "2DS": "2º DS",  "2HOS": "2º HOS", "2ENF": "2º ENF", "2MOD": "2º MOD",
            "3DS": "3º DS",  "3HOS": "3º HOS", "3ENF": "3º ENF", "3MOD": "3º MOD",
        }
        CORES_ANO = {"1": "#1565c0", "2": "#6a1b9a", "3": "#c62828"}

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        hd = tk.Frame(aba_ref, bg=T["BG_CARD"]); hd.pack(fill="x", padx=20, pady=(15, 5))
        tk.Label(hd, text="Controle de Refeições - Hoje", font=("Segoe UI", 16, "bold"),
                 bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        tk.Label(hd, text=f"  ({_hoje()})", font=("Segoe UI", 11),
                 bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left")
        tk.Frame(aba_ref, bg=T["BORDER_GRID"], height=2).pack(fill="x", padx=20, pady=(5, 0))

        # ── Linha de filtros por texto / combo ────────────────────────────────
        container_filtros = tk.Frame(aba_ref, bg=T["OBS_BG"])
        container_filtros.pack(fill="x", padx=20, pady=(6, 0))

        filtro_frame = tk.Frame(container_filtros, bg=T["OBS_BG"])
        filtro_frame.pack(fill="x", padx=10, pady=(0, 0))

        tk.Label(filtro_frame, text="Filtrar por:", font=("Segoe UI", 10, "bold"),
                 bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(side="left", padx=(10, 8), pady=8)

        fv_serie = tk.StringVar()
        fv_curso  = tk.StringVar()
        fv_nome   = tk.StringVar()

        tk.Label(filtro_frame, text="Série:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        cb_serie = ttk.Combobox(filtro_frame, textvariable=fv_serie, state="readonly",
                                values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                width=9, font=("Segoe UI", 10))
        cb_serie.set("(Todas)")
        cb_serie.pack(side="left", padx=(4, 14), pady=6)

        tk.Label(filtro_frame, text="Curso:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        cursos_opcoes = [
            "(Todos)",
            "Desenvolvimento de Sistemas",
            "Hospedagem",
            "Enfermagem",
            "Modelagem do Vestuario",
        ]
        cb_curso = ttk.Combobox(filtro_frame, textvariable=fv_curso, state="readonly",
                                values=cursos_opcoes, width=26, font=("Segoe UI", 10))
        cb_curso.set("(Todos)")
        cb_curso.pack(side="left", padx=(4, 14), pady=6)

        tk.Label(filtro_frame, text="Nome:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        entry_nome = tk.Entry(filtro_frame, textvariable=fv_nome, font=("Segoe UI", 10),
                              bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1,
                              insertbackground=T["ACCENT_VIBRANT"], width=20)
        entry_nome.pack(side="left", ipady=4, padx=(4, 14), pady=6)

        btn_limpar = tk.Button(filtro_frame, text="✕ Limpar", font=("Segoe UI", 9, "bold"),
                               bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"], bd=0,
                               cursor="hand2", padx=8, pady=4)
        btn_limpar.pack(side="left", padx=(0, 10), pady=6)

        # ── Botões de atalho por sala ──────────────────────────────────────────
        sala_frame = tk.Frame(aba_ref, bg=T["BG_CARD"])
        sala_frame.pack(fill="x", padx=20, pady=(6, 0))
        tk.Label(sala_frame, text="Sala rápida:", font=("Segoe UI", 9, "bold"),
                 bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left", padx=(4, 8))

        btn_sala_refs = {}

        def _filtrar_por_sigla(sigla):
            info = SIGLAS[sigla]
            fv_serie.set(info["serie"])
            fv_curso.set(info["curso"])
            fv_nome.set("")
            # destaca botão ativo, reseta os demais
            for s, b in btn_sala_refs.items():
                b.configure(relief="flat", bg=CORES_ANO[s[0]], fg="white")
            btn_sala_refs[sigla].configure(relief="solid", bg="white",
                                           fg=CORES_ANO[sigla[0]])
            atualizar_ref()

        for sigla, label in SIGLAS_LABEL.items():
            ano = sigla[0]
            cor = CORES_ANO[ano]
            b = tk.Button(sala_frame, text=label, font=("Segoe UI", 8, "bold"),
                          bg=cor, fg="white", bd=0, padx=9, pady=4,
                          cursor="hand2",
                          command=lambda s=sigla: _filtrar_por_sigla(s))
            b.pack(side="left", padx=2, pady=4)
            btn_sala_refs[sigla] = b

        # ── Barra de botões e contadores ──────────────────────────────────────
        btn_row = tk.Frame(aba_ref, bg=T["BG_CARD"])
        btn_row.pack(fill="x", padx=20, pady=(8, 4))

        lbl_total_ref = tk.Label(btn_row, text="Total: 0", font=("Segoe UI", 12, "bold"),
                                 bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"], padx=14, pady=6)
        lbl_total_ref.pack(side="left", padx=(0, 10))

        lbl_filtro_ativo = tk.Label(btn_row, text="", font=("Segoe UI", 9, "italic"),
                                    bg=T["BG_CARD"], fg=T["FRASE_FG"])
        lbl_filtro_ativo.pack(side="left", padx=(0, 16))

        def exportar_csv_ref():
            try:
                nome_arq = f"refeitorio_{_hoje().replace('/', '_')}.csv"
                _escrever_csv(
                    nome_arq,
                    ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
                    _ler_refeitorio_todos_db(),
                )
                messagebox.showinfo("Exportado", f"Arquivo salvo como:\n{nome_arq}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")

        def apagar_hoje():
            if messagebox.askyesno("Confirmar", f"Apagar todos os registros de hoje ({_hoje()})?"):
                try:
                    _apagar_refeitorio_data_db(_hoje())
                    messagebox.showinfo("Sucesso", "Registros de hoje apagados do banco.")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao apagar registros:\n{e}")
                atualizar_ref()

        tk.Button(btn_row, text="↻  Atualizar", command=lambda: atualizar_ref(),
                  bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="⬇  Exportar CSV", command=exportar_csv_ref,
                  bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="🗑  Apagar hoje", command=apagar_hoje,
                  bg="#c62828", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="right")

        # ── Tabela: NOME | SÉRIE | CURSO | STATUS ─────────────────────────────
        ft = tk.Frame(aba_ref, bg=T["BORDER_GRID"], padx=1, pady=1)
        ft.pack(expand=True, fill="both", padx=20, pady=(4, 20))

        tref = ttk.Treeview(ft, columns=("N", "S", "C", "AL", "A"), show="headings")
        tref.heading("N",  text="ALUNO");  tref.column("N",  width=320, anchor="w")
        tref.heading("S",  text="SÉRIE");  tref.column("S",  width=90,  anchor="center")
        tref.heading("C",  text="CURSO");  tref.column("C",  width=220, anchor="w")
        tref.heading("AL", text="ALMOÇO"); tref.column("AL", width=130, anchor="center")
        tref.heading("A",  text="AULA");   tref.column("A",  width=120, anchor="center")

        sb_ref = ttk.Scrollbar(ft, orient="vertical", command=tref.yview)
        tref.configure(yscrollcommand=sb_ref.set)
        tref.pack(side="left", expand=True, fill="both")
        sb_ref.pack(side="right", fill="y")

        tref.tag_configure("almoca",
                           foreground="#1b5e20", background="#e8f5e9",
                           font=("Segoe UI", 11, "bold"))
        tref.tag_configure("nao_almoca",
                           foreground="#b71c1c", background="#ffebee",
                           font=("Segoe UI", 11, "bold"))

        # ======================================================================
        # CORREÇÃO: normalização e mapas de equivalência para série e curso.
        # O JSON dos alunos usa "1º","2º","3º" e siglas "DS","ENF","HOSP","MOD".
        # O cliente pode gravar no CSV qualquer uma dessas variações.
        # Os filtros do combo usam nomes completos ("1 Ano", "Enfermagem", etc.).
        # As funções abaixo resolvem todas as variantes para um mesmo grupo,
        # permitindo que o filtro funcione independentemente do formato gravado.
        # ======================================================================
        import unicodedata as _ud

        def _norm(texto):
            """Remove acentos e converte para minúsculas."""
            return _ud.normalize("NFD", str(texto).lower()).encode("ascii", "ignore").decode("ascii")

        # Grupos de série: chave = identificador canônico, valor = lista de variantes normalizadas
        _GRUPOS_SERIE = {
            "1": ["1º", "1o", "1 ano", "primeiro ano", "primeiro", "1"],
            "2": ["2º", "2o", "2 ano", "segundo ano",  "segundo",  "2"],
            "3": ["3º", "3o", "3 ano", "terceiro ano", "terceiro", "3"],
        }

        # Grupos de curso
        _GRUPOS_CURSO = {
            "ds":   ["ds", "desenvolvimento de sistemas", "dev. sistemas", "dev sistemas", "desenv. sistemas"],
            "enf":  ["enf", "enfermagem"],
            "hosp": ["hosp", "hospedagem"],
            "mod":  ["mod", "modelagem do vestuario", "modelagem", "vestuario"],
        }

        def _grupo_serie(texto):
            """Retorna a chave canônica do grupo de série, ou None."""
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_SERIE.items():
                for v in variantes:
                    if t == v or t.startswith(v) or v.startswith(t):
                        return chave
            return None

        def _grupo_curso(texto):
            """Retorna a chave canônica do grupo de curso, ou None."""
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_CURSO.items():
                for v in variantes:
                    if t == v or t in v or v in t:
                        return chave
            return None

        def _serie_bate(serie_aluno, filtro_serie):
            if filtro_serie in ("(Todas)", ""):
                return True
            g_aluno  = _grupo_serie(serie_aluno)
            g_filtro = _grupo_serie(filtro_serie)
            if g_aluno is None or g_filtro is None:
                # fallback: substring normalizado
                return _norm(filtro_serie) in _norm(serie_aluno)
            return g_aluno == g_filtro

        def _curso_bate(curso_aluno, filtro_curso):
            if filtro_curso in ("(Todos)", ""):
                return True
            g_aluno  = _grupo_curso(curso_aluno)
            g_filtro = _grupo_curso(filtro_curso)
            if g_aluno is None or g_filtro is None:
                return _norm(filtro_curso) in _norm(curso_aluno)
            return g_aluno == g_filtro

        def _canonizar_serie(serie):
            grupo = _grupo_serie(serie)
            if grupo == "1":
                return "1 Ano"
            if grupo == "2":
                return "2 Ano"
            if grupo == "3":
                return "3 Ano"
            return serie.strip() if serie else "Sem Serie"

        def _canonizar_curso(curso):
            grupo = _grupo_curso(curso)
            if grupo == "ds":
                return "Desenvolvimento de Sistemas"
            if grupo == "enf":
                return "Enfermagem"
            if grupo == "hosp":
                return "Hospedagem"
            if grupo == "mod":
                return "Modelagem do Vestuario"
            return curso.strip() if curso else "Sem Curso"
        # ======================================================================

        # ── Atualização da tabela com filtros aplicados ───────────────────────
        def atualizar_ref():
            for i in tref.get_children(): tref.delete(i)
            registros = _registros_hoje()

            # r = [Data, HoraEntrada, Matricula, Nome, Serie, Curso, Refeicao]
            alunos = {}
            for r in registros:
                mat   = r[2]
                nome  = r[3]
                serie = r[4] if len(r) > 4 else ""
                curso = r[5] if len(r) > 5 else ""
                ref   = r[6].strip().lower() if len(r) > 6 else ""
                hora  = r[1] if len(r) > 1 else ""
                if mat not in alunos:
                    aula = _aula_por_hora(hora)
                    alunos[mat] = {"nome": nome, "serie": serie,
                                   "curso": curso, "almoca": False,
                                   "hora": hora, "aula": aula}
                else:
                    if hora and alunos[mat].get("hora", "") and hora < alunos[mat]["hora"]:
                        alunos[mat]["hora"] = hora
                        alunos[mat]["aula"] = _aula_por_hora(hora)
                if ref == "almoco":
                    alunos[mat]["almoca"] = True

            f_serie = fv_serie.get()
            f_curso  = fv_curso.get()
            f_nome   = fv_nome.get().strip().lower()

            exibidos = []
            for info in alunos.values():
                if not _serie_bate(info["serie"], f_serie): continue
                if not _curso_bate(info["curso"], f_curso): continue
                if f_nome and f_nome not in info["nome"].lower(): continue
                exibidos.append(info)

            cnt_total = len(exibidos)
            cnt_sim   = sum(1 for a in exibidos if a["almoca"])
            cnt_nao   = cnt_total - cnt_sim
            lbl_total_ref.config(
                text=f"Total: {cnt_total}   ✔ {cnt_sim}   ✘ {cnt_nao}")

            partes_filtro = []
            if f_serie not in ("(Todas)", ""): partes_filtro.append(f_serie)
            if f_curso  not in ("(Todos)", ""): partes_filtro.append(f_curso)
            if f_nome:                          partes_filtro.append(f'"{f_nome}"')
            lbl_filtro_ativo.config(
                text=("Filtro ativo: " + " | ".join(partes_filtro)) if partes_filtro else "")

            for info in sorted(exibidos, key=lambda a: a["nome"].lower()):
                serie_exib = info["serie"] if info["serie"] else "-"
                curso_exib = info["curso"] if info["curso"] else "-"
                status = "✔ Sim" if info["almoca"] else "✘ Não"
                aula    = info.get("aula", "Fora do horário")
                if info["almoca"]:
                    tref.insert("", "end",
                                values=(info["nome"], serie_exib, curso_exib, status, aula),
                                tags=("almoca",))
                else:
                    tref.insert("", "end",
                                values=(info["nome"], serie_exib, curso_exib, status, aula),
                                tags=("nao_almoca",))

        # ── Limpar todos os filtros ────────────────────────────────────────────
        def limpar_filtros():
            fv_serie.set("(Todas)")
            fv_curso.set("(Todos)")
            fv_nome.set("")
            for sigla, b in btn_sala_refs.items():
                b.configure(relief="flat", bg=CORES_ANO[sigla[0]], fg="white")
            atualizar_ref()

        btn_limpar.configure(command=limpar_filtros)

        # Gatilhos automáticos de filtro
        cb_serie.bind("<<ComboboxSelected>>", lambda e: atualizar_ref())
        cb_curso.bind("<<ComboboxSelected>>", lambda e: atualizar_ref())
        fv_nome.trace_add("write", lambda *_: atualizar_ref())
        
        # ── Painel de Taxa de Adesão por Turma ─────────────────────────────────
        taxa_frame = tk.Frame(aba_ref, bg=T["BG_CARD"])
        taxa_frame.pack(fill="x", padx=20, pady=(10, 0))
        
        tk.Label(taxa_frame, text="Taxa de Adesão por Turma (Hoje)", font=("Segoe UI", 12, "bold"),
                bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", pady=(0, 10))
        
        def _desenhar_taxa_adesao():
            # Limpar widgets anteriores
            for w in taxa_frame.winfo_children()[1:]:
                w.destroy()
            
            # Ler lista_alunos.json
            lista = []
            if os.path.exists(LISTA_ALUNOS_FILE):
                with open(LISTA_ALUNOS_FILE, "r", encoding="utf-8") as f:
                    lista = json.load(f)
            
            if not lista:
                tk.Label(taxa_frame, text="Cadastre alunos na aba QR Codes para ver a taxa de adesão.",
                        font=("Segoe UI", 10), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(pady=10)
                return
            
            # Inicializar todas as turmas padrão para exibir todas elas mesmo sem alunos
            grupos = {}
            turma_ordem = []
            for sigla, label in SIGLAS_LABEL.items():
                info = SIGLAS[sigla]
                chave = f"{info['serie']} - {info['curso']}"
                grupos[chave] = []
                turma_ordem.append((chave, label))
            
            # Agrupar alunos por serie + curso canônicos
            for al in lista:
                serie = _canonizar_serie(al.get("serie", ""))
                curso = _canonizar_curso(al.get("curso", ""))
                chave = f"{serie} - {curso}"
                if chave not in grupos:
                    grupos[chave] = []
                    turma_ordem.append((chave, chave))
                grupos[chave].append(al)
            
            # Contar quem almoçou hoje
            registros_hoje = _registros_hoje()
            almocou = set()
            for r in registros_hoje:
                if len(r) > 6 and r[6].strip().lower() == "almoco":
                    almocou.add(r[2])  # matricula
            
            # Criar cards por turma em grid para manter o layout
            cards_frame = tk.Frame(taxa_frame, bg=T["BG_CARD"])
            cards_frame.pack(fill="x", pady=(0, 10))
            for col in range(4):
                cards_frame.grid_columnconfigure(col, weight=1, uniform="cards")
            
            idx = 0
            for chave, label in turma_ordem:
                total = len(grupos.get(chave, []))
                almocou_count = sum(1 for al in grupos.get(chave, []) if al.get("matricula", "") in almocou)
                pct = (almocou_count / total * 100) if total > 0 else 0
                
                # Determinar cor
                if pct > 70:
                    cor = "#4caf50"  # Verde
                elif pct > 40:
                    cor = "#fdd835"  # Amarelo
                else:
                    cor = "#f44336"  # Vermelho
                
                card = tk.Frame(cards_frame, bg=T["BG_CARD"],
                              highlightbackground=T["BORDER_GRID"], highlightthickness=1, padx=12, pady=10)
                card.grid(row=idx // 4, column=idx % 4, sticky="nsew", padx=6, pady=6)
                idx += 1
                
                tk.Label(card, text=label, font=("Segoe UI", 9, "bold"),
                        bg=T["BG_CARD"], fg=T["FG_TEXT"]).pack(anchor="w")
                
                tk.Label(card, text=f"{almocou_count} / {total}", font=("Segoe UI", 11, "bold"),
                        bg=T["BG_CARD"], fg=cor).pack(anchor="w", pady=(4, 0))
                
                canvas_barra = tk.Canvas(card, bg=T["ENTRY_BG"], highlightthickness=0, height=8)
                canvas_barra.pack(fill="x", pady=(6, 0))
                canvas_barra.create_rectangle(0, 0, 180, 8, fill=T["ENTRY_BG"], outline=T["BORDER_GRID"])
                largura_preenchida = (pct / 100) * 180
                canvas_barra.create_rectangle(0, 0, largura_preenchida, 8, fill=cor, outline="")
                
                tk.Label(card, text=f"{pct:.1f}%", font=("Segoe UI", 9),
                        bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(anchor="w", pady=(4, 0))
        
        def abrir_ausentes():
            win_aus = tk.Toplevel(jd)
            win_aus.title("Alunos Ausentes Hoje")
            win_aus.state("zoomed")
            win_aus.configure(bg=T["BG_MAIN"])
            
            criar_barra_topo(win_aus, "👥  Alunos Ausentes", cmd_voltar=win_aus.destroy)
            
            # Ler lista de alunos
            lista_todos = []
            if os.path.exists(LISTA_ALUNOS_FILE):
                with open(LISTA_ALUNOS_FILE, "r", encoding="utf-8") as f:
                    lista_todos = json.load(f)
            
            # Ler registros de hoje
            registros_hoje = _registros_hoje()
            matriculas_presentes = set(r[2] for r in registros_hoje if len(r) > 2)
            
            # Filtrar ausentes
            ausentes = [al for al in lista_todos if al.get("matricula", "") not in matriculas_presentes]
            
            # Filtros
            filt_frame = tk.Frame(win_aus, bg=T["OBS_BG"])
            filt_frame.pack(fill="x", padx=20, pady=(10, 0))
            
            fv_serie_aus = tk.StringVar()
            fv_curso_aus = tk.StringVar()
            
            tk.Label(filt_frame, text="Filtrar:", font=("Segoe UI", 10, "bold"),
                    bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(side="left", padx=(10, 8), pady=8)
            
            tk.Label(filt_frame, text="Série:", font=("Segoe UI", 9),
                    bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
            cb_serie_aus = ttk.Combobox(filt_frame, textvariable=fv_serie_aus, state="readonly",
                        values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                        width=9, font=("Segoe UI", 10))
            cb_serie_aus.pack(side="left", padx=(4, 14), pady=6)
            cb_serie_aus.set("(Todas)")
            
            tk.Label(filt_frame, text="Curso:", font=("Segoe UI", 9),
                    bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
            cb_curso_aus = ttk.Combobox(filt_frame, textvariable=fv_curso_aus, state="readonly",
                        values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem",
                               "Enfermagem", "Modelagem do Vestuario"],
                        width=26, font=("Segoe UI", 10))
            cb_curso_aus.pack(side="left", padx=(4, 14), pady=6)
            cb_curso_aus.set("(Todos)")
            
            # Treeview ausentes
            ft_aus = tk.Frame(win_aus, bg=T["BORDER_GRID"], padx=1, pady=1)
            ft_aus.pack(expand=True, fill="both", padx=20, pady=(10, 15))
            
            taus = ttk.Treeview(ft_aus, columns=("Nome", "Série", "Curso"), show="headings")
            taus.heading("Nome", text="NOME")
            taus.column("Nome", width=300, anchor="w")
            taus.heading("Série", text="SÉRIE")
            taus.column("Série", width=100, anchor="center")
            taus.heading("Curso", text="CURSO")
            taus.column("Curso", width=250, anchor="w")
            
            sb_aus = ttk.Scrollbar(ft_aus, orient="vertical", command=taus.yview)
            taus.configure(yscrollcommand=sb_aus.set)
            taus.pack(side="left", expand=True, fill="both")
            sb_aus.pack(side="right", fill="y")
            
            taus.tag_configure("even", background=T["ENTRY_BG"])
            taus.tag_configure("odd", background=T["BG_CARD"])
            
            def atualizar_ausentes():
                for i in taus.get_children(): taus.delete(i)
                
                f_serie = fv_serie_aus.get()
                f_curso = fv_curso_aus.get()
                
                exibidos = []
                for al in ausentes:
                    if not _serie_bate(al.get("serie", ""), f_serie): continue
                    if not _curso_bate(al.get("curso", ""), f_curso): continue
                    exibidos.append(al)
                
                for idx, al in enumerate(sorted(exibidos, key=lambda a: a.get("nome", ""))):
                    tag = "even" if idx % 2 == 0 else "odd"
                    taus.insert("", "end", values=(al.get("nome", ""), al.get("serie", ""), al.get("curso", "")), tags=(tag,))
                
                lbl_aus_cnt.config(text=f"{len(exibidos)} ausente(s) de {len(lista_todos)} cadastrado(s)")
            
            def exportar_ausentes():
                try:
                    nome_arq = f"ausentes_marwin_{_hoje().replace('/', '_')}.csv"
                    with open(nome_arq, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        w.writerow(["Nome", "Série", "Curso"])
                        for item in taus.get_children():
                            vals = taus.item(item)["values"]
                            w.writerow(vals)
                    messagebox.showinfo("Sucesso", f"CSV exportado:\n{nome_arq}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")
            
            btn_aus_frame = tk.Frame(win_aus, bg=T["BG_CARD"])
            btn_aus_frame.pack(fill="x", padx=20, pady=(0, 10))
            
            tk.Button(btn_aus_frame, text="Exportar CSV", command=exportar_ausentes,
                     bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI", 10, "bold"),
                     padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
            
            lbl_aus_cnt = tk.Label(btn_aus_frame, text="0 ausente(s)", font=("Segoe UI", 10),
                                   bg=T["BG_CARD"], fg=T["FRASE_FG"])
            lbl_aus_cnt.pack(side="left")
            
            # Gatilhos
            fv_serie_aus.trace_add("write", lambda *_: atualizar_ausentes())
            fv_curso_aus.trace_add("write", lambda *_: atualizar_ausentes())
            
            atualizar_ausentes()

        atualizar_ref()
        _desenhar_taxa_adesao()

        # Auto-refresh a cada 10 segundos
        def _auto_update():
            if aba_ref.winfo_exists():
                atualizar_ref()
                _desenhar_taxa_adesao()
                aba_ref.after(10000, _auto_update)
        aba_ref.after(10000, _auto_update)

    # ── ABA QR CODES ─────────────────────────────────────────────────────────
    def setup_qrcodes():
        for w in aba_qr.winfo_children(): w.destroy()

        LISTA_FILE = os.path.join(DADOS_DIR, "lista_alunos.json")
        QR_DIR     = "qrcodes_marwin"
        os.makedirs(QR_DIR, exist_ok=True)

        def _ler_lista():
            if os.path.exists(LISTA_FILE):
                with open(LISTA_FILE,"r",encoding="utf-8") as f:
                    return json.load(f)
            return []

        def _salvar_lista(lista):
            with open(LISTA_FILE,"w",encoding="utf-8") as f:
                json.dump(lista, f, ensure_ascii=False, indent=2)

        def _extrair_ano_serie(serie):
            if not serie:
                return ""
            serie_strip = serie.strip()
            por_extenso = {
                "primeiro": "1", "primera": "1", "1o": "1",
                "segundo":  "2", "segunda":  "2", "2o": "2",
                "terceiro": "3", "terceira": "3", "3o": "3",
            }
            serie_lower = serie_strip.lower()
            for chave, val in por_extenso.items():
                if chave in serie_lower:
                    return val
            for ch in serie_strip:
                if ch.isdigit():
                    return ch
            return ""

        def _limpar_texto_pasta(texto):
            invalidos = r'\/:*?"<>|'
            resultado = ""
            for ch in str(texto):
                resultado += "_" if ch in invalidos else ch
            return resultado.strip()

        def _limpar_texto_arquivo(texto):
            return (str(texto)
                    .replace(" ", "_")
                    .replace("/", "_")
                    .replace("\\", "_")
                    .replace("º", "")
                    .replace("°", "")
                    .replace(":", "_")
                    .replace("*", "_")
                    .replace("?", "_")
                    .replace('"', "_")
                    .replace("<", "_")
                    .replace(">", "_")
                    .replace("|", "_")
                    .strip("_"))

        def _pasta_turma(al):
            serie = al.get("serie", "").strip()
            curso = al.get("curso", "").strip()
            ano = _extrair_ano_serie(serie)
            nome_serie = f"{ano} Ano" if ano else "Sem Serie"
            nome_curso = _limpar_texto_pasta(curso) if curso else "Sem Curso"
            pasta = os.path.join(QR_DIR, nome_serie, nome_curso)
            os.makedirs(pasta, exist_ok=True)
            return pasta

        def _nome_arquivo(al):
            matricula = _limpar_texto_arquivo(al.get("matricula", "sem_matricula"))
            nome      = _limpar_texto_arquivo(al.get("nome",      "sem_nome"))
            serie     = _limpar_texto_arquivo(al.get("serie",     ""))
            curso     = _limpar_texto_arquivo(al.get("curso",     ""))
            partes = [matricula, nome]
            if serie: partes.append(serie)
            if curso: partes.append(curso)
            nome_arq = "_".join(partes) + ".png"
            pasta    = _pasta_turma(al)
            return os.path.join(pasta, nome_arq)

        def _gerar_png(al):
            payload = json.dumps({
                "matricula": al["matricula"],
                "nome":      al["nome"],
                "serie":     al.get("serie", ""),
                "curso":     al.get("curso", "")
            }, ensure_ascii=False)
            qr = qrcode.QRCode(
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10, border=4)
            qr.add_data(payload)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            caminho = _nome_arquivo(al)
            img.save(caminho)
            return caminho

        corpo = tk.Frame(aba_qr, bg=T["BG_CARD"]); corpo.pack(fill="both", expand=True)

        esq = tk.Frame(corpo, bg=T["OBS_BG"], width=290)
        esq.pack(side="left", fill="y"); esq.pack_propagate(False)

        tk.Label(esq, text="Novo Aluno", font=("Segoe UI",13,"bold"),
                 bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(pady=(18,2))
        tk.Frame(esq, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=18, pady=(10,12))

        campos = {}
        defs = [("Matricula *","matricula","2026001"),
                ("Nome *","nome","Nome do Aluno"),
                ("Serie  (ex: 1 Ano, 2 Ano, 3 Ano)","serie","1 Ano"),
                ("Curso","curso","Desenvolvimento de Sistemas")]
        for lbl_txt, key, ph in defs:
            tk.Label(esq, text=lbl_txt, font=("Segoe UI",9,"bold"),
                     bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=18)
            e = tk.Entry(esq, font=("Segoe UI",11), relief="solid", bd=1,
                         bg=T["BG_CARD"], fg=T["FG_TEXT"],
                         insertbackground=T["ACCENT_VIBRANT"])
            e.insert(0, ph)
            e.pack(fill="x", padx=18, ipady=5, pady=(2,8))
            campos[key] = e

        tk.Frame(esq, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=18, pady=(4,8))
        dica_frame = tk.Frame(esq, bg=T["ENTRY_BG"],
                              highlightbackground=T["BORDER_GRID"], highlightthickness=1)
        dica_frame.pack(fill="x", padx=18, pady=(0,8))
        tk.Label(dica_frame, text="Pasta gerada:", font=("Segoe UI",8,"bold"),
                 bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=8, pady=(6,0))
        dica_lbl = tk.Label(dica_frame,
                            text="qrcodes_marwin/\n  1 Ano/\n    Desenvolvimento de Sistemas/",
                            font=("Courier",8), bg=T["ENTRY_BG"], fg=T["FRASE_FG"],
                            justify="left")
        dica_lbl.pack(anchor="w", padx=8, pady=(2,6))

        def _atualizar_dica(*_):
            serie = campos["serie"].get().strip()
            curso = campos["curso"].get().strip()
            ano   = _extrair_ano_serie(serie)
            nome_s = f"{ano} Ano" if ano else "Sem Serie"
            nome_c = _limpar_texto_pasta(curso) if curso else "Sem Curso"
            dica_lbl.config(text=f"qrcodes_marwin/\n  {nome_s}/\n    {nome_c}/")

        campos["serie"].bind("<KeyRelease>", _atualizar_dica)
        campos["curso"].bind("<KeyRelease>", _atualizar_dica)

        preview_lbl = tk.Label(esq, text="QR Code preview", font=("Segoe UI",9),
                       bg=T["OBS_BG"], fg=T["FRASE_FG"],
                       relief="solid", bd=1, justify="center", anchor="center",
                       wraplength=260)
        preview_lbl.pack(pady=6)
        _img_ref = {}

        def _preview_qr(al):
            payload = json.dumps({
                "matricula": al["matricula"], "nome": al["nome"],
                "serie": al.get("serie",""), "curso": al.get("curso","")
            }, ensure_ascii=False)
            qr2 = qrcode.QRCode(
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=6, border=2)
            qr2.add_data(payload); qr2.make(fit=True)
            img_pil = qr2.make_image(fill_color="black", back_color="white").convert("RGB")
            img_pil = img_pil.resize((250,250), Image.NEAREST)
            img_tk  = ImageTk.PhotoImage(img_pil)
            _img_ref["img"] = img_tk
            preview_lbl.config(image=img_tk, text="")

        def _campos_para_dict():
            return {
                "matricula": campos["matricula"].get().strip(),
                "nome":      campos["nome"].get().strip(),
                "serie":     campos["serie"].get().strip(),
                "curso":     campos["curso"].get().strip(),
            }

        def adicionar_aluno():
            al = _campos_para_dict()
            if not al["matricula"] or not al["nome"]:
                messagebox.showwarning("Campos obrigatorios","Informe Matricula e Nome."); return
            lista = _ler_lista()
            if any(a["matricula"]==al["matricula"] for a in lista):
                messagebox.showwarning("Duplicado",
                    f"Matricula {al['matricula']} ja esta cadastrada."); return
            lista.append(al)
            _salvar_lista(lista)
            caminho = _gerar_png(al)
            _preview_qr(al)
            _atualizar_dica()
            carregar_lista()
            messagebox.showinfo("Aluno adicionado",
                f"QR Code gerado em:\n{caminho}")

        def reemitir_selecionado():
            sel = tlista.selection()
            if not sel:
                messagebox.showwarning("Aviso","Selecione um aluno na lista."); return
            val = tlista.item(sel[0],"values")
            lista = _ler_lista()
            al = next((a for a in lista if a["matricula"]==val[1]), None)
            if not al: return
            caminho = _gerar_png(al)
            _preview_qr(al)
            for key, idx in [("matricula",1),("nome",2),("serie",3),("curso",4)]:
                campos[key].delete(0,"end"); campos[key].insert(0, val[idx] if idx < len(val) else "")
            _atualizar_dica()
            messagebox.showinfo("QR Reemitido",
                f"QR Code de {al['nome']} regenerado em:\n{caminho}")

        tk.Button(esq, text="Adicionar e Gerar QR",
                  command=adicionar_aluno,
                  bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI",10,"bold"),
                  pady=9, bd=0, cursor="hand2").pack(fill="x", padx=18, pady=(0,4))
        tk.Button(esq, text="Reemitir QR Selecionado",
                  command=reemitir_selecionado,
                  bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI",10,"bold"),
                  pady=9, bd=0, cursor="hand2").pack(fill="x", padx=18, pady=(0,4))

        mid = tk.Frame(corpo, bg=T["BG_CARD"]); mid.pack(side="left", fill="both", expand=True)
        _configurar_layout_portrait(corpo, esq, mid)

        bar = tk.Frame(mid, bg=T["BG_CARD"]); bar.pack(fill="x", padx=16, pady=(14,6))
        tk.Label(bar, text="Alunos Cadastrados",
                 font=("Segoe UI",13,"bold"), bg=T["BG_CARD"],
                 fg=T["ACCENT_VIBRANT"]).pack(side="left")
        lbl_cnt2 = tk.Label(bar, text="", font=("Segoe UI",9),
                            bg=T["BG_CARD"], fg=T["FRASE_FG"])
        lbl_cnt2.pack(side="left", padx=10)

        busca_row = tk.Frame(mid, bg=T["BG_CARD"]); busca_row.pack(fill="x", padx=16, pady=(0,6))
        tk.Label(busca_row, text="Buscar:", font=("Segoe UI",11),
                 bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left")
        busca_var = tk.StringVar()
        entry_busca = tk.Entry(busca_row, textvariable=busca_var, font=("Segoe UI",11),
                               bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1,
                               insertbackground=T["ACCENT_VIBRANT"])
        entry_busca.pack(side="left", fill="x", expand=True, padx=(6,0), ipady=5)

        tk.Frame(mid, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=16, pady=(0,6))

        ft2 = tk.Frame(mid, bg=T["BORDER_GRID"], padx=1, pady=1)
        ft2.pack(expand=True, fill="both", padx=16)

        cols_t = ("QR","M","N","S","A","C")
        hdrs_t = ("QR?","MATRIC.","NOME","SERIE","ANO","CURSO")
        wids_t = (45,   90,       200,   120,    60,   180)
        tlista  = ttk.Treeview(ft2, columns=cols_t, show="headings")
        for c,h,w in zip(cols_t, hdrs_t, wids_t):
            tlista.heading(c, text=h)
            tlista.column(c, width=w, anchor="center" if w<=90 else "w")
        tlista.column("A", anchor="center")
        sb2 = ttk.Scrollbar(ft2, orient="vertical", command=tlista.yview)
        sbx = ttk.Scrollbar(ft2, orient="horizontal", command=tlista.xview)
        tlista.configure(yscrollcommand=sb2.set, xscrollcommand=sbx.set)
        tlista.pack(side="left", expand=True, fill="both")
        sb2.pack(side="right", fill="y")
        sbx.pack(side="bottom", fill="x")
        tlista.tag_configure("even",  background=T["ENTRY_BG"])
        tlista.tag_configure("odd",   background=T["BG_CARD"])
        tlista.tag_configure("semqr", background="#fff9c4")

        btn_row3 = tk.Frame(mid, bg=T["BG_CARD"]); btn_row3.pack(fill="x", padx=16, pady=8)

        def remover_aluno():
            sel = tlista.selection()
            if not sel:
                messagebox.showwarning("Aviso","Selecione um aluno."); return
            val = tlista.item(sel[0],"values")
            if not messagebox.askyesno("Confirmar",
                    f"Remover {val[2]} da lista?\n(O arquivo PNG nao sera apagado.)"):
                return
            lista = _ler_lista()
            lista = [a for a in lista if a["matricula"] != val[1]]
            _salvar_lista(lista); carregar_lista()

        def gerar_lote():
            lista = _ler_lista()
            if not lista:
                messagebox.showwarning("Lista vazia","Nenhum aluno cadastrado."); return
            if not messagebox.askyesno("Confirmar",
                    f"Gerar/atualizar QR Codes para {len(lista)} aluno(s)?\n\n"
                    f"Estrutura de pastas:\n"
                    f"  qrcodes_marwin/\n"
                    f"    1 Ano/\n"
                    f"      Desenvolvimento de Sistemas/\n"
                    f"    2 Ano/\n"
                    f"      Redes de Computadores/\n"
                    f"    ..."):
                return
            erros = 0
            pastas_criadas = set()
            for al in lista:
                try:
                    _gerar_png(al)
                    pastas_criadas.add(_pasta_turma(al))
                except Exception:
                    erros += 1
            carregar_lista()
            msg = (f"{len(lista)-erros} QR Code(s) gerados em "
                   f"{len(pastas_criadas)} pasta(s) dentro de '{QR_DIR}/'.")
            if erros: msg += f"\n{erros} erro(s)."
            messagebox.showinfo("Lote concluido", msg)

        def abrir_pasta():
            import subprocess as sp
            try:
                pasta = os.path.abspath(QR_DIR)
                if sys.platform == "win32":
                    sp.Popen(["explorer", pasta])
                elif sys.platform == "darwin":
                    sp.Popen(["open", pasta])
                else:
                    sp.Popen(["xdg-open", pasta])
            except Exception as e:
                messagebox.showerror("Erro", str(e))

        tk.Button(btn_row3, text="Remover", command=remover_aluno,
                  bg="#c62828", fg="white", font=("Segoe UI",9,"bold"),
                  padx=12, pady=7, bd=0, cursor="hand2").pack(side="left", padx=(0,6))
        tk.Button(btn_row3, text="Gerar QRs (lote)", command=gerar_lote,
                  bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI",9,"bold"),
                  padx=12, pady=7, bd=0, cursor="hand2").pack(side="left", padx=(0,6))
        tk.Button(btn_row3, text="Abrir Pasta QRs", command=abrir_pasta,
                  bg=T["BTN_VOLTAR"], fg="white", font=("Segoe UI",9,"bold"),
                  padx=12, pady=7, bd=0, cursor="hand2").pack(side="left")

        imp = tk.Frame(corpo, bg=T["OBS_BG"], width=280)
        imp.pack(side="right", fill="y"); imp.pack_propagate(False)
        _configurar_layout_portrait(corpo, imp, esq, lado="right")

        tk.Label(imp, text="Importar Planilha", font=("Segoe UI",13,"bold"),
                 bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(pady=(18,2))
        tk.Label(imp, text="XLSX ou CSV com os alunos",
                 font=("Segoe UI",9), bg=T["OBS_BG"], fg=T["FRASE_FG"]).pack(pady=(2,0))
        tk.Frame(imp, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=18, pady=(10,12))

        frame_inst = tk.Frame(imp, bg=T["ENTRY_BG"],
                              highlightbackground=T["BORDER_GRID"], highlightthickness=1)
        frame_inst.pack(fill="x", padx=18, pady=(0,10))
        tk.Label(frame_inst, text="Colunas esperadas:",
                 font=("Segoe UI",8,"bold"), bg=T["ENTRY_BG"],
                 fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=10, pady=(6,2))
        for col_txt in ["matricula  (obrigatorio)",
                    "nome       (obrigatorio)",
                    "serie      (ex: 1 Ano, 2 Ano)",
                    "curso      (ex: Des. Sistemas)"]:
            tk.Label(frame_inst, text=f"  * {col_txt}",
                     font=("Segoe UI",8), bg=T["ENTRY_BG"],
                     fg=T["FG_TEXT"], justify="left").pack(anchor="w", padx=10)
        tk.Label(frame_inst,
                 text="\nEstrutura gerada:\nqrcodes_marwin/\n"
                      "  1 Ano/\n"
                      "    Desenvolvimento de Sistemas/\n"
                      "  2 Ano/\n"
                      "    Redes de Computadores/\n"
                      "  3 Ano/\n"
                      "    Informatica/",
                 font=("Segoe UI",8,"italic"), bg=T["ENTRY_BG"],
                 fg=T["FRASE_FG"], justify="left").pack(anchor="w", padx=10, pady=(4,8))

        tk.Label(imp, text="Mapear colunas (opcional):",
                 font=("Segoe UI",8,"bold"), bg=T["OBS_BG"],
                 fg=T["ACCENT_VIBRANT"]).pack(anchor="w", padx=18)
        tk.Label(imp, text="Deixe em branco p/ deteccao automatica.",
                 font=("Segoe UI",8,"italic"), bg=T["OBS_BG"],
                 fg=T["FRASE_FG"]).pack(anchor="w", padx=18, pady=(0,6))

        map_vars = {}
        for campo_mk, rotulo_mk in [("matricula","Col. Matricula"),
                                    ("nome","Col. Nome"),
                                    ("serie","Col. Serie"),
                                    ("curso","Col. Curso")]:
            fr = tk.Frame(imp, bg=T["OBS_BG"]); fr.pack(fill="x", padx=18, pady=(0,4))
            tk.Label(fr, text=rotulo_mk, font=("Segoe UI",8),
                     bg=T["OBS_BG"], fg=T["FG_TEXT"], width=12, anchor="w").pack(side="left")
            v = tk.StringVar()
            tk.Entry(fr, textvariable=v, font=("Segoe UI",9),
                     bg=T["BG_CARD"], fg=T["FG_TEXT"], relief="solid", bd=1,
                     width=14).pack(side="left", ipady=3)
            map_vars[campo_mk] = v

        tk.Frame(imp, bg=T["BORDER_GRID"], height=1).pack(fill="x", padx=18, pady=(10,10))

        gerar_qr_import_var = tk.BooleanVar(value=True)
        tk.Checkbutton(imp, text="Gerar QR Codes ao importar",
                       variable=gerar_qr_import_var,
                       font=("Segoe UI",9,"bold"), bg=T["OBS_BG"],
                       fg=T["ACCENT_VIBRANT"], activebackground=T["OBS_BG"],
                       selectcolor=T["ENTRY_BG"], cursor="hand2").pack(anchor="w", padx=18)

        sobreescrever_var = tk.BooleanVar(value=False)
        tk.Checkbutton(imp, text="Atualizar duplicatas",
                       variable=sobreescrever_var,
                       font=("Segoe UI",9), bg=T["OBS_BG"],
                       fg=T["FG_TEXT"], activebackground=T["OBS_BG"],
                       selectcolor=T["ENTRY_BG"], cursor="hand2").pack(anchor="w", padx=18)

        lbl_import_status = tk.Label(imp, text="", font=("Segoe UI",9,"bold"),
                                     bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"],
                                     wraplength=240, justify="center")
        lbl_import_status.pack(pady=8, padx=18)

        prog_frame = tk.Frame(imp, bg=T["OBS_BG"]); prog_frame.pack(fill="x", padx=18, pady=(0,10))
        prog_bar = ttk.Progressbar(prog_frame, orient="horizontal",
                                   mode="determinate", length=220)
        prog_bar.pack(fill="x")

        def _normalizar_cabecalho(cabecalho):
            mapa_auto = {}
            sinonimos = {
                "matricula": ["matricula","matricula","mat","mat.","codigo","id","registro"],
                "nome":      ["nome","aluno","estudante","discente","name","nomecompleto"],
                "serie":     ["serie","serie","turma","ano","class","classe","periodo","ano/serie"],
                "curso":     ["curso","habilitacao","area","modalidade","formacao"],
            }
            cab_lower = [str(c).strip().lower() for c in cabecalho]
            for campo, sinonimos_lista in sinonimos.items():
                manual = map_vars[campo].get().strip()
                if manual and manual in cabecalho:
                    mapa_auto[campo] = cabecalho.index(manual); continue
                for s in sinonimos_lista:
                    for idx_c, c in enumerate(cab_lower):
                        if s in c:
                            mapa_auto[campo] = idx_c; break
                    if campo in mapa_auto: break
            return mapa_auto

        def importar_planilha():
            from tkinter import filedialog
            caminho_pl = filedialog.askopenfilename(
                title="Selecionar planilha de alunos",
                filetypes=[("Planilhas","*.xlsx *.xls *.csv *.tsv"),
                           ("Excel","*.xlsx *.xls"),
                           ("CSV / TSV","*.csv *.tsv"),
                           ("Todos","*.*")])
            if not caminho_pl:
                return

            lbl_import_status.config(text="Lendo arquivo...", fg=T["FRASE_FG"])
            aba_qr.update()

            ext = os.path.splitext(caminho_pl)[1].lower()
            linhas_raw = []

            try:
                if ext in (".xlsx", ".xls"):
                    try:
                        import openpyxl
                    except ImportError:
                        lbl_import_status.config(text="Instalando openpyxl...", fg=T["FRASE_FG"])
                        aba_qr.update()
                        subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
                        import openpyxl
                    wb = openpyxl.load_workbook(caminho_pl, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        linhas_raw.append([str(c).strip() if c is not None else "" for c in row])
                    wb.close()

                elif ext in (".csv", ".tsv"):
                    sep = "\t" if ext == ".tsv" else None
                    for enc in ("utf-8-sig","utf-8","latin-1","cp1252"):
                        try:
                            with open(caminho_pl,"r",encoding=enc,newline="") as f:
                                sample = f.read(4096); f.seek(0)
                                if sep is None:
                                    try:    sep = csv.Sniffer().sniff(sample).delimiter
                                    except: sep = ","
                                reader_pl = csv.reader(f, delimiter=sep)
                                linhas_raw = [[c.strip() for c in row] for row in reader_pl]
                            break
                        except (UnicodeDecodeError, Exception):
                            continue
                else:
                    messagebox.showerror("Formato nao suportado","Use .xlsx, .xls, .csv ou .tsv"); return

            except Exception as e:
                messagebox.showerror("Erro ao ler arquivo", str(e)); return

            if len(linhas_raw) < 2:
                messagebox.showwarning("Arquivo vazio","O arquivo nao tem dados suficientes."); return

            cabecalho = linhas_raw[0]
            mapa = _normalizar_cabecalho(cabecalho)

            if "matricula" not in mapa or "nome" not in mapa:
                messagebox.showerror("Colunas nao encontradas",
                    "Nao foi possivel identificar as colunas de Matricula e Nome.\n"
                    "Use os campos de mapeamento manual ou renomeie as colunas\n"
                    "para 'matricula' e 'nome'."); return

            lista_atual = _ler_lista()
            mats_existentes = {a["matricula"]: i for i, a in enumerate(lista_atual)}

            novos = 0; atualizados = 0; ignorados = 0
            dados_importados = []

            linhas_dados = [l for l in linhas_raw[1:] if any(c for c in l)]
            prog_bar["maximum"] = max(len(linhas_dados), 1)
            prog_bar["value"]   = 0

            for idx_linha, linha in enumerate(linhas_dados):
                prog_bar["value"] = idx_linha + 1
                aba_qr.update_idletasks()

                def _cel(campo, _linha=linha):
                    idx_c = mapa.get(campo)
                    if idx_c is None or idx_c >= len(_linha): return ""
                    return str(_linha[idx_c]).strip()

                matricula = _cel("matricula")
                nome      = _cel("nome")
                serie     = _cel("serie")
                curso     = _cel("curso")

                if not matricula or not nome:
                    ignorados += 1; continue

                al = {"matricula": matricula, "nome": nome,
                      "serie": serie, "curso": curso}

                if matricula in mats_existentes:
                    if sobreescrever_var.get():
                        lista_atual[mats_existentes[matricula]] = al
                        atualizados += 1
                        dados_importados.append(al)
                    else:
                        ignorados += 1
                else:
                    lista_atual.append(al)
                    mats_existentes[matricula] = len(lista_atual) - 1
                    novos += 1
                    dados_importados.append(al)

            _salvar_lista(lista_atual)

            if gerar_qr_import_var.get() and dados_importados:
                lbl_import_status.config(
                    text=f"Gerando {len(dados_importados)} QR Code(s)...",
                    fg=T["FRASE_FG"])
                prog_bar["maximum"] = len(dados_importados)
                prog_bar["value"]   = 0
                erros_qr = 0
                pastas_criadas = set()
                for idx_al, al in enumerate(dados_importados):
                    prog_bar["value"] = idx_al + 1
                    aba_qr.update_idletasks()
                    try:
                        _gerar_png(al)
                        pastas_criadas.add(_pasta_turma(al))
                    except Exception:
                        erros_qr += 1
                msg_qr = (f"\n{len(dados_importados)-erros_qr} QR Code(s) gerados "
                          f"em {len(pastas_criadas)} pasta(s) dentro de '{QR_DIR}/'.")
                if erros_qr: msg_qr += f" ({erros_qr} erros)"
            else:
                msg_qr = ""

            prog_bar["value"] = 0
            carregar_lista()

            resumo = (f"Importacao concluida!\n"
                      f"Novos: {novos}  |  Atualizados: {atualizados}"
                      f"  |  Ignorados: {ignorados}{msg_qr}")
            lbl_import_status.config(text=resumo, fg=T["ACCENT_VIBRANT"])
            messagebox.showinfo("Importacao concluida", resumo)

        tk.Button(imp, text="Selecionar Arquivo e Importar",
                  command=importar_planilha,
                  bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI",11,"bold"),
                  pady=12, bd=0, cursor="hand2",
                  wraplength=220).pack(fill="x", padx=18, pady=(0,6))

        tk.Button(imp, text="Abrir Pasta dos QR Codes",
                  command=abrir_pasta,
                  bg=T["BTN_VOLTAR"], fg="white", font=("Segoe UI",9,"bold"),
                  pady=8, bd=0, cursor="hand2").pack(fill="x", padx=18)

        def carregar_lista(filtro=""):
            for i in tlista.get_children(): tlista.delete(i)
            lista = _ler_lista()
            filtro_lower = filtro.lower()
            exibidos = 0
            for al in lista:
                serie = al.get("serie", "")
                curso = al.get("curso", "")
                if filtro_lower and filtro_lower not in al["nome"].lower() \
                        and filtro_lower not in al["matricula"].lower() \
                        and filtro_lower not in curso.lower() \
                        and filtro_lower not in serie.lower():
                    continue
                tem_png = os.path.exists(_nome_arquivo(al))
                icone_qr = "OK" if tem_png else "X"
                tag = ("even" if exibidos%2==0 else "odd") if tem_png else "semqr"
                ano = _extrair_ano_serie(serie)
                ano_exib = f"{ano}º Ano" if ano else "-"
                tlista.insert("","end",
                    values=(icone_qr,
                            al.get("matricula",""),
                            al.get("nome",""),
                            serie,
                            ano_exib,
                            curso),
                    tags=(tag,))
                exibidos += 1
            total = len(lista)
            lbl_cnt2.config(text=f"({exibidos} de {total} aluno(s))")

        busca_var.trace_add("write", lambda *_: carregar_lista(busca_var.get()))

        def preencher_campos_ao_clicar(event):
            sel = tlista.selection()
            if not sel: return
            val = tlista.item(sel[0],"values")
            mapa_campos = [("matricula",1),("nome",2),("serie",3),("curso",5)]
            for key, vi in mapa_campos:
                campos[key].delete(0,"end")
                campos[key].insert(0, val[vi] if vi < len(val) else "")
            _atualizar_dica()
            al_dict = {k: campos[k].get() for k in ["matricula","nome","serie","curso"]}
            _preview_qr(al_dict)

        tlista.bind("<<TreeviewSelect>>", preencher_campos_ao_clicar)
        carregar_lista()

    # ── Eventos de troca de aba ───────────────────────────────────────────────
    def on_tab(event):
        idx = nb.index("current")
        if   idx == 1: gerar_relatorio()
        elif idx == 2: setup_cardapio()
        elif idx == 3: setup_eventos()
        elif idx == 4: setup_refeitorio()
        elif idx == 5: setup_frequencia()
        elif idx == 6: setup_historico()
        elif idx == 7: setup_qrcodes()
        elif idx == 8: setup_logs()

    nb.bind("<<NotebookTabChanged>>", on_tab)
    setup_cardapio(); setup_eventos()

    # ── ABA FREQUÊNCIA ────────────────────────────────────────────────────────
    def setup_frequencia():
        for w in aba_freq.winfo_children(): w.destroy()

        SIGLAS = {
            "1DS":  {"serie": "1 Ano", "curso": "Desenvolvimento de Sistemas"},
            "1HOS": {"serie": "1 Ano", "curso": "Hospedagem"},
            "1ENF": {"serie": "1 Ano", "curso": "Enfermagem"},
            "1MOD": {"serie": "1 Ano", "curso": "Modelagem do Vestuario"},
            "2DS":  {"serie": "2 Ano", "curso": "Desenvolvimento de Sistemas"},
            "2HOS": {"serie": "2 Ano", "curso": "Hospedagem"},
            "2ENF": {"serie": "2 Ano", "curso": "Enfermagem"},
            "2MOD": {"serie": "2 Ano", "curso": "Modelagem do Vestuario"},
            "3DS":  {"serie": "3 Ano", "curso": "Desenvolvimento de Sistemas"},
            "3HOS": {"serie": "3 Ano", "curso": "Hospedagem"},
            "3ENF": {"serie": "3 Ano", "curso": "Enfermagem"},
            "3MOD": {"serie": "3 Ano", "curso": "Modelagem do Vestuario"},
        }
        SIGLAS_LABEL = {
            "1DS": "1º DS",  "1HOS": "1º HOS", "1ENF": "1º ENF", "1MOD": "1º MOD",
            "2DS": "2º DS",  "2HOS": "2º HOS", "2ENF": "2º ENF", "2MOD": "2º MOD",
            "3DS": "3º DS",  "3HOS": "3º HOS", "3ENF": "3º ENF", "3MOD": "3º MOD",
        }
        CORES_ANO = {"1": "#1565c0", "2": "#6a1b9a", "3": "#c62828"}

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        hd = tk.Frame(aba_freq, bg=T["BG_CARD"]); hd.pack(fill="x", padx=20, pady=(15, 5))
        tk.Label(hd, text="Controle de Frequência - Hoje", font=("Segoe UI", 16, "bold"),
                 bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        tk.Label(hd, text=f"  ({_hoje()})", font=("Segoe UI", 11),
                 bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left")
        tk.Frame(aba_freq, bg=T["BORDER_GRID"], height=2).pack(fill="x", padx=20, pady=(5, 0))

        # ── Linha de filtros por texto / combo ────────────────────────────────
        container_filtros = tk.Frame(aba_freq, bg=T["OBS_BG"])
        container_filtros.pack(fill="x", padx=20, pady=(6, 0))

        filtro_frame = tk.Frame(container_filtros, bg=T["OBS_BG"])
        filtro_frame.pack(fill="x", padx=10, pady=(0, 0))

        tk.Label(filtro_frame, text="Filtrar por:", font=("Segoe UI", 10, "bold"),
                 bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(side="left", padx=(10, 8), pady=8)

        fv_serie_f = tk.StringVar()
        fv_curso_f  = tk.StringVar()
        fv_nome_f   = tk.StringVar()

        tk.Label(filtro_frame, text="Série:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        cb_serie_f = ttk.Combobox(filtro_frame, textvariable=fv_serie_f, state="readonly",
                                values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                width=9, font=("Segoe UI", 10))
        cb_serie_f.set("(Todas)")
        cb_serie_f.pack(side="left", padx=(4, 14), pady=6)

        tk.Label(filtro_frame, text="Curso:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        cursos_opcoes = [
            "(Todos)",
            "Desenvolvimento de Sistemas",
            "Hospedagem",
            "Enfermagem",
            "Modelagem do Vestuario",
        ]
        cb_curso_f = ttk.Combobox(filtro_frame, textvariable=fv_curso_f, state="readonly",
                                values=cursos_opcoes, width=26, font=("Segoe UI", 10))
        cb_curso_f.set("(Todos)")
        cb_curso_f.pack(side="left", padx=(4, 14), pady=6)

        tk.Label(filtro_frame, text="Nome:", font=("Segoe UI", 9),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
        entry_nome_f = tk.Entry(filtro_frame, textvariable=fv_nome_f, font=("Segoe UI", 10),
                              bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1,
                              insertbackground=T["ACCENT_VIBRANT"], width=20)
        entry_nome_f.pack(side="left", ipady=4, padx=(4, 14), pady=6)

        btn_limpar_f = tk.Button(filtro_frame, text="✕ Limpar", font=("Segoe UI", 9, "bold"),
                               bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"], bd=0,
                               cursor="hand2", padx=8, pady=4)
        btn_limpar_f.pack(side="left", padx=(0, 10), pady=6)

        # ── Botões de atalho por sala ──────────────────────────────────────────
        sala_frame_f = tk.Frame(aba_freq, bg=T["BG_CARD"])
        sala_frame_f.pack(fill="x", padx=20, pady=(6, 0))
        tk.Label(sala_frame_f, text="Sala rápida:", font=("Segoe UI", 9, "bold"),
                 bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(side="left", padx=(4, 8))

        btn_sala_refs_f = {}

        def _filtrar_por_sigla_f(sigla):
            info = SIGLAS[sigla]
            fv_serie_f.set(info["serie"])
            fv_curso_f.set(info["curso"])
            fv_nome_f.set("")
            for s, b in btn_sala_refs_f.items():
                b.configure(relief="flat", bg=CORES_ANO[s[0]], fg="white")
            btn_sala_refs_f[sigla].configure(relief="solid", bg="white",
                                           fg=CORES_ANO[sigla[0]])
            atualizar_freq()

        for sigla, label in SIGLAS_LABEL.items():
            ano = sigla[0]
            cor = CORES_ANO[ano]
            b = tk.Button(sala_frame_f, text=label, font=("Segoe UI", 8, "bold"),
                          bg=cor, fg="white", bd=0, padx=9, pady=4,
                          cursor="hand2",
                          command=lambda s=sigla: _filtrar_por_sigla_f(s))
            b.pack(side="left", padx=2, pady=4)
            btn_sala_refs_f[sigla] = b

        # ── Barra de botões e contadores ──────────────────────────────────────
        btn_row_f = tk.Frame(aba_freq, bg=T["BG_CARD"])
        btn_row_f.pack(fill="x", padx=20, pady=(8, 4))

        lbl_total_freq = tk.Label(btn_row_f, text="Total: 0", font=("Segoe UI", 12, "bold"),
                                 bg=T["ENTRY_BG"], fg=T["ACCENT_VIBRANT"], padx=14, pady=6)
        lbl_total_freq.pack(side="left", padx=(0, 10))

        def exportar_csv_freq():
            try:
                nome_arq = f"frequencia_{_hoje().replace('/', '_')}.csv"
                _escrever_csv(
                    nome_arq,
                    ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
                    _ler_frequencia_todos_db(),
                )
                messagebox.showinfo("Exportado", f"Arquivo salvo como:\n{nome_arq}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")

        def apagar_hoje_freq():
            if messagebox.askyesno("Confirmar", f"Apagar todos os registros de hoje ({_hoje()})?"):
                try:
                    _apagar_frequencia_data_db(_hoje())
                    messagebox.showinfo("Sucesso", "Registros de hoje apagados do banco.")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao apagar registros:\n{e}")
                atualizar_freq()

        def abrir_ausentes_freq():
            win_aus = tk.Toplevel(jd)
            win_aus.title("Alunos Ausentes Hoje")
            win_aus.state("zoomed")
            win_aus.configure(bg=T["BG_MAIN"])
            criar_barra_topo(win_aus, "👥  Alunos Ausentes", cmd_voltar=win_aus.destroy)

            lista_todos = []
            if os.path.exists(LISTA_ALUNOS_FILE):
                with open(LISTA_ALUNOS_FILE, "r", encoding="utf-8") as f:
                    lista_todos = json.load(f)

            registros_hoje = _registros_freq_hoje()
            presentes = set(r[2] for r in registros_hoje if len(r) > 2)
            ausentes = [al for al in lista_todos if al.get("matricula", "") not in presentes]

            filt_frame = tk.Frame(win_aus, bg=T["OBS_BG"])
            filt_frame.pack(fill="x", padx=20, pady=(10, 0))

            fv_serie_aus = tk.StringVar(value="(Todas)")
            fv_curso_aus = tk.StringVar(value="(Todos)")

            tk.Label(filt_frame, text="Filtrar:", font=("Segoe UI", 10, "bold"),
                    bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"]).pack(side="left", padx=(10, 8), pady=8)
            tk.Label(filt_frame, text="Série:", font=("Segoe UI", 9),
                    bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
            cb_serie_aus = ttk.Combobox(filt_frame, textvariable=fv_serie_aus, state="readonly",
                        values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                        width=9, font=("Segoe UI", 10))
            cb_serie_aus.pack(side="left", padx=(4, 14), pady=6)

            tk.Label(filt_frame, text="Curso:", font=("Segoe UI", 9),
                    bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left")
            cb_curso_aus = ttk.Combobox(filt_frame, textvariable=fv_curso_aus, state="readonly",
                        values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem",
                                "Enfermagem", "Modelagem do Vestuario"],
                        width=26, font=("Segoe UI", 10))
            cb_curso_aus.pack(side="left", padx=(4, 14), pady=6)

            ft_aus = tk.Frame(win_aus, bg=T["BORDER_GRID"], padx=1, pady=1)
            ft_aus.pack(expand=True, fill="both", padx=20, pady=(10, 15))

            taus = ttk.Treeview(ft_aus, columns=("Nome", "Série", "Curso"), show="headings")
            taus.heading("Nome", text="NOME")
            taus.column("Nome", width=300, anchor="w")
            taus.heading("Série", text="SÉRIE")
            taus.column("Série", width=100, anchor="center")
            taus.heading("Curso", text="CURSO")
            taus.column("Curso", width=250, anchor="w")

            sb_aus = ttk.Scrollbar(ft_aus, orient="vertical", command=taus.yview)
            taus.configure(yscrollcommand=sb_aus.set)
            taus.pack(side="left", expand=True, fill="both")
            sb_aus.pack(side="right", fill="y")

            taus.tag_configure("even", background=T["ENTRY_BG"])
            taus.tag_configure("odd", background=T["BG_CARD"])

            def atualizar_ausentes():
                for i in taus.get_children(): taus.delete(i)
                f_serie = fv_serie_aus.get()
                f_curso = fv_curso_aus.get()
                exibidos = []
                for al in ausentes:
                    if f_serie not in ("(Todas)", "") and not _serie_bate_f(al.get("serie", ""), f_serie):
                        continue
                    if f_curso not in ("(Todos)", "") and not _curso_bate_f(al.get("curso", ""), f_curso):
                        continue
                    exibidos.append(al)
                for idx, al in enumerate(sorted(exibidos, key=lambda a: a.get("nome", ""))):
                    tag = "even" if idx % 2 == 0 else "odd"
                    taus.insert("", "end",
                                values=(al.get("nome", ""), al.get("serie", ""), al.get("curso", "")),
                                tags=(tag,))
                lbl_aus_cnt.config(text=f"{len(exibidos)} ausente(s) de {len(lista_todos)} cadastrado(s)")

            btn_aus_frame = tk.Frame(win_aus, bg=T["BG_CARD"])
            btn_aus_frame.pack(fill="x", padx=20, pady=(0, 10))
            tk.Button(btn_aus_frame, text="⬇  Exportar CSV", command=lambda: exportar_ausentes(),
                     bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI", 10, "bold"),
                     padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
            lbl_aus_cnt = tk.Label(btn_aus_frame, text="0 ausente(s)", font=("Segoe UI", 10),
                                   bg=T["BG_CARD"], fg=T["FRASE_FG"])
            lbl_aus_cnt.pack(side="left")

            def exportar_ausentes():
                try:
                    nome_arq = f"ausentes_frequencia_{_hoje().replace('/', '_')}.csv"
                    with open(nome_arq, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        w.writerow(["Nome", "Série", "Curso"])
                        for item in taus.get_children():
                            vals = taus.item(item)["values"]
                            w.writerow(vals)
                    messagebox.showinfo("Sucesso", f"CSV exportado:\n{nome_arq}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")

            fv_serie_aus.trace_add("write", lambda *_: atualizar_ausentes())
            fv_curso_aus.trace_add("write", lambda *_: atualizar_ausentes())
            atualizar_ausentes()

        tk.Button(btn_row_f, text="👁  Ver Ausentes", command=abrir_ausentes_freq,
                  bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_row_f, text="↻  Atualizar", command=lambda: atualizar_freq(),
                  bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_row_f, text="⬇  Exportar CSV", command=exportar_csv_freq,
                  bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_row_f, text="🗑  Apagar hoje", command=apagar_hoje_freq,
                  bg="#c62828", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, bd=0, cursor="hand2").pack(side="right")

        # ── Tabela: NOME | SÉRIE | CURSO | HORA ENTRADA | AULA ────────────────
        ft_f = tk.Frame(aba_freq, bg=T["BORDER_GRID"], padx=1, pady=1)
        ft_f.pack(expand=True, fill="both", padx=20, pady=(4, 20))

        tfreq = ttk.Treeview(ft_f, columns=("N", "S", "C", "HE", "A"), show="headings")
        tfreq.heading("N",  text="ALUNO");  tfreq.column("N",  width=320, anchor="w")
        tfreq.heading("S",  text="SÉRIE");  tfreq.column("S",  width=90,  anchor="center")
        tfreq.heading("C",  text="CURSO");  tfreq.column("C",  width=220, anchor="w")
        tfreq.heading("HE", text="HORA ENTRADA"); tfreq.column("HE", width=130, anchor="center")
        tfreq.heading("A",  text="AULA");   tfreq.column("A",  width=120, anchor="center")

        sb_freq = ttk.Scrollbar(ft_f, orient="vertical", command=tfreq.yview)
        tfreq.configure(yscrollcommand=sb_freq.set)
        tfreq.pack(side="left", expand=True, fill="both")
        sb_freq.pack(side="right", fill="y")

        tfreq.tag_configure("presente",
                           foreground="#1b5e20", background="#e8f5e9",
                           font=("Segoe UI", 11, "bold"))

        import unicodedata as _ud

        def _norm(texto):
            return _ud.normalize("NFD", str(texto).lower()).encode("ascii", "ignore").decode("ascii")

        _GRUPOS_SERIE = {
            "1": ["1º", "1o", "1 ano", "primeiro ano", "primeiro", "1"],
            "2": ["2º", "2o", "2 ano", "segundo ano",  "segundo",  "2"],
            "3": ["3º", "3o", "3 ano", "terceiro ano", "terceiro", "3"],
        }

        _GRUPOS_CURSO = {
            "ds":   ["ds", "desenvolvimento de sistemas", "dev. sistemas", "dev sistemas", "desenv. sistemas"],
            "enf":  ["enf", "enfermagem"],
            "hosp": ["hosp", "hospedagem"],
            "mod":  ["mod", "modelagem do vestuario", "modelagem", "vestuario"],
        }

        def _grupo_serie_f(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_SERIE.items():
                for v in variantes:
                    if t == v or t.startswith(v) or v.startswith(t):
                        return chave
            return None

        def _grupo_curso_f(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_CURSO.items():
                for v in variantes:
                    if t == v or t in v or v in t:
                        return chave
            return None

        def _serie_bate_f(serie_aluno, filtro_serie):
            if filtro_serie in ("(Todas)", ""):
                return True
            g_aluno  = _grupo_serie_f(serie_aluno)
            g_filtro = _grupo_serie_f(filtro_serie)
            if g_aluno is None or g_filtro is None:
                return _norm(filtro_serie) in _norm(serie_aluno)
            return g_aluno == g_filtro

        def _curso_bate_f(curso_aluno, filtro_curso):
            if filtro_curso in ("(Todos)", ""):
                return True
            g_aluno  = _grupo_curso_f(curso_aluno)
            g_filtro = _grupo_curso_f(filtro_curso)
            if g_aluno is None or g_filtro is None:
                return _norm(filtro_curso) in _norm(curso_aluno)
            return g_aluno == g_filtro

        def _canonizar_serie_f(serie):
            grupo = _grupo_serie_f(serie)
            if grupo == "1":
                return "1 Ano"
            if grupo == "2":
                return "2 Ano"
            if grupo == "3":
                return "3 Ano"
            return serie.strip() if serie else "Sem Serie"

        def _canonizar_curso_f(curso):
            grupo = _grupo_curso_f(curso)
            if grupo == "ds":
                return "Desenvolvimento de Sistemas"
            if grupo == "enf":
                return "Enfermagem"
            if grupo == "hosp":
                return "Hospedagem"
            if grupo == "mod":
                return "Modelagem do Vestuario"
            return curso.strip() if curso else "Sem Curso"

        def atualizar_freq():
            for i in tfreq.get_children(): tfreq.delete(i)
            registros = _registros_freq_hoje()

            alunos_unicos = {}
            for r in registros:
                mat   = r[2]
                nome  = r[3]
                serie = r[4] if len(r) > 4 else ""
                curso = r[5] if len(r) > 5 else ""
                hora  = r[1] if len(r) > 1 else ""
                aula  = r[6] if len(r) > 6 else _aula_por_hora(hora)
                if mat not in alunos_unicos:
                    alunos_unicos[mat] = {"nome": nome, "serie": serie,
                                          "curso": curso, "hora": hora, "aula": aula}
                else:
                    if hora and alunos_unicos[mat].get("hora", "") and hora < alunos_unicos[mat]["hora"]:
                        alunos_unicos[mat]["hora"] = hora
                        alunos_unicos[mat]["aula"] = aula

            f_serie = fv_serie_f.get()
            f_curso  = fv_curso_f.get()
            f_nome   = fv_nome_f.get().strip().lower()

            exibidos = []
            for info in alunos_unicos.values():
                if not _serie_bate_f(info["serie"], f_serie): continue
                if not _curso_bate_f(info["curso"], f_curso): continue
                if f_nome and f_nome not in info["nome"].lower(): continue
                exibidos.append(info)

            cnt_total = len(exibidos)
            lbl_total_freq.config(text=f"Total: {cnt_total}")

            for info in sorted(exibidos, key=lambda a: a["nome"].lower()):
                serie_exib = info["serie"] if info["serie"] else "-"
                curso_exib = info["curso"] if info["curso"] else "-"
                hora_exib = info.get("hora", "")
                aula = info.get("aula", "Fora do horário")
                tfreq.insert("", "end",
                            values=(info["nome"], serie_exib, curso_exib, hora_exib, aula),
                            tags=("presente",))

        def limpar_filtros_f():
            fv_serie_f.set("(Todas)")
            fv_curso_f.set("(Todos)")
            fv_nome_f.set("")
            for sigla, b in btn_sala_refs_f.items():
                b.configure(relief="flat", bg=CORES_ANO[sigla[0]], fg="white")
            atualizar_freq()

        btn_limpar_f.configure(command=limpar_filtros_f)

        cb_serie_f.bind("<<ComboboxSelected>>", lambda e: atualizar_freq())
        cb_curso_f.bind("<<ComboboxSelected>>", lambda e: atualizar_freq())
        fv_nome_f.trace_add("write", lambda *_: atualizar_freq())
        
        atualizar_freq()

    # ── ABA HISTÓRICO ─────────────────────────────────────────────────────────
    def setup_historico():
        for w in aba_hist.winfo_children(): w.destroy()
        
        # Cabeçalho com seletor Refeitório / Frequência
        hd = tk.Frame(aba_hist, bg=T["BG_CARD"]); hd.pack(fill="x", padx=20, pady=(15, 5))
        hd_txt_frame = tk.Frame(hd, bg=T["BG_CARD"]); hd_txt_frame.pack(side="left")
        tk.Label(hd_txt_frame, text="Histórico de ", font=("Segoe UI", 16, "bold"),
                 bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        
        hist_tipo_var = tk.StringVar(value="refeitorio")
        
        def _atualizar_titulo(*args):
            tipo = hist_tipo_var.get()
            titulo_texto = "Refeições" if tipo == "refeitorio" else "Frequência"
            lbl_titulo.config(text=titulo_texto)
            _buscar_historico()
        
        lbl_titulo = tk.Label(hd_txt_frame, text="Refeições", font=("Segoe UI", 16, "bold"),
                             bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"])
        lbl_titulo.pack(side="left")
        
        hd_radio_frame = tk.Frame(hd, bg=T["BG_CARD"]); hd_radio_frame.pack(side="right")
        tk.Radiobutton(hd_radio_frame, text="●  Refeitório", variable=hist_tipo_var, value="refeitorio",
                      bg=T["BG_CARD"], fg=T["FG_TEXT"], selectcolor=T["ACCENT_SOFT"],
                      activebackground=T["BG_CARD"], font=("Segoe UI", 9),
                      command=lambda: _atualizar_titulo()).pack(side="left", padx=(10, 8))
        tk.Radiobutton(hd_radio_frame, text="●  Frequência", variable=hist_tipo_var, value="frequencia",
                      bg=T["BG_CARD"], fg=T["FG_TEXT"], selectcolor=T["ACCENT_SOFT"],
                      activebackground=T["BG_CARD"], font=("Segoe UI", 9),
                      command=lambda: _atualizar_titulo()).pack(side="left", padx=(0, 10))
        
        # Filtros
        filtro_frame = tk.Frame(aba_hist, bg=T["OBS_BG"])
        filtro_frame.pack(fill="x", padx=20, pady=(5, 10))
        
        tk.Label(filtro_frame, text="Período:", font=("Segoe UI", 9, "bold"),
                 bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(10, 6))
        
        periodo_var = tk.StringVar(value="hoje")
        for opt, val in [("Hoje", "hoje"), ("Esta semana", "semana"), ("Este mês", "mes"), ("Personalizado", "custom")]:
            tk.Radiobutton(filtro_frame, text=opt, variable=periodo_var, value=val,
                          bg=T["OBS_BG"], fg=T["FG_TEXT"], selectcolor=T["ACCENT_SOFT"],
                          activebackground=T["OBS_BG"], font=("Segoe UI", 9)).pack(side="left", padx=4)
        
        # Campos de data personalizada
        custom_frame = tk.Frame(filtro_frame, bg=T["OBS_BG"])
        custom_frame.pack(side="left", padx=(10, 0))
        
        data_inicio_var = tk.StringVar(value="01/01/2026")
        data_fim_var = tk.StringVar(value=datetime.date.today().strftime("%d/%m/%Y"))
        
        tk.Label(custom_frame, text="De:", font=("Segoe UI", 9),
                bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(4, 2))
        tk.Entry(custom_frame, textvariable=data_inicio_var, font=("Segoe UI", 9),
                bg=T["BG_CARD"], fg=T["FG_TEXT"], relief="solid", bd=1, width=12).pack(side="left", ipady=2, padx=(0, 8))
        
        tk.Label(custom_frame, text="Até:", font=("Segoe UI", 9),
                bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(4, 2))
        tk.Entry(custom_frame, textvariable=data_fim_var, font=("Segoe UI", 9),
                bg=T["BG_CARD"], fg=T["FG_TEXT"], relief="solid", bd=1, width=12).pack(side="left", ipady=2)
        
        # Filtro série/curso
        serie_var = tk.StringVar()
        curso_var = tk.StringVar()
        
        tk.Label(filtro_frame, text="  |  Série:", font=("Segoe UI", 9, "bold"),
                bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(10, 4))
        ttk.Combobox(filtro_frame, textvariable=serie_var, state="readonly",
                    values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                    width=10, font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))
        
        tk.Label(filtro_frame, text="Curso:", font=("Segoe UI", 9),
                bg=T["OBS_BG"], fg=T["FG_TEXT"]).pack(side="left", padx=(0, 4))
        ttk.Combobox(filtro_frame, textvariable=curso_var, state="readonly",
                    values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem", 
                           "Enfermagem", "Modelagem do Vestuario"],
                    width=20, font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
        
        def _buscar_historico():
            for i in th.get_children(): th.delete(i)
            
            # Determinar intervalo de datas
            hoje = datetime.date.today()
            if periodo_var.get() == "hoje":
                d_inicio = d_fim = hoje
            elif periodo_var.get() == "semana":
                d_inicio = hoje - datetime.timedelta(days=hoje.weekday())
                d_fim = hoje
            elif periodo_var.get() == "mes":
                d_inicio = hoje.replace(day=1)
                d_fim = hoje
            else:
                try:
                    d_inicio = datetime.datetime.strptime(data_inicio_var.get(), "%d/%m/%Y").date()
                    d_fim = datetime.datetime.strptime(data_fim_var.get(), "%d/%m/%Y").date()
                except Exception:
                    messagebox.showerror("Erro", "Data inválida. Use DD/MM/YYYY")
                    return
            
            eh_frequencia = hist_tipo_var.get() == "frequencia"

            dados_por_data = {}
            try:
                linhas_hist = _ler_frequencia_todos_db() if eh_frequencia else _ler_refeitorio_todos_db()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao carregar histórico do banco:\n{e}")
                return
            for row in linhas_hist:
                if len(row) >= 7:
                    try:
                        data_obj = datetime.datetime.strptime(row[0], "%d/%m/%Y").date()
                        if d_inicio <= data_obj <= d_fim:
                            serie = row[4] if len(row) > 4 else ""
                            curso = row[5] if len(row) > 5 else ""

                            if serie_var.get() != "(Todas)" and serie_var.get() not in serie:
                                continue
                            if curso_var.get() != "(Todos)" and curso_var.get() not in curso:
                                continue

                            if data_obj not in dados_por_data:
                                dados_por_data[data_obj] = {"total": 0, "presentes": 0, "ausentes": 0, "almocou": 0, "nao_almocou": 0}

                            if eh_frequencia:
                                dados_por_data[data_obj]["presentes"] += 1
                                dados_por_data[data_obj]["total"] += 1
                            else:
                                refeicao = row[6] if len(row) > 6 else ""
                                if refeicao.lower() == "almoco":
                                    dados_por_data[data_obj]["almocou"] += 1
                                else:
                                    dados_por_data[data_obj]["nao_almocou"] += 1
                                dados_por_data[data_obj]["total"] += 1
                    except Exception:
                        pass
            
            # Preencher tabela conforme tipo
            if eh_frequencia:
                th.heading("Total", text="TOTAL ALUNOS")
                th.heading("Almoços", text="PRESENTES")
                th.heading("Não Almoços", text="AUSENTES")
                th.heading("% Adesão", text="% PRESENÇA")
                
                total_geral = 0
                presentes_geral = 0
                for data_obj in sorted(dados_por_data.keys(), reverse=True):
                    dados = dados_por_data[data_obj]
                    data_str = data_obj.strftime("%d/%m/%Y")
                    total = dados["total"]
                    presentes = dados["presentes"]
                    ausentes = total - presentes
                    pct = (presentes / total * 100) if total > 0 else 0
                    
                    th.insert("", "end", values=(data_str, total, presentes, ausentes, f"{pct:.1f}%"),
                             tags=("presente" if pct >= 70 else ("warning" if pct >= 40 else "ausente"),))
                    total_geral += total
                    presentes_geral += presentes
                
                pct_geral = (presentes_geral / total_geral * 100) if total_geral > 0 else 0
                lbl_hist_cnt.config(text=f"Presença geral: {pct_geral:.1f}% ({presentes_geral}/{total_geral})")
            else:
                th.heading("Total", text="TOTAL REGISTROS")
                th.heading("Almoços", text="TOTAL ALMOÇOS")
                th.heading("Não Almoços", text="TOTAL NÃO ALMOÇOS")
                th.heading("% Adesão", text="% ADESÃO")
                
                total_geral = 0
                almocou_geral = 0
                for data_obj in sorted(dados_por_data.keys(), reverse=True):
                    dados = dados_por_data[data_obj]
                    data_str = data_obj.strftime("%d/%m/%Y")
                    total = dados["total"]
                    almocou = dados["almocou"]
                    nao_almocou = dados["nao_almocou"]
                    pct = (almocou / total * 100) if total > 0 else 0
                    
                    th.insert("", "end", values=(data_str, total, almocou, nao_almocou, f"{pct:.1f}%"))
                    total_geral += total
                    almocou_geral += almocou
                
                pct_geral = (almocou_geral / total_geral * 100) if total_geral > 0 else 0
                lbl_hist_cnt.config(text=f"Adesão geral: {pct_geral:.1f}% ({almocou_geral}/{total_geral})")
            
            # Desenhar gráfico
            desenhar_grafico_historico(dados_por_data)
        
        # Tabela
        f_hist = tk.Frame(aba_hist, bg=T["BORDER_GRID"], padx=1, pady=1)
        f_hist.pack(expand=True, fill="both", padx=20, pady=(5, 15))
        
        th = ttk.Treeview(f_hist, columns=("Data", "Total", "Almoços", "Não Almoços", "% Adesão"), show="headings")
        th.heading("Data", text="DATA")
        th.column("Data", width=100, anchor="center")
        th.heading("Total", text="TOTAL REGISTROS")
        th.column("Total", width=120, anchor="center")
        th.heading("Almoços", text="TOTAL ALMOÇOS")
        th.column("Almoços", width=120, anchor="center")
        th.heading("Não Almoços", text="TOTAL NÃO ALMOÇOS")
        th.column("Não Almoços", width=140, anchor="center")
        th.heading("% Adesão", text="% ADESÃO")
        th.column("% Adesão", width=100, anchor="center")
        
        th.pack(expand=True, fill="both")
        th.tag_configure("even", background=T["ENTRY_BG"])
        th.tag_configure("odd", background=T["BG_CARD"])
        # Tags para frequência
        th.tag_configure("presente", foreground="#1b5e20", background="#e8f5e9")
        th.tag_configure("warning", foreground="#f57f17", background="#fff3cd")
        th.tag_configure("ausente", foreground="#c62828", background="#ffebee")
        
        lbl_hist_cnt = tk.Label(aba_hist, text="Adesão geral: 0%", font=("Segoe UI", 10),
                               bg=T["BG_CARD"], fg=T["FRASE_FG"])
        lbl_hist_cnt.pack(anchor="e", padx=25, pady=(2, 10))
        
        # Gráfico
        canvas_grafico = tk.Canvas(aba_hist, bg=T["BG_CARD"], highlightthickness=0, height=200)
        canvas_grafico.pack(fill="x", padx=20, pady=(0, 10))
        
        def desenhar_grafico_historico(dados):
            canvas_grafico.delete("all")
            if not dados:
                canvas_grafico.create_text(400, 100, text="Nenhum dado para o período selecionado",
                                          font=("Segoe UI", 12), fill=T["FRASE_FG"])
                return
            
            datas_ordenadas = sorted(dados.keys(), reverse=True)[:30]
            
            if hist_tipo_var.get() == "frequencia":
                # Para frequência: mostrar presentes
                max_valor = max((dados[d]["presentes"] for d in datas_ordenadas), default=1)
            else:
                # Para refeitório: mostrar almoços
                max_valor = max((dados[d]["almocou"] for d in datas_ordenadas), default=1)
            
            if max_valor == 0: max_valor = 1
            
            w = canvas_grafico.winfo_width()
            if w <= 1: w = 800
            h = canvas_grafico.winfo_height()
            if h <= 1: h = 200
            
            margin = 40
            grafico_w = w - 2 * margin
            grafico_h = h - 40
            
            x_unit = grafico_w / max(len(datas_ordenadas), 1)
            y_unit = grafico_h / max_valor
            
            for idx, data_obj in enumerate(reversed(datas_ordenadas)):
                if hist_tipo_var.get() == "frequencia":
                    valor = dados[data_obj]["presentes"]
                    total = dados[data_obj]["total"]
                    pct = (valor / total * 100) if total > 0 else 0
                else:
                    valor = dados[data_obj]["almocou"]
                    total = dados[data_obj]["total"]
                    pct = (valor / total * 100) if total > 0 else 0
                
                x_start = margin + idx * x_unit
                x_end = x_start + x_unit * 0.8
                y_end = h - 30
                y_start = y_end - (valor * y_unit)
                
                cor = "#4caf50" if pct >= 70 else ("#fdd835" if pct >= 40 else "#f44336")
                canvas_grafico.create_rectangle(x_start, y_start, x_end, y_end, fill=cor, outline="white", width=2)
                
                data_label = data_obj.strftime("%d/%m")
                canvas_grafico.create_text(x_start + (x_end - x_start) / 2, h - 10,
                                         text=data_label, font=("Segoe UI", 8), fill=T["FG_TEXT"])
        
        def exportar_pdf_historico():
            if not th.get_children():
                messagebox.showwarning("Aviso", "Nenhum dado de histórico para exportar.")
                return
            try:
                from fpdf import FPDF
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", "B", 16)
                pdf.cell(0, 10, "HISTORICO DE REFEICOES - EEEP MARWIN", ln=True, align="C")
                pdf.set_font("Arial", "", 12)
                pdf.cell(0, 8, f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
                pdf.ln(6)
                
                # Ler dados da tabela
                pdf.set_font("Arial", "B", 11)
                pdf.cell(40, 8, "Data", border=1)
                pdf.cell(40, 8, "Total", border=1)
                pdf.cell(50, 8, "Almocos", border=1)
                pdf.cell(50, 8, "Nao Almocos", border=1)
                pdf.cell(0, 8, "% Adesao", border=1, ln=True)
                
                pdf.set_font("Arial", "", 10)
                for item in th.get_children():
                    vals = th.item(item)["values"]
                    for val in vals[:-1]:
                        pdf.cell(40, 7, str(val), border=1)
                    pdf.cell(0, 7, str(vals[-1]), border=1, ln=True)
                
                nome_arq = f"historico_marwin_{datetime.datetime.now().strftime('%d_%m_%Y')}.pdf"
                pdf.output(nome_arq)
                messagebox.showinfo("Sucesso", f"PDF gerado:\n{nome_arq}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao gerar PDF:\n{e}")
        
        def exportar_csv_historico():
            try:
                nome_arq = f"historico_marwin_{datetime.datetime.now().strftime('%d_%m_%Y')}.csv"
                with open(nome_arq, "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow(["Data", "Total", "Almocos", "Nao Almocos", "% Adesao"])
                    for item in th.get_children():
                        vals = th.item(item)["values"]
                        w.writerow(vals)
                messagebox.showinfo("Sucesso", f"CSV gerado:\n{nome_arq}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao gerar CSV:\n{e}")
        
        # Botões
        btn_frame = tk.Frame(aba_hist, bg=T["BG_CARD"])
        btn_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        tk.Button(btn_frame, text="Buscar", command=_buscar_historico,
                 bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
                 padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="Exportar PDF", command=exportar_pdf_historico,
                 bg=T["BTN_CARDAPIO"], fg="white", font=("Segoe UI", 10, "bold"),
                 padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="Exportar CSV", command=exportar_csv_historico,
                 bg=T["ACCENT_VIBRANT"], fg="white", font=("Segoe UI", 10, "bold"),
                 padx=14, pady=6, bd=0, cursor="hand2").pack(side="left")
        
        # Listeners para filtros
        periodo_var.trace_add("write", lambda *_: _buscar_historico())
        serie_var.trace_add("write", lambda *_: _buscar_historico())
        curso_var.trace_add("write", lambda *_: _buscar_historico())
        
        _buscar_historico()
    
    # ── ABA LOGS ──────────────────────────────────────────────────────────────
    def setup_logs():
        for w in aba_logs.winfo_children(): w.destroy()
        
        # Cabeçalho
        hd = tk.Frame(aba_logs, bg=T["BG_CARD"]); hd.pack(fill="x", padx=20, pady=(15, 10))
        tk.Label(hd, text="Logs do Sistema", font=("Segoe UI", 16, "bold"),
                 bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(side="left")
        
        # Botões
        btn_frame = tk.Frame(aba_logs, bg=T["BG_CARD"])
        btn_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        def atualizar_logs():
            text_logs.config(state="normal")
            text_logs.delete("1.0", "end")
            
            if os.path.exists(LOG_FILE):
                with open(LOG_FILE, "r", encoding="utf-8") as f:
                    linhas = f.readlines()
                    ultimas_linhas = linhas[-200:]
                    text_logs.insert("end", "".join(ultimas_linhas))
            else:
                text_logs.insert("end", "Nenhum arquivo de log encontrado ainda.")
            
            text_logs.config(state="disabled")
            text_logs.see("end")
        
        def limpar_logs():
            if messagebox.askyesno("Confirmar", "Limpar todos os logs?"):
                try:
                    open(LOG_FILE, "w").close()
                    atualizar_logs()
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao limpar logs:\n{e}")
        
        tk.Button(btn_frame, text="↻  Atualizar", command=atualizar_logs,
                 bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
                 padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(btn_frame, text="🗑  Limpar logs", command=limpar_logs,
                 bg="#c62828", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=14, pady=6, bd=0, cursor="hand2").pack(side="left")
        
        # Text widget com scrollbar
        f_logs = tk.Frame(aba_logs, bg=T["BORDER_GRID"], padx=1, pady=1)
        f_logs.pack(expand=True, fill="both", padx=20, pady=(0, 15))
        
        text_logs = tk.Text(f_logs, bg=T["BG_CARD"], fg=T["FG_TEXT"],
                           font=("Courier New", 9), relief="flat", state="disabled")
        sb_logs = ttk.Scrollbar(f_logs, orient="vertical", command=text_logs.yview)
        text_logs.configure(yscrollcommand=sb_logs.set)
        text_logs.pack(side="left", expand=True, fill="both")
        sb_logs.pack(side="right", fill="y")
        
        atualizar_logs()


def aplicar_tema(parent):
    for w in parent.winfo_children():
        cls = w.winfo_class()
        try:
            if cls in ("Frame", "Labelframe"):
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"]); break
                    if bg == nome_tema["ENTRY_BG"]:       w.configure(bg=T["ENTRY_BG"]); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"]); break
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BORDER_GRID"]:    w.configure(bg=T["BORDER_GRID"]); break
            elif cls == "Label":
                bg = w.cget("bg"); fg = w.cget("fg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"]); break
                    if bg == nome_tema["ENTRY_BG"]:       w.configure(bg=T["ENTRY_BG"]); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"]); break
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"]); break
                for nome_tema in TEMAS.values():
                    if fg == nome_tema["ACCENT_VIBRANT"]: w.configure(fg=T["ACCENT_VIBRANT"]); break
                    if fg == nome_tema["FRASE_FG"]:       w.configure(fg=T["FRASE_FG"]); break
                    if fg == nome_tema["FG_TEXT"]:        w.configure(fg=T["FG_TEXT"]); break
            elif cls == "Button":
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"], fg="white"); break
                    if bg == nome_tema["ACCENT_SOFT"]:    w.configure(bg=T["ACCENT_SOFT"],    fg="white"); break
                    if bg == nome_tema["BTN_CARDAPIO"]:   w.configure(bg=T["BTN_CARDAPIO"],   fg="white"); break
                    if bg == nome_tema["BTN_VOLTAR"]:     w.configure(bg=T["BTN_VOLTAR"],     fg="white"); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"],  fg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"], fg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"], fg=T["FRASE_FG"]); break
            elif cls == "Entry":
                w.configure(bg=T["ENTRY_BG"], fg=T["FG_TEXT"], insertbackground=T["ACCENT_VIBRANT"])
            elif cls == "Canvas":
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]: w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]: w.configure(bg=T["BG_MAIN"]); break
            elif cls == "Checkbutton":
                w.configure(bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"],
                            activebackground=T["BG_CARD"], selectcolor=T["ENTRY_BG"])
        except Exception:
            pass
        if w.winfo_children():
            aplicar_tema(w)

def alternar_tema():
    global tema_atual, T
    tema_atual = "escuro" if tema_atual=="claro" else "claro"
    T = TEMAS[tema_atual]; salvar_tema_tk(tema_atual)
    janela.configure(bg=T["BG_MAIN"])
    btn_tema.configure(text="☀️ Tema" if tema_atual=="escuro" else "🌙 Tema",
                       bg=T["ACCENT_VIBRANT"], fg="#b2dfb4")
    aplicar_tema(janela)
    for w in janela.winfo_children():
        if w.winfo_class() == "Toplevel":
            w.configure(bg=T["BG_MAIN"]); aplicar_tema(w)

def iniciar_tkinter(url_publica):
    global janela, T, tema_atual, url_ngrok_global, btn_tema
    url_ngrok_global = url_publica
    tema_atual = carregar_tema()
    T = TEMAS[tema_atual]
    janela = tk.Tk()
    janela.title("EEEP MARWIN — Servidor Admin")
    janela.state("zoomed"); janela.configure(bg=T["BG_MAIN"])
    janela.bind("<F11>", lambda e: janela.attributes("-fullscreen", not janela.attributes("-fullscreen")))

    # ── Barra de topo institucional ──────────────────────────────────────────
    barra_topo = tk.Frame(janela, bg=T["ACCENT_VIBRANT"], height=64)
    barra_topo.pack(fill="x"); barra_topo.pack_propagate(False)
    tk.Label(barra_topo, text="EEEP MARWIN",
             font=("Segoe UI", 15, "bold"), bg=T["ACCENT_VIBRANT"], fg="white").pack(side="left", padx=28, pady=16)
    tk.Label(barra_topo, text="Painel do Servidor Admin",
             font=("Segoe UI", 9), bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="left", padx=(0, 0), pady=22)

    # URL + botão copiar à direita
    url_frame = tk.Frame(barra_topo, bg=T["ACCENT_VIBRANT"]); url_frame.pack(side="right", padx=16)
    tk.Label(url_frame, text=f"● {url_publica}", font=("Segoe UI", 9),
             bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="left", padx=(0, 8))
    tk.Button(url_frame, text="Copiar URL", font=("Segoe UI", 8, "bold"),
              bg=T["ACCENT_SOFT"], fg="white", bd=0, padx=10, pady=4, cursor="hand2",
              command=lambda: [janela.clipboard_clear(),
                               janela.clipboard_append(url_publica),
                               messagebox.showinfo("Copiado", "URL copiada para a área de transferência!")]
              ).pack(side="left")

    # Botão tema na barra de topo
    btn_tema = tk.Button(barra_topo, text="🌙 Tema", font=("Segoe UI", 10),
                         bg=T["ACCENT_VIBRANT"], fg="#b2dfb4", bd=0,
                         cursor="hand2", activebackground=T["ACCENT_SOFT"],
                         command=alternar_tema)
    btn_tema.pack(side="right", padx=4)

    # Linha amarela accent
    tk.Frame(janela, bg=T["HIGHLIGHT_YELLOW"], height=4).pack(fill="x")

    # ── Conteúdo central ────────────────────────────────────────────────────
    main = tk.Frame(janela, bg=T["BG_MAIN"]); main.pack(expand=True, fill="both")
    center_wrap = tk.Frame(main, bg=T["BG_MAIN"])
    center_wrap.place(relx=0.5, rely=0.5, anchor="center")

    # Logo / ícone
    try:
        from PIL import Image as _PilImg, ImageTk as _PilImgTk
        _path_marwin = _buscar_logo_png()
        if not _path_marwin:
            _pasta = os.path.dirname(os.path.abspath(__file__))
            _path_marwin = os.path.join(_pasta, "logo_marwin.png")
        if os.path.exists(_path_marwin):
            _img = _PilImg.open(_path_marwin).convert("RGBA")
            _img.thumbnail((520, 520), _PilImg.LANCZOS)
            _logo_tk = _PilImgTk.PhotoImage(_img, master=janela)
            tk.Label(center_wrap, image=_logo_tk, bg=T["BG_MAIN"]).pack(pady=(0, 10))
            center_wrap._logo_ref = _logo_tk
            center_wrap._logo_ref = _logo_tk   # evitar GC
    except Exception:
        tk.Label(center_wrap, text="🏫", font=("Segoe UI", 44), bg=T["BG_MAIN"]).pack(pady=(0, 8))

    tk.Label(center_wrap, text="Escola Estadual de Educação Profissional",
             font=("Segoe UI", 13), bg=T["BG_MAIN"], fg=T["FRASE_FG"]).pack()
    tk.Label(center_wrap, text="MARWIN",
             font=("Segoe UI", 44, "bold"), bg=T["BG_MAIN"], fg=T["ACCENT_VIBRANT"]).pack(pady=(2, 4))
    tk.Label(center_wrap, text="Painel do Servidor",
             font=("Segoe UI", 11), bg=T["BG_MAIN"], fg=T["FRASE_FG"]).pack()

    tk.Frame(center_wrap, bg=T["BORDER_GRID"], height=1).pack(fill="x", pady=(18, 24))

    tk.Button(center_wrap, text="  ABRIR PAINEL ADMIN  ",
              font=("Segoe UI", 17, "bold"),
              bg=T["ACCENT_VIBRANT"], fg="white", bd=0,
              padx=50, pady=20, cursor="hand2",
              activebackground=T["ACCENT_SOFT"], activeforeground="white",
              command=abrir_painel_admin).pack(fill="x")

    tk.Label(center_wrap,
             text="💚 Servidor ativo — aguardando conexões dos clientes.",
             bg=T["BG_MAIN"], fg=T["FRASE_FG"],
             font=("Segoe UI", 10, "italic")).pack(pady=(24, 0))

    janela.mainloop()

# ==============================================================================
# INICIALIZACAO
# ==============================================================================
if __name__ == "__main__":
    _iniciar_pool_pg()
    try:
        _criar_tabelas_neon()
    except Exception:
        pass
    if _cloud_api_url():
        threading.Thread(target=_sincronizar_tudo_nuvem, daemon=True).start()

    threading.Thread(
        target=lambda: app.run(host="0.0.0.0", port=5000, use_reloader=False, debug=False),
        daemon=True
    ).start()
    time.sleep(1)

    local_ip = _get_local_ip()
    url = f"http://{local_ip}:5000"
    logger.info(f"Servidor MARWIN iniciado em {url}")
    print(f"\n[CLIENTE] Acesse localmente em: {url}")
    print(f"[CLIENTE] Ou use http://localhost:5000 no proprio computador")
    
    iniciar_tkinter(url)
