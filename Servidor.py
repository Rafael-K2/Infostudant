"""
SERVIDOR MARWIN - ServidorV15.py
======================================
Rode ESTE arquivo no PC da escola.

O que ele faz ao iniciar:
  1. Prepara o servidor Flask na porta 5000
  2. Conecta-se ao PostgreSQL Neon como fonte de dados
  3. Abre o painel admin Tkinter com IP local e botão copiar URL

Instalação (uma vez só):
  pip install flask flask-cors fpdf2 qrcode pillow psycopg2-binary

O index.html pode usar a API na nuvem (ApiNuvem.py) — configure dados/cloud_config.json.
O painel admin sincroniza cardápio/eventos/config com a nuvem automaticamente.

Organização dos QR Codes: qrcodes_marwin/
    1 Ano - Desenvolvimento de Sistemas/
      2026001_Joao_Silva_1_Ano_Desenvolvimento_de_Sistemas.png
      ...
    2 Ano - Redes de Computadores/
      ...
    3 Ano - Informatica/
      ...
"""

import sys, subprocess

for pkg in ["flask", "flask-cors", "fpdf2", "qrcode", "pillow", "psycopg2-binary"]:
    try:
        __import__(pkg.replace("-", "_").replace("pillow","PIL"))
    except ImportError:
        print(f"Instalando {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

import os, json, csv, datetime, threading, calendar, time, io, base64, re, logging, smtplib, shutil, zipfile, unicodedata, socket
from collections import Counter
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
import datetime

import qrcode
from PIL import Image

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from logging.handlers import TimedRotatingFileHandler

# ── Fuso horário oficial do Brasil ───────────────────────────────────────────
# Horário de Brasília = UTC-3, fixo (o Brasil não usa mais horário de verão
# desde 2019). Definido explicitamente para que o horário/data registrados
# ao ler o QR Code NÃO dependam do fuso horário configurado no sistema
# operacional/servidor onde este script está rodando (ex.: servidores em
# nuvem costumam usar UTC por padrão, o que atrasava/avançava os horários).
FUSO_BRASIL = datetime.timezone(datetime.timedelta(hours=-3))

def _agora_br():
    """Retorna o datetime atual já corrigido para o horário de Brasília (UTC-3)."""
    return datetime.datetime.now(FUSO_BRASIL)

app = Flask(__name__)
CORS(app)

DADOS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "dados"))
os.makedirs(DADOS_DIR, exist_ok=True)

# ── Configuração de Logging ───────────────────────────────────────────────────
LOG_FILE = os.path.join(DADOS_DIR, "marwin.log")
logger = logging.getLogger("marwin")
logger.setLevel(logging.DEBUG)
try:
    log_handler = TimedRotatingFileHandler(
        LOG_FILE, when="midnight", interval=1, backupCount=7, encoding="utf-8"
    )
    log_handler.setLevel(logging.DEBUG)
    log_format = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s",
                                     datefmt="%d/%m/%Y %H:%M:%S")
    log_handler.setFormatter(log_format)
    logger.addHandler(log_handler)
except Exception as e:
    print(f"[LOG] Erro ao configurar logging: {e}")

def _buscar_logo_png():
    if os.path.isdir(DADOS_DIR):
        logos = [n for n in os.listdir(DADOS_DIR) if n.lower().endswith(".png")]
        if logos:
            for nome in logos:
                if "logo" in nome.lower():
                    return os.path.join(DADOS_DIR, nome)
            return os.path.join(DADOS_DIR, logos[0])
    return None

CARDAPIO_FILE    = os.path.join(DADOS_DIR, "cardapio.json")
EVENTOS_FILE     = os.path.join(DADOS_DIR, "eventos.json")
CONFIG_FILE      = os.path.join(DADOS_DIR, "config_sistema.json")
TEMA_FILE        = os.path.join(DADOS_DIR, "tema_config.json")
LISTA_ALUNOS_FILE = os.path.join(DADOS_DIR, "lista_alunos.json")
DB_CONFIG_FILE   = os.path.join(DADOS_DIR, "db_config.json")
CLOUD_CONFIG_FILE = os.path.join(DADOS_DIR, "cloud_config.json")
MESES_PT = (
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
)
DEFAULT_ADMIN_PLAIN = "Marwin2026"
# Prefer env var MARWIN_ADMIN_PASS (pode conter hash bcrypt). Fallback para DEFAULT_ADMIN_PLAIN.
ADMIN_PASSWORD = os.getenv("MARWIN_ADMIN_PASS", DEFAULT_ADMIN_PLAIN)
# Senha em texto claro para envio no header X-Senha ao sincronizar com a nuvem.
# Se MARWIN_ADMIN_PASS for um hash bcrypt, use MARWIN_ADMIN_PLAIN_PASS com a senha real.
ADMIN_PLAIN_PASS = os.getenv("MARWIN_ADMIN_PLAIN_PASS", DEFAULT_ADMIN_PLAIN)
if ADMIN_PASSWORD == DEFAULT_ADMIN_PLAIN:
    print(
        "\n[AVISO DE SEGURANÇA] A variável de ambiente MARWIN_ADMIN_PASS não está definida.\n"
        "O servidor está usando a senha padrão. Defina MARWIN_ADMIN_PASS para maior segurança.\n"
    )

# Try to import bcrypt if available (optional). If ADMIN_PASSWORD is a bcrypt hash, we'll use it.
try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except Exception:
    BCRYPT_AVAILABLE = False

def ler_json(path, padrao):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return padrao

def salvar_json(path, dados):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def _ler_cloud_config():
    return ler_json(CLOUD_CONFIG_FILE, {"api_url": "", "sincronizar_automatico": True})

def _cloud_api_url():
    return _ler_cloud_config().get("api_url", "").strip().rstrip("/")

def _sync_nuvem(rota, metodo, dados=None):

    cfg = _ler_cloud_config()
    base = cfg.get("api_url", "").strip().rstrip("/")
    if not base:
        logger.debug(f"Sync nuvem ignorado ({rota}): api_url não configurada em cloud_config.json")
        return False
    if not cfg.get("sincroniz   ar_automatico", True):
        logger.debug(f"Sync nuvem ignorado ({rota}): sincronizar_automatico=false")
        return False
    import urllib.request
    url = base + rota
    body = json.dumps(dados, ensure_ascii=False).encode("utf-8") if dados is not None else None
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json", "X-Senha": ADMIN_PLAIN_PASS},
        method=metodo,
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            ok = 200 <= resp.status < 300
            if not ok:
                logger.warning(f"Sync nuvem ({rota}): HTTP {resp.status}")

            return ok
    except urllib.error.HTTPError as e:
        corpo = ""
        try:
            corpo = e.read().decode("utf-8", errors="replace")[:200]
        except Exception:
            pass
        logger.warning(f"Sync nuvem falhou ({rota}): HTTP {e.code} — {corpo}")
        return False
    except Exception as e:
        logger.warning(f"Sync nuvem falhou ({rota}): {e}")

        return False

def _sincronizar_tudo_nuvem():
    ok = True
    if not _sync_nuvem("/admin/cardapio", "PUT", ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO)):
        ok = False
    if not _sync_nuvem("/admin/eventos", "PUT", ler_json(EVENTOS_FILE, EVENTOS_PADRAO)):
        ok = False
    if not _sync_nuvem("/admin/config", "PUT", ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"})):
        ok = False
    if ok and _cloud_api_url():
        logger.info("Configurações sincronizadas com a API na nuvem")

    return ok

# ── Credenciais do banco — use variável de ambiente para evitar exposição ──────
# Defina MARWIN_DB_URL no sistema antes de rodar:
#   Windows : set MARWIN_DB_URL=postgresql://usuario:senha@host/neondb?sslmode=require
#   Linux   : export MARWIN_DB_URL=postgresql://...
_DB_FALLBACK = (
    "postgresql://neondb_owner:npg_ydP7rqBR0ZoQ@"
    "ep-broad-mountain-apatc77i-pooler.c-7.us-east-1.aws.neon.tech"
    "/neondb?sslmode=require&channel_binding=require"
)
DB_DEFAULT_CONNECTION = os.getenv("MARWIN_DB_URL", _DB_FALLBACK)
if DB_DEFAULT_CONNECTION == _DB_FALLBACK:
    print(
        "\n[AVISO DE SEGURANÇA] A variável de ambiente MARWIN_DB_URL não está definida.\n"
        "O servidor está usando a connection string padrão embutida no código.\n"
        "Defina MARWIN_DB_URL para evitar expor credenciais.\n"
    )
PG_POOL = None

def _carregar_db_config():
    cfg = ler_json(DB_CONFIG_FILE, {})
    if not cfg.get("connection_string"):
        cfg["connection_string"] = DB_DEFAULT_CONNECTION
        salvar_json(DB_CONFIG_FILE, cfg)
    return cfg

def _iniciar_pool_pg():
    global PG_POOL
    if PG_POOL is not None:
        return
    cfg = _carregar_db_config()
    try:
        import psycopg2
        from psycopg2 import pool
        PG_POOL = pool.ThreadedConnectionPool(1, 10, cfg["connection_string"])
        logger.info("Pool PostgreSQL inicializado")
    except Exception as e:
        PG_POOL = None
        logger.warning("PostgreSQL indisponível")

def get_pg_conn():
    global PG_POOL
    if PG_POOL is None:
        _iniciar_pool_pg()
    if PG_POOL is None:
        return None
    try:
        return PG_POOL.getconn()
    except Exception as e:
        logger.warning("Falha ao obter conexão PostgreSQL")
        return None

def _release_pg_conn(conn):
    global PG_POOL
    if not conn:
        return
    try:
        if PG_POOL:
            PG_POOL.putconn(conn)
        else:
            conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

def _executar_pg(sql, params=None, fetch=False):
    conn = get_pg_conn()
    if conn is None:
        raise RuntimeError("PostgreSQL indisponível")
    cur = None
    try:
        cur = conn.cursor()
        cur.execute(sql, params or ())
        if fetch:
            cols = [d[0] for d in cur.description] if cur.description else []
            rows = cur.fetchall()
            return [dict(zip(cols, row)) for row in rows]
        conn.commit()
        return []
    finally:
        if cur is not None:
            try:
                cur.close()
            except Exception:
                pass
        _release_pg_conn(conn)

def _get_local_ip():
    try:
        ip = socket.gethostbyname(socket.gethostname())
        if ip.startswith("127.") or ip == "0.0.0.0":
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect(("8.8.8.8", 80))
                ip = s.getsockname()[0]
            finally:
                s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _criar_tabelas_neon():
    try:
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS refeitorio (
                data TEXT NOT NULL,
                horaentrada TEXT NOT NULL,
                matricula TEXT NOT NULL,
                nome TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                refeicao TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS frequencia (
                data TEXT NOT NULL,
                horaentrada TEXT NOT NULL,
                matricula TEXT NOT NULL,
                nome TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                aula TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS avaliacoes (
                data TEXT NOT NULL,
                aluno TEXT NOT NULL,
                serie TEXT,
                curso TEXT,
                estagio TEXT,
                item TEXT,
                nota TEXT
            );
            """
        )
        _executar_pg(
            """
            CREATE TABLE IF NOT EXISTS sistema_config (
                chave TEXT PRIMARY KEY,
                valor JSONB NOT NULL
            );
            """
        )
        logger.info("Tabelas Neon verificadas/criadas")
    except Exception as e:
        logger.warning(f"Nao foi possivel criar tabelas Neon: {e}")


def _ler_refeitorio_hoje_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, refeicao FROM refeitorio WHERE data = %s",
        (_hoje(),),
        fetch=True
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["refeicao"]]
        for row in rows
    ]


def _inserir_refeitorio_db(registro):
    _executar_pg(
        "INSERT INTO refeitorio (data, horaentrada, matricula, nome, serie, curso, refeicao) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _apagar_refeitorio_data_db(data_alvo):
    _executar_pg("DELETE FROM refeitorio WHERE data = %s", (data_alvo,))


def _refeitorio_duplicado_db(matricula, refeicao):
    try:
        rows = _executar_pg(
            "SELECT horaentrada, nome FROM refeitorio WHERE data = %s AND matricula = %s AND refeicao = %s LIMIT 1",
            (_hoje(), matricula, refeicao),
            fetch=True
        )
        if not rows:
            return None

        total_hoje = _executar_pg(
            "SELECT COUNT(*) AS total FROM refeitorio WHERE data = %s",
            (_hoje(),),
            fetch=True
        )
        total_refeicao = _executar_pg(
            "SELECT COUNT(*) AS total FROM refeitorio WHERE data = %s AND refeicao = %s",
            (_hoje(), refeicao),
            fetch=True
        )
        return {
            "hora": rows[0]["horaentrada"],
            "nome": rows[0]["nome"],
            "total_hoje": total_hoje[0].get("total", 0) if total_hoje else 0,
            "total_refeicao": total_refeicao[0].get("total", 0) if total_refeicao else 0,
        }
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade de refeitorio: {e}")
        return None


def _ler_frequencia_hoje_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, aula FROM frequencia WHERE data = %s",
        (_hoje(),),
        fetch=True
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["aula"]]
        for row in rows
    ]


def _inserir_frequencia_db(registro):
    _executar_pg(
        "INSERT INTO frequencia (data, horaentrada, matricula, nome, serie, curso, aula) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _apagar_frequencia_data_db(data_alvo):
    _executar_pg("DELETE FROM frequencia WHERE data = %s", (data_alvo,))


def _frequencia_duplicado_db(matricula):
    try:
        rows = _executar_pg(
            "SELECT horaentrada, nome FROM frequencia WHERE data = %s AND matricula = %s LIMIT 1",
            (_hoje(), matricula),
            fetch=True
        )
        if not rows:
            return None

        total_hoje = _executar_pg(
            "SELECT COUNT(*) AS total FROM frequencia WHERE data = %s",
            (_hoje(),),
            fetch=True
        )
        return {
            "hora": rows[0]["horaentrada"],
            "nome": rows[0]["nome"],
            "total_hoje": total_hoje[0].get("total", 0) if total_hoje else 0,
        }
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade de frequencia: {e}")
        return None


def _inserir_avaliacao_db(registro):
    _executar_pg(
        "INSERT INTO avaliacoes (data, aluno, serie, curso, estagio, item, nota) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (registro[0], registro[1], registro[2], registro[3], registro[4], registro[5], registro[6])
    )


def _ler_avaliacoes_db():
    rows = _executar_pg(
        "SELECT data, aluno, serie, curso, estagio, item, nota FROM avaliacoes ORDER BY data DESC",
        (),
        fetch=True
    )
    if not rows:
        return []
    return [
        {
            "Data": row["data"],
            "Aluno": row["aluno"],
            "Serie": row["serie"],
            "Curso": row["curso"],
            "Estagio": row["estagio"],
            "Item": row["item"],
            "Nota": row["nota"],
        }
        for row in rows
    ]


def _ler_refeitorio_todos_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, refeicao FROM refeitorio ORDER BY data DESC, horaentrada DESC",
        (),
        fetch=True,
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["refeicao"]]
        for row in rows
    ]


def _ler_frequencia_todos_db():
    rows = _executar_pg(
        "SELECT data, horaentrada, matricula, nome, serie, curso, aula FROM frequencia ORDER BY data DESC, horaentrada DESC",
        (),
        fetch=True,
    )
    if not rows:
        return []
    return [
        [row["data"], row["horaentrada"], row["matricula"], row["nome"], row["serie"], row["curso"], row["aula"]]
        for row in rows
    ]


def _avaliacoes_para_linhas():
    return [
        [r["Data"], r["Aluno"], r["Serie"], r["Curso"], r["Estagio"], r["Item"], r["Nota"]]
        for r in _ler_avaliacoes_db()
    ]


def _escrever_csv(caminho, header, linhas):
    with open(caminho, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(linhas)


def _csv_bytes(header, linhas):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    w.writerows(linhas)
    out = io.BytesIO(buf.getvalue().encode("utf-8"))
    out.seek(0)
    return out


def _nome_backup_mes_sugerido():
    hoje = _agora_br().date()
    return f"backup_{MESES_PT[hoje.month - 1]}_{hoje.year}"


def _exportar_backup_mensal_csv(nome_base):
    """Exporta todas as tabelas do banco para CSVs em dados/backups/YYYY_MM/."""
    nome_base = re.sub(r'[<>:"/\\|?*]', "_", (nome_base or "").strip())
    if not nome_base:
        raise ValueError("Nome do arquivo inválido")
    pasta = os.path.join(DADOS_DIR, "backups", _agora_br().date().strftime("%Y_%m"))
    os.makedirs(pasta, exist_ok=True)
    arquivos = []
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_avaliacoes_escola.csv"),
        ["Data", "Aluno", "Serie", "Curso", "Estagio", "Item", "Nota"],
        _avaliacoes_para_linhas(),
    )
    arquivos.append(f"{nome_base}_avaliacoes_escola.csv")
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_refeitorio.csv"),
        ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
        _ler_refeitorio_todos_db(),
    )
    arquivos.append(f"{nome_base}_refeitorio.csv")
    _escrever_csv(
        os.path.join(pasta, f"{nome_base}_frequencia.csv"),
        ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
        _ler_frequencia_todos_db(),
    )
    arquivos.append(f"{nome_base}_frequencia.csv")
    return pasta, arquivos


def _avaliacao_ja_existe_db(nome, semana_iso, ano_iso):
    if not nome:
        return False
    try:
        rows = _executar_pg(
            "SELECT data FROM avaliacoes WHERE lower(aluno) = lower(%s)",
            (nome,),
            fetch=True
        )
    except Exception as e:
        logger.warning(f"PostgreSQL indisponível para verificação de duplicidade: {e}")
        return False

    if not rows:
        return False
    for row in rows:
        try:
            data_str = row["data"].split(" ")[0]
            data_obj = datetime.datetime.strptime(data_str, "%d/%m/%Y").date()
            if data_obj.isocalendar()[1] == semana_iso and data_obj.isocalendar()[0] == ano_iso:
                return True
        except Exception:
            continue
    return False


def _apagar_avaliacoes_db():
    _executar_pg("DELETE FROM avaliacoes", ())


def _limpar_tabelas_neon():
    _executar_pg("DELETE FROM refeitorio", ())
    _executar_pg("DELETE FROM frequencia", ())
    _executar_pg("DELETE FROM avaliacoes", ())
    logger.info("Tabelas Neon refeitorio, frequencia e avaliacoes limpas")


# ── Colunas de cada tabela para detecção automática ──────────────────────────
_COLS_REFEITORIO  = {"Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"}
_COLS_FREQUENCIA  = {"Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"}
_COLS_AVALIACOES  = {"Data", "Aluno", "Serie", "Curso", "Estagio", "Item", "Nota"}


def _detectar_tabela_csv(colunas_csv):
    """Retorna 'refeitorio', 'frequencia', 'avaliacoes' ou None."""
    cols = set(colunas_csv)
    if "Refeicao" in cols and _COLS_REFEITORIO.issubset(cols):
        return "refeitorio"
    if "Aula" in cols and _COLS_FREQUENCIA.issubset(cols):
        return "frequencia"
    if "Nota" in cols and _COLS_AVALIACOES.issubset(cols):
        return "avaliacoes"
    return None


def _duplicata_existe(tabela, linha):
    """Verifica se o registro já existe no banco (comparação por campos-chave)."""
    try:
        if tabela == "refeitorio":
            rows = _executar_pg(
                "SELECT 1 FROM refeitorio WHERE data=%s AND horaentrada=%s AND matricula=%s AND refeicao=%s LIMIT 1",
                (linha.get("Data",""), linha.get("HoraEntrada",""),
                 linha.get("Matricula",""), linha.get("Refeicao","")),
                fetch=True,
            )
        elif tabela == "frequencia":
            rows = _executar_pg(
                "SELECT 1 FROM frequencia WHERE data=%s AND horaentrada=%s AND matricula=%s AND aula=%s LIMIT 1",
                (linha.get("Data",""), linha.get("HoraEntrada",""),
                 linha.get("Matricula",""), linha.get("Aula","")),
                fetch=True,
            )
        elif tabela == "avaliacoes":
            rows = _executar_pg(
                "SELECT 1 FROM avaliacoes WHERE data=%s AND aluno=%s AND estagio=%s AND item=%s LIMIT 1",
                (linha.get("Data",""), linha.get("Aluno",""),
                 linha.get("Estagio",""), linha.get("Item","")),
                fetch=True,
            )
        else:
            return False
        return bool(rows)
    except Exception:
        return False


def _importar_csv_para_banco(caminho_csv, ignorar_duplicatas=True, callback_progresso=None):
    """
    Lê o CSV, detecta a tabela pelo conjunto de colunas e insere os registros.

    Parâmetros
    ----------
    caminho_csv        : str  – caminho completo do arquivo CSV
    ignorar_duplicatas : bool – se True, pula linhas já existentes no banco
    callback_progresso : callable(atual, total) | None – chamado a cada linha processada

    Retorna
    -------
    dict com chaves: tabela, inseridos, ignorados, erros
    Levanta ValueError se o arquivo não puder ser identificado.
    """
    with open(caminho_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        linhas = list(reader)

    if not linhas:
        return {"tabela": "?", "inseridos": 0, "ignorados": 0, "erros": 0}

    tabela = _detectar_tabela_csv(list(linhas[0].keys()))

    if tabela is None:
        raise ValueError(
            "Colunas do CSV não correspondem a nenhuma tabela conhecida.\n"
            "Esperado:\n"
            f"  refeitorio  → {sorted(_COLS_REFEITORIO)}\n"
            f"  frequencia  → {sorted(_COLS_FREQUENCIA)}\n"
            f"  avaliacoes  → {sorted(_COLS_AVALIACOES)}"
        )

    total = len(linhas)
    inseridos = ignorados = erros = 0

    for i, linha in enumerate(linhas, 1):
        if callback_progresso:
            try:
                callback_progresso(i, total)
            except Exception:
                pass

        try:
            if ignorar_duplicatas and _duplicata_existe(tabela, linha):
                ignorados += 1
                continue

            if tabela == "refeitorio":
                _inserir_refeitorio_db([
                    linha.get("Data",""), linha.get("HoraEntrada",""),
                    linha.get("Matricula",""), linha.get("Nome",""),
                    linha.get("Serie",""), linha.get("Curso",""), linha.get("Refeicao",""),
                ])
            elif tabela == "frequencia":
                _inserir_frequencia_db([
                    linha.get("Data",""), linha.get("HoraEntrada",""),
                    linha.get("Matricula",""), linha.get("Nome",""),
                    linha.get("Serie",""), linha.get("Curso",""), linha.get("Aula",""),
                ])
            elif tabela == "avaliacoes":
                _inserir_avaliacao_db([
                    linha.get("Data",""), linha.get("Aluno",""),
                    linha.get("Serie",""), linha.get("Curso",""),
                    linha.get("Estagio",""), linha.get("Item",""), linha.get("Nota",""),
                ])
            inseridos += 1
        except Exception as e_linha:
            erros += 1
            logger.warning(f"_importar_csv_para_banco ({os.path.basename(caminho_csv)}) linha {i}: {e_linha}")

    logger.info(
        f"Importação CSV concluída — tabela={tabela} arquivo={os.path.basename(caminho_csv)} "
        f"inseridos={inseridos} ignorados={ignorados} erros={erros}"
    )
    return {"tabela": tabela, "inseridos": inseridos, "ignorados": ignorados, "erros": erros}


def _importar_csv_para_banco_forcado(caminho_csv, tabela, ignorar_duplicatas=True, callback_progresso=None):
    """
    Versão de _importar_csv_para_banco com tabela de destino forçada manualmente.
    Usada quando a detecção automática falha e o usuário escolheu a tabela no diálogo.
    """
    with open(caminho_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        linhas = list(reader)

    if not linhas:
        return {"tabela": tabela, "inseridos": 0, "ignorados": 0, "erros": 0}

    total = len(linhas)
    inseridos = ignorados = erros = 0

    for i, linha in enumerate(linhas, 1):
        if callback_progresso:
            try:
                callback_progresso(i, total)
            except Exception:
                pass
        try:
            if ignorar_duplicatas and _duplicata_existe(tabela, linha):
                ignorados += 1
                continue
            if tabela == "refeitorio":
                _inserir_refeitorio_db([
                    linha.get("Data", ""), linha.get("HoraEntrada", ""),
                    linha.get("Matricula", ""), linha.get("Nome", ""),
                    linha.get("Serie", ""), linha.get("Curso", ""), linha.get("Refeicao", ""),
                ])
            elif tabela == "frequencia":
                _inserir_frequencia_db([
                    linha.get("Data", ""), linha.get("HoraEntrada", ""),
                    linha.get("Matricula", ""), linha.get("Nome", ""),
                    linha.get("Serie", ""), linha.get("Curso", ""), linha.get("Aula", ""),
                ])
            elif tabela == "avaliacoes":
                _inserir_avaliacao_db([
                    linha.get("Data", ""), linha.get("Aluno", ""),
                    linha.get("Serie", ""), linha.get("Curso", ""),
                    linha.get("Estagio", ""), linha.get("Item", ""), linha.get("Nota", ""),
                ])
            inseridos += 1
        except Exception as e_linha:
            erros += 1
            logger.warning(f"_importar_csv_para_banco_forcado ({os.path.basename(caminho_csv)}) linha {i}: {e_linha}")

    logger.info(
        f"Importação forçada — tabela={tabela} arquivo={os.path.basename(caminho_csv)} "
        f"inseridos={inseridos} ignorados={ignorados} erros={erros}"
    )
    return {"tabela": tabela, "inseridos": inseridos, "ignorados": ignorados, "erros": erros}


CARDAPIO_PADRAO = {
    "SEGUNDA": ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "TERCA":   ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "QUARTA":  ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "QUINTA":  ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
    "SEXTA":   ["Cuzcuz com ovo e cafe fresco", "Arroz, feijao, frango cozido e batata doce", "Pao com manteiga e vitamina de goiaba"],
}
EVENTOS_PADRAO = [
    {"data": "03/02", "evento": "Inicio das Aulas"},
    {"data": "01/05", "evento": "Feriado: Dia do Trabalhador"},
    {"data": "24/06", "evento": "Arraia do Marwin"},
    {"data": "07/09", "evento": "Feriado: Independencia do Brasil"},
    {"data": "15/10", "evento": "Dia do Professor"},
    {"data": "25/12", "evento": "Natal"},
]

# ==============================================================================
# ROTAS FLASK
# ==============================================================================
import secrets

def checar_senha(req):
    hdr = req.headers.get("X-Senha", "")
    if not hdr:
        return False
    # If ADMIN_PASSWORD looks like a bcrypt hash and bcrypt is available, verify accordingly
    if BCRYPT_AVAILABLE and isinstance(ADMIN_PASSWORD, str) and ADMIN_PASSWORD.startswith("$2"):
        try:
            return bcrypt.checkpw(hdr.encode("utf-8"), ADMIN_PASSWORD.encode("utf-8"))
        except Exception:
            return False
    # Fallback: constant-time compare with configured plaintext
    return secrets.compare_digest(hdr, ADMIN_PASSWORD)

# Rota para servir o cliente web (index.html)
@app.route('/')
def serve_cliente():
    pasta_raiz = os.path.dirname(os.path.abspath(__file__))
    for nome in ('index.html', 'Index.html'):
        arquivo_index = os.path.join(pasta_raiz, nome)
        if os.path.isfile(arquivo_index):
            return send_from_directory(pasta_raiz, nome)
    return jsonify({'erro': 'index.html não encontrado'}), 404

# Rota para servir arquivos estáticos (CSS, JS, imagens, etc) da pasta raiz
@app.route('/<path:path>')
def serve_static(path):
    try:
        pasta_raiz = os.path.dirname(os.path.abspath(__file__))
        arquivo = os.path.join(pasta_raiz, path)
        # Segurança: evitar acesso fora da pasta raiz
        caminho_absoluto = os.path.abspath(arquivo)
        if not caminho_absoluto.startswith(os.path.abspath(pasta_raiz)):
            return jsonify({'erro': 'Acesso negado'}), 403
        if os.path.isfile(caminho_absoluto):
            return send_from_directory(pasta_raiz, path)
        return jsonify({'erro': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route("/cardapio",  methods=["GET"])
def get_cardapio(): return jsonify(ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO))

@app.route("/eventos",   methods=["GET"])
def get_eventos():  return jsonify(ler_json(EVENTOS_FILE, EVENTOS_PADRAO))

@app.route("/config",    methods=["GET"])
def get_config():   return jsonify(ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"}))

@app.route("/avaliacoes", methods=["GET"])
def get_avaliacoes_publicas():
    """Retorna todos os registros de avaliações em formato JSON"""
    try:
        registros = _ler_avaliacoes_db()
        return jsonify({"avaliacoes": registros, "total": len(registros)})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao carregar avaliacoes: {e}")
        return jsonify({"erro": "Erro ao carregar avaliações"}), 500

@app.route("/avaliacao", methods=["POST"])
def post_avaliacao():
    dados = request.get_json()
    if not dados: return jsonify({"erro": "JSON invalido"}), 400
    nome, serie, curso = dados.get("nome","Anonimo"), dados.get("serie","N/A"), dados.get("curso","N/A")
    respostas = dados.get("respostas", {})

    # Verificar se já avaliou esta semana (apenas se não for anônimo)
    if nome and nome.strip().lower() not in {"anonimo", "anônimo"}:
        hoje = _agora_br().date()
        semana_iso = hoje.isocalendar()[1]
        ano_iso = hoje.isocalendar()[0]

        try:
            if _avaliacao_ja_existe_db(nome, semana_iso, ano_iso):
                logger.warning(f"Avaliação duplicada detectada: {nome} - Semana {semana_iso}/{ano_iso}")
                return jsonify({"status": "ja_avaliou", "mensagem": "Você já avaliou esta semana"}), 200
        except Exception as e:
            logger.error(f"Erro ao verificar duplicidade de avaliação: {e}")
            return jsonify({"erro": "Banco de dados indisponível"}), 503

    data_hora = _agora_br().strftime("%d/%m/%Y %H:%M:%S")
    registros = []
    for chave, nota in respostas.items():
        estagio, item = chave.split("|", 1)
        registros.append([data_hora, nome, serie, curso, estagio, item, nota])

    try:
        for registro in registros:
            _inserir_avaliacao_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao salvar avaliação: {e}")
        return jsonify({"erro": "Erro ao salvar avaliação"}), 500
    logger.info(f"Avaliação recebida de {nome} ({len(respostas)} itens)")
    return jsonify({"status": "ok"})

@app.route("/avaliacao/verificar", methods=["GET"])
def verificar_avaliacao():
    nome = request.args.get("nome", "").strip()
    if not nome or nome.lower() in {"anonimo", "anônimo"}:
        return jsonify({"ja_avaliou": False}), 200

    hoje = _agora_br().date()
    semana_iso = hoje.isocalendar()[1]
    ano_iso = hoje.isocalendar()[0]

    try:
        ja_avaliou = _avaliacao_ja_existe_db(nome, semana_iso, ano_iso)
        return jsonify({"ja_avaliou": ja_avaliou}), 200
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao verificar avaliação: {e}")
        return jsonify({"erro": "Erro ao verificar avaliação"}), 500

@app.route("/admin/avaliacoes", methods=["GET"])
def get_avaliacoes():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_avaliacoes_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao carregar avaliacoes: {e}")
        return jsonify({"erro": "Erro ao carregar avaliações"}), 500

@app.route("/admin/avaliacoes", methods=["DELETE"])
def del_avaliacoes():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        _apagar_avaliacoes_db()
        return jsonify({"status": "apagado"})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar avaliações: {e}")
        return jsonify({"erro": "Erro ao apagar avaliações"}), 500

@app.route("/admin/cardapio", methods=["PUT"])
def put_cardapio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    dados = request.get_json()
    salvar_json(CARDAPIO_FILE, dados)
    _sync_nuvem("/admin/cardapio", "PUT", dados)
    return jsonify({"status": "ok"})

@app.route("/admin/eventos", methods=["POST"])
def add_evento():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    ev = request.get_json()
    eventos = ler_json(EVENTOS_FILE, EVENTOS_PADRAO); eventos.append(ev)
    salvar_json(EVENTOS_FILE, eventos)
    _sync_nuvem("/admin/eventos", "PUT", eventos)
    return jsonify({"status": "ok"})

@app.route("/admin/eventos", methods=["DELETE"])
def del_evento():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    ev = request.get_json()
    eventos = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
    eventos = [e for e in eventos if not (e["data"]==ev["data"] and e["evento"]==ev["evento"])]
    salvar_json(EVENTOS_FILE, eventos)
    _sync_nuvem("/admin/eventos", "PUT", eventos)
    return jsonify({"status": "ok"})

@app.route("/admin/config", methods=["PUT"])
def put_config():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    dados = request.get_json()
    salvar_json(CONFIG_FILE, dados)
    _sync_nuvem("/admin/config", "PUT", dados)
    return jsonify({"status": "ok"})

# ==============================================================================
# ROTAS FLASK — REFEITÓRIO
# ==============================================================================
def _hoje():
    return _agora_br().strftime("%d/%m/%Y")

def _aula_por_hora(hora):
    """Retorna a aula ou intervalo correspondente ao horário de entrada."""
    try:
        hora_limpa = hora.strip().split(" ")[0]
        if len(hora_limpa) == 5:
            hora_limpa += ":00"
        t = datetime.datetime.strptime(hora_limpa, "%H:%M:%S").time()
    except Exception:
        return "Fora do horário"

    periodos = [
        ("07:10:00", "08:00:00", "Aula 1"),
        ("08:00:00", "08:50:00", "Aula 2"),
        ("08:50:00", "09:10:00", "Intervalo"),
        ("09:10:00", "10:00:00", "Aula 3"),
        ("10:00:00", "10:50:00", "Aula 4"),
        ("10:50:00", "11:40:00", "Aula 5"),
        ("11:40:00", "13:00:00", "Intervalo"),
        ("13:00:00", "13:50:00", "Aula 6"),
        ("13:50:00", "14:40:00", "Aula 7"),
        ("14:40:00", "15:00:00", "Intervalo"),
        ("15:00:00", "15:50:00", "Aula 8"),
        ("15:50:00", "16:40:00", "Aula 9"),
    ]
    for inicio, fim, nome_aula in periodos:
        if datetime.time.fromisoformat(inicio) <= t < datetime.time.fromisoformat(fim):
            return nome_aula
    return "Fora do horário"

def _registros_hoje():
    try:
        return _ler_refeitorio_hoje_db()
    except Exception as e:
        logger.error(f"Erro ao ler refeitorio do banco: {e}")
        return []

@app.route("/refeitorio/registrar", methods=["POST"])
def registrar_refeicao():
    dados = request.get_json()
    if not dados:
        return jsonify({"erro": "JSON invalido"}), 400
    matricula = dados.get("matricula", "").strip()
    refeicao  = dados.get("refeicao", "almoco").strip().lower()
    nome      = dados.get("nome", "Desconhecido").strip()
    serie     = dados.get("serie", "N/A").strip()
    curso     = dados.get("curso", "N/A").strip()

    # Limpa dados malformados do leitor USB (não use locals())
    nome = nome or "Desconhecido"
    serie = serie or "N/A"
    curso = curso or "N/A"
    matricula = matricula or ""

    def _limpar_campo(valor: str) -> str:
        if not valor:
            return valor
        v = valor
        if "^" in v or "Ç" in v:
            v = v.replace("^Ç", ":").replace("^", "").replace("Ç", "")
        v = re.sub(r"[^a-zA-Z0-9À-ÿ\s.'-]", "", v).strip()
        return v

    nome = _limpar_campo(nome)
    serie = _limpar_campo(serie)
    curso = _limpar_campo(curso)
    matricula = _limpar_campo(matricula)

    today      = _hoje()
    if not matricula:
        return jsonify({"erro": "Matricula nao informada"}), 400

    db_duplicado = _refeitorio_duplicado_db(matricula, refeicao)
    if db_duplicado:
        return jsonify({
            "status": "ja_registrado",
            "nome": db_duplicado["nome"],
            "hora": db_duplicado["hora"],
            "total_hoje": db_duplicado["total_hoje"],
            "total_refeicao": db_duplicado["total_refeicao"],
        }), 200

    hora = _agora_br().strftime("%H:%M:%S")
    periodo = _aula_por_hora(hora)
    registro = [today, hora, matricula, nome, serie, curso, refeicao]

    try:
        _inserir_refeitorio_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao registrar refeição: {e}")
        return jsonify({"erro": "Erro ao registrar refeição"}), 500
    logger.info(f"Refeição registrada: {nome} ({matricula}) - {refeicao}")
    # Contagens após inserir
    registros = _ler_refeitorio_hoje_db()
    total_hoje = len(registros)
    total_refeicao = sum(1 for r in registros if r[6] == refeicao)
    return jsonify({"status": "ok", "nome": nome, "hora": hora, "aula": periodo,
                    "total_hoje": total_hoje, "total_refeicao": total_refeicao}), 200

@app.route("/refeitorio/hoje", methods=["GET"])
def get_refeitorio_hoje():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_refeitorio_hoje_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao ler refeitorio: {e}")
        return jsonify({"erro": "Erro ao ler dados de refeitório"}), 500

@app.route("/refeitorio/exportar", methods=["GET"])
def exportar_refeitorio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        linhas = _ler_refeitorio_todos_db()
        buf = _csv_bytes(
            ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
            linhas,
        )
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"refeitorio_{_hoje().replace('/', '_')}.csv",
            mimetype="text/csv",
        )
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao exportar refeitorio: {e}")
        return jsonify({"erro": "Erro ao exportar refeitório"}), 500

@app.route("/refeitorio/apagar", methods=["DELETE"])
def apagar_refeitorio():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    data_alvo = request.args.get("data", _hoje())
    try:
        _apagar_refeitorio_data_db(data_alvo)
        return jsonify({"status": "apagado", "data": data_alvo})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar refeitorio: {e}")
        return jsonify({"erro": "Erro ao apagar refeitório"}), 500

@app.route("/refeitorio/qrcode/<matricula>", methods=["GET"])
def gerar_qrcode_img(matricula):
    nome  = request.args.get("nome", "")
    serie = request.args.get("serie", "")
    curso = request.args.get("curso", "")
    payload = json.dumps({"matricula": matricula, "nome": nome, "serie": serie, "curso": curso}, ensure_ascii=False)
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
    qr.add_data(payload); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
    return send_file(buf, mimetype="image/png", download_name=f"qrcode_{matricula}.png")

# ==============================================================================
# ROTAS FLASK — FREQUÊNCIA
# ==============================================================================

def _registros_freq_hoje():
    try:
        return _ler_frequencia_hoje_db()
    except Exception as e:
        logger.error(f"Erro ao ler frequencia do banco: {e}")
        return []

@app.route("/frequencia/registrar", methods=["POST"])
def registrar_frequencia():
    dados = request.get_json()
    if not dados:
        return jsonify({"erro": "JSON invalido"}), 400
    matricula = dados.get("matricula", "").strip()
    nome      = dados.get("nome", "Desconhecido").strip()
    serie     = dados.get("serie", "N/A").strip()
    curso     = dados.get("curso", "N/A").strip()

    # Limpa dados malformados do leitor USB
    def _limpar_campo(valor: str) -> str:
        if not valor:
            return valor
        v = valor
        if "^" in v or "Ç" in v:
            v = v.replace("^Ç", ":").replace("^", "").replace("Ç", "")
        v = re.sub(r"[^a-zA-Z0-9À-ÿ\s.'-]", "", v).strip()
        return v

    nome = _limpar_campo(nome)
    serie = _limpar_campo(serie)
    curso = _limpar_campo(curso)
    matricula = _limpar_campo(matricula)

    nome = nome or "Desconhecido"
    serie = serie or "N/A"
    curso = curso or "N/A"

    hoje = _hoje()
    if not matricula:
        return jsonify({"erro": "Matricula nao informada"}), 400

    db_duplicado = _frequencia_duplicado_db(matricula)
    if db_duplicado:
        return jsonify({"status": "ja_registrado", "nome": db_duplicado["nome"], "hora": db_duplicado["hora"], "total_hoje": db_duplicado["total_hoje"]}), 200

    hora = _agora_br().strftime("%H:%M:%S")
    aula = _aula_por_hora(hora)
    registro = [hoje, hora, matricula, nome, serie, curso, aula]

    try:
        _inserir_frequencia_db(registro)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao registrar frequência: {e}")
        return jsonify({"erro": "Erro ao registrar frequência"}), 500

    logger.info(f"Frequência registrada: {nome} ({matricula})")

    registros = _ler_frequencia_hoje_db()
    total_hoje = len(registros)
    return jsonify({"status": "ok", "nome": nome, "hora": hora, "aula": aula, "total_hoje": total_hoje}), 200

@app.route("/frequencia/hoje", methods=["GET"])
def get_frequencia_hoje():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        registros = _ler_frequencia_hoje_db()
        return jsonify(registros)
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao ler frequencia: {e}")
        return jsonify({"erro": "Erro ao ler dados de frequência"}), 500

@app.route("/frequencia/exportar", methods=["GET"])
def exportar_frequencia():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    try:
        linhas = _ler_frequencia_todos_db()
        buf = _csv_bytes(
            ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
            linhas,
        )
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"frequencia_{_hoje().replace('/', '_')}.csv",
            mimetype="text/csv",
        )
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao exportar frequencia: {e}")
        return jsonify({"erro": "Erro ao exportar frequência"}), 500

@app.route("/frequencia/apagar", methods=["DELETE"])
def apagar_frequencia():
    if not checar_senha(request): return jsonify({"erro": "Acesso negado"}), 403
    data_alvo = request.args.get("data", _hoje())
    try:
        _apagar_frequencia_data_db(data_alvo)
        return jsonify({"status": "apagado", "data": data_alvo})
    except RuntimeError as e:
        logger.error(f"PostgreSQL indisponível: {e}")
        return jsonify({"erro": "Banco de dados indisponível"}), 503
    except Exception as e:
        logger.error(f"Erro ao apagar frequencia: {e}")
        return jsonify({"erro": "Erro ao apagar frequência"}), 500

@app.route("/admin/limpar_neon", methods=["DELETE"])
def admin_limpar_neon():
    if not checar_senha(request):
        return jsonify({"erro": "Acesso negado"}), 403
    try:
        _limpar_tabelas_neon()
        return jsonify({"status": "neon_limpo"}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

# ==============================================================================
# BACKUP POR E-MAIL (opcional)
# ==============================================================================
def _enviar_backup_email():
    """Envia backup compactado por e-mail."""
    try:
        cfg = ler_json(CONFIG_FILE, {})
        if not cfg.get("email_backup_ativo", False):
            return

        email_remetente = cfg.get("email_remetente", "")
        email_senha = cfg.get("email_senha_app", "")
        email_destino = cfg.get("email_destino", "")
        email_smtp = cfg.get("email_smtp", "smtp.gmail.com")
        email_porta = cfg.get("email_porta", 587)

        if not all([email_remetente, email_senha, email_destino]):
            logger.warning("Backup por e-mail desativado: credenciais incompletas")
            return

        ts = _agora_br().strftime("%Y%m%d_%H%M%S")
        zip_name = f"backup_marwin_{ts}.zip"
        pasta_exp, arquivos_csv = _exportar_backup_mensal_csv(f"email_{ts}")
        with zipfile.ZipFile(zip_name, "w") as zf:
            for nome_arq in arquivos_csv:
                zf.write(os.path.join(pasta_exp, nome_arq), arcname=nome_arq)
            if os.path.exists(LISTA_ALUNOS_FILE):
                zf.write(LISTA_ALUNOS_FILE, arcname="lista_alunos.json")

        # Enviar por SMTP
        msg = MIMEMultipart()
        msg["From"] = email_remetente
        msg["To"] = email_destino
        msg["Subject"] = f"Backup MARWIN - {_agora_br().strftime('%d/%m/%Y %H:%M')}"

        body = f"Backup automático do sistema MARWIN - {_agora_br().strftime('%d/%m/%Y %H:%M:%S')}"
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with open(zip_name, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename= {zip_name}")
            msg.attach(part)

        server = smtplib.SMTP(email_smtp, email_porta)
        server.starttls()
        server.login(email_remetente, email_senha)
        server.send_message(msg)
        server.quit()

        os.remove(zip_name)
        logger.info(f"Backup enviado por e-mail para {email_destino}")
    except Exception as e:
        logger.error(f"Erro ao enviar backup por e-mail: {e}")

def _agendar_backup_email():
    """Agenda backup automático para 23h30 todos os dias."""
    def _loop_backup():
        while True:
            agora = _agora_br()
            proxima_exec = agora.replace(hour=23, minute=30, second=0, microsecond=0)
            if agora >= proxima_exec:
                proxima_exec += datetime.timedelta(days=1)
            tempo_espera = (proxima_exec - agora).total_seconds()
            time.sleep(tempo_espera)
            _enviar_backup_email()

    thread = threading.Thread(target=_loop_backup, daemon=True)
    thread.start()

def _iniciar_fluxo_backup_mes(btn_backup=None):
    """Fluxo guiado: nome do arquivo → exportar CSV → confirmação dupla → limpar banco."""
    def _habilitar_btn(estado):
        if btn_backup is not None:
            try:
                btn_backup.config(state=estado)
            except Exception:
                pass

    _habilitar_btn("disabled")

    dlg_nome = tk.Toplevel(janela)
    dlg_nome.title("Backup do Mês")
    dlg_nome.configure(bg=T["BG_CARD"])
    dlg_nome.transient(janela)
    dlg_nome.grab_set()
    dlg_nome.resizable(False, False)

    tk.Label(dlg_nome, text="Nome do arquivo de backup",
             font=("Segoe UI", 12, "bold"), bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"]).pack(padx=20, pady=(16, 6))
    tk.Label(dlg_nome, text="Os CSVs serão salvos em dados/backups/AAAA_MM/",
             font=("Segoe UI", 9), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(padx=20, pady=(0, 10))

    nome_var = tk.StringVar(value=_nome_backup_mes_sugerido())
    entry_nome = tk.Entry(dlg_nome, textvariable=nome_var, font=("Segoe UI", 11),
                          bg=T["ENTRY_BG"], fg=T["FG_TEXT"], width=36)
    entry_nome.pack(padx=20, pady=(0, 16))
    entry_nome.focus_set()
    entry_nome.select_range(0, "end")

    resultado = {"cancelado": True, "nome": ""}

    def _cancelar_inicio():
        resultado["cancelado"] = True
        dlg_nome.destroy()
        _habilitar_btn("normal")

    def _confirmar_nome():
        nome = nome_var.get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "Informe um nome para o backup.", parent=dlg_nome)
            return
        resultado["cancelado"] = False
        resultado["nome"] = nome
        dlg_nome.destroy()

    btn_row = tk.Frame(dlg_nome, bg=T["BG_CARD"])
    btn_row.pack(pady=(0, 16))
    tk.Button(btn_row, text="Cancelar", command=_cancelar_inicio,
              bg=T["FRASE_FG"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)
    tk.Button(btn_row, text="Continuar", command=_confirmar_nome,
              bg=T["ACCENT_SOFT"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)

    dlg_nome.wait_window()
    if resultado["cancelado"]:
        return

    try:
        pasta, arquivos = _exportar_backup_mensal_csv(resultado["nome"])
        lista = "\n".join(f"• {a}" for a in arquivos)
        messagebox.showinfo(
            "Backup exportado",
            f"CSVs gerados com sucesso em:\n{pasta}\n\n{lista}",
        )
        logger.info(f"Backup do mês exportado: {pasta}")
    except Exception as e:
        logger.error(f"Falha ao exportar backup do mês: {e}")
        messagebox.showerror("Erro", f"Falha ao exportar backup:\n{e}")
        _habilitar_btn("normal")
        return

    if not messagebox.askyesno(
        "Confirmar exclusão",
        "Tem certeza que deseja apagar todos os dados do banco?\n"
        "Esta ação não pode ser desfeita.",
        icon="warning",
    ):
        messagebox.showinfo("Cancelado", "Nenhum dado foi removido do banco. O backup CSV foi mantido.")
        _habilitar_btn("normal")
        return

    dlg_final = tk.Toplevel(janela)
    dlg_final.title("Confirmação final")
    dlg_final.configure(bg=T["BG_CARD"])
    dlg_final.transient(janela)
    dlg_final.grab_set()
    dlg_final.resizable(False, False)

    tk.Label(
        dlg_final,
        text="Confirmação final: os dados serão permanentemente removidos do banco.\n"
             "Digite CONFIRMAR para prosseguir.",
        font=("Segoe UI", 10),
        bg=T["BG_CARD"],
        fg=T["FG_TEXT"],
        justify="left",
    ).pack(padx=20, pady=(16, 10))

    conf_var = tk.StringVar()
    entry_conf = tk.Entry(dlg_final, textvariable=conf_var, font=("Segoe UI", 11),
                          bg=T["ENTRY_BG"], fg=T["FG_TEXT"], width=28)
    entry_conf.pack(padx=20, pady=(0, 12))

    btn_apagar = tk.Button(
        dlg_final, text="Apagar dados do banco", state="disabled",
        bg="#c62828", fg="white", font=("Segoe UI", 10, "bold"),
        padx=14, pady=6, bd=0,
    )
    confirmado = {"ok": False}

    def _atualizar_btn(*_):
        if conf_var.get() == "CONFIRMAR":
            btn_apagar.config(state="normal", cursor="hand2")
        else:
            btn_apagar.config(state="disabled", cursor="")

    def _executar_apagar():
        confirmado["ok"] = True
        dlg_final.destroy()

    def _cancelar_final():
        confirmado["ok"] = False
        dlg_final.destroy()

    conf_var.trace_add("write", _atualizar_btn)
    btn_apagar.config(command=_executar_apagar)

    btn_row2 = tk.Frame(dlg_final, bg=T["BG_CARD"])
    btn_row2.pack(pady=(0, 16))
    tk.Button(btn_row2, text="Cancelar", command=_cancelar_final,
              bg=T["FRASE_FG"], fg="white", font=("Segoe UI", 10, "bold"),
              padx=14, pady=6, bd=0, cursor="hand2").pack(side="left", padx=6)
    btn_apagar.pack(in_=btn_row2, side="left", padx=6)

    dlg_final.wait_window()

    if not confirmado["ok"]:
        messagebox.showinfo("Cancelado", "Nenhum dado foi removido do banco. O backup CSV foi mantido.")
        _habilitar_btn("normal")
        return

    try:
        _limpar_tabelas_neon()
        messagebox.showinfo("Concluído", "Dados do banco removidos com sucesso. O backup CSV foi mantido.")
        logger.info("Backup do mês concluído: banco limpo após exportação")
    except Exception as e:
        logger.error(f"Falha ao limpar banco após backup: {e}")
        messagebox.showerror(
            "Erro",
            f"O backup CSV foi salvo, mas falhou ao apagar o banco:\n{e}",
        )
    finally:
        _habilitar_btn("normal")

# ==============================================================================
# PAINEL ADMIN TKINTER
# ==============================================================================
TEMAS = {
    "claro": {
        "BG_MAIN":"#f2f8f2","BG_CARD":"#ffffff","ACCENT_VIBRANT":"#1b5e20",
        "ACCENT_SOFT":"#388e3c","FG_TEXT":"#1a2a1a","BORDER_GRID":"#c8e6c9",
        "HIGHLIGHT_YELLOW":"#f9a825","OBS_BG":"#e8f5e9","BTN_CARDAPIO":"#2e7d32",
        "FRASE_FG":"#4a7a4a","ENTRY_BG":"#f0f8f0","BTN_VOLTAR":"#1b5e20"
    },
    "escuro": {
        "BG_MAIN":"#0a1a0a","BG_CARD":"#142014","ACCENT_VIBRANT":"#4caf50",
        "ACCENT_SOFT":"#388e3c","FG_TEXT":"#c8e6c9","BORDER_GRID":"#2e5e2e",
        "HIGHLIGHT_YELLOW":"#f9a825","OBS_BG":"#1a3a1a","BTN_CARDAPIO":"#2e7d32",
        "FRASE_FG":"#81c784","ENTRY_BG":"#1a2e1a","BTN_VOLTAR":"#1b5e20"
    }
}

janela = None
T = None
tema_atual = None
url_ngrok_global = ""
btn_tema = None

def carregar_tema():
    if os.path.exists(TEMA_FILE):
        with open(TEMA_FILE) as f: return json.load(f).get("tema","claro")
    return "claro"

def salvar_tema_tk(tema):
    with open(TEMA_FILE,"w") as f: json.dump({"tema":tema},f)

def pedir_senha_tk():
    win = tk.Toplevel(janela); win.title("Acesso Administrativo"); win.resizable(False, False)
    win.configure(bg=T["BG_MAIN"]); win.transient(janela); win.grab_set()
    win.update_idletasks()
    w, h = 420, 300
    sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # Barra verde no topo do diálogo
    barra = tk.Frame(win, bg=T["ACCENT_VIBRANT"], height=44)
    barra.pack(fill="x"); barra.pack_propagate(False)
    tk.Label(barra, text="🔒  Acesso Administrativo", font=("Segoe UI", 12, "bold"),
             bg=T["ACCENT_VIBRANT"], fg="white").pack(side="left", padx=16, pady=12)
    tk.Frame(win, bg=T["HIGHLIGHT_YELLOW"], height=3).pack(fill="x")

    frame = tk.Frame(win, bg=T["BG_CARD"], padx=36, pady=28)
    frame.pack(expand=True, fill="both", padx=24, pady=20)
    tk.Label(frame, text="Digite a senha para continuar",
             font=("Segoe UI", 11), bg=T["BG_CARD"], fg=T["FRASE_FG"]).pack(pady=(0, 16))
    senha_var = tk.StringVar()
    entry = tk.Entry(frame, textvariable=senha_var, show="*", font=("Segoe UI", 13),
                     bg=T["ENTRY_BG"], fg=T["FG_TEXT"], relief="solid", bd=1,
                     insertbackground=T["ACCENT_VIBRANT"])
    entry.pack(fill="x", ipady=8, pady=(0, 20)); entry.focus()
    resultado = {"senha": None}
    def confirmar(): resultado["senha"] = senha_var.get(); win.destroy()
    def cancelar(): win.destroy()
    bf = tk.Frame(frame, bg=T["BG_CARD"]); bf.pack(fill="x")
    tk.Button(bf, text="Cancelar", bg=T["OBS_BG"], fg=T["ACCENT_VIBRANT"],
              font=("Segoe UI", 11, "bold"), bd=0, padx=15, pady=10,
              cursor="hand2", command=cancelar).pack(side="left", expand=True, fill="x", padx=(0, 6))
    tk.Button(bf, text="Entrar", bg=T["ACCENT_VIBRANT"], fg="white",
              font=("Segoe UI", 11, "bold"), bd=0, padx=15, pady=10,
              cursor="hand2", activebackground=T["ACCENT_SOFT"],
              command=confirmar).pack(side="right", expand=True, fill="x", padx=(6, 0))
    win.bind("<Return>", lambda e: confirmar()); win.wait_window(); return resultado["senha"]
# ── Cores do novo tema ──────────────────────────────────────────────────────
VERDE_ESCURO   = "#0F3D2E"
VERDE_VIBRANTE = "#1F6B45"
VERDE_SIDEBAR  = "#123D2C"
VERDE_SELECAO  = "#1F6B45"
VERDE_HOVER    = "#1A4D38"
VERDE_CLARO    = "#E8F5E9"
AZUL_CLARO     = "#E3F2FD"
ROXO_CLARO     = "#F3E5F5"
LARANJA_CLARO  = "#FFF3E0"
VERMELHO_CLARO = "#FFEBEE"
CINZA_BG       = "#F5F6F8"
BRANCO         = "#FFFFFF"
TEXTO_CINZA    = "#6B7280"
TEXTO_ESCURO   = "#1F2937"
TEXTO_CLARO    = "#D7E8DE"


def abrir_painel_admin_ctk(event=None):
    """Abre o Painel Administrativo com o novo visual (CustomTkinter).
 
    Mantém a MESMA verificação de senha e as MESMAS funções de leitura
    de dados do painel original — só muda a interface.
    """
    senha = pedir_senha_tk()

    def _local_check(pw: str) -> bool:
        if not pw:
            return False
        if BCRYPT_AVAILABLE and isinstance(ADMIN_PASSWORD, str) and ADMIN_PASSWORD.startswith("$2"):
            try:
                return bcrypt.checkpw(pw.encode("utf-8"), ADMIN_PASSWORD.encode("utf-8"))
            except Exception:
                return False
        return secrets.compare_digest(pw, ADMIN_PASSWORD)

    if not _local_check(senha):
        if senha is not None:
            logger.warning("Tentativa de acesso admin com senha incorreta")
            messagebox.showerror("Acesso negado", "Senha incorreta!")
        return

    logger.info("Painel admin (CTk) aberto com sucesso")

    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("green")

    jd = ctk.CTkToplevel(janela)
    jd.title("Painel Administrativo — EEEP MARWIN")
    jd.geometry("1400x850")
    jd.state("zoomed")
    jd.configure(fg_color=CINZA_BG)
    jd.attributes("-topmost", True)
    jd.lift()
    jd.focus_force()
    jd.after(100, lambda: jd.attributes("-topmost", False))

    jd.grid_columnconfigure(0, weight=0)
    jd.grid_columnconfigure(1, weight=1)
    jd.grid_rowconfigure(0, weight=1)

    paginas = {}

    # ──────────────────────────────────────────────────────────────────
    # SIDEBAR
    # ──────────────────────────────────────────────────────────────────
    sidebar = ctk.CTkFrame(jd, width=230, corner_radius=0, fg_color=VERDE_SIDEBAR)
    sidebar.grid(row=0, column=0, sticky="nsew")
    sidebar.grid_propagate(False)

    topo = ctk.CTkFrame(sidebar, fg_color="transparent")
    topo.pack(fill="x", padx=20, pady=(24, 20))

    logo_path = _buscar_logo_png()
    logo_ref = None
    if logo_path:
        try:
            from PIL import Image
            img = Image.open(logo_path).convert("RGBA")
            img.thumbnail((40, 40), Image.LANCZOS)
            logo_ref = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            ctk.CTkLabel(topo, image=logo_ref, text="").pack(side="left")
        except Exception:
            ctk.CTkLabel(topo, text="🏫", font=("Segoe UI", 28), width=40).pack(side="left")
    else:
        ctk.CTkLabel(topo, text="🏫", font=("Segoe UI", 28), width=40).pack(side="left")
    jd._logo_ref = logo_ref  # mantém referência viva

    textos_logo = ctk.CTkFrame(topo, fg_color="transparent")
    textos_logo.pack(side="left", padx=(8, 0))
    ctk.CTkLabel(textos_logo, text="Painel Administrativo",
                  font=("Segoe UI", 14, "bold"), text_color="white").pack(anchor="w")
    ctk.CTkLabel(textos_logo, text="EEEP MARWIN",
                  font=("Segoe UI", 11), text_color=TEXTO_CLARO).pack(anchor="w")

    itens_menu = [
        ("🏠", "Visão Geral"),
        ("📋", "Avaliações"),
        ("📅", "Relatório Semanal"),
        ("🍽️", "Editar Cardápio"),
        ("🗓️", "Editar Eventos"),
        ("👤", "Refeitório"),
        ("⏱️", "Frequência"),
        ("🕘", "Histórico"),
        ("🔲", "QR Codes"),
        ("📄", "Logs"),
    ]

    botoes_menu = {}

    def selecionar_botao(nome):
        for n, b in botoes_menu.items():
            if n == nome:
                b.configure(fg_color=VERDE_SELECAO, text_color="white", hover_color=VERDE_SELECAO)
            else:
                b.configure(fg_color="transparent", text_color=TEXTO_CLARO, hover_color=VERDE_HOVER)

    # Páginas com muitos CTkOptionMenu que consomem handles de menu do Tkinter.
    # Essas páginas são destruídas e recriadas a cada navegação para evitar o
    # erro "No more menus can be allocated".
    PAGINAS_RECRIAR = {"Refeitório", "Frequência", "Histórico", "QR Codes",
                        "Editar Cardápio", "Editar Eventos", "Relatório Semanal"}

    def mostrar_pagina(nome):
        selecionar_botao(nome)
        _scroll_canvas.yview_moveto(0)  # Volta ao topo ao trocar de página

        # Destrói página anterior para liberar handles GDI/Tk
        if nome in paginas:
            try:
                plt.close("all")
                paginas[nome].destroy()
            except Exception:
                pass
            del paginas[nome]

        if nome not in paginas:
            try:
                if nome == "Visão Geral":
                    paginas[nome] = criar_pagina_visao_geral()
                elif nome == "Avaliações":
                    paginas[nome] = criar_pagina_avaliacoes()
                elif nome == "Relatório Semanal":
                    paginas[nome] = criar_pagina_relatorio_semanal()
                elif nome == "Editar Cardápio":
                    paginas[nome] = criar_pagina_cardapio()
                elif nome == "Editar Eventos":
                     paginas[nome] = criar_pagina_eventos()
                elif nome == "Refeitório":
                    paginas[nome] = criar_pagina_refeitorio()
                elif nome == "Frequência":
                    paginas[nome] = criar_pagina_frequencia()
                elif nome == "Histórico":
                    paginas[nome] = criar_pagina_historico()
                elif nome == "QR Codes":
                    paginas[nome] = criar_pagina_qrcodes()
                elif nome == "Logs":
                    paginas[nome] = criar_pagina_logs()
                else:
                    paginas[nome] = criar_pagina_em_construcao(nome)
            except Exception as e:
                import traceback
                paginas[nome] = criar_pagina_erro(nome, traceback.format_exc())
                print(traceback.format_exc())
        for pg in paginas.values():
            pg.grid_remove()
        paginas[nome].grid(row=0, column=0, sticky="nsew")

    def criar_pagina_erro(nome, erro_texto):
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(page, text=f"Erro ao abrir '{nome}'",
                      font=("Segoe UI", 16, "bold"), text_color="#F44336"
                      ).grid(row=0, column=0, sticky="w", padx=24, pady=(24, 8))
        box = ctk.CTkTextbox(page, font=("Consolas", 11))
        box.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 24))
        box.insert("1.0", erro_texto)
        box.configure(state="disabled")
        return page

    for icone, nome in itens_menu:
        btn = ctk.CTkButton(
            sidebar, text=f"  {icone}   {nome}", anchor="w", font=("Segoe UI", 13),
            fg_color="transparent", text_color=TEXTO_CLARO, hover_color=VERDE_HOVER,
            corner_radius=8, height=38,
            command=lambda n=nome: mostrar_pagina(n),
        )
        btn.pack(fill="x", padx=12, pady=2)
        botoes_menu[nome] = btn

    rodape = ctk.CTkFrame(sidebar, fg_color=VERDE_SELECAO, corner_radius=10)
    rodape.pack(fill="x", padx=12, pady=16, side="bottom")
    ctk.CTkLabel(rodape, text="👤", font=("Segoe UI", 22), text_color="white"
                   ).pack(side="left", padx=(12, 8), pady=12)
    info_rodape = ctk.CTkFrame(rodape, fg_color="transparent")
    info_rodape.pack(side="left", pady=12)
    ctk.CTkLabel(info_rodape, text="Administrador", font=("Segoe UI", 12, "bold"),
                  text_color="white").pack(anchor="w")
    ctk.CTkLabel(info_rodape, text="Sistema", font=("Segoe UI", 10),
                  text_color="#B2DFB4").pack(anchor="w")
    ctk.CTkLabel(info_rodape, text="● Online", font=("Segoe UI", 10),
                  text_color="#7BE08C").pack(anchor="w")

    # ──────────────────────────────────────────────────────────────────
    # ÁREA DE CONTEÚDO + HEADER
    # ──────────────────────────────────────────────────────────────────
    container = ctk.CTkFrame(jd, fg_color=CINZA_BG, corner_radius=0)
    container.grid(row=0, column=1, sticky="nsew")
    container.grid_rowconfigure(1, weight=1)
    container.grid_columnconfigure(0, weight=1)
    container.grid_columnconfigure(1, weight=0)

    header = ctk.CTkFrame(container, fg_color=VERDE_VIBRANTE, height=56, corner_radius=0)
    header.grid(row=0, column=0, sticky="ew")
    header.grid_propagate(False)
    ctk.CTkLabel(header, text="", fg_color="transparent").pack(side="left", expand=True)
    status = ctk.CTkFrame(header, fg_color="transparent")
    status.pack(side="right", padx=20)
    ctk.CTkLabel(status, text="●", font=("Segoe UI", 14), text_color="#4CAF50").pack(side="left")
    servidor_txt = f"  Servidor: {url_ngrok_global}" if url_ngrok_global else "  Servidor: indisponível"
    ctk.CTkLabel(status, text=servidor_txt, font=("Segoe UI", 12), text_color="white").pack(side="left")

    # Canvas de scroll único — compartilhado por todas as páginas.
    # Cada página é um CTkFrame normal colocado dentro deste canvas.
    # Isso evita o esgotamento de handles GDI que ocorre quando cada
    # página cria seu próprio CTkScrollableFrame com Canvas interno.
    _scroll_canvas = tk.Canvas(container, bg=CINZA_BG, highlightthickness=0)
    _scroll_canvas.grid(row=1, column=0, sticky="nsew")
    _vsb = ctk.CTkScrollbar(container, orientation="vertical",
                              command=_scroll_canvas.yview)
    _vsb.grid(row=1, column=1, sticky="ns")
    _scroll_canvas.configure(yscrollcommand=_vsb.set)

    _scroll_inner = ctk.CTkFrame(_scroll_canvas, fg_color=CINZA_BG, corner_radius=0)
    _scroll_inner.grid_columnconfigure(0, weight=1)
    _scroll_inner_id = _scroll_canvas.create_window((0, 0), window=_scroll_inner, anchor="nw")

    def _on_inner_configure(event):
        _scroll_canvas.configure(scrollregion=_scroll_canvas.bbox("all"))

    def _on_canvas_configure(event):
        _scroll_canvas.itemconfig(_scroll_inner_id, width=event.width)

    _scroll_inner.bind("<Configure>", _on_inner_configure)
    _scroll_canvas.bind("<Configure>", _on_canvas_configure)

    def _on_mousewheel(event):
        _scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    _scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    # ──────────────────────────────────────────────────────────────────
    # HELPERS DE LAYOUT (cards / tabelas)
    # ──────────────────────────────────────────────────────────────────
    def card_resumo(parent, row, col, icone, cor_fundo, cor_icone, titulo, valor, subtitulo):
        card = ctk.CTkFrame(parent, fg_color=BRANCO, corner_radius=12)
        card.grid(row=row, column=col, sticky="nsew", padx=8, pady=4)
        conteudo = ctk.CTkFrame(card, fg_color="transparent")
        conteudo.pack(fill="x", padx=18, pady=18)
        icone_box = ctk.CTkFrame(conteudo, fg_color=cor_fundo, corner_radius=10, width=48, height=48)
        icone_box.pack(side="left", padx=(0, 14))
        icone_box.pack_propagate(False)
        ctk.CTkLabel(icone_box, text=icone, font=("Segoe UI", 20),
                      text_color=cor_icone).place(relx=0.5, rely=0.5, anchor="center")
        textos = ctk.CTkFrame(conteudo, fg_color="transparent")
        textos.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(textos, text=titulo, font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")
        valor_lbl = ctk.CTkLabel(textos, text=valor, font=("Segoe UI", 26, "bold"), text_color=TEXTO_ESCURO)
        valor_lbl.pack(anchor="w")
        sub_lbl = ctk.CTkLabel(textos, text=subtitulo, font=("Segoe UI", 10), text_color=TEXTO_CINZA)
        sub_lbl.pack(anchor="w")
        return valor_lbl, sub_lbl

    def card_tabela(parent, titulo, colunas, linhas, rodape="", larguras=None):
        card = ctk.CTkFrame(parent, fg_color=BRANCO, corner_radius=12)
        topo = ctk.CTkFrame(card, fg_color="transparent")
        topo.pack(fill="x", padx=18, pady=(16, 10))
        ctk.CTkLabel(topo, text=titulo, font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")

        header_t = ctk.CTkFrame(card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_t.pack(fill="x", padx=18)
        n = len(colunas)
        for i, c in enumerate(colunas):
            w = larguras.get(i) if larguras else None
            ctk.CTkLabel(header_t, text=c, font=("Segoe UI", 10, "bold"), text_color="white",
                          width=w or 0, anchor="w").pack(side="left", expand=(i == n - 1),
                                                            fill="x", padx=8, pady=8)

        if not linhas:
            ctk.CTkLabel(card, text="Nenhum registro encontrado.", font=("Segoe UI", 11),
                          text_color=TEXTO_CINZA).pack(anchor="w", padx=18, pady=12)
        else:
            for linha in linhas:
                linha_frame = ctk.CTkFrame(card, fg_color="transparent")
                linha_frame.pack(fill="x", padx=18)
                for i, valor in enumerate(linha):
                    w = larguras.get(i) if larguras else None
                    ctk.CTkLabel(linha_frame, text=str(valor), font=("Segoe UI", 11),
                                  width=w or 0, anchor="w",
                                  text_color="#374151").pack(side="left", expand=(i == n - 1),
                                                                fill="x", padx=8, pady=8)
                ctk.CTkFrame(card, fg_color="#F0F0F0", height=1).pack(fill="x", padx=18)

        if rodape:
            ctk.CTkLabel(card, text=rodape, font=("Segoe UI", 10),
                          text_color="#2E7D32").pack(anchor="w", padx=18, pady=(10, 16))
        else:
            ctk.CTkFrame(card, fg_color="transparent", height=10).pack()
        return card

    # ──────────────────────────────────────────────────────────────────
    # PÁGINA: VISÃO GERAL (ligada ao banco)
    # ──────────────────────────────────────────────────────────────────
    def criar_pagina_visao_geral():
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure((0, 1, 2), weight=1)

        # Cabeçalho com data de hoje
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 16))
        cab.grid_columnconfigure(0, weight=1)

        textos = ctk.CTkFrame(cab, fg_color="transparent")
        textos.grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(textos, text="Bom dia, Administrador! 👋",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(textos, text="Aqui está um resumo das atividades de hoje.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        hoje_dt = _agora_br().date()
        dias_semana_pt = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
                            "Sexta-feira", "Sábado", "Domingo"]
        data_box = ctk.CTkFrame(cab, fg_color=BRANCO, corner_radius=10)
        data_box.grid(row=0, column=1, sticky="e", padx=4)
        ctk.CTkLabel(data_box, text="📅", font=("Segoe UI", 18)).pack(side="left", padx=(14, 6), pady=10)
        txt_data = ctk.CTkFrame(data_box, fg_color="transparent")
        txt_data.pack(side="left", padx=(0, 16), pady=8)
        ctk.CTkLabel(txt_data, text=_hoje(), font=("Segoe UI", 12, "bold"),
                      text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(txt_data, text=dias_semana_pt[hoje_dt.weekday()], font=("Segoe UI", 10),
                      text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Cards de resumo ──────────────────────────────────────────
        try:
            refeitorio_hoje = _ler_refeitorio_hoje_db()
        except Exception as e:
            logger.error(f"Erro ao ler refeitorio do banco: {e}")
            refeitorio_hoje = []

        try:
            freq_hoje = _ler_frequencia_hoje_db()
        except Exception as e:
            logger.error(f"Erro ao ler frequencia do banco: {e}")
            freq_hoje = []

        try:
            avaliacoes_todas = _ler_avaliacoes_db()
        except Exception as e:
            logger.error(f"Erro ao ler avaliacoes do banco: {e}")
            avaliacoes_todas = []

        # Conta avaliações da semana atual (segunda a domingo)
        inicio_semana = hoje_dt - datetime.timedelta(days=hoje_dt.weekday())
        fim_semana = inicio_semana + datetime.timedelta(days=6)
        avaliacoes_semana = 0
        for av in avaliacoes_todas:
            try:
                d = datetime.datetime.strptime(av["Data"], "%d/%m/%Y").date()
                if inicio_semana <= d <= fim_semana:
                    avaliacoes_semana += 1
            except Exception:
                continue

        card_resumo(page, 1, 0, "📋", VERDE_CLARO, VERDE_VIBRANTE,
                       "Refeições hoje", str(len(refeitorio_hoje)), "registros no refeitório")
        card_resumo(page, 1, 1, "👥", AZUL_CLARO, "#2196F3",
                       "Presenças hoje", str(len(freq_hoje)), "alunos registrados")
        card_resumo(page, 1, 2, "⭐", ROXO_CLARO, "#9C27B0",
                       "Avaliações (semana)", str(avaliacoes_semana), "respostas esta semana")

        # ── Últimas entradas no refeitório ─────────────────────────────
        ultimas_refeicoes = refeitorio_hoje[-5:][::-1]  # 5 mais recentes
        linhas_ref = [
            (r[1], r[3], r[4], r[5], r[6])  # hora, nome, serie, curso, refeicao
            for r in ultimas_refeicoes
        ]
        tabela1 = card_tabela(page, "📥  Últimas entradas no refeitório hoje",
                                 ["HORA", "NOME", "SÉRIE", "CURSO", "REFEIÇÃO"],
                                 linhas_ref, rodape=f"Total: {len(refeitorio_hoje)} registros")
        tabela1.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=(0, 8), pady=(8, 0))

        # ── Últimas presenças registradas hoje ──────────────────────────
        ultimas_presencas = freq_hoje[-5:][::-1]
        linhas_freq = [
            (f[1], f[3], f[4], f[5], f[6])  # hora, nome, serie, curso, aula
            for f in ultimas_presencas
        ]
        tabela2 = card_tabela(page, "👥  Últimas presenças registradas hoje",
                                 ["HORA", "NOME", "SÉRIE", "CURSO", "AULA"],
                                 linhas_freq, rodape=f"Total: {len(freq_hoje)} registros")
        tabela2.grid(row=2, column=2, columnspan=1, sticky="nsew", padx=(8, 0), pady=(8, 0))

        # ── Eventos cadastrados ───────────────────────────────────────
        try:
            eventos = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
        except Exception:
            eventos = []
        linhas_eventos = [(ev.get("data", ""), ev.get("evento", "")) for ev in eventos]
        tabela_eventos = card_tabela(page, "📅  Eventos cadastrados",
                                         ["DATA", "EVENTO"], linhas_eventos, larguras={0: 80})
        tabela_eventos.grid(row=3, column=0, columnspan=3, sticky="nsew", pady=(16, 0))

        # ── Importar CSV(s) para o banco de dados ──────────────────────
        def _dialogo_mapeamento_manual(nome_arquivo, colunas_csv):
            """Pede ao usuário a tabela de destino quando a detecção automática falha.
            Retorna 'refeitorio', 'frequencia', 'avaliacoes' ou None (cancelado)."""
            resultado = {"tabela": None}
            dlg = ctk.CTkToplevel(jd)
            dlg.title("Mapeamento manual de tabela")
            dlg.configure(fg_color=BRANCO)
            dlg.transient(jd)
            dlg.grab_set()
            dlg.resizable(False, False)
            dlg.geometry("440x320")

            ctk.CTkLabel(dlg, text="Tabela de destino não reconhecida",
                          font=("Segoe UI", 14, "bold"), text_color=VERDE_VIBRANTE
                          ).pack(padx=20, pady=(20, 4))
            ctk.CTkLabel(dlg, text=f"Arquivo: {nome_arquivo}",
                          font=("Segoe UI", 10, "italic"), text_color=TEXTO_CINZA
                          ).pack(padx=20)
            ctk.CTkLabel(dlg, text=f"Colunas detectadas: {', '.join(colunas_csv)}",
                          font=("Segoe UI", 9), text_color=TEXTO_ESCURO,
                          wraplength=400, justify="left").pack(padx=20, pady=(4, 12))
            ctk.CTkLabel(dlg, text="Selecione a tabela de destino para esta importação:",
                          font=("Segoe UI", 10), text_color=TEXTO_ESCURO).pack(padx=20, pady=(0, 8))

            tabela_var = ctk.StringVar(value="")
            opcoes = [
                ("refeitorio", "Refeitório  (Data, HoraEntrada, Matricula, Nome, Serie, Curso, Refeicao)"),
                ("frequencia", "Frequência  (Data, HoraEntrada, Matricula, Nome, Serie, Curso, Aula)"),
                ("avaliacoes", "Avaliações  (Data, Aluno, Serie, Curso, Estagio, Item, Nota)"),
            ]
            for val, texto in opcoes:
                ctk.CTkRadioButton(dlg, text=texto, variable=tabela_var, value=val,
                                    font=("Segoe UI", 10), fg_color=VERDE_VIBRANTE,
                                    hover_color=VERDE_ESCURO, wraplength=380
                                    ).pack(anchor="w", padx=30, pady=4)

            def _confirmar():
                if not tabela_var.get():
                    messagebox.showwarning("Aviso", "Selecione uma tabela de destino.", parent=dlg)
                    return
                resultado["tabela"] = tabela_var.get()
                dlg.destroy()

            def _cancelar():
                dlg.destroy()

            btn_r = ctk.CTkFrame(dlg, fg_color="transparent")
            btn_r.pack(pady=(16, 16))
            ctk.CTkButton(btn_r, text="Cancelar", command=_cancelar,
                           fg_color="#6B7280", hover_color="#4B5563",
                           font=("Segoe UI", 11, "bold")).pack(side="left", padx=6)
            ctk.CTkButton(btn_r, text="Importar", command=_confirmar,
                           fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                           font=("Segoe UI", 11, "bold")).pack(side="left", padx=6)

            dlg.wait_window()
            return resultado["tabela"]

        import_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        import_card.grid(row=4, column=0, columnspan=3, sticky="nsew", pady=(16, 0))

        ctk.CTkLabel(import_card, text="💾  Importar CSV(s) para o banco de dados",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO
                      ).pack(anchor="w", padx=18, pady=(16, 4))
        ctk.CTkLabel(import_card,
                      text="Selecione um ou mais arquivos CSV (refeitório, frequência ou avaliações) "
                           "para importar/restaurar no banco. A tabela de destino é detectada "
                           "automaticamente pelas colunas do arquivo; se não for reconhecida, "
                           "você poderá escolher manualmente.",
                      font=("Segoe UI", 11), text_color=TEXTO_CINZA,
                      wraplength=1000, justify="left").pack(anchor="w", padx=18, pady=(0, 10))

        linha_import = ctk.CTkFrame(import_card, fg_color="transparent")
        linha_import.pack(fill="x", padx=18, pady=(0, 6))

        var_ignorar_dup = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(linha_import, text="Ignorar duplicatas (recomendado)",
                          variable=var_ignorar_dup, fg_color=VERDE_VIBRANTE,
                          hover_color=VERDE_ESCURO, font=("Segoe UI", 11)
                          ).pack(side="left", padx=(0, 16))

        prog_import = ctk.CTkProgressBar(linha_import, width=240)
        prog_import.set(0)

        lbl_import_status = ctk.CTkLabel(import_card, text="", font=("Segoe UI", 10),
                                          text_color=TEXTO_CINZA, wraplength=1000, justify="left")

        def _selecionar_e_importar():
            from tkinter import filedialog
            pasta_backups = os.path.join(DADOS_DIR, "backups")
            inicial = pasta_backups if os.path.isdir(pasta_backups) else DADOS_DIR
            caminhos = filedialog.askopenfilenames(
                parent=jd,
                title="Selecionar CSV(s) para importar",
                initialdir=inicial,
                filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")],
            )
            if not caminhos:
                return

            # Pré-leitura e detecção da tabela de cada arquivo (rápido, local)
            tarefas = []
            for caminho in caminhos:
                nome_arq = os.path.basename(caminho)
                try:
                    with open(caminho, "r", encoding="utf-8") as f:
                        primeiras = list(csv.DictReader(f))
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao ler {nome_arq}:\n{e}")
                    continue
                if not primeiras:
                    continue
                colunas_csv = list(primeiras[0].keys())
                tabela = _detectar_tabela_csv(colunas_csv)
                if tabela is None:
                    tabela = _dialogo_mapeamento_manual(nome_arq, colunas_csv)
                    if tabela is None:
                        continue
                tarefas.append((caminho, tabela))

            if not tarefas:
                return

            ignorar_dup = var_ignorar_dup.get()
            total_arquivos = len(tarefas)

            btn_import.configure(state="disabled")
            prog_import.set(0)
            prog_import.pack(side="left", padx=(0, 12))
            lbl_import_status.configure(text="Iniciando importação...")
            lbl_import_status.pack(anchor="w", padx=18, pady=(0, 12))

            def _thread_body():
                resumo_total = {"inseridos": 0, "ignorados": 0, "erros": 0}
                detalhes = []

                for idx_arq, (caminho, tabela) in enumerate(tarefas, 1):
                    nome_arq = os.path.basename(caminho)

                    def _progresso(atual, total, _a=idx_arq, _t=total_arquivos, _nome=nome_arq):
                        frac = (_a - 1 + atual / max(total, 1)) / _t
                        page.after(0, lambda: (
                            prog_import.set(frac),
                            lbl_import_status.configure(text=f"[{_a}/{_t}] {_nome} — {atual}/{total}")
                        ))

                    try:
                        resultado = _importar_csv_para_banco_forcado(caminho, tabela, ignorar_dup, _progresso)
                    except Exception as e_imp:
                        resumo_total["erros"] += 1
                        detalhes.append(f"❌ {nome_arq}: {e_imp}")
                        continue

                    resumo_total["inseridos"] += resultado["inseridos"]
                    resumo_total["ignorados"] += resultado["ignorados"]
                    resumo_total["erros"] += resultado["erros"]
                    detalhes.append(
                        f"✅ {nome_arq} → {resultado['tabela']}: "
                        f"{resultado['inseridos']} inserido(s), "
                        f"{resultado['ignorados']} ignorado(s), "
                        f"{resultado['erros']} erro(s)."
                    )

                def _finalizar():
                    prog_import.set(1)
                    btn_import.configure(state="normal")
                    lbl_import_status.configure(text="Importação concluída.")
                    msg = (
                        f"Importação concluída!\n\n"
                        f"✅ Inseridos: {resumo_total['inseridos']}\n"
                        f"⏭ Ignorados: {resumo_total['ignorados']}\n"
                        f"❌ Erros: {resumo_total['erros']}\n\n"
                        + "\n".join(detalhes)
                    )
                    messagebox.showinfo("Resumo da importação", msg)
                    page.after(2500, lambda: (prog_import.pack_forget(), lbl_import_status.pack_forget()))

                page.after(0, _finalizar)

            threading.Thread(target=_thread_body, daemon=True).start()

        btn_import = ctk.CTkButton(linha_import, text="📂  Selecionar CSV(s) e importar para o banco",
                                     fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                                     font=("Segoe UI", 11, "bold"), height=34,
                                     command=_selecionar_e_importar)
        btn_import.pack(side="left")

        ctk.CTkFrame(import_card, fg_color="transparent", height=8).pack()

        return page

    def criar_pagina_em_construcao(nome):
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(0, weight=1)
        ctk.CTkLabel(page, text=f"Página '{nome}' — em construção\n(em breve será migrada para o novo visual)",
                      font=("Segoe UI", 16), text_color=TEXTO_CINZA, justify="center"
                      ).grid(row=0, column=0)
        return page

    mostrar_pagina("Visão Geral")
    # ──────────────────────────────────────────────────────────────────
    # PÁGINA: AVALIAÇÕES (ligada ao banco)
    # ──────────────────────────────────────────────────────────────────
    def criar_pagina_avaliacoes():
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(3, weight=1)

        # Título
        bloco = ctk.CTkFrame(page, fg_color="transparent")
        bloco.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 12))
        ctk.CTkLabel(bloco, text="📋  Avaliações", font=("Segoe UI", 20, "bold"),
                      text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(bloco, text="Acompanhe e filtre as avaliações dos alunos.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Card de filtros ────────────────────────────────────────────
        filtros_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        filtros_card.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 12))

        linha1 = ctk.CTkFrame(filtros_card, fg_color="transparent")
        linha1.pack(fill="x", padx=18, pady=(18, 6))

        def _campo(parent, label, widget_cls, **kwargs):
            blk = ctk.CTkFrame(parent, fg_color="transparent")
            blk.pack(side="left", padx=(0, 16))
            ctk.CTkLabel(blk, text=label, font=("Segoe UI", 11),
                          text_color=TEXTO_CINZA).pack(anchor="w")
            w = widget_cls(blk, width=170, height=34, **kwargs)
            w.pack(anchor="w", pady=(4, 0))
            return w

        ent_nome = _campo(linha1, "Nome", ctk.CTkEntry, placeholder_text="Buscar por nome...")
        ent_data = _campo(linha1, "Data (dd/mm/aaaa)", ctk.CTkEntry, placeholder_text="dd/mm/aaaa")

        combo_est = _campo(linha1, "Estágio", ctk.CTkComboBox,
                              values=["Todos", "Comida", "Limpeza", "Ensino", "Semana"])
        combo_est.set("Todos")

        # ── Estatísticas (preenchidas em carregar_dados) ────────────────
        stats_frame = ctk.CTkFrame(page, fg_color="transparent")
        stats_frame.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 12))
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        lbl_total, _ = card_resumo(stats_frame, 0, 0, "📋", AZUL_CLARO, "#2196F3",
                                       "Total de avaliações", "0", "")
        lbl_pos, sub_pos = card_resumo(stats_frame, 0, 1, "🙂", VERDE_CLARO, "#4CAF50",
                                           "Positivas", "0", "")
        lbl_neu, sub_neu = card_resumo(stats_frame, 0, 2, "😐", "#FFF8E1", "#FBC02D",
                                           "Neutras", "0", "")
        lbl_neg, sub_neg = card_resumo(stats_frame, 0, 3, "🙁", VERMELHO_CLARO, "#F44336",
                                           "Negativas", "0", "")

        # ── Tabela (scrollable) ──────────────────────────────────────────
        tabela_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=3, column=0, sticky="nsew", padx=24, pady=(0, 8))
        tabela_card.grid_rowconfigure(1, weight=1)
        tabela_card.grid_columnconfigure(0, weight=1)

        colunas = ["DATA", "ALUNO", "SÉRIE", "CURSO", "ESTÁGIO", "CATEGORIA", "ITEM", "RESPOSTA"]
        larguras = {0: 110, 1: 140, 2: 60, 3: 60, 4: 70, 5: 110, 6: 0, 7: 90}

        header_t = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_t.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        n = len(colunas)
        for i, c in enumerate(colunas):
            ctk.CTkLabel(header_t, text=c, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=larguras.get(i, 0),
                          anchor="w").pack(side="left", expand=(i == n - 1), fill="x", padx=8, pady=8)

        corpo = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo.grid(row=1, column=0, sticky="nsew", padx=18)

        lbl_cnt = ctk.CTkLabel(tabela_card, text="0 registro(s)", font=("Segoe UI", 10),
                                  text_color=TEXTO_CINZA)
        lbl_cnt.grid(row=2, column=0, sticky="w", padx=18, pady=(4, 14))
        _ativo_aval = {"vivo": True}
        page.bind("<Destroy>", lambda e: _ativo_aval.update({"vivo": False}))

        # ── Lógica (igual ao painel antigo) ──────────────────────────────
        MAPA_ESTAGIO = {"Comida": "1", "Limpeza": "2", "Ensino": "3", "Semana": "4"}

        def _eh_almoco_favorito(item):
            return "almoco favorito" in item.lower() or "almoço favorito" in item.lower()

        def carregar_dados():
            for w in corpo.winfo_children():
                w.destroy()
            lbl_cnt.configure(text="Carregando...")

            # Captura os filtros antes de sair da thread principal
            nome_f = ent_nome.get().strip().lower()
            data_f = ent_data.get().strip()
            est_f = combo_est.get()
            est_f = MAPA_ESTAGIO.get(est_f) if est_f and est_f != "Todos" else None

            def _buscar():
                try:
                    return _avaliacoes_para_linhas(), None
                except Exception as e:
                    return None, e

            def _renderizar(linhas, erro):
                if not _ativo_aval["vivo"] or not page.winfo_exists():
                    return
                if erro is not None:
                    messagebox.showerror("Erro", f"Falha ao carregar avaliações do banco:\n{erro}")
                    lbl_cnt.configure(text="0 registro(s)")
                    return

                total = pos = neu = neg = 0

                for idx, r in enumerate(linhas):
                    if len(r) < 7:
                        continue
                    data, aluno, serie, curso, estagio, item, nota = r

                    if nome_f and nome_f not in str(aluno).lower():
                        continue
                    if data_f and data_f not in str(data):
                        continue
                    if est_f and str(estagio) != est_f:
                        continue

                    categoria = "Almoço favorito" if _eh_almoco_favorito(item) else "Avaliação"
                    resposta = nota
                    if str(estagio) == "4" and not _eh_almoco_favorito(item):
                        try:
                            n = float(nota)
                            if n <= 1:
                                resposta = "Ruim"
                            elif n <= 3:
                                resposta = "Medio"
                            else:
                                resposta = "Bom"
                        except Exception:
                            pass

                    total += 1
                    if resposta == "Bom":
                        pos += 1
                    elif resposta == "Medio":
                        neu += 1
                    elif resposta == "Ruim":
                        neg += 1

                    linha_frame = ctk.CTkFrame(corpo, fg_color="transparent")
                    linha_frame.pack(fill="x")
                    valores = [data, aluno, serie, curso, estagio, categoria, item, resposta]
                    n = len(valores)
                    for i, valor in enumerate(valores):
                        ctk.CTkLabel(linha_frame, text=str(valor), font=("Segoe UI", 11),
                                      width=larguras.get(i, 0), anchor="w",
                                      text_color="#374151", wraplength=320 if i == 6 else 0
                                      ).pack(side="left", expand=(i == n - 1), fill="x", padx=8, pady=6)
                    ctk.CTkFrame(corpo, fg_color="#F0F0F0", height=1).pack(fill="x")

                lbl_cnt.configure(text=f"{total} registro(s)")
                lbl_total.configure(text=str(total))

                def _pct(v):
                    return f"{(v / total * 100):.1f}%" if total else "0%"

                lbl_pos.configure(text=str(pos)); sub_pos.configure(text=_pct(pos))
                lbl_neu.configure(text=str(neu)); sub_neu.configure(text=_pct(neu))
                lbl_neg.configure(text=str(neg)); sub_neg.configure(text=_pct(neg))

            def _thread_body():
                linhas, erro = _buscar()
                corpo.after(0, lambda: _renderizar(linhas, erro))

            threading.Thread(target=_thread_body, daemon=True).start()

        # Botões Buscar / Exportar PDF
        ctk.CTkButton(linha1, text="🔍 Buscar", fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                        width=100, height=34, font=("Segoe UI", 11, "bold"),
                        command=carregar_dados).pack(side="left", padx=(20, 4), pady=(18, 0))

        def exportar_pdf():
            def _gerar():
                try:
                    reader = [["Data", "Aluno", "Serie", "Curso", "Estagio", "Item", "Nota"]] + _avaliacoes_para_linhas()
                except Exception as e:
                    corpo.after(0, lambda: messagebox.showerror("Erro", f"Falha ao ler avaliações do banco:\n{e}"))
                    return
                if len(reader) <= 1:
                    corpo.after(0, lambda: messagebox.showwarning("Aviso", "Não há avaliações para exportar."))
                    return
                try:
                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.set_auto_page_break(True, margin=15)
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 16)
                    pdf.cell(0, 10, "AVALIACOES ESCOLARES - EEEP MARWIN", ln=True, align="C")
                    pdf.set_font("Arial", "", 12)
                    pdf.cell(0, 8, f"Data: {_agora_br().strftime('%d/%m/%Y %H:%M')}", ln=True)
                    pdf.ln(6)
                    pdf.set_font("Arial", "B", 11)
                    pdf.cell(0, 8, "Registros de avaliacao:", ln=True)
                    pdf.ln(3)
                    pdf.set_font("Arial", "", 10)
                    for row in reader[1:]:
                        if len(row) < 7:
                            continue
                        pdf.multi_cell(0, 6, f"Data: {row[0]} | Aluno: {row[1]} | Serie: {row[2]} | Curso: {row[3]}")
                        pdf.multi_cell(0, 6, f"Estagio: {row[4]} | Item: {row[5]} | Nota: {row[6]}")
                        pdf.ln(2)
                    nome_arq = f"avaliacoes_marwin_{_agora_br().strftime('%d_%m_%Y')}.pdf"
                    pdf.output(nome_arq)
                    corpo.after(0, lambda: messagebox.showinfo("Sucesso", f"PDF gerado com sucesso:\n{nome_arq}"))
                except Exception as e:
                    corpo.after(0, lambda: messagebox.showerror("Erro ao exportar", f"Falha ao gerar PDF:\n{e}"))

            threading.Thread(target=_gerar, daemon=True).start()

        ctk.CTkButton(linha1, text="⭳ Exportar PDF", fg_color="#374151", hover_color="#1F2937",
                        width=130, height=34, font=("Segoe UI", 11, "bold"),
                        command=exportar_pdf).pack(side="left", padx=4, pady=(18, 0))

        # ── Linha 2: Avaliações ativas / modo de leitura / apagar tudo ───
        linha2 = ctk.CTkFrame(filtros_card, fg_color="transparent")
        linha2.pack(fill="x", padx=18, pady=(4, 18))

        cfg_sys = ler_json(CONFIG_FILE, {"avaliacoes_ativas": True, "modo_leitura": "camera"})

        var_ativas = ctk.BooleanVar(value=cfg_sys.get("avaliacoes_ativas", True))
        var_modo = ctk.StringVar(value=cfg_sys.get("modo_leitura", "camera"))

        def salvar_config():
            cfg_sys["avaliacoes_ativas"] = var_ativas.get()
            cfg_sys["modo_leitura"] = var_modo.get()
            salvar_json(CONFIG_FILE, cfg_sys)
            try:
                _sync_nuvem("/admin/config", "PUT", cfg_sys)
            except Exception:
                pass

        ctk.CTkCheckBox(linha2, text="Avaliações Ativas", variable=var_ativas,
                          fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                          command=salvar_config).pack(side="left", padx=(0, 16))

        ctk.CTkLabel(linha2, text="Modo de leitura:", font=("Segoe UI", 11),
                      text_color=TEXTO_CINZA).pack(side="left", padx=(0, 8))
        ctk.CTkRadioButton(linha2, text="Webcam", variable=var_modo, value="camera",
                              fg_color=VERDE_VIBRANTE, command=salvar_config).pack(side="left", padx=6)
        ctk.CTkRadioButton(linha2, text="Leitor USB", variable=var_modo, value="usb",
                              fg_color=VERDE_VIBRANTE, command=salvar_config).pack(side="left", padx=6)

        def apagar_tudo():
            if messagebox.askyesno("Aviso", "Apagar todos os dados?"):
                try:
                    _apagar_avaliacoes_db()
                    carregar_dados()
                    messagebox.showinfo("Sucesso", "Avaliações apagadas do banco.")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao apagar avaliações:\n{e}")

        ctk.CTkButton(linha2, text="Apagar tudo", fg_color="#C62828", hover_color="#8E1F1F",
                        height=32, font=("Segoe UI", 11, "bold"),
                        command=apagar_tudo).pack(side="right")

        carregar_dados()
        return page

    def criar_pagina_relatorio_semanal():

        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)

        # ── Cabeçalho ────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", pady=(4, 16))
        ctk.CTkLabel(cab, text="Relatório Semanal 📊",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Refeições e presenças dos últimos 7 dias.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Cards de resumo semanal ───────────────────────────────────
        cards_row = ctk.CTkFrame(page, fg_color="transparent")
        cards_row.grid(row=1, column=0, sticky="ew", pady=(0, 16))
        cards_row.grid_columnconfigure((0, 1, 2), weight=1)

        lbl_ref_sem,  sub_ref_sem  = card_resumo(cards_row, 0, 0, "🍽️", VERDE_CLARO,  VERDE_VIBRANTE, "Refeições (semana)", "...", "registros")
        lbl_freq_sem, sub_freq_sem = card_resumo(cards_row, 0, 1, "⏱️", AZUL_CLARO,   "#1565C0",       "Presenças (semana)", "...", "alunos únicos")
        lbl_aval_sem, sub_aval_sem = card_resumo(cards_row, 0, 2, "📋", ROXO_CLARO,   "#6A1B9A",       "Avaliações (semana)", "...", "respostas")

        # ── Gráfico de linha matplotlib ───────────────────────────────
        grafico_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        grafico_card.grid(row=2, column=0, sticky="ew", pady=(0, 16))

        topo_graf = ctk.CTkFrame(grafico_card, fg_color="transparent")
        topo_graf.pack(fill="x", padx=18, pady=(16, 4))
        ctk.CTkLabel(topo_graf, text="Atividade diária — últimos 7 dias",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")
        lbl_atualizado = ctk.CTkLabel(topo_graf, text="", font=("Segoe UI", 10),
                                       text_color=TEXTO_CINZA)
        lbl_atualizado.pack(side="right")

        # Placeholder enquanto carrega
        frame_graf = ctk.CTkFrame(grafico_card, fg_color="transparent")
        frame_graf.pack(fill="x", padx=18, pady=(0, 16))
        lbl_carregando = ctk.CTkLabel(frame_graf, text="Carregando gráfico...",
                                       font=("Segoe UI", 12), text_color=TEXTO_CINZA)
        lbl_carregando.pack(pady=40)

        canvas_ref = {}  # guarda referência do canvas matplotlib
        _ativo = {"vivo": True}  # flag de cancelamento da thread

        def _renderizar_grafico(dados_ref, dados_freq, dados_aval):
            """Roda na thread principal — cria/atualiza o gráfico."""
            if not _ativo["vivo"] or not page.winfo_exists():
                return
            lbl_carregando.pack_forget()
            if "canvas" in canvas_ref:
                try:
                    canvas_ref["canvas"].get_tk_widget().pack_forget()
                    canvas_ref["canvas"].get_tk_widget().destroy()
                except Exception:
                    pass
                try:
                    plt.close(canvas_ref["fig"])
                except Exception:
                    pass
                canvas_ref.clear()

            datas_str = sorted(set(list(dados_ref.keys()) + list(dados_freq.keys()) + list(dados_aval.keys())))

            if not datas_str:
                ctk.CTkLabel(frame_graf, text="Sem dados na semana.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(pady=40)
                return

            try:
                datas_dt = [datetime.datetime.strptime(d, "%d/%m/%Y") for d in datas_str]
            except Exception:
                datas_dt = list(range(len(datas_str)))

            y_ref  = [dados_ref.get(d, 0)  for d in datas_str]
            y_freq = [dados_freq.get(d, 0) for d in datas_str]
            y_aval = [dados_aval.get(d, 0) for d in datas_str]

            fig, ax = plt.subplots(figsize=(10, 3.8))
            fig.patch.set_facecolor(BRANCO)
            ax.set_facecolor(CINZA_BG)

            ax.plot(datas_dt, y_ref,  marker="o", linewidth=2.5, color=VERDE_VIBRANTE,
                    label="Refeições",  markersize=7)
            ax.plot(datas_dt, y_freq, marker="s", linewidth=2.5, color="#1565C0",
                    label="Presenças", markersize=7)
            ax.plot(datas_dt, y_aval, marker="^", linewidth=2.5, color="#6A1B9A",
                    label="Avaliações", markersize=7, linestyle="--")

            for x, y in zip(datas_dt, y_ref):
                if y: ax.annotate(str(y), (x, y), textcoords="offset points",
                                  xytext=(0, 8), ha="center", fontsize=9, color=VERDE_VIBRANTE)
            for x, y in zip(datas_dt, y_freq):
                if y: ax.annotate(str(y), (x, y), textcoords="offset points",
                                  xytext=(0, 8), ha="center", fontsize=9, color="#1565C0")

            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m") if datas_dt and isinstance(datas_dt[0], datetime.datetime) else plt.FuncFormatter(lambda v, _: datas_str[int(v)] if int(v) < len(datas_str) else ""))
            ax.xaxis.set_major_locator(mdates.DayLocator() if datas_dt and isinstance(datas_dt[0], datetime.datetime) else plt.MaxNLocator(integer=True))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=0, ha="center", fontsize=9)

            ax.set_ylabel("Quantidade", fontsize=10)
            ax.legend(loc="upper left", fontsize=10, frameon=False)
            ax.spines[["top", "right"]].set_visible(False)
            ax.grid(axis="y", linestyle="--", alpha=0.4)
            ax.set_ylim(bottom=0)
            fig.tight_layout()

            canvas_tk = FigureCanvasTkAgg(fig, master=frame_graf)
            canvas_tk.draw()
            canvas_tk.get_tk_widget().pack(fill="x")
            canvas_ref["canvas"] = canvas_tk
            canvas_ref["fig"]    = fig

            lbl_atualizado.configure(
                text=f"Atualizado em {_agora_br().strftime('%H:%M:%S')}")

        # ── Tabela diária detalhada ───────────────────────────────────
        tabela_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=3, column=0, sticky="ew", pady=(0, 16))
        topo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        topo_tab.pack(fill="x", padx=18, pady=(16, 10))
        ctk.CTkLabel(topo_tab, text="Detalhe por dia",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")

        # Cabeçalho da tabela
        header_tab = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_tab.pack(fill="x", padx=18)
        COLS_TAB = [("Data", 120), ("Dia da Semana", 160), ("Refeições", 120),
                    ("Presenças", 120), ("Avaliações", 120)]
        for i, (col, w) in enumerate(COLS_TAB):
            ctk.CTkLabel(header_tab, text=col, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=w, anchor="w"
                          ).pack(side="left", expand=(i == len(COLS_TAB)-1),
                                 fill="x", padx=8, pady=8)

        corpo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo_tab.pack(fill="x", padx=18)

        DIAS_SEMANA = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

        # ── Função de carga de dados ──────────────────────────────────
        def _carregar():
            hoje_d  = _agora_br().date()
            dias    = [hoje_d - datetime.timedelta(days=i) for i in range(6, -1, -1)]
            datas_s = [d.strftime("%d/%m/%Y") for d in dias]

            dados_ref  = {}
            dados_freq = {}
            dados_aval = {}

            try:
                for row in _ler_refeitorio_todos_db():
                    d = row[0]
                    if d in datas_s:
                        dados_ref[d] = dados_ref.get(d, 0) + 1
            except Exception:
                pass

            try:
                vistos = {}
                for row in _ler_frequencia_todos_db():
                    d, mat = row[0], row[2]
                    if d in datas_s:
                        key = (d, mat)
                        if key not in vistos:
                            vistos[key] = True
                            dados_freq[d] = dados_freq.get(d, 0) + 1
            except Exception:
                pass

            try:
                for a in _ler_avaliacoes_db():
                    d = a.get("Data", "").split(" ")[0]
                    if d in datas_s:
                        dados_aval[d] = dados_aval.get(d, 0) + 1
            except Exception:
                pass

            total_ref  = sum(dados_ref.values())
            total_freq = sum(dados_freq.values())
            total_aval = sum(dados_aval.values())

            # Atualiza na thread principal
            def _atualizar_ui():
                # Se a página foi destruída antes da thread terminar, aborta
                if not _ativo["vivo"] or not page.winfo_exists():
                    return
                lbl_ref_sem.configure(text=str(total_ref))
                lbl_freq_sem.configure(text=str(total_freq))
                lbl_aval_sem.configure(text=str(total_aval))

                # Limpa corpo da tabela
                for w in corpo_tab.winfo_children():
                    w.destroy()

                for i, (data_s, dia_d) in enumerate(zip(datas_s, dias)):
                    r_ref  = dados_ref.get(data_s, 0)
                    r_freq = dados_freq.get(data_s, 0)
                    r_aval = dados_aval.get(data_s, 0)
                    dia_nome = DIAS_SEMANA[dia_d.weekday()]

                    bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                    linha_f = ctk.CTkFrame(corpo_tab, fg_color=bg, corner_radius=0)
                    linha_f.pack(fill="x")

                    valores = [(data_s, 120), (dia_nome, 160),
                               (str(r_ref),  120), (str(r_freq), 120), (str(r_aval), 120)]
                    n = len(valores)
                    for j, (val, w) in enumerate(valores):
                        cor_val = VERDE_VIBRANTE if j >= 2 and int(val) > 0 else "#374151"
                        ctk.CTkLabel(linha_f, text=val, font=("Segoe UI", 11),
                                      width=w, anchor="w",
                                      text_color=cor_val
                                      ).pack(side="left", expand=(j == n-1),
                                             fill="x", padx=8, pady=7)
                    ctk.CTkFrame(corpo_tab, fg_color="#F0F0F0", height=1).pack(fill="x")

                # Gráfico
                _renderizar_grafico(dados_ref, dados_freq, dados_aval)

            page.after(0, _atualizar_ui)

        def _limpar_ao_destruir(event=None):
            _ativo["vivo"] = False  # sinaliza threads para pararem
            if "canvas" in canvas_ref:
                try:
                    canvas_ref["canvas"].get_tk_widget().pack_forget()
                    canvas_ref["canvas"].get_tk_widget().destroy()
                except Exception:
                    pass
                try:
                    plt.close(canvas_ref["fig"])
                except Exception:
                    pass
                canvas_ref.clear()

        page.bind("<Destroy>", _limpar_ao_destruir)

        threading.Thread(target=_carregar, daemon=True).start()

        # Botão Atualizar
        btn_row = ctk.CTkFrame(page, fg_color="transparent")
        btn_row.grid(row=4, column=0, sticky="w", pady=(0, 24))
        ctk.CTkButton(btn_row, text="↻  Atualizar", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, width=130, height=36,
                       font=("Segoe UI", 11, "bold"),
                       command=lambda: threading.Thread(target=_carregar, daemon=True).start()
                       ).pack(side="left")
        return page
    def criar_pagina_cardapio():
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", pady=(4, 8))
        ctk.CTkLabel(cab, text="Editar Cardápio 🍽️",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Edite as refeições de cada dia e clique em Salvar.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Botão Salvar (topo) ───────────────────────────────────────────────
        btn_bar = ctk.CTkFrame(page, fg_color="transparent")
        btn_bar.grid(row=1, column=0, sticky="ew", pady=(0, 16))

        lbl_status = ctk.CTkLabel(btn_bar, text="", font=("Segoe UI", 11),
                                   text_color=VERDE_VIBRANTE)
        lbl_status.pack(side="right", padx=(12, 0))

        # Constantes dos dias
        DIAS = ["SEGUNDA", "TERCA", "QUARTA", "QUINTA", "SEXTA"]
        DIAS_LABEL = {
            "SEGUNDA": "Segunda-feira",
            "TERCA":   "Terça-feira",
            "QUARTA":  "Quarta-feira",
            "QUINTA":  "Quinta-feira",
            "SEXTA":   "Sexta-feira",
        }
        REFEICOES = ["Merenda Manhã", "Almoço", "Merenda Tarde"]
        CORES_DIA = {
            "SEGUNDA": VERDE_VIBRANTE,
            "TERCA":   "#1565C0",
            "QUARTA":  "#6A1B9A",
            "QUINTA":  "#E65100",
            "SEXTA":   "#C62828",
        }

        # Carrega cardápio atual
        ca = ler_json(CARDAPIO_FILE, CARDAPIO_PADRAO)
        ents = {}  # dia -> [Entry, Entry, Entry]

        # ── Grade de cards ─────────────────────────────────────────────────────
        grade = ctk.CTkFrame(page, fg_color="transparent")
        grade.grid(row=2, column=0, sticky="ew")
        grade.grid_columnconfigure(0, weight=1, uniform="col")
        grade.grid_columnconfigure(1, weight=1, uniform="col")

        def _criar_card_dia(parent, dia, row, col, colspan=1):
            cor = CORES_DIA[dia]
            card = ctk.CTkFrame(parent, fg_color=BRANCO, corner_radius=12)
            if colspan == 2:
                card.grid(row=row, column=col, columnspan=2, sticky="ew",
                          padx=6, pady=6)
            else:
                card.grid(row=row, column=col, sticky="nsew", padx=6, pady=6)

            # Cabeçalho colorido
            hd = ctk.CTkFrame(card, fg_color=cor, corner_radius=8,
                               height=44)
            hd.pack(fill="x", padx=8, pady=(8, 0))
            hd.pack_propagate(False)
            ctk.CTkLabel(hd, text=f"  {DIAS_LABEL[dia]}",
                          font=("Segoe UI", 13, "bold"),
                          text_color="white").place(relx=0, rely=0.5, anchor="w", x=8)

            # Campos de refeição
            ents[dia] = []
            valores_dia = ca.get(dia, ["", "", ""])
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=14, pady=(10, 14))

            if colspan == 2:
                # Sexta: 3 campos lado a lado
                for j, nome_ref in enumerate(REFEICOES):
                    col_frame = ctk.CTkFrame(inner, fg_color="transparent")
                    col_frame.pack(side="left", expand=True, fill="both",
                                   padx=(0, 12 if j < 2 else 0))
                    ctk.CTkLabel(col_frame, text=nome_ref,
                                  font=("Segoe UI", 10, "bold"),
                                  text_color=TEXTO_CINZA).pack(anchor="w", pady=(0, 4))
                    e = ctk.CTkEntry(col_frame, font=("Segoe UI", 11),
                                      height=38, border_color=cor,
                                      border_width=2)
                    e.insert(0, valores_dia[j] if j < len(valores_dia) else "")
                    e.pack(fill="x")
                    ents[dia].append(e)
            else:
                # Seg–Qui: 3 campos empilhados
                for j, nome_ref in enumerate(REFEICOES):
                    ctk.CTkLabel(inner, text=nome_ref,
                                  font=("Segoe UI", 10, "bold"),
                                  text_color=TEXTO_CINZA).pack(anchor="w",
                                                               pady=(6 if j > 0 else 0, 4))
                    e = ctk.CTkEntry(inner, font=("Segoe UI", 11),
                                      height=38, border_color=cor,
                                      border_width=2)
                    e.insert(0, valores_dia[j] if j < len(valores_dia) else "")
                    e.pack(fill="x")
                    ents[dia].append(e)

        # Seg / Ter  (linha 0)
        _criar_card_dia(grade, "SEGUNDA", row=0, col=0)
        _criar_card_dia(grade, "TERCA",   row=0, col=1)
        # Qua / Qui  (linha 1)
        _criar_card_dia(grade, "QUARTA",  row=1, col=0)
        _criar_card_dia(grade, "QUINTA",  row=1, col=1)
        # Sexta em linha cheia (linha 2)
        _criar_card_dia(grade, "SEXTA",   row=2, col=0, colspan=2)

        # ── Salvar ────────────────────────────────────────────────────────────
        def salvar():
            novo = {dia: [e.get() for e in ents[dia]] for dia in DIAS}
            salvar_json(CARDAPIO_FILE, novo)
            try:
                _sync_nuvem("/admin/cardapio", "PUT", novo)
            except Exception:
                pass
            lbl_status.configure(text="✔  Cardápio salvo com sucesso!")
            page.after(3000, lambda: lbl_status.configure(text=""))

        ctk.CTkButton(btn_bar, text="💾  Salvar Alterações",
                       fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                       font=("Segoe UI", 12, "bold"), height=40, width=200,
                       command=salvar).pack(side="left")
        return page
    def criar_pagina_eventos():
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=0)
        page.grid_columnconfigure(1, weight=1)
        page.grid_rowconfigure(1, weight=1)

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=(4, 16))
        ctk.CTkLabel(cab, text="Editar Eventos 📅",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Adicione datas e descrições ao calendário escolar.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Coluna esquerda: formulário de novo evento ──────────────────────────
        form_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12, width=300)
        form_card.grid(row=1, column=0, sticky="ns", padx=(0, 16))
        form_card.grid_propagate(False)

        # Logo (se existir)
        logo_path = _buscar_logo_png()
        if logo_path:
            try:
                from PIL import Image as _PilImg
                _img_logo = _PilImg.open(logo_path).convert("RGBA")
                _img_logo.thumbnail((110, 110), _PilImg.LANCZOS)
                _ctk_logo = ctk.CTkImage(light_image=_img_logo, dark_image=_img_logo,
                                          size=_img_logo.size)
                ctk.CTkLabel(form_card, image=_ctk_logo, text="").pack(pady=(24, 8))
            except Exception:
                pass

        ctk.CTkLabel(form_card, text="Novo Evento",
                      font=("Segoe UI", 15, "bold"), text_color=TEXTO_ESCURO).pack(pady=(8, 2))
        ctk.CTkLabel(form_card, text="Cadastre datas especiais\ndo calendário escolar",
                      font=("Segoe UI", 10), text_color=TEXTO_CINZA,
                      justify="center").pack(pady=(0, 16))

        linha_div = ctk.CTkFrame(form_card, fg_color="#E5E7EB", height=1)
        linha_div.pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(form_card, text="Data", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(anchor="w", padx=20)
        ctk.CTkLabel(form_card, text="Formato: DD/MM  (ex: 15/07)",
                      font=("Segoe UI", 9), text_color=TEXTO_CINZA).pack(anchor="w", padx=20, pady=(0, 4))
        ent_data = ctk.CTkEntry(form_card, font=("Segoe UI", 13), height=38,
                                 border_color=VERDE_VIBRANTE, border_width=2,
                                 placeholder_text="01/01")
        ent_data.pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(form_card, text="Descrição", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(anchor="w", padx=20)
        ctk.CTkLabel(form_card, text="Nome do evento ou feriado",
                      font=("Segoe UI", 9), text_color=TEXTO_CINZA).pack(anchor="w", padx=20, pady=(0, 4))
        ent_desc = ctk.CTkEntry(form_card, font=("Segoe UI", 13), height=38,
                                 border_color=VERDE_VIBRANTE, border_width=2,
                                 placeholder_text="Ex: Feira Científica")
        ent_desc.pack(fill="x", padx=20, pady=(0, 8))

        lbl_form_status = ctk.CTkLabel(form_card, text="", font=("Segoe UI", 10),
                                        text_color=VERDE_VIBRANTE)
        lbl_form_status.pack(pady=(0, 4))

        # ── Coluna direita: lista de eventos cadastrados ────────────────────────
        lista_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        lista_card.grid(row=1, column=1, sticky="nsew")
        lista_card.grid_rowconfigure(2, weight=1)
        lista_card.grid_columnconfigure(0, weight=1)

        topo_lista = ctk.CTkFrame(lista_card, fg_color="transparent")
        topo_lista.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 4))
        ctk.CTkLabel(topo_lista, text="Eventos Cadastrados",
                      font=("Segoe UI", 14, "bold"), text_color=TEXTO_ESCURO).pack(side="left")
        lbl_total = ctk.CTkLabel(topo_lista, text="", font=("Segoe UI", 10),
                                  text_color=TEXTO_CINZA)
        lbl_total.pack(side="right")

        # Cabeçalho da tabela
        header_ev = ctk.CTkFrame(lista_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_ev.grid(row=1, column=0, sticky="ew", padx=18)
        ctk.CTkLabel(header_ev, text="Data", font=("Segoe UI", 10, "bold"),
                      text_color="white", width=90, anchor="w"
                      ).pack(side="left", padx=8, pady=8)
        ctk.CTkLabel(header_ev, text="Evento", font=("Segoe UI", 10, "bold"),
                      text_color="white", anchor="w"
                      ).pack(side="left", expand=True, fill="x", padx=8, pady=8)
        ctk.CTkLabel(header_ev, text="", font=("Segoe UI", 10, "bold"),
                      text_color="white", width=40
                      ).pack(side="right", padx=8, pady=8)

        # Corpo scrollável da lista
        corpo_ev = ctk.CTkFrame(lista_card, fg_color="transparent")
        corpo_ev.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 16))
        corpo_ev.grid_columnconfigure(0, weight=1)

        # ── Funções de carregar / adicionar / remover ───────────────────────────
        def _ordenar_eventos(evs):
            try:
                return sorted(evs, key=lambda x: (int(x["data"].split("/")[1]),
                                                    int(x["data"].split("/")[0])))
            except Exception:
                return evs

        def carregar_eventos():
            for w in corpo_ev.winfo_children():
                w.destroy()

            evs = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
            evs = _ordenar_eventos(evs)
            lbl_total.configure(text=f"{len(evs)} evento(s) cadastrado(s)")

            if not evs:
                ctk.CTkLabel(corpo_ev, text="Nenhum evento cadastrado.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA
                              ).grid(row=0, column=0, pady=24)
                return

            for i, ev in enumerate(evs):
                bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                linha = ctk.CTkFrame(corpo_ev, fg_color=bg, corner_radius=6)
                linha.grid(row=i, column=0, sticky="ew", pady=2)
                linha.grid_columnconfigure(1, weight=1)

                ctk.CTkLabel(linha, text=ev["data"], font=("Segoe UI", 11, "bold"),
                              text_color=VERDE_VIBRANTE, width=90, anchor="w"
                              ).grid(row=0, column=0, padx=8, pady=8, sticky="w")
                ctk.CTkLabel(linha, text=ev["evento"], font=("Segoe UI", 11),
                              text_color="#374151", anchor="w"
                              ).grid(row=0, column=1, padx=8, pady=8, sticky="ew")

                def _remover(e=ev):
                    if not messagebox.askyesno("Confirmar", f"Remover o evento '{e['evento']}' ({e['data']})?"):
                        return
                    atuais = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
                    atuais = [x for x in atuais if not (x["data"] == e["data"] and x["evento"] == e["evento"])]
                    salvar_json(EVENTOS_FILE, atuais)
                    try:
                        _sync_nuvem("/admin/eventos", "PUT", atuais)
                    except Exception:
                        pass
                    carregar_eventos()

                ctk.CTkButton(linha, text="🗑", width=36, height=28,
                               fg_color="#FEE2E2", hover_color="#FCA5A5",
                               text_color="#C62828", font=("Segoe UI", 12),
                               command=_remover
                               ).grid(row=0, column=2, padx=8, pady=8)

        def adicionar_evento():
            data = ent_data.get().strip()
            desc = ent_desc.get().strip()

            if not data or not desc:
                lbl_form_status.configure(text="⚠ Preencha a data e a descrição.",
                                           text_color="#C62828")
                return
            if "/" not in data or len(data) != 5:
                lbl_form_status.configure(text="⚠ Use o formato DD/MM (ex: 15/07).",
                                           text_color="#C62828")
                return

            evs = ler_json(EVENTOS_FILE, EVENTOS_PADRAO)
            evs.append({"data": data, "evento": desc})
            salvar_json(EVENTOS_FILE, evs)
            try:
                _sync_nuvem("/admin/eventos", "PUT", evs)
            except Exception:
                pass

            lbl_form_status.configure(text="✔ Evento adicionado!", text_color=VERDE_VIBRANTE)
            ent_data.delete(0, "end")
            ent_desc.delete(0, "end")
            ent_data.focus()
            carregar_eventos()
            page.after(2500, lambda: lbl_form_status.configure(text=""))

        ctk.CTkButton(form_card, text="➕  Adicionar Evento",
                       fg_color=VERDE_VIBRANTE, hover_color=VERDE_ESCURO,
                       font=("Segoe UI", 12, "bold"), height=40,
                       command=adicionar_evento
                       ).pack(fill="x", padx=20, pady=(4, 24))

        ent_desc.bind("<Return>", lambda e: adicionar_evento())
        ent_data.bind("<Return>", lambda e: ent_desc.focus())

        carregar_eventos()
        return page


    def criar_pagina_refeitorio():
        import unicodedata as _ud

        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(4, weight=1)

        # ── Mapas de siglas / cursos / normalização (mesma lógica do painel antigo) ──
        SIGLAS_LABEL = {
            "1DS": "1º DS",  "1HOS": "1º HOS", "1ENF": "1º ENF", "1MOD": "1º MOD",
            "2DS": "2º DS",  "2HOS": "2º HOS", "2ENF": "2º ENF", "2MOD": "2º MOD",
            "3DS": "3º DS",  "3HOS": "3º HOS", "3ENF": "3º ENF", "3MOD": "3º MOD",
        }
        SIGLAS = {
            "1DS":  {"serie": "1 Ano", "curso": "Desenvolvimento de Sistemas"},
            "1HOS": {"serie": "1 Ano", "curso": "Hospedagem"},
            "1ENF": {"serie": "1 Ano", "curso": "Enfermagem"},
            "1MOD": {"serie": "1 Ano", "curso": "Modelagem do Vestuario"},
            "2DS":  {"serie": "2 Ano", "curso": "Desenvolvimento de Sistemas"},
            "2HOS": {"serie": "2 Ano", "curso": "Hospedagem"},
            "2ENF": {"serie": "2 Ano", "curso": "Enfermagem"},
            "2MOD": {"serie": "2 Ano", "curso": "Modelagem do Vestuario"},
            "3DS":  {"serie": "3 Ano", "curso": "Desenvolvimento de Sistemas"},
            "3HOS": {"serie": "3 Ano", "curso": "Hospedagem"},
            "3ENF": {"serie": "3 Ano", "curso": "Enfermagem"},
            "3MOD": {"serie": "3 Ano", "curso": "Modelagem do Vestuario"},
        }
        CORES_ANO = {"1": VERDE_VIBRANTE, "2": "#6A1B9A", "3": "#C62828"}

        def _norm(texto):
            return _ud.normalize("NFD", str(texto).lower()).encode("ascii", "ignore").decode("ascii")

        _GRUPOS_SERIE = {
            "1": ["1º", "1o", "1 ano", "primeiro ano", "primeiro", "1"],
            "2": ["2º", "2o", "2 ano", "segundo ano",  "segundo",  "2"],
            "3": ["3º", "3o", "3 ano", "terceiro ano", "terceiro", "3"],
        }
        _GRUPOS_CURSO = {
            "ds":   ["ds", "desenvolvimento de sistemas", "dev. sistemas", "dev sistemas", "desenv. sistemas"],
            "enf":  ["enf", "enfermagem"],
            "hosp": ["hosp", "hospedagem"],
            "mod":  ["mod", "modelagem do vestuario", "modelagem", "vestuario"],
        }

        def _grupo_serie(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_SERIE.items():
                for v in variantes:
                    if t == v or t.startswith(v) or v.startswith(t):
                        return chave
            return None

        def _grupo_curso(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_CURSO.items():
                for v in variantes:
                    if t == v or t in v or v in t:
                        return chave
            return None

        def _serie_bate(serie_aluno, filtro_serie):
            if filtro_serie in ("(Todas)", ""):
                return True
            g_a, g_f = _grupo_serie(serie_aluno), _grupo_serie(filtro_serie)
            if g_a is None or g_f is None:
                return _norm(filtro_serie) in _norm(serie_aluno)
            return g_a == g_f

        def _curso_bate(curso_aluno, filtro_curso):
            if filtro_curso in ("(Todos)", ""):
                return True
            g_a, g_f = _grupo_curso(curso_aluno), _grupo_curso(filtro_curso)
            if g_a is None or g_f is None:
                return _norm(filtro_curso) in _norm(curso_aluno)
            return g_a == g_f

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 12))
        ctk.CTkLabel(cab, text="Refeitório 🍽️",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text=f"Controle de refeições — hoje ({_hoje()})",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Filtros ───────────────────────────────────────────────────────────
        filtro_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        filtro_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        filtro_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        filtro_row.pack(fill="x", padx=14, pady=(12, 8))

        ctk.CTkLabel(filtro_row, text="Série:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_serie = ctk.CTkOptionMenu(filtro_row, values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                      width=110, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_serie.set("(Todas)")
        cb_serie.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(filtro_row, text="Curso:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_curso = ctk.CTkOptionMenu(filtro_row,
                                      values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem",
                                              "Enfermagem", "Modelagem do Vestuario"],
                                      width=220, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_curso.set("(Todos)")
        cb_curso.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(filtro_row, text="Nome:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_nome = ctk.CTkEntry(filtro_row, width=180, height=30,
                                 placeholder_text="Buscar por nome...")
        ent_nome.pack(side="left", padx=(0, 14))

        # Salas rápidas
        sala_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        sala_row.pack(fill="x", padx=14, pady=(0, 12))
        ctk.CTkLabel(sala_row, text="Sala rápida:", font=("Segoe UI", 10, "bold"),
                      text_color=TEXTO_CINZA).pack(side="left", padx=(0, 8))

        btn_sala_refs = {}

        # ── Cards de resumo ───────────────────────────────────────────────────
        cards_row = ctk.CTkFrame(page, fg_color="transparent")
        cards_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        cards_row.grid_columnconfigure((0, 1, 2, 3), weight=1)

        lbl_total, sub_total = card_resumo(cards_row, 0, 0, "👥", VERDE_CLARO, VERDE_VIBRANTE,
                                            "Total Filtrado", "0", "alunos")
        lbl_sim,   sub_sim   = card_resumo(cards_row, 0, 1, "✔️", "#E8F5E9", "#2E7D32",
                                            "Almoçaram", "0", "confirmados")
        lbl_nao,   sub_nao   = card_resumo(cards_row, 0, 2, "✘", "#FFEBEE", "#C62828",
                                            "Não Almoçaram", "0", "pendentes")
        lbl_pct,   sub_pct   = card_resumo(cards_row, 0, 3, "📊", ROXO_CLARO, "#6A1B9A",
                                            "Adesão", "0%", "geral hoje")

        # ── Linha: gráfico de rosca + tabela ────────────────────────────────────
        meio_row = ctk.CTkFrame(page, fg_color="transparent")
        meio_row.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        meio_row.grid_columnconfigure(0, weight=0)
        meio_row.grid_columnconfigure(1, weight=1)

        # Donut
        donut_card = ctk.CTkFrame(meio_row, fg_color=BRANCO, corner_radius=12, width=260)
        donut_card.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        donut_card.grid_propagate(False)
        ctk.CTkLabel(donut_card, text="Adesão Geral",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(pady=(16, 4))

        cv_donut = tk.Canvas(donut_card, bg=BRANCO, highlightthickness=0, height=180)
        cv_donut.pack(fill="x", padx=20, pady=(4, 8))

        legenda_donut = ctk.CTkFrame(donut_card, fg_color="transparent")
        legenda_donut.pack(pady=(0, 16))

        # Tabela de alunos
        tabela_card = ctk.CTkFrame(meio_row, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=0, column=1, sticky="nsew")
        tabela_card.grid_rowconfigure(2, weight=1)
        tabela_card.grid_columnconfigure(0, weight=1)

        topo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        topo_tab.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        ctk.CTkLabel(topo_tab, text="Alunos — hoje",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")
        lbl_filtro_ativo = ctk.CTkLabel(topo_tab, text="", font=("Segoe UI", 9),
                                         text_color=TEXTO_CINZA)
        lbl_filtro_ativo.pack(side="right")

        header_tab = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_tab.grid(row=1, column=0, sticky="ew", padx=14)
        COLS = [("Aluno", 220), ("Série", 80), ("Curso", 200), ("Almoço", 90), ("Aula", 110)]
        for i, (col, w) in enumerate(COLS):
            ctk.CTkLabel(header_tab, text=col, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=w, anchor="w"
                          ).pack(side="left", expand=(i == len(COLS)-1),
                                 fill="x", padx=6, pady=8)

        corpo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo_tab.grid(row=2, column=0, sticky="nsew", padx=14, pady=(0, 14))
        corpo_tab.grid_columnconfigure(0, weight=1)

        # ── Botões de ação ────────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(page, fg_color="transparent")
        btn_row.grid(row=4, column=0, sticky="sw", pady=(0, 4))

        lbl_acao_status = ctk.CTkLabel(btn_row, text="", font=("Segoe UI", 10),
                                        text_color=VERDE_VIBRANTE)

        # ── Função de desenho do donut ───────────────────────────────────────
        def _desenhar_donut(sim, nao):
            cv_donut.delete("all")
            cv_donut.update_idletasks()
            w = cv_donut.winfo_width() or 220
            h = cv_donut.winfo_height() or 180
            cx, cy = w / 2, h / 2
            r_ext, r_int = min(w, h) / 2 - 10, (min(w, h) / 2 - 10) * 0.55

            total = sim + nao
            if total == 0:
                cv_donut.create_oval(cx - r_ext, cy - r_ext, cx + r_ext, cy + r_ext,
                                      fill="#F0F0F0", outline="")
                cv_donut.create_oval(cx - r_int, cy - r_int, cx + r_int, cy + r_int,
                                      fill=BRANCO, outline="")
                cv_donut.create_text(cx, cy, text="Sem\ndados", font=("Segoe UI", 11),
                                      fill=TEXTO_CINZA)
                return

            angulo = 90
            for valor, cor in [(sim, "#4CAF50"), (nao, "#F44336")]:
                if valor <= 0:
                    continue
                ext_ang = -(valor / total) * 360
                cv_donut.create_arc(cx - r_ext, cy - r_ext, cx + r_ext, cy + r_ext,
                                     start=angulo, extent=ext_ang,
                                     fill=cor, outline=BRANCO, width=2, style="pieslice")
                angulo += ext_ang

            cv_donut.create_oval(cx - r_int, cy - r_int, cx + r_int, cy + r_int,
                                  fill=BRANCO, outline="")
            pct = (sim / total * 100) if total else 0
            cv_donut.create_text(cx, cy - 8, text=f"{pct:.0f}%",
                                  font=("Segoe UI", 16, "bold"), fill="#2E7D32")
            cv_donut.create_text(cx, cy + 12, text="almoçaram",
                                  font=("Segoe UI", 8), fill=TEXTO_CINZA)

        def _atualizar_legenda(sim, nao):
            for w in legenda_donut.winfo_children():
                w.destroy()
            total = sim + nao
            for label_l, val_l, cor_l in [("Almoçaram", sim, "#4CAF50"), ("Não almoçaram", nao, "#F44336")]:
                pct_l = (val_l / total * 100) if total else 0
                row_l = ctk.CTkFrame(legenda_donut, fg_color="transparent")
                row_l.pack(fill="x", pady=1, padx=4)
                ctk.CTkFrame(row_l, fg_color=cor_l, width=12, height=12, corner_radius=2).pack(side="left", padx=(0, 6))
                ctk.CTkLabel(row_l, text=label_l, font=("Segoe UI", 10),
                              text_color=TEXTO_ESCURO).pack(side="left")
                ctk.CTkLabel(row_l, text=f"{val_l} ({pct_l:.0f}%)", font=("Segoe UI", 10, "bold"),
                              text_color=TEXTO_CINZA).pack(side="right")

        # ── Atualização principal ────────────────────────────────────────────
        def atualizar_ref():
            f_serie = cb_serie.get()
            f_curso = cb_curso.get()
            f_nome = ent_nome.get().strip().lower()

            def _thread_body():
                registros = _registros_hoje()
                page.after(0, lambda: _renderizar(registros, f_serie, f_curso, f_nome))

            threading.Thread(target=_thread_body, daemon=True).start()

        def _renderizar(registros, f_serie, f_curso, f_nome):
            alunos = {}
            for r in registros:
                mat   = r[2]
                nome  = r[3]
                serie = r[4] if len(r) > 4 else ""
                curso = r[5] if len(r) > 5 else ""
                ref   = r[6].strip().lower() if len(r) > 6 else ""
                hora  = r[1] if len(r) > 1 else ""
                if mat not in alunos:
                    aula = _aula_por_hora(hora)
                    alunos[mat] = {"nome": nome, "serie": serie, "curso": curso,
                                   "almoca": False, "hora": hora, "aula": aula}
                else:
                    if hora and alunos[mat].get("hora", "") and hora < alunos[mat]["hora"]:
                        alunos[mat]["hora"] = hora
                        alunos[mat]["aula"] = _aula_por_hora(hora)
                if ref == "almoco":
                    alunos[mat]["almoca"] = True

            exibidos = []
            for info in alunos.values():
                if not _serie_bate(info["serie"], f_serie): continue
                if not _curso_bate(info["curso"], f_curso): continue
                if f_nome and f_nome not in info["nome"].lower(): continue
                exibidos.append(info)

            cnt_total = len(exibidos)
            cnt_sim = sum(1 for a in exibidos if a["almoca"])
            cnt_nao = cnt_total - cnt_sim
            pct = (cnt_sim / cnt_total * 100) if cnt_total else 0

            lbl_total.configure(text=str(cnt_total))
            lbl_sim.configure(text=str(cnt_sim))
            lbl_nao.configure(text=str(cnt_nao))
            lbl_pct.configure(text=f"{pct:.0f}%")

            partes_filtro = []
            if f_serie not in ("(Todas)", ""): partes_filtro.append(f_serie)
            if f_curso not in ("(Todos)", ""): partes_filtro.append(f_curso)
            if f_nome: partes_filtro.append(f'"{f_nome}"')
            lbl_filtro_ativo.configure(
                text=("Filtro: " + " | ".join(partes_filtro)) if partes_filtro else "")

            _desenhar_donut(cnt_sim, cnt_nao)
            _atualizar_legenda(cnt_sim, cnt_nao)

            # Tabela
            for w in corpo_tab.winfo_children():
                w.destroy()

            if not exibidos:
                ctk.CTkLabel(corpo_tab, text="Nenhum registro encontrado.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA
                              ).grid(row=0, column=0, pady=24)
                return

            for i, info in enumerate(sorted(exibidos, key=lambda a: a["nome"].lower())):
                bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                cor_status = "#2E7D32" if info["almoca"] else "#C62828"
                texto_status = "✔ Sim" if info["almoca"] else "✘ Não"

                linha = ctk.CTkFrame(corpo_tab, fg_color=bg, corner_radius=4)
                linha.grid(row=i, column=0, sticky="ew", pady=1)

                valores = [(info["nome"] or "-", 220, "#374151"),
                           (info["serie"] or "-", 80, "#374151"),
                           (info["curso"] or "-", 200, "#374151"),
                           (texto_status, 90, cor_status),
                           (info.get("aula", "Fora do horário"), 110, "#374151")]
                n = len(valores)
                for j, (val, w, cor_t) in enumerate(valores):
                    ctk.CTkLabel(linha, text=val, font=("Segoe UI", 11),
                                  width=w, anchor="w", text_color=cor_t
                                  ).pack(side="left", expand=(j == n-1), fill="x", padx=6, pady=6)

        # ── Filtro por sala rápida ────────────────────────────────────────────
        def _filtrar_por_sigla(sigla):
            info = SIGLAS[sigla]
            cb_serie.set(info["serie"])
            cb_curso.set(info["curso"])
            ent_nome.delete(0, "end")
            for s, b in btn_sala_refs.items():
                b.configure(fg_color=CORES_ANO[s[0]], text_color="white")
            btn_sala_refs[sigla].configure(fg_color="white", text_color=CORES_ANO[sigla[0]])
            atualizar_ref()

        for sigla, label in SIGLAS_LABEL.items():
            ano = sigla[0]
            cor = CORES_ANO[ano]
            b = ctk.CTkButton(sala_row, text=label, font=("Segoe UI", 9, "bold"),
                               fg_color=cor, hover_color=VERDE_ESCURO,
                               text_color="white", width=64, height=26,
                               command=lambda s=sigla: _filtrar_por_sigla(s))
            b.pack(side="left", padx=2)
            btn_sala_refs[sigla] = b

        # ── Limpar filtros ────────────────────────────────────────────────────
        def limpar_filtros():
            cb_serie.set("(Todas)")
            cb_curso.set("(Todos)")
            ent_nome.delete(0, "end")
            for sigla, b in btn_sala_refs.items():
                b.configure(fg_color=CORES_ANO[sigla[0]], text_color="white")
            atualizar_ref()

        ctk.CTkButton(filtro_row, text="✕ Limpar", font=("Segoe UI", 10, "bold"),
                       fg_color="transparent", hover_color="#F0F0F0",
                       text_color=VERDE_VIBRANTE, width=80, height=30,
                       command=limpar_filtros).pack(side="left")

        # ── Exportar CSV ──────────────────────────────────────────────────────
        def exportar_csv_ref():
            try:
                nome_arq = f"refeitorio_{_hoje().replace('/', '_')}.csv"
                _escrever_csv(
                    nome_arq,
                    ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Refeicao"],
                    _ler_refeitorio_todos_db(),
                )
                lbl_acao_status.configure(text=f"✔ Exportado: {nome_arq}", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_acao_status.configure(text=f"⚠ Falha ao exportar: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_acao_status.configure(text=""))

        # ── Apagar registros de hoje ──────────────────────────────────────────
        def apagar_hoje():
            if not messagebox.askyesno("Confirmar", f"Apagar todos os registros de hoje ({_hoje()})?"):
                return
            try:
                _apagar_refeitorio_data_db(_hoje())
                lbl_acao_status.configure(text="✔ Registros de hoje apagados.", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_acao_status.configure(text=f"⚠ Falha ao apagar: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_acao_status.configure(text=""))
            atualizar_ref()

        ctk.CTkButton(btn_row, text="↻  Atualizar", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=36, width=130, command=atualizar_ref).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="⬇  Exportar CSV", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=36, width=150, command=exportar_csv_ref).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🗑  Apagar hoje", fg_color="#C62828",
                       hover_color="#8E1010", font=("Segoe UI", 11, "bold"),
                       height=36, width=140, command=apagar_hoje).pack(side="left", padx=(0, 12))
        lbl_acao_status.pack(side="left")

        # Gatilhos automáticos de filtro
        cb_serie.configure(command=lambda _: atualizar_ref())
        cb_curso.configure(command=lambda _: atualizar_ref())
        ent_nome.bind("<KeyRelease>", lambda e: atualizar_ref())

        # Carrega na primeira abertura
        page.after(150, atualizar_ref)
        cv_donut.bind("<Configure>", lambda e: atualizar_ref())

        return page
    def criar_pagina_frequencia():
        import unicodedata as _ud

        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(3, weight=1)

        # ── Mapas de siglas / cursos / normalização ──────────────────────────
        SIGLAS_LABEL = {
            "1DS": "1º DS",  "1HOS": "1º HOS", "1ENF": "1º ENF", "1MOD": "1º MOD",
            "2DS": "2º DS",  "2HOS": "2º HOS", "2ENF": "2º ENF", "2MOD": "2º MOD",
            "3DS": "3º DS",  "3HOS": "3º HOS", "3ENF": "3º ENF", "3MOD": "3º MOD",
        }
        SIGLAS = {
            "1DS":  {"serie": "1 Ano", "curso": "Desenvolvimento de Sistemas"},
            "1HOS": {"serie": "1 Ano", "curso": "Hospedagem"},
            "1ENF": {"serie": "1 Ano", "curso": "Enfermagem"},
            "1MOD": {"serie": "1 Ano", "curso": "Modelagem do Vestuario"},
            "2DS":  {"serie": "2 Ano", "curso": "Desenvolvimento de Sistemas"},
            "2HOS": {"serie": "2 Ano", "curso": "Hospedagem"},
            "2ENF": {"serie": "2 Ano", "curso": "Enfermagem"},
            "2MOD": {"serie": "2 Ano", "curso": "Modelagem do Vestuario"},
            "3DS":  {"serie": "3 Ano", "curso": "Desenvolvimento de Sistemas"},
            "3HOS": {"serie": "3 Ano", "curso": "Hospedagem"},
            "3ENF": {"serie": "3 Ano", "curso": "Enfermagem"},
            "3MOD": {"serie": "3 Ano", "curso": "Modelagem do Vestuario"},
        }
        CORES_ANO = {"1": VERDE_VIBRANTE, "2": "#6A1B9A", "3": "#C62828"}

        def _norm(texto):
            return _ud.normalize("NFD", str(texto).lower()).encode("ascii", "ignore").decode("ascii")

        _GRUPOS_SERIE = {
            "1": ["1º", "1o", "1 ano", "primeiro ano", "primeiro", "1"],
            "2": ["2º", "2o", "2 ano", "segundo ano",  "segundo",  "2"],
            "3": ["3º", "3o", "3 ano", "terceiro ano", "terceiro", "3"],
        }
        _GRUPOS_CURSO = {
            "ds":   ["ds", "desenvolvimento de sistemas", "dev. sistemas", "dev sistemas", "desenv. sistemas"],
            "enf":  ["enf", "enfermagem"],
            "hosp": ["hosp", "hospedagem"],
            "mod":  ["mod", "modelagem do vestuario", "modelagem", "vestuario"],
        }

        def _grupo_serie(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_SERIE.items():
                for v in variantes:
                    if t == v or t.startswith(v) or v.startswith(t):
                        return chave
            return None

        def _grupo_curso(texto):
            t = _norm(texto).strip()
            for chave, variantes in _GRUPOS_CURSO.items():
                for v in variantes:
                    if t == v or t in v or v in t:
                        return chave
            return None

        def _serie_bate(serie_aluno, filtro_serie):
            if filtro_serie in ("(Todas)", ""):
                return True
            g_a, g_f = _grupo_serie(serie_aluno), _grupo_serie(filtro_serie)
            if g_a is None or g_f is None:
                return _norm(filtro_serie) in _norm(serie_aluno)
            return g_a == g_f

        def _curso_bate(curso_aluno, filtro_curso):
            if filtro_curso in ("(Todos)", ""):
                return True
            g_a, g_f = _grupo_curso(curso_aluno), _grupo_curso(filtro_curso)
            if g_a is None or g_f is None:
                return _norm(filtro_curso) in _norm(curso_aluno)
            return g_a == g_f

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 12))
        ctk.CTkLabel(cab, text="Frequência ⏱️",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text=f"Controle de presença — hoje ({_hoje()})",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Filtros ───────────────────────────────────────────────────────────
        filtro_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        filtro_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        filtro_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        filtro_row.pack(fill="x", padx=14, pady=(12, 8))

        ctk.CTkLabel(filtro_row, text="Série:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_serie = ctk.CTkOptionMenu(filtro_row, values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                      width=110, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_serie.set("(Todas)")
        cb_serie.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(filtro_row, text="Curso:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_curso = ctk.CTkOptionMenu(filtro_row,
                                      values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem",
                                              "Enfermagem", "Modelagem do Vestuario"],
                                      width=220, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_curso.set("(Todos)")
        cb_curso.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(filtro_row, text="Nome:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_nome = ctk.CTkEntry(filtro_row, width=180, height=30,
                                 placeholder_text="Buscar por nome...")
        ent_nome.pack(side="left", padx=(0, 14))

        # Salas rápidas
        sala_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        sala_row.pack(fill="x", padx=14, pady=(0, 12))
        ctk.CTkLabel(sala_row, text="Sala rápida:", font=("Segoe UI", 10, "bold"),
                      text_color=TEXTO_CINZA).pack(side="left", padx=(0, 8))

        btn_sala_refs = {}

        # ── Cards de resumo ───────────────────────────────────────────────────
        cards_row = ctk.CTkFrame(page, fg_color="transparent")
        cards_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        cards_row.grid_columnconfigure((0, 1, 2), weight=1)

        lbl_total, sub_total = card_resumo(cards_row, 0, 0, "👥", VERDE_CLARO, VERDE_VIBRANTE,
                                            "Presentes (filtro)", "0", "alunos")
        lbl_cad,   sub_cad   = card_resumo(cards_row, 0, 1, "📋", AZUL_CLARO, "#1565C0",
                                            "Total Cadastrados", "0", "alunos na lista")
        lbl_aus,   sub_aus   = card_resumo(cards_row, 0, 2, "🚫", "#FFEBEE", "#C62828",
                                            "Ausentes", "0", "ainda não vieram")

        # ── Tabela de presentes ──────────────────────────────────────────────
        tabela_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=3, column=0, sticky="nsew")
        tabela_card.grid_rowconfigure(2, weight=1)
        tabela_card.grid_columnconfigure(0, weight=1)

        topo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        topo_tab.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        ctk.CTkLabel(topo_tab, text="Alunos Presentes — hoje",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")
        lbl_filtro_ativo = ctk.CTkLabel(topo_tab, text="", font=("Segoe UI", 9),
                                         text_color=TEXTO_CINZA)
        lbl_filtro_ativo.pack(side="right")

        header_tab = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_tab.grid(row=1, column=0, sticky="ew", padx=14)
        COLS = [("Aluno", 240), ("Série", 80), ("Curso", 220), ("Hora Entrada", 120), ("Aula", 110)]
        for i, (col, w) in enumerate(COLS):
            ctk.CTkLabel(header_tab, text=col, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=w, anchor="w"
                          ).pack(side="left", expand=(i == len(COLS)-1),
                                 fill="x", padx=6, pady=8)

        corpo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo_tab.grid(row=2, column=0, sticky="nsew", padx=14, pady=(0, 14))
        corpo_tab.grid_columnconfigure(0, weight=1)

        # ── Botões de ação ────────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(page, fg_color="transparent")
        btn_row.grid(row=4, column=0, sticky="w", pady=(12, 4))

        lbl_acao_status = ctk.CTkLabel(btn_row, text="", font=("Segoe UI", 10),
                                        text_color=VERDE_VIBRANTE)

        # ── Estado interno ────────────────────────────────────────────────────
        _estado = {"alunos_unicos": {}, "lista_todos": []}

        # ── Atualização principal ────────────────────────────────────────────
        def atualizar_freq():
            f_serie = cb_serie.get()
            f_curso = cb_curso.get()
            f_nome = ent_nome.get().strip().lower()

            def _thread_body():
                registros = _registros_freq_hoje()
                lista_todos = []
                if os.path.exists(LISTA_ALUNOS_FILE):
                    try:
                        with open(LISTA_ALUNOS_FILE, "r", encoding="utf-8") as f:
                            lista_todos = json.load(f)
                    except Exception:
                        lista_todos = []
                page.after(0, lambda: _renderizar(registros, lista_todos, f_serie, f_curso, f_nome))

            threading.Thread(target=_thread_body, daemon=True).start()

        def _renderizar(registros, lista_todos, f_serie, f_curso, f_nome):
            alunos_unicos = {}
            for r in registros:
                mat   = r[2]
                nome  = r[3]
                serie = r[4] if len(r) > 4 else ""
                curso = r[5] if len(r) > 5 else ""
                hora  = r[1] if len(r) > 1 else ""
                aula  = r[6] if len(r) > 6 else _aula_por_hora(hora)
                if mat not in alunos_unicos:
                    alunos_unicos[mat] = {"nome": nome, "serie": serie,
                                           "curso": curso, "hora": hora, "aula": aula}
                else:
                    if hora and alunos_unicos[mat].get("hora", "") and hora < alunos_unicos[mat]["hora"]:
                        alunos_unicos[mat]["hora"] = hora
                        alunos_unicos[mat]["aula"] = aula

            _estado["alunos_unicos"] = alunos_unicos
            _estado["lista_todos"] = lista_todos

            exibidos = []
            for info in alunos_unicos.values():
                if not _serie_bate(info["serie"], f_serie): continue
                if not _curso_bate(info["curso"], f_curso): continue
                if f_nome and f_nome not in info["nome"].lower(): continue
                exibidos.append(info)

            presentes_total = len(alunos_unicos)
            cadastrados_total = len(lista_todos)
            ausentes_total = max(0, cadastrados_total - presentes_total) if cadastrados_total else 0

            lbl_total.configure(text=str(len(exibidos)))
            lbl_cad.configure(text=str(cadastrados_total))
            lbl_aus.configure(text=str(ausentes_total))

            partes_filtro = []
            if f_serie not in ("(Todas)", ""): partes_filtro.append(f_serie)
            if f_curso not in ("(Todos)", ""): partes_filtro.append(f_curso)
            if f_nome: partes_filtro.append(f'"{f_nome}"')
            lbl_filtro_ativo.configure(
                text=("Filtro: " + " | ".join(partes_filtro)) if partes_filtro else "")

            for w in corpo_tab.winfo_children():
                w.destroy()

            if not exibidos:
                ctk.CTkLabel(corpo_tab, text="Nenhum registro encontrado.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA
                              ).grid(row=0, column=0, pady=24)
                return

            for i, info in enumerate(sorted(exibidos, key=lambda a: a["nome"].lower())):
                bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                linha = ctk.CTkFrame(corpo_tab, fg_color=bg, corner_radius=4)
                linha.grid(row=i, column=0, sticky="ew", pady=1)

                valores = [(info["nome"] or "-", 240),
                           (info["serie"] or "-", 80),
                           (info["curso"] or "-", 220),
                           (info.get("hora", "") or "-", 120),
                           (info.get("aula", "Fora do horário"), 110)]
                n = len(valores)
                for j, (val, w) in enumerate(valores):
                    cor_t = "#2E7D32" if j == 0 else "#374151"
                    fonte = ("Segoe UI", 11, "bold") if j == 0 else ("Segoe UI", 11)
                    ctk.CTkLabel(linha, text=val, font=fonte,
                                  width=w, anchor="w", text_color=cor_t
                                  ).pack(side="left", expand=(j == n-1), fill="x", padx=6, pady=6)

        # ── Filtro por sala rápida ────────────────────────────────────────────
        def _filtrar_por_sigla(sigla):
            info = SIGLAS[sigla]
            cb_serie.set(info["serie"])
            cb_curso.set(info["curso"])
            ent_nome.delete(0, "end")
            for s, b in btn_sala_refs.items():
                b.configure(fg_color=CORES_ANO[s[0]], text_color="white")
            btn_sala_refs[sigla].configure(fg_color="white", text_color=CORES_ANO[sigla[0]])
            atualizar_freq()

        for sigla, label in SIGLAS_LABEL.items():
            ano = sigla[0]
            cor = CORES_ANO[ano]
            b = ctk.CTkButton(sala_row, text=label, font=("Segoe UI", 9, "bold"),
                               fg_color=cor, hover_color=VERDE_ESCURO,
                               text_color="white", width=64, height=26,
                               command=lambda s=sigla: _filtrar_por_sigla(s))
            b.pack(side="left", padx=2)
            btn_sala_refs[sigla] = b

        # ── Limpar filtros ────────────────────────────────────────────────────
        def limpar_filtros():
            cb_serie.set("(Todas)")
            cb_curso.set("(Todos)")
            ent_nome.delete(0, "end")
            for sigla, b in btn_sala_refs.items():
                b.configure(fg_color=CORES_ANO[sigla[0]], text_color="white")
            atualizar_freq()

        ctk.CTkButton(filtro_row, text="✕ Limpar", font=("Segoe UI", 10, "bold"),
                       fg_color="transparent", hover_color="#F0F0F0",
                       text_color=VERDE_VIBRANTE, width=80, height=30,
                       command=limpar_filtros).pack(side="left")

        # ── Ver ausentes (janela secundária CTk) ─────────────────────────────
        def abrir_ausentes():
            lista_todos = _estado["lista_todos"]
            presentes = set(_estado["alunos_unicos"].keys())
            ausentes = [al for al in lista_todos if al.get("matricula", "") not in presentes]

            _sel_aus = {"selecionado": None, "linha_widgets": {}}

            win = ctk.CTkToplevel(page)
            win.title("Alunos Ausentes Hoje")
            win.geometry("760x560")
            win.configure(fg_color=CINZA_BG)

            ctk.CTkLabel(win, text="👥  Alunos Ausentes Hoje",
                          font=("Segoe UI", 16, "bold"), text_color=TEXTO_ESCURO
                          ).pack(anchor="w", padx=20, pady=(16, 4))

            filt = ctk.CTkFrame(win, fg_color="transparent")
            filt.pack(fill="x", padx=20, pady=(0, 8))
            ctk.CTkLabel(filt, text="Série:", font=("Segoe UI", 11, "bold"),
                          text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
            cb_serie_aus = ctk.CTkOptionMenu(filt, values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                              width=100, fg_color=VERDE_VIBRANTE,
                                              button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
            cb_serie_aus.set("(Todas)")
            cb_serie_aus.pack(side="left", padx=(0, 14))

            ctk.CTkLabel(filt, text="Curso:", font=("Segoe UI", 11, "bold"),
                          text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
            cb_curso_aus = ctk.CTkOptionMenu(filt, values=["(Todos)", "Desenvolvimento de Sistemas",
                                                            "Hospedagem", "Enfermagem", "Modelagem do Vestuario"],
                                              width=200, fg_color=VERDE_VIBRANTE,
                                              button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
            cb_curso_aus.set("(Todos)")
            cb_curso_aus.pack(side="left")

            lista_card = ctk.CTkFrame(win, fg_color=BRANCO, corner_radius=12)
            lista_card.pack(fill="both", expand=True, padx=20, pady=(0, 8))
            lista_card.grid_rowconfigure(1, weight=1)
            lista_card.grid_columnconfigure(0, weight=1)

            hdr = ctk.CTkFrame(lista_card, fg_color=VERDE_ESCURO, corner_radius=6)
            hdr.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 0))
            ctk.CTkLabel(hdr, text="Nome", font=("Segoe UI", 10, "bold"),
                          text_color="white", width=280, anchor="w").pack(side="left", padx=6, pady=8)
            ctk.CTkLabel(hdr, text="Série", font=("Segoe UI", 10, "bold"),
                          text_color="white", width=100, anchor="w").pack(side="left", padx=6, pady=8)
            ctk.CTkLabel(hdr, text="Curso", font=("Segoe UI", 10, "bold"),
                          text_color="white", anchor="w").pack(side="left", expand=True, fill="x", padx=6, pady=8)

            corpo_aus = ctk.CTkFrame(lista_card, fg_color="transparent")
            corpo_aus.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
            corpo_aus.grid_columnconfigure(0, weight=1)

            lbl_cnt_aus = ctk.CTkLabel(win, text="", font=("Segoe UI", 10), text_color=TEXTO_CINZA)
            lbl_cnt_aus.pack(anchor="w", padx=20, pady=(0, 4))

            lbl_sel_aus = ctk.CTkLabel(win, text="Nenhum aluno selecionado.",
                                         font=("Segoe UI", 10, "bold"), text_color=VERDE_ESCURO)
            lbl_sel_aus.pack(anchor="w", padx=20, pady=(0, 8))

            def _selecionar_aus(al):
                _sel_aus["selecionado"] = al
                mat_sel = al.get("matricula", "")
                for mat, (frame, bg_original) in _sel_aus["linha_widgets"].items():
                    frame.configure(fg_color=(bg_original if mat != mat_sel else "#E8F5E9"))
                lbl_sel_aus.configure(text=f"Selecionado: {al.get('nome', '-')}")
                btn_presenca.pack(side="left")

            def _preencher():
                _sel_aus["selecionado"] = None
                _sel_aus["linha_widgets"] = {}
                lbl_sel_aus.configure(text="Nenhum aluno selecionado.")
                btn_presenca.pack_forget()
                for w in corpo_aus.winfo_children():
                    w.destroy()
                f_serie = cb_serie_aus.get()
                f_curso = cb_curso_aus.get()
                exibidos = []
                for al in ausentes:
                    if not _serie_bate(al.get("serie", ""), f_serie): continue
                    if not _curso_bate(al.get("curso", ""), f_curso): continue
                    exibidos.append(al)

                lbl_cnt_aus.configure(text=f"{len(exibidos)} ausente(s) de {len(lista_todos)} cadastrado(s)")

                if not exibidos:
                    ctk.CTkLabel(corpo_aus, text="Nenhum ausente encontrado.",
                                  font=("Segoe UI", 12), text_color=TEXTO_CINZA
                                  ).grid(row=0, column=0, pady=24)
                    return

                for i, al in enumerate(sorted(exibidos, key=lambda a: a.get("nome", "").lower())):
                    bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                    linha = ctk.CTkFrame(corpo_aus, fg_color=bg, corner_radius=4, cursor="hand2")
                    linha.grid(row=i, column=0, sticky="ew", pady=1)

                    lbl_n = ctk.CTkLabel(linha, text=al.get("nome", "-"), font=("Segoe UI", 11),
                                  width=280, anchor="w", text_color="#C62828")
                    lbl_n.pack(side="left", padx=6, pady=6)
                    lbl_s = ctk.CTkLabel(linha, text=al.get("serie", "-"), font=("Segoe UI", 11),
                                  width=100, anchor="w", text_color="#374151")
                    lbl_s.pack(side="left", padx=6, pady=6)
                    lbl_c = ctk.CTkLabel(linha, text=al.get("curso", "-"), font=("Segoe UI", 11),
                                  anchor="w", text_color="#374151")
                    lbl_c.pack(side="left", expand=True, fill="x", padx=6, pady=6)

                    _sel_aus["linha_widgets"][al.get("matricula", "")] = (linha, bg)

                    def _bind_click_aus(widget, a=al):
                        widget.bind("<Button-1>", lambda e, _a=a: _selecionar_aus(_a))

                    for w in (linha, lbl_n, lbl_s, lbl_c):
                        _bind_click_aus(w)

            def _exportar():
                try:
                    nome_arq = f"ausentes_frequencia_{_hoje().replace('/', '_')}.csv"
                    f_serie = cb_serie_aus.get()
                    f_curso = cb_curso_aus.get()
                    exibidos = [al for al in ausentes
                                 if _serie_bate(al.get("serie", ""), f_serie)
                                 and _curso_bate(al.get("curso", ""), f_curso)]
                    with open(nome_arq, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        w.writerow(["Nome", "Série", "Curso"])
                        for al in sorted(exibidos, key=lambda a: a.get("nome", "").lower()):
                            w.writerow([al.get("nome", ""), al.get("serie", ""), al.get("curso", "")])
                    messagebox.showinfo("Sucesso", f"CSV exportado:\n{nome_arq}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")

            lbl_status_aus = ctk.CTkLabel(win, text="", font=("Segoe UI", 10), text_color=VERDE_VIBRANTE)

            def _colocar_presenca():
                sel = _sel_aus.get("selecionado")
                if not sel:
                    lbl_status_aus.configure(text="⚠ Selecione um aluno na lista de ausentes.",
                                              text_color="#C62828")
                    win.after(3000, lambda: lbl_status_aus.configure(text=""))
                    return

                matricula = sel.get("matricula", "")
                nome = sel.get("nome", "Desconhecido")
                serie = sel.get("serie", "N/A")
                curso = sel.get("curso", "N/A")

                if not messagebox.askyesno("Confirmar", f"Marcar presença de {nome} agora?"):
                    return

                try:
                    if _frequencia_duplicado_db(matricula):
                        messagebox.showinfo("Aviso", f"{nome} já está marcado como presente hoje.")
                    else:
                        hora = _agora_br().strftime("%H:%M:%S")
                        aula = _aula_por_hora(hora)
                        registro = [_hoje(), hora, matricula, nome, serie, curso, aula]
                        _inserir_frequencia_db(registro)
                        lbl_status_aus.configure(text=f"✔ Presença registrada: {nome}",
                                                  text_color=VERDE_VIBRANTE)
                        win.after(4000, lambda: lbl_status_aus.configure(text=""))
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao registrar presença:\n{e}")
                    return

                ausentes[:] = [a for a in ausentes if a.get("matricula", "") != matricula]
                _preencher()
                atualizar_freq()

            btn_row_aus = ctk.CTkFrame(win, fg_color="transparent")
            btn_row_aus.pack(anchor="w", padx=20, pady=(0, 4))
            ctk.CTkButton(btn_row_aus, text="⬇  Exportar CSV", fg_color="#1565C0",
                           hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                           height=34, command=_exportar).pack(side="left", padx=(0, 8))

            btn_presenca = ctk.CTkButton(btn_row_aus, text="✔  Colocar Presença", fg_color=VERDE_VIBRANTE,
                                          hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                                          height=34, command=_colocar_presenca)
            # Só aparece depois que um aluno é selecionado na lista (ver _selecionar_aus / _preencher)

            lbl_status_aus.pack(anchor="w", padx=20, pady=(0, 16))

            cb_serie_aus.configure(command=lambda _: _preencher())
            cb_curso_aus.configure(command=lambda _: _preencher())
            _preencher()

        # ── Exportar CSV ──────────────────────────────────────────────────────
        def exportar_csv_freq():
            try:
                nome_arq = f"frequencia_{_hoje().replace('/', '_')}.csv"
                _escrever_csv(
                    nome_arq,
                    ["Data", "HoraEntrada", "Matricula", "Nome", "Serie", "Curso", "Aula"],
                    _ler_frequencia_todos_db(),
                )
                lbl_acao_status.configure(text=f"✔ Exportado: {nome_arq}", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_acao_status.configure(text=f"⚠ Falha ao exportar: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_acao_status.configure(text=""))

        # ── Apagar registros de hoje ──────────────────────────────────────────
        def apagar_hoje_freq():
            if not messagebox.askyesno("Confirmar", f"Apagar todos os registros de hoje ({_hoje()})?"):
                return
            try:
                _apagar_frequencia_data_db(_hoje())
                lbl_acao_status.configure(text="✔ Registros de hoje apagados.", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_acao_status.configure(text=f"⚠ Falha ao apagar: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_acao_status.configure(text=""))
            atualizar_freq()

        ctk.CTkButton(btn_row, text="👁  Ver Ausentes", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=36, width=140, command=abrir_ausentes).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="↻  Atualizar", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=36, width=130, command=atualizar_freq).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="⬇  Exportar CSV", fg_color="#6A1B9A",
                       hover_color="#4A148C", font=("Segoe UI", 11, "bold"),
                       height=36, width=150, command=exportar_csv_freq).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🗑  Apagar hoje", fg_color="#C62828",
                       hover_color="#8E1010", font=("Segoe UI", 11, "bold"),
                       height=36, width=140, command=apagar_hoje_freq).pack(side="left", padx=(0, 12))
        lbl_acao_status.pack(side="left")

        # Gatilhos automáticos de filtro
        cb_serie.configure(command=lambda _: atualizar_freq())
        cb_curso.configure(command=lambda _: atualizar_freq())
        ent_nome.bind("<KeyRelease>", lambda e: atualizar_freq())

        atualizar_freq()
        return page
    def criar_pagina_historico():
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)

        # ── Estado ────────────────────────────────────────────────────────────
        _estado = {"tipo": "refeitorio", "periodo": "hoje", "dados_por_data": {}}

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", pady=(4, 12))
        ctk.CTkLabel(cab, text="Histórico 📚",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Consulte o histórico de refeições ou frequência por período.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Card de filtros ───────────────────────────────────────────────────
        filtro_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        filtro_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        # Linha 1: seletor de tipo (segmented button)
        topo_filtro = ctk.CTkFrame(filtro_card, fg_color="transparent")
        topo_filtro.pack(fill="x", padx=14, pady=(14, 8))
        ctk.CTkLabel(topo_filtro, text="Histórico de:", font=("Segoe UI", 12, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 10))

        seg_tipo = ctk.CTkSegmentedButton(topo_filtro, values=["Refeitório", "Frequência"],
                                           fg_color="#F0F0F0", selected_color=VERDE_VIBRANTE,
                                           selected_hover_color=VERDE_ESCURO,
                                           unselected_color="#F0F0F0",
                                           text_color=TEXTO_ESCURO)
        seg_tipo.set("Refeitório")
        seg_tipo.pack(side="left")

        # Linha 2: período
        periodo_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        periodo_row.pack(fill="x", padx=14, pady=(0, 8))
        ctk.CTkLabel(periodo_row, text="Período:", font=("Segoe UI", 12, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 10))

        seg_periodo = ctk.CTkSegmentedButton(periodo_row,
                                              values=["Hoje", "Esta semana", "Este mês", "Personalizado"],
                                              fg_color="#F0F0F0", selected_color="#1565C0",
                                              selected_hover_color="#0D47A1",
                                              unselected_color="#F0F0F0",
                                              text_color=TEXTO_ESCURO)
        seg_periodo.set("Hoje")
        seg_periodo.pack(side="left")

        # Linha 3: período personalizado (oculta por padrão)
        custom_row = ctk.CTkFrame(filtro_card, fg_color="transparent")

        ctk.CTkLabel(custom_row, text="De:", font=("Segoe UI", 11),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_data_ini = ctk.CTkEntry(custom_row, width=110, height=30, placeholder_text="01/01/2026")
        ent_data_ini.insert(0, "01/01/2026")
        ent_data_ini.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(custom_row, text="Até:", font=("Segoe UI", 11),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_data_fim = ctk.CTkEntry(custom_row, width=110, height=30)
        ent_data_fim.insert(0, _agora_br().date().strftime("%d/%m/%Y"))
        ent_data_fim.pack(side="left")

        # Linha 4: série / curso
        serie_curso_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
        serie_curso_row.pack(fill="x", padx=14, pady=(0, 14))

        ctk.CTkLabel(serie_curso_row, text="Série:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_serie = ctk.CTkOptionMenu(serie_curso_row, values=["(Todas)", "1 Ano", "2 Ano", "3 Ano"],
                                      width=110, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_serie.set("(Todas)")
        cb_serie.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(serie_curso_row, text="Curso:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_curso = ctk.CTkOptionMenu(serie_curso_row,
                                      values=["(Todos)", "Desenvolvimento de Sistemas", "Hospedagem",
                                              "Enfermagem", "Modelagem do Vestuario"],
                                      width=220, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_curso.set("(Todos)")
        cb_curso.pack(side="left", padx=(0, 14))

        ctk.CTkButton(serie_curso_row, text="🔍  Buscar", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=32, width=110,
                       command=lambda: buscar_historico()).pack(side="left")

        # ── Resumo geral ──────────────────────────────────────────────────────
        resumo_row = ctk.CTkFrame(page, fg_color="transparent")
        resumo_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        resumo_row.grid_columnconfigure((0, 1, 2), weight=1)

        lbl_total, sub_total = card_resumo(resumo_row, 0, 0, "📋", AZUL_CLARO, "#1565C0",
                                            "Total Geral", "0", "registros")
        lbl_pos,   sub_pos   = card_resumo(resumo_row, 0, 1, "✔️", "#E8F5E9", "#2E7D32",
                                            "Almoços / Presentes", "0", "no período")
        lbl_pct,   sub_pct   = card_resumo(resumo_row, 0, 2, "📊", ROXO_CLARO, "#6A1B9A",
                                            "% Geral", "0%", "adesão / presença")

        # ── Gráfico ───────────────────────────────────────────────────────────
        grafico_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        grafico_card.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        ctk.CTkLabel(grafico_card, text="Evolução diária (últimos 30 dias do período)",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO
                      ).pack(anchor="w", padx=18, pady=(16, 4))
        frame_graf = ctk.CTkFrame(grafico_card, fg_color="transparent")
        frame_graf.pack(fill="x", padx=18, pady=(0, 16))
        canvas_ref = {}
        _ativo_hist = {"vivo": True}

        # ── Tabela ────────────────────────────────────────────────────────────
        tabela_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=4, column=0, sticky="ew", pady=(0, 12))

        topo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        topo_tab.pack(fill="x", padx=14, pady=(14, 6))
        ctk.CTkLabel(topo_tab, text="Detalhe por dia",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")

        header_tab = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_tab.pack(fill="x", padx=14)
        col_labels = {}
        COLS = [("Data", 110), ("Total", 130), ("Almoços", 130), ("Não Almoços", 150), ("% Adesão", 110)]
        for i, (col, w) in enumerate(COLS):
            lbl = ctk.CTkLabel(header_tab, text=col, font=("Segoe UI", 10, "bold"),
                                text_color="white", width=w, anchor="w")
            lbl.pack(side="left", expand=(i == len(COLS)-1), fill="x", padx=6, pady=8)
            col_labels[col] = lbl

        corpo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo_tab.pack(fill="x", padx=14, pady=(0, 14))

        # ── Botões inferiores ─────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(page, fg_color="transparent")
        btn_row.grid(row=5, column=0, sticky="w", pady=(0, 24))
        lbl_status = ctk.CTkLabel(btn_row, text="", font=("Segoe UI", 10), text_color=VERDE_VIBRANTE)

        # ── Lógica de busca ───────────────────────────────────────────────────
        def buscar_historico():
            tipo_label = seg_tipo.get()
            eh_frequencia = (tipo_label == "Frequência")
            _estado["tipo"] = "frequencia" if eh_frequencia else "refeitorio"

            periodo_label = seg_periodo.get()
            hoje = _agora_br().date()
            if periodo_label == "Hoje":
                d_inicio = d_fim = hoje
            elif periodo_label == "Esta semana":
                d_inicio = hoje - datetime.timedelta(days=hoje.weekday())
                d_fim = hoje
            elif periodo_label == "Este mês":
                d_inicio = hoje.replace(day=1)
                d_fim = hoje
            else:
                try:
                    d_inicio = datetime.datetime.strptime(ent_data_ini.get().strip(), "%d/%m/%Y").date()
                    d_fim = datetime.datetime.strptime(ent_data_fim.get().strip(), "%d/%m/%Y").date()
                except Exception:
                    lbl_status.configure(text="⚠ Data inválida. Use DD/MM/AAAA.", text_color="#C62828")
                    return

            f_serie = cb_serie.get()
            f_curso = cb_curso.get()

            def _thread_body():
                try:
                    linhas = _ler_frequencia_todos_db() if eh_frequencia else _ler_refeitorio_todos_db()
                except Exception as e:
                    page.after(0, lambda: lbl_status.configure(
                        text=f"⚠ Falha ao carregar: {e}", text_color="#C62828"))
                    return

                dados_por_data = {}
                for row in linhas:
                    if len(row) < 7:
                        continue
                    try:
                        data_obj = datetime.datetime.strptime(row[0], "%d/%m/%Y").date()
                    except Exception:
                        continue
                    if not (d_inicio <= data_obj <= d_fim):
                        continue

                    serie = row[4] if len(row) > 4 else ""
                    curso = row[5] if len(row) > 5 else ""
                    if f_serie != "(Todas)" and f_serie not in serie:
                        continue
                    if f_curso != "(Todos)" and f_curso not in curso:
                        continue

                    if data_obj not in dados_por_data:
                        dados_por_data[data_obj] = {"total": 0, "presentes": 0,
                                                       "almocou": 0, "nao_almocou": 0}

                    if eh_frequencia:
                        dados_por_data[data_obj]["presentes"] += 1
                        dados_por_data[data_obj]["total"] += 1
                    else:
                        refeicao = row[6] if len(row) > 6 else ""
                        if refeicao.strip().lower() == "almoco":
                            dados_por_data[data_obj]["almocou"] += 1
                        else:
                            dados_por_data[data_obj]["nao_almocou"] += 1
                        dados_por_data[data_obj]["total"] += 1

                page.after(0, lambda: _renderizar(dados_por_data, eh_frequencia))

            def _limpar_hist(event=None):
                _ativo_hist["vivo"] = False
                if "canvas" in canvas_ref:
                    try:
                        canvas_ref["canvas"].get_tk_widget().pack_forget()
                        canvas_ref["canvas"].get_tk_widget().destroy()
                    except Exception:
                        pass
                    try:
                        plt.close(canvas_ref.get("fig"))
                    except Exception:
                        pass
                    canvas_ref.clear()

            page.bind("<Destroy>", _limpar_hist)

        threading.Thread(target=_thread_body, daemon=True).start()

        def _renderizar(dados_por_data, eh_frequencia):
            if not _ativo_hist["vivo"] or not page.winfo_exists():
                return
            _estado["dados_por_data"] = dados_por_data

            if eh_frequencia:
                col_labels["Total"].configure(text="Total Alunos")
                col_labels["Almoços"].configure(text="Presentes")
                col_labels["Não Almoços"].configure(text="Ausentes")
                col_labels["% Adesão"].configure(text="% Presença")
                sub_pos.configure(text="presentes")
                sub_pct.configure(text="presença")
            else:
                col_labels["Total"].configure(text="Total Registros")
                col_labels["Almoços"].configure(text="Total Almoços")
                col_labels["Não Almoços"].configure(text="Não Almoços")
                col_labels["% Adesão"].configure(text="% Adesão")
                sub_pos.configure(text="almoços")
                sub_pct.configure(text="adesão")

            for w in corpo_tab.winfo_children():
                w.destroy()

            if not dados_por_data:
                ctk.CTkLabel(corpo_tab, text="Nenhum dado para o período selecionado.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(pady=20)
                lbl_total.configure(text="0")
                lbl_pos.configure(text="0")
                lbl_pct.configure(text="0%")
                _renderizar_grafico({}, eh_frequencia)
                return

            total_geral = 0
            pos_geral = 0  # presentes ou almoços, dependendo do tipo

            for i, data_obj in enumerate(sorted(dados_por_data.keys(), reverse=True)):
                dados = dados_por_data[data_obj]
                data_str = data_obj.strftime("%d/%m/%Y")
                total = dados["total"]

                if eh_frequencia:
                    valor_pos = dados["presentes"]
                    valor_neg = total - valor_pos
                else:
                    valor_pos = dados["almocou"]
                    valor_neg = dados["nao_almocou"]

                pct = (valor_pos / total * 100) if total else 0
                total_geral += total
                pos_geral += valor_pos

                if pct >= 70:
                    cor_pct = "#2E7D32"
                elif pct >= 40:
                    cor_pct = "#E65100"
                else:
                    cor_pct = "#C62828"

                bg = BRANCO if i % 2 == 0 else "#F8F9FA"
                linha = ctk.CTkFrame(corpo_tab, fg_color=bg, corner_radius=4)
                linha.pack(fill="x", pady=1)

                valores = [(data_str, 110, "#374151"), (str(total), 130, "#374151"),
                           (str(valor_pos), 130, "#2E7D32"), (str(valor_neg), 150, "#C62828"),
                           (f"{pct:.1f}%", 110, cor_pct)]
                n = len(valores)
                for j, (val, w, cor_t) in enumerate(valores):
                    ctk.CTkLabel(linha, text=val, font=("Segoe UI", 11, "bold" if j == 4 else "normal"),
                                  width=w, anchor="w", text_color=cor_t
                                  ).pack(side="left", expand=(j == n-1), fill="x", padx=6, pady=6)

            pct_geral = (pos_geral / total_geral * 100) if total_geral else 0
            lbl_total.configure(text=str(total_geral))
            lbl_pos.configure(text=str(pos_geral))
            lbl_pct.configure(text=f"{pct_geral:.1f}%")

            _renderizar_grafico(dados_por_data, eh_frequencia)

        # ── Gráfico matplotlib ────────────────────────────────────────────────
        def _renderizar_grafico(dados_por_data, eh_frequencia):
            if not _ativo_hist["vivo"] or not page.winfo_exists():
                return
            if "canvas" in canvas_ref:
                try:
                    canvas_ref["canvas"].get_tk_widget().pack_forget()
                    canvas_ref["canvas"].get_tk_widget().destroy()
                except Exception:
                    pass
                try:
                    plt.close(canvas_ref["fig"])
                except Exception:
                    pass
                del canvas_ref["canvas"]

            for w in frame_graf.winfo_children():
                w.destroy()

            if not dados_por_data:
                ctk.CTkLabel(frame_graf, text="Sem dados para exibir.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(pady=40)
                return

            datas_ordenadas = sorted(dados_por_data.keys())[-30:]
            labels = [d.strftime("%d/%m") for d in datas_ordenadas]

            if eh_frequencia:
                valores = [dados_por_data[d]["presentes"] for d in datas_ordenadas]
                totais  = [dados_por_data[d]["total"] for d in datas_ordenadas]
                cor_label = "Presentes"
            else:
                valores = [dados_por_data[d]["almocou"] for d in datas_ordenadas]
                totais  = [dados_por_data[d]["total"] for d in datas_ordenadas]
                cor_label = "Almoços"

            cores = []
            for v, t in zip(valores, totais):
                pct = (v / t * 100) if t else 0
                if pct >= 70:
                    cores.append("#4CAF50")
                elif pct >= 40:
                    cores.append("#FDD835")
                else:
                    cores.append("#F44336")

            fig, ax = plt.subplots(figsize=(10, 3.6))
            fig.patch.set_facecolor(BRANCO)
            ax.set_facecolor(CINZA_BG)

            bars = ax.bar(labels, valores, color=cores, edgecolor="white", linewidth=1.5)
            for b, v in zip(bars, valores):
                if v > 0:
                    ax.annotate(str(v), (b.get_x() + b.get_width() / 2, b.get_height()),
                                 textcoords="offset points", xytext=(0, 4),
                                 ha="center", fontsize=9, color="#374151")

            ax.set_ylabel(cor_label, fontsize=10)
            ax.spines[["top", "right"]].set_visible(False)
            ax.grid(axis="y", linestyle="--", alpha=0.4)
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=8)
            fig.tight_layout()

            canvas_tk = FigureCanvasTkAgg(fig, master=frame_graf)
            canvas_tk.draw()
            canvas_tk.get_tk_widget().pack(fill="x")
            canvas_ref["canvas"] = canvas_tk
            canvas_ref["fig"] = fig

        # ── Exportações ───────────────────────────────────────────────────────
        def exportar_csv_historico():
            dados = _estado["dados_por_data"]
            if not dados:
                lbl_status.configure(text="⚠ Nenhum dado para exportar.", text_color="#C62828")
                page.after(4000, lambda: lbl_status.configure(text=""))
                return
            try:
                nome_arq = f"historico_marwin_{_agora_br().strftime('%d_%m_%Y')}.csv"
                eh_frequencia = _estado["tipo"] == "frequencia"
                with open(nome_arq, "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow(["Data", "Total", "Almocos/Presentes", "NaoAlmocos/Ausentes", "%Adesao/Presenca"])
                    for data_obj in sorted(dados.keys(), reverse=True):
                        d = dados[data_obj]
                        total = d["total"]
                        if eh_frequencia:
                            pos, neg = d["presentes"], total - d["presentes"]
                        else:
                            pos, neg = d["almocou"], d["nao_almocou"]
                        pct = (pos / total * 100) if total else 0
                        w.writerow([data_obj.strftime("%d/%m/%Y"), total, pos, neg, f"{pct:.1f}%"])
                lbl_status.configure(text=f"✔ CSV gerado: {nome_arq}", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_status.configure(text=f"⚠ Falha: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_status.configure(text=""))

        def exportar_pdf_historico():
            dados = _estado["dados_por_data"]
            if not dados:
                lbl_status.configure(text="⚠ Nenhum dado para exportar.", text_color="#C62828")
                page.after(4000, lambda: lbl_status.configure(text=""))
                return
            try:
                from fpdf import FPDF
                eh_frequencia = _estado["tipo"] == "frequencia"
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", "B", 16)
                titulo = "HISTORICO DE FREQUENCIA" if eh_frequencia else "HISTORICO DE REFEICOES"
                pdf.cell(0, 10, f"{titulo} - EEEP MARWIN", ln=True, align="C")
                pdf.set_font("Arial", "", 12)
                pdf.cell(0, 8, f"Gerado em: {_agora_br().strftime('%d/%m/%Y %H:%M')}", ln=True)
                pdf.ln(4)
                pdf.set_font("Arial", "B", 11)
                pdf.cell(40, 8, "Data", border=1)
                pdf.cell(40, 8, "Total", border=1)
                pdf.cell(50, 8, "Almocos/Pres.", border=1)
                pdf.cell(50, 8, "Nao Alm./Aus.", border=1)
                pdf.cell(0, 8, "% Adesao/Pres.", border=1, ln=True)
                pdf.set_font("Arial", "", 10)
                for data_obj in sorted(dados.keys(), reverse=True):
                    d = dados[data_obj]
                    total = d["total"]
                    if eh_frequencia:
                        pos, neg = d["presentes"], total - d["presentes"]
                    else:
                        pos, neg = d["almocou"], d["nao_almocou"]
                    pct = (pos / total * 100) if total else 0
                    pdf.cell(40, 7, data_obj.strftime("%d/%m/%Y"), border=1)
                    pdf.cell(40, 7, str(total), border=1)
                    pdf.cell(50, 7, str(pos), border=1)
                    pdf.cell(50, 7, str(neg), border=1)
                    pdf.cell(0, 7, f"{pct:.1f}%", border=1, ln=True)
                nome_arq = f"historico_marwin_{_agora_br().strftime('%d_%m_%Y')}.pdf"
                pdf.output(nome_arq)
                lbl_status.configure(text=f"✔ PDF gerado: {nome_arq}", text_color=VERDE_VIBRANTE)
            except Exception as e:
                lbl_status.configure(text=f"⚠ Falha: {e}", text_color="#C62828")
            page.after(4000, lambda: lbl_status.configure(text=""))

        ctk.CTkButton(btn_row, text="⬇  Exportar CSV", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=36, width=150, command=exportar_csv_historico).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="📄  Exportar PDF", fg_color="#6A1B9A",
                       hover_color="#4A148C", font=("Segoe UI", 11, "bold"),
                       height=36, width=150, command=exportar_pdf_historico).pack(side="left", padx=(0, 12))
        lbl_status.pack(side="left")

        # ── Mostrar/ocultar período personalizado ────────────────────────────
        def _on_periodo_change(valor):
            if valor == "Personalizado":
                custom_row.pack(fill="x", padx=14, pady=(0, 8), after=periodo_row)
            else:
                custom_row.pack_forget()
            buscar_historico()

        def _on_tipo_change(valor):
            buscar_historico()

        seg_periodo.configure(command=_on_periodo_change)
        seg_tipo.configure(command=_on_tipo_change)
        cb_serie.configure(command=lambda _: buscar_historico())
        cb_curso.configure(command=lambda _: buscar_historico())

        buscar_historico()
        return page

    def criar_pagina_qrcodes():
        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=0)
        page.grid_columnconfigure(1, weight=1)
        page.grid_columnconfigure(2, weight=0)
        page.grid_rowconfigure(1, weight=1)

        # ── Helpers / lógica reaproveitada do painel antigo ─────────────────────
        LISTA_FILE = os.path.join(DADOS_DIR, "lista_alunos.json")
        QR_DIR     = "qrcodes_marwin"
        os.makedirs(QR_DIR, exist_ok=True)

        def _ler_lista():
            if os.path.exists(LISTA_FILE):
                with open(LISTA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            return []

        def _salvar_lista(lista):
            with open(LISTA_FILE, "w", encoding="utf-8") as f:
                json.dump(lista, f, ensure_ascii=False, indent=2)

        def _extrair_ano_serie(serie):
            if not serie:
                return ""
            serie_strip = serie.strip()
            por_extenso = {
                "primeiro": "1", "primera": "1", "1o": "1",
                "segundo":  "2", "segunda":  "2", "2o": "2",
                "terceiro": "3", "terceira": "3", "3o": "3",
            }
            serie_lower = serie_strip.lower()
            for chave, val in por_extenso.items():
                if chave in serie_lower:
                    return val
            for ch in serie_strip:
                if ch.isdigit():
                    return ch
            return ""

        def _limpar_texto_pasta(texto):
            invalidos = r'\/:*?"<>|'
            resultado = ""
            for ch in str(texto):
                resultado += "_" if ch in invalidos else ch
            return resultado.strip()

        def _limpar_texto_arquivo(texto):
            return (str(texto)
                    .replace(" ", "_")
                    .replace("/", "_")
                    .replace("\\", "_")
                    .replace("º", "")
                    .replace("°", "")
                    .replace(":", "_")
                    .replace("*", "_")
                    .replace("?", "_")
                    .replace('"', "_")
                    .replace("<", "_")
                    .replace(">", "_")
                    .replace("|", "_")
                    .strip("_"))

        def _pasta_turma(al):
            serie = al.get("serie", "").strip()
            curso = al.get("curso", "").strip()
            ano = _extrair_ano_serie(serie)
            nome_serie = f"{ano} Ano" if ano else "Sem Serie"
            nome_curso = _limpar_texto_pasta(curso) if curso else "Sem Curso"
            pasta = os.path.join(QR_DIR, nome_serie, nome_curso)
            os.makedirs(pasta, exist_ok=True)
            return pasta

        def _nome_arquivo(al):
            matricula = _limpar_texto_arquivo(al.get("matricula", "sem_matricula"))
            nome      = _limpar_texto_arquivo(al.get("nome",      "sem_nome"))
            serie     = _limpar_texto_arquivo(al.get("serie",     ""))
            curso     = _limpar_texto_arquivo(al.get("curso",     ""))
            partes = [matricula, nome]
            if serie: partes.append(serie)
            if curso: partes.append(curso)
            nome_arq = "_".join(partes) + ".png"
            pasta    = _pasta_turma(al)
            return os.path.join(pasta, nome_arq)

        def _gerar_png(al):
            payload = json.dumps({
                "matricula": al["matricula"],
                "nome":      al["nome"],
                "serie":     al.get("serie", ""),
                "curso":     al.get("curso", "")
            }, ensure_ascii=False)
            qr = qrcode.QRCode(
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10, border=4)
            qr.add_data(payload)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            caminho = _nome_arquivo(al)
            img.save(caminho)
            return caminho

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, columnspan=3, sticky="ew", padx=4, pady=(4, 12))
        ctk.CTkLabel(cab, text="QR Codes 🔳",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Cadastre alunos e gere QR Codes individuais ou em lote.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ══════════════════════════════════════════════════════════════════════
        # COLUNA ESQUERDA — Novo aluno + preview QR
        # ══════════════════════════════════════════════════════════════════════
        esq = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12, width=300)
        esq.grid(row=1, column=0, sticky="ns", padx=(0, 12))
        esq.grid_propagate(False)

        ctk.CTkLabel(esq, text="Novo Aluno", font=("Segoe UI", 15, "bold"),
                      text_color=TEXTO_ESCURO).pack(pady=(18, 2))
        ctk.CTkFrame(esq, fg_color="#E5E7EB", height=1).pack(fill="x", padx=18, pady=(10, 12))

        campos = {}
        defs = [("Matrícula *", "matricula", "2026001"),
                ("Nome *", "nome", "Nome do Aluno"),
                ("Série (ex: 1 Ano, 2 Ano, 3 Ano)", "serie", "1 Ano"),
                ("Curso", "curso", "Desenvolvimento de Sistemas")]
        for lbl_txt, key, ph in defs:
            ctk.CTkLabel(esq, text=lbl_txt, font=("Segoe UI", 10, "bold"),
                          text_color=VERDE_VIBRANTE).pack(anchor="w", padx=18)
            e = ctk.CTkEntry(esq, font=("Segoe UI", 11), height=34,
                              placeholder_text=ph)
            e.pack(fill="x", padx=18, pady=(2, 8))
            campos[key] = e

        ctk.CTkFrame(esq, fg_color="#E5E7EB", height=1).pack(fill="x", padx=18, pady=(4, 8))

        dica_frame = ctk.CTkFrame(esq, fg_color="#F8F9FA", corner_radius=6)
        dica_frame.pack(fill="x", padx=18, pady=(0, 8))
        ctk.CTkLabel(dica_frame, text="Pasta gerada:", font=("Segoe UI", 9, "bold"),
                      text_color=VERDE_VIBRANTE).pack(anchor="w", padx=8, pady=(6, 0))
        dica_lbl = ctk.CTkLabel(dica_frame,
                                 text="qrcodes_marwin/\n  1 Ano/\n    Desenvolvimento de Sistemas/",
                                 font=("Courier", 9), text_color=TEXTO_CINZA, justify="left")
        dica_lbl.pack(anchor="w", padx=8, pady=(2, 6))

        def _atualizar_dica(*_):
            serie = campos["serie"].get().strip()
            curso = campos["curso"].get().strip()
            ano   = _extrair_ano_serie(serie)
            nome_s = f"{ano} Ano" if ano else "Sem Serie"
            nome_c = _limpar_texto_pasta(curso) if curso else "Sem Curso"
            dica_lbl.configure(text=f"qrcodes_marwin/\n  {nome_s}/\n    {nome_c}/")

        campos["serie"].bind("<KeyRelease>", _atualizar_dica)
        campos["curso"].bind("<KeyRelease>", _atualizar_dica)

        # Preview QR
        preview_frame = ctk.CTkFrame(esq, fg_color="#F8F9FA", corner_radius=8,
                                       width=250, height=250)
        preview_frame.pack(pady=6)
        preview_frame.pack_propagate(False)
        preview_lbl = ctk.CTkLabel(preview_frame, text="QR Code preview\nserá exibido aqui",
                                    font=("Segoe UI", 10), text_color=TEXTO_CINZA)
        preview_lbl.pack(expand=True)
        _img_ref = {}

        def _preview_qr(al):
            payload = json.dumps({
                "matricula": al["matricula"], "nome": al["nome"],
                "serie": al.get("serie", ""), "curso": al.get("curso", "")
            }, ensure_ascii=False)
            qr2 = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M,
                                 box_size=6, border=2)
            qr2.add_data(payload); qr2.make(fit=True)
            img_pil = qr2.make_image(fill_color="black", back_color="white").convert("RGB")
            img_pil = img_pil.resize((230, 230), Image.NEAREST)
            img_ctk = ctk.CTkImage(light_image=img_pil, dark_image=img_pil, size=(230, 230))
            _img_ref["img"] = img_ctk
            preview_lbl.configure(image=img_ctk, text="")

        lbl_form_status = ctk.CTkLabel(esq, text="", font=("Segoe UI", 10),
                                        text_color=VERDE_VIBRANTE, wraplength=260, justify="center")
        lbl_form_status.pack(pady=(4, 4), padx=18)

        def _campos_para_dict():
            return {
                "matricula": campos["matricula"].get().strip(),
                "nome":      campos["nome"].get().strip(),
                "serie":     campos["serie"].get().strip(),
                "curso":     campos["curso"].get().strip(),
            }

        def adicionar_aluno():
            al = _campos_para_dict()
            if not al["matricula"] or not al["nome"]:
                lbl_form_status.configure(text="⚠ Informe Matrícula e Nome.", text_color="#C62828")
                return
            lista = _ler_lista()
            if any(a["matricula"] == al["matricula"] for a in lista):
                lbl_form_status.configure(text=f"⚠ Matrícula {al['matricula']} já cadastrada.",
                                           text_color="#C62828")
                return
            lista.append(al)
            _salvar_lista(lista)
            caminho = _gerar_png(al)
            _preview_qr(al)
            _atualizar_dica()
            carregar_lista()
            lbl_form_status.configure(text=f"✔ Aluno adicionado!\nQR salvo em:\n{caminho}",
                                       text_color=VERDE_VIBRANTE)

        def reemitir_selecionado():
            sel = _estado.get("selecionado")
            if not sel:
                lbl_form_status.configure(text="⚠ Selecione um aluno na lista.", text_color="#C62828")
                return
            lista = _ler_lista()
            al = next((a for a in lista if a["matricula"] == sel["matricula"]), None)
            if not al:
                return
            caminho = _gerar_png(al)
            _preview_qr(al)
            for key in ("matricula", "nome", "serie", "curso"):
                campos[key].delete(0, "end")
                campos[key].insert(0, al.get(key, ""))
            _atualizar_dica()
            lbl_form_status.configure(text=f"✔ QR reemitido!\nSalvo em:\n{caminho}",
                                       text_color=VERDE_VIBRANTE)

        ctk.CTkButton(esq, text="➕  Adicionar e Gerar QR", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=38, command=adicionar_aluno).pack(fill="x", padx=18, pady=(4, 4))
        ctk.CTkButton(esq, text="🔁  Reemitir QR Selecionado", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=38, command=reemitir_selecionado).pack(fill="x", padx=18, pady=(0, 16))

        # ══════════════════════════════════════════════════════════════════════
        # COLUNA CENTRAL — Lista de alunos cadastrados
        # ══════════════════════════════════════════════════════════════════════
        mid = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        mid.grid(row=1, column=1, sticky="nsew", padx=(0, 12))
        mid.grid_rowconfigure(3, weight=1)
        mid.grid_columnconfigure(0, weight=1)

        topo_mid = ctk.CTkFrame(mid, fg_color="transparent")
        topo_mid.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 4))
        ctk.CTkLabel(topo_mid, text="Alunos Cadastrados", font=("Segoe UI", 14, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left")
        lbl_cnt = ctk.CTkLabel(topo_mid, text="", font=("Segoe UI", 10), text_color=TEXTO_CINZA)
        lbl_cnt.pack(side="left", padx=10)

        busca_row = ctk.CTkFrame(mid, fg_color="transparent")
        busca_row.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 8))
        ctk.CTkLabel(busca_row, text="Buscar:", font=("Segoe UI", 11),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_busca = ctk.CTkEntry(busca_row, height=32,
                                  placeholder_text="Nome, matrícula, série ou curso...")
        ent_busca.pack(side="left", fill="x", expand=True)

        header_lista = ctk.CTkFrame(mid, fg_color=VERDE_ESCURO, corner_radius=6)
        header_lista.grid(row=2, column=0, sticky="ew", padx=14)
        COLS = [("QR", 50), ("Matrícula", 100), ("Nome", 220), ("Série", 110), ("Curso", 200)]
        for i, (col, w) in enumerate(COLS):
            ctk.CTkLabel(header_lista, text=col, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=w, anchor="w"
                          ).pack(side="left", expand=(i == len(COLS)-1), fill="x", padx=6, pady=8)

        corpo_lista = ctk.CTkFrame(mid, fg_color="transparent")
        corpo_lista.grid(row=3, column=0, sticky="nsew", padx=14, pady=(0, 8))
        corpo_lista.grid_columnconfigure(0, weight=1)

        _estado = {"selecionado": None, "linha_widgets": {}}

        # Botões inferiores da coluna central
        btn_row_mid = ctk.CTkFrame(mid, fg_color="transparent")
        btn_row_mid.grid(row=4, column=0, sticky="ew", padx=14, pady=(0, 14))
        lbl_mid_status = ctk.CTkLabel(btn_row_mid, text="", font=("Segoe UI", 10),
                                       text_color=VERDE_VIBRANTE)

        def remover_aluno():
            sel = _estado.get("selecionado")
            if not sel:
                lbl_mid_status.configure(text="⚠ Selecione um aluno.", text_color="#C62828")
                page.after(3000, lambda: lbl_mid_status.configure(text=""))
                return
            if not messagebox.askyesno("Confirmar",
                    f"Remover {sel['nome']} da lista?\n(O arquivo PNG não será apagado.)"):
                return
            lista = _ler_lista()
            lista = [a for a in lista if a["matricula"] != sel["matricula"]]
            _salvar_lista(lista)
            _estado["selecionado"] = None
            carregar_lista()

        def gerar_lote():
            lista = _ler_lista()
            if not lista:
                lbl_mid_status.configure(text="⚠ Nenhum aluno cadastrado.", text_color="#C62828")
                page.after(3000, lambda: lbl_mid_status.configure(text=""))
                return
            if not messagebox.askyesno("Confirmar",
                    f"Gerar/atualizar QR Codes para {len(lista)} aluno(s)?\n\n"
                    f"Estrutura de pastas:\n"
                    f"  qrcodes_marwin/\n"
                    f"    1 Ano/\n"
                    f"      Desenvolvimento de Sistemas/\n"
                    f"    2 Ano/\n"
                    f"      ...\n"
                    f"    3 Ano/\n"
                    f"      ..."):
                return

            def _thread_body():
                erros = 0
                pastas_criadas = set()
                for al in lista:
                    try:
                        _gerar_png(al)
                        pastas_criadas.add(_pasta_turma(al))
                    except Exception:
                        erros += 1
                msg = (f"✔ {len(lista)-erros} QR Code(s) gerados em "
                       f"{len(pastas_criadas)} pasta(s) dentro de '{QR_DIR}/'.")
                if erros:
                    msg += f" ({erros} erro(s))"
                page.after(0, lambda: (lbl_mid_status.configure(text=msg, text_color=VERDE_VIBRANTE),
                                        carregar_lista(ent_busca.get())))
                page.after(6000, lambda: lbl_mid_status.configure(text=""))

            threading.Thread(target=_thread_body, daemon=True).start()

        def abrir_pasta():
            import subprocess as sp
            try:
                pasta = os.path.abspath(QR_DIR)
                if sys.platform == "win32":
                    sp.Popen(["explorer", pasta])
                elif sys.platform == "darwin":
                    sp.Popen(["open", pasta])
                else:
                    sp.Popen(["xdg-open", pasta])
            except Exception as e:
                lbl_mid_status.configure(text=f"⚠ {e}", text_color="#C62828")
                page.after(4000, lambda: lbl_mid_status.configure(text=""))

        ctk.CTkButton(btn_row_mid, text="🗑  Remover", fg_color="#C62828",
                       hover_color="#8E1010", font=("Segoe UI", 11, "bold"),
                       height=36, width=110, command=remover_aluno).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row_mid, text="🔳  Gerar QRs (lote)", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=36, width=160, command=gerar_lote).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row_mid, text="📁  Abrir Pasta QRs", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=36, width=160, command=abrir_pasta).pack(side="left", padx=(0, 12))
        lbl_mid_status.pack(side="left")

        # ── Carregamento da lista ────────────────────────────────────────────
        def _selecionar(al):
            _estado["selecionado"] = al
            for mat, frame in _estado["linha_widgets"].items():
                frame.configure(fg_color=(BRANCO if mat != al["matricula"] else "#E8F5E9"))

        def carregar_lista(filtro=""):
            for w in corpo_lista.winfo_children():
                w.destroy()
            _estado["linha_widgets"] = {}

            lista = _ler_lista()
            filtro_lower = filtro.lower()
            exibidos = []
            for al in lista:
                serie = al.get("serie", "")
                curso = al.get("curso", "")
                if filtro_lower and filtro_lower not in al.get("nome", "").lower() \
                        and filtro_lower not in al.get("matricula", "").lower() \
                        and filtro_lower not in curso.lower() \
                        and filtro_lower not in serie.lower():
                    continue
                exibidos.append(al)

            lbl_cnt.configure(text=f"({len(exibidos)} de {len(lista)} aluno(s))")

            if not exibidos:
                ctk.CTkLabel(corpo_lista, text="Nenhum aluno encontrado.",
                              font=("Segoe UI", 12), text_color=TEXTO_CINZA
                              ).grid(row=0, column=0, pady=24)
                return

            for i, al in enumerate(exibidos):
                serie = al.get("serie", "")
                curso = al.get("curso", "")
                tem_png = os.path.exists(_nome_arquivo(al))
                ano = _extrair_ano_serie(serie)
                ano_exib = f"{ano}º Ano" if ano else "-"

                bg_normal = BRANCO if i % 2 == 0 else "#F8F9FA"
                if not tem_png:
                    bg_normal = "#FFF9C4"

                linha = ctk.CTkFrame(corpo_lista, fg_color=bg_normal, corner_radius=4, cursor="hand2")
                linha.grid(row=i, column=0, sticky="ew", pady=1)
                _estado["linha_widgets"][al["matricula"]] = linha

                icone_qr = "✔" if tem_png else "✘"
                cor_icone = "#2E7D32" if tem_png else "#C62828"

                valores = [(icone_qr, 50, cor_icone), (al.get("matricula", ""), 100, "#374151"),
                           (al.get("nome", ""), 220, "#374151"), (ano_exib, 110, "#374151"),
                           (curso or "-", 200, "#374151")]
                n = len(valores)
                widgets_linha = []
                for j, (val, w, cor_t) in enumerate(valores):
                    lbl = ctk.CTkLabel(linha, text=val, font=("Segoe UI", 11),
                                        width=w, anchor="w", text_color=cor_t)
                    lbl.pack(side="left", expand=(j == n-1), fill="x", padx=6, pady=6)
                    widgets_linha.append(lbl)

                def _bind_click(widget, a=al):
                    widget.bind("<Button-1>", lambda e, _a=a: _selecionar(_a))

                _bind_click(linha)
                for w in widgets_linha:
                    _bind_click(w)

        ent_busca.bind("<KeyRelease>", lambda e: carregar_lista(ent_busca.get()))

        # ══════════════════════════════════════════════════════════════════════
        # COLUNA DIREITA — Importar planilha
        # ══════════════════════════════════════════════════════════════════════
        dir_col = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12, width=300)
        dir_col.grid(row=1, column=2, sticky="ns")

        ctk.CTkLabel(dir_col, text="Importar Planilha", font=("Segoe UI", 15, "bold"),
                      text_color=TEXTO_ESCURO).pack(pady=(18, 2))
        ctk.CTkLabel(dir_col, text="XLSX ou CSV com os alunos",
                      font=("Segoe UI", 10), text_color=TEXTO_CINZA).pack(pady=(0, 8))
        ctk.CTkFrame(dir_col, fg_color="#E5E7EB", height=1).pack(fill="x", padx=18, pady=(0, 12))

        frame_inst = ctk.CTkFrame(dir_col, fg_color="#F8F9FA", corner_radius=8)
        frame_inst.pack(fill="x", padx=18, pady=(0, 10))
        ctk.CTkLabel(frame_inst, text="Colunas esperadas:", font=("Segoe UI", 10, "bold"),
                      text_color=VERDE_VIBRANTE).pack(anchor="w", padx=10, pady=(8, 2))
        for col_txt in ["matrícula  (obrigatório)", "nome       (obrigatório)",
                        "série      (ex: 1 Ano, 2 Ano)", "curso      (ex: Des. Sistemas)"]:
            ctk.CTkLabel(frame_inst, text=f"• {col_txt}", font=("Segoe UI", 9),
                          text_color=TEXTO_ESCURO, justify="left").pack(anchor="w", padx=10)
        ctk.CTkLabel(frame_inst,
                      text="\nEstrutura gerada:\nqrcodes_marwin/\n"
                           "  1 Ano/\n    Desenvolvimento de Sistemas/\n"
                           "  2 Ano/\n    ...\n  3 Ano/\n    ...",
                      font=("Courier", 8), text_color=TEXTO_CINZA, justify="left"
                      ).pack(anchor="w", padx=10, pady=(4, 8))

        ctk.CTkLabel(dir_col, text="Mapear colunas (opcional)", font=("Segoe UI", 10, "bold"),
                      text_color=VERDE_VIBRANTE).pack(anchor="w", padx=18, pady=(4, 0))
        ctk.CTkLabel(dir_col, text="Deixe em branco p/ detecção automática.",
                      font=("Segoe UI", 9), text_color=TEXTO_CINZA).pack(anchor="w", padx=18, pady=(0, 6))

        map_vars = {}
        for campo_mk, rotulo_mk in [("matricula", "Col. Matrícula"), ("nome", "Col. Nome"),
                                     ("serie", "Col. Série"), ("curso", "Col. Curso")]:
            fr = ctk.CTkFrame(dir_col, fg_color="transparent")
            fr.pack(fill="x", padx=18, pady=(0, 4))
            ctk.CTkLabel(fr, text=rotulo_mk, font=("Segoe UI", 9),
                          text_color=TEXTO_ESCURO, width=110, anchor="w").pack(side="left")
            ent = ctk.CTkEntry(fr, height=28, font=("Segoe UI", 9))
            ent.pack(side="left", fill="x", expand=True)
            map_vars[campo_mk] = ent

        ctk.CTkFrame(dir_col, fg_color="#E5E7EB", height=1).pack(fill="x", padx=18, pady=(10, 10))

        gerar_qr_import_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(dir_col, text="Gerar QR Codes ao importar",
                         variable=gerar_qr_import_var, font=("Segoe UI", 10, "bold"),
                         text_color=VERDE_VIBRANTE, fg_color=VERDE_VIBRANTE,
                         hover_color=VERDE_ESCURO).pack(anchor="w", padx=18, pady=(0, 4))

        sobreescrever_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(dir_col, text="Atualizar duplicatas",
                         variable=sobreescrever_var, font=("Segoe UI", 10),
                         text_color=TEXTO_ESCURO, fg_color=VERDE_VIBRANTE,
                         hover_color=VERDE_ESCURO).pack(anchor="w", padx=18, pady=(0, 10))

        lbl_import_status = ctk.CTkLabel(dir_col, text="", font=("Segoe UI", 10, "bold"),
                                          text_color=VERDE_VIBRANTE, wraplength=250, justify="center")
        lbl_import_status.pack(pady=4, padx=18)

        prog_bar = ctk.CTkProgressBar(dir_col, height=10)
        prog_bar.set(0)
        prog_bar.pack(fill="x", padx=18, pady=(0, 10))

        def _normalizar_cabecalho(cabecalho):
            mapa_auto = {}
            sinonimos = {
                "matricula": ["matricula", "mat", "mat.", "codigo", "id", "registro"],
                "nome":      ["nome", "aluno", "estudante", "discente", "name", "nomecompleto"],
                "serie":     ["serie", "turma", "ano", "class", "classe", "periodo", "ano/serie"],
                "curso":     ["curso", "habilitacao", "area", "modalidade", "formacao"],
            }
            cab_lower = [str(c).strip().lower() for c in cabecalho]
            for campo, sinonimos_lista in sinonimos.items():
                manual = map_vars[campo].get().strip()
                if manual and manual in cabecalho:
                    mapa_auto[campo] = cabecalho.index(manual)
                    continue
                for s in sinonimos_lista:
                    for idx_c, c in enumerate(cab_lower):
                        if s in c:
                            mapa_auto[campo] = idx_c
                            break
                    if campo in mapa_auto:
                        break
            return mapa_auto

        def importar_planilha():
            from tkinter import filedialog
            caminho_pl = filedialog.askopenfilename(
                title="Selecionar planilha de alunos",
                filetypes=[("Planilhas", "*.xlsx *.xls *.csv *.tsv"),
                           ("Excel", "*.xlsx *.xls"),
                           ("CSV / TSV", "*.csv *.tsv"),
                           ("Todos", "*.*")])
            if not caminho_pl:
                return

            lbl_import_status.configure(text="Lendo arquivo...", text_color=TEXTO_CINZA)
            page.update()

            ext = os.path.splitext(caminho_pl)[1].lower()
            linhas_raw = []

            try:
                if ext in (".xlsx", ".xls"):
                    try:
                        import openpyxl
                    except ImportError:
                        lbl_import_status.configure(text="Instalando openpyxl...", text_color=TEXTO_CINZA)
                        page.update()
                        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
                        import openpyxl
                    wb = openpyxl.load_workbook(caminho_pl, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        linhas_raw.append([str(c).strip() if c is not None else "" for c in row])
                    wb.close()
                elif ext in (".csv", ".tsv"):
                    sep = "\t" if ext == ".tsv" else None
                    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                        try:
                            with open(caminho_pl, "r", encoding=enc, newline="") as f:
                                sample = f.read(4096); f.seek(0)
                                if sep is None:
                                    try: sep = csv.Sniffer().sniff(sample).delimiter
                                    except: sep = ","
                                reader_pl = csv.reader(f, delimiter=sep)
                                linhas_raw = [[c.strip() for c in row] for row in reader_pl]
                            break
                        except (UnicodeDecodeError, Exception):
                            continue
                else:
                    lbl_import_status.configure(text="⚠ Use .xlsx, .xls, .csv ou .tsv", text_color="#C62828")
                    return
            except Exception as e:
                lbl_import_status.configure(text=f"⚠ Erro ao ler arquivo: {e}", text_color="#C62828")
                return

            if len(linhas_raw) < 2:
                lbl_import_status.configure(text="⚠ Arquivo sem dados suficientes.", text_color="#C62828")
                return

            cabecalho = linhas_raw[0]
            mapa = _normalizar_cabecalho(cabecalho)

            if "matricula" not in mapa or "nome" not in mapa:
                lbl_import_status.configure(
                    text="⚠ Não foi possível identificar colunas de Matrícula e Nome.\n"
                         "Use o mapeamento manual.",
                    text_color="#C62828")
                return

            lista_atual = _ler_lista()
            mats_existentes = {a["matricula"]: i for i, a in enumerate(lista_atual)}

            novos = 0; atualizados = 0; ignorados = 0
            dados_importados = []

            linhas_dados = [l for l in linhas_raw[1:] if any(c for c in l)]
            total_linhas = max(len(linhas_dados), 1)

            for idx_linha, linha in enumerate(linhas_dados):
                prog_bar.set((idx_linha + 1) / total_linhas)
                page.update_idletasks()

                def _cel(campo, _linha=linha):
                    idx_c = mapa.get(campo)
                    if idx_c is None or idx_c >= len(_linha):
                        return ""
                    return str(_linha[idx_c]).strip()

                matricula = _cel("matricula")
                nome      = _cel("nome")
                serie     = _cel("serie")
                curso     = _cel("curso")

                if not matricula or not nome:
                    ignorados += 1
                    continue

                al = {"matricula": matricula, "nome": nome, "serie": serie, "curso": curso}

                if matricula in mats_existentes:
                    if sobreescrever_var.get():
                        lista_atual[mats_existentes[matricula]] = al
                        atualizados += 1
                        dados_importados.append(al)
                    else:
                        ignorados += 1
                else:
                    lista_atual.append(al)
                    mats_existentes[matricula] = len(lista_atual) - 1
                    novos += 1
                    dados_importados.append(al)

            _salvar_lista(lista_atual)

            msg_qr = ""
            if gerar_qr_import_var.get() and dados_importados:
                lbl_import_status.configure(text=f"Gerando {len(dados_importados)} QR Code(s)...",
                                             text_color=TEXTO_CINZA)
                prog_bar.set(0)
                erros_qr = 0
                pastas_criadas = set()
                total_al = max(len(dados_importados), 1)
                for idx_al, al in enumerate(dados_importados):
                    prog_bar.set((idx_al + 1) / total_al)
                    page.update_idletasks()
                    try:
                        _gerar_png(al)
                        pastas_criadas.add(_pasta_turma(al))
                    except Exception:
                        erros_qr += 1
                msg_qr = (f"\n{len(dados_importados)-erros_qr} QR Code(s) gerados "
                          f"em {len(pastas_criadas)} pasta(s) dentro de '{QR_DIR}/'.")
                if erros_qr:
                    msg_qr += f" ({erros_qr} erros)"

            prog_bar.set(0)
            carregar_lista(ent_busca.get())

            resumo = (f"✔ Importação concluída!\n"
                      f"Novos: {novos}  |  Atualizados: {atualizados}  |  "
                      f"Ignorados: {ignorados}{msg_qr}")
            lbl_import_status.configure(text=resumo, text_color=VERDE_VIBRANTE)

        ctk.CTkButton(dir_col, text="📂  Selecionar Arquivo e Importar", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=42, command=importar_planilha).pack(fill="x", padx=18, pady=(0, 6))
        ctk.CTkButton(dir_col, text="📁  Abrir Pasta dos QR Codes", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 10, "bold"),
                       height=34, command=abrir_pasta).pack(fill="x", padx=18, pady=(0, 18))

        carregar_lista()
        return page

    # ════════════════════════════════════════════════════════════════════
    # PÁGINA: LOGS DO SISTEMA
    # ════════════════════════════════════════════════════════════════════
    def criar_pagina_logs():
        import re as _re

        page = ctk.CTkFrame(_scroll_inner, fg_color=CINZA_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(3, weight=1)

        CORES_NIVEL = {
            "DEBUG":   "#9E9E9E",
            "INFO":    "#2196F3",
            "WARNING": "#FB8C00",
            "ERROR":   "#F44336",
            "CRITICAL": "#B71C1C",
        }

        # ── Cabeçalho ─────────────────────────────────────────────────────────
        cab = ctk.CTkFrame(page, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", pady=(4, 12))
        ctk.CTkLabel(cab, text="Logs do Sistema 📄",
                      font=("Segoe UI", 22, "bold"), text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(cab, text="Acompanhe os eventos registrados pelo servidor.",
                      font=("Segoe UI", 12), text_color=TEXTO_CINZA).pack(anchor="w")

        # ── Card de filtros ───────────────────────────────────────────────────
        filtro_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        filtro_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        linha_f = ctk.CTkFrame(filtro_card, fg_color="transparent")
        linha_f.pack(fill="x", padx=14, pady=14)

        ctk.CTkLabel(linha_f, text="Nível:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        cb_nivel = ctk.CTkOptionMenu(linha_f,
                                      values=["Todos", "INFO", "WARNING", "ERROR", "DEBUG", "CRITICAL"],
                                      width=120, fg_color=VERDE_VIBRANTE,
                                      button_color=VERDE_ESCURO, button_hover_color=VERDE_ESCURO)
        cb_nivel.set("Todos")
        cb_nivel.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(linha_f, text="Buscar:", font=("Segoe UI", 11, "bold"),
                      text_color=TEXTO_ESCURO).pack(side="left", padx=(0, 6))
        ent_busca = ctk.CTkEntry(linha_f, width=260, height=32,
                                  placeholder_text="Filtrar por texto na mensagem...")
        ent_busca.pack(side="left", padx=(0, 14))

        ctk.CTkButton(linha_f, text="🔍  Buscar", fg_color=VERDE_VIBRANTE,
                       hover_color=VERDE_ESCURO, font=("Segoe UI", 11, "bold"),
                       height=32, width=100,
                       command=lambda: atualizar_logs()).pack(side="left", padx=(0, 8))
        ctk.CTkButton(linha_f, text="↻  Atualizar", fg_color="#1565C0",
                       hover_color="#0D47A1", font=("Segoe UI", 11, "bold"),
                       height=32, width=110,
                       command=lambda: atualizar_logs()).pack(side="left", padx=(0, 8))
        ctk.CTkButton(linha_f, text="🗑  Limpar logs", fg_color="#C62828",
                       hover_color="#8E1F1F", font=("Segoe UI", 11, "bold"),
                       height=32, width=120,
                       command=lambda: limpar_logs()).pack(side="left")

        # ── Cards de resumo ───────────────────────────────────────────────────
        resumo_row = ctk.CTkFrame(page, fg_color="transparent")
        resumo_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        resumo_row.grid_columnconfigure((0, 1, 2, 3), weight=1)

        lbl_total, _ = card_resumo(resumo_row, 0, 0, "📄", AZUL_CLARO, "#1565C0",
                                    "Total exibido", "0", "linhas")
        lbl_info, _ = card_resumo(resumo_row, 0, 1, "ℹ️", VERDE_CLARO, "#2196F3",
                                   "INFO", "0", "")
        lbl_warn, _ = card_resumo(resumo_row, 0, 2, "⚠️", LARANJA_CLARO, "#FB8C00",
                                   "WARNING", "0", "")
        lbl_err, _ = card_resumo(resumo_row, 0, 3, "🚨", VERMELHO_CLARO, "#F44336",
                                  "ERROR / CRITICAL", "0", "")

        # ── Tabela de logs ─────────────────────────────────────────────────────
        tabela_card = ctk.CTkFrame(page, fg_color=BRANCO, corner_radius=12)
        tabela_card.grid(row=3, column=0, sticky="nsew", pady=(0, 4))
        tabela_card.grid_rowconfigure(2, weight=1)
        tabela_card.grid_columnconfigure(0, weight=1)

        topo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        topo_tab.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        ctk.CTkLabel(topo_tab, text="Eventos recentes (mais novos primeiro)",
                      font=("Segoe UI", 13, "bold"), text_color=TEXTO_ESCURO).pack(side="left")

        header_tab = ctk.CTkFrame(tabela_card, fg_color=VERDE_ESCURO, corner_radius=6)
        header_tab.grid(row=1, column=0, sticky="ew", padx=14)
        COLS = [("DATA/HORA", 150), ("NÍVEL", 100), ("MENSAGEM", 0)]
        for i, (col, w) in enumerate(COLS):
            ctk.CTkLabel(header_tab, text=col, font=("Segoe UI", 10, "bold"),
                          text_color="white", width=w, anchor="w"
                          ).pack(side="left", expand=(i == len(COLS) - 1), fill="x", padx=8, pady=8)

        corpo_tab = ctk.CTkFrame(tabela_card, fg_color="transparent")
        corpo_tab.grid(row=2, column=0, sticky="nsew", padx=14, pady=(4, 14))
        corpo_tab.grid_columnconfigure(0, weight=1)

        lbl_status = ctk.CTkLabel(page, text="", font=("Segoe UI", 10), text_color=TEXTO_CINZA)
        lbl_status.grid(row=4, column=0, sticky="w", pady=(8, 16))

        PADRAO_LOG = _re.compile(r"^(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}) - (\w+) - (.*)$")

        def _ler_linhas_log(max_linhas=500):
            if not os.path.exists(LOG_FILE):
                return []
            try:
                with open(LOG_FILE, "r", encoding="utf-8") as f:
                    linhas = f.readlines()
            except Exception:
                return []
            return linhas[-max_linhas:]

        def limpar_logs():
            if not messagebox.askyesno("Confirmar", "Limpar todos os logs do sistema?\nEsta ação não pode ser desfeita."):
                return
            try:
                open(LOG_FILE, "w", encoding="utf-8").close()
                logger.info("Logs do sistema limpos via painel administrativo")
                atualizar_logs()
                messagebox.showinfo("Sucesso", "Logs limpos com sucesso.")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao limpar logs:\n{e}")

        def atualizar_logs():
            for w in corpo_tab.winfo_children():
                w.destroy()

            nivel_f = cb_nivel.get()
            busca_f = ent_busca.get().strip().lower()

            linhas_raw = _ler_linhas_log(500)

            registros = []
            for linha in linhas_raw:
                linha = linha.rstrip("\n")
                if not linha:
                    continue
                m = PADRAO_LOG.match(linha)
                if m:
                    data_hora, nivel, msg = m.group(1), m.group(2), m.group(3)
                else:
                    data_hora, nivel, msg = "", "INFO", linha

                if nivel_f != "Todos" and nivel.upper() != nivel_f:
                    continue
                if busca_f and busca_f not in linha.lower():
                    continue

                registros.append((data_hora, nivel.upper(), msg))

            registros.reverse()  # mais recentes primeiro

            cnt_info = sum(1 for r in registros if r[1] == "INFO")
            cnt_warn = sum(1 for r in registros if r[1] == "WARNING")
            cnt_err = sum(1 for r in registros if r[1] in ("ERROR", "CRITICAL"))

            lbl_total.configure(text=str(len(registros)))
            lbl_info.configure(text=str(cnt_info))
            lbl_warn.configure(text=str(cnt_warn))
            lbl_err.configure(text=str(cnt_err))

            if not registros:
                ctk.CTkLabel(corpo_tab, text="Nenhum log encontrado para o filtro selecionado.",
                              font=("Segoe UI", 11), text_color=TEXTO_CINZA
                              ).grid(row=0, column=0, sticky="w", pady=12, padx=4)
            else:
                for i, (data_hora, nivel, msg) in enumerate(registros):
                    cor = CORES_NIVEL.get(nivel, TEXTO_ESCURO)
                    linha_frame = ctk.CTkFrame(corpo_tab, fg_color=("#FAFAFA" if i % 2 else "transparent"))
                    linha_frame.grid(row=i, column=0, sticky="ew")
                    linha_frame.grid_columnconfigure(2, weight=1)

                    ctk.CTkLabel(linha_frame, text=data_hora or "—", font=("Segoe UI", 11),
                                  text_color="#374151", width=150, anchor="w"
                                  ).grid(row=0, column=0, padx=8, pady=6, sticky="w")
                    ctk.CTkLabel(linha_frame, text=nivel, font=("Segoe UI", 10, "bold"),
                                  text_color=cor, width=100, anchor="w"
                                  ).grid(row=0, column=1, padx=8, pady=6, sticky="w")
                    ctk.CTkLabel(linha_frame, text=msg, font=("Segoe UI", 11),
                                  text_color="#374151", anchor="w", justify="left",
                                  wraplength=700
                                  ).grid(row=0, column=2, padx=8, pady=6, sticky="ew")

            agora_txt = _agora_br().strftime("%d/%m/%Y %H:%M:%S")
            lbl_status.configure(text=f"Atualizado em {agora_txt}  ·  exibindo até 500 linhas mais recentes do arquivo de log.")

        atualizar_logs()
        return page


def aplicar_tema(parent):
    for w in parent.winfo_children():
        cls = w.winfo_class()
        try:
            if cls in ("Frame", "Labelframe"):
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"]); break
                    if bg == nome_tema["ENTRY_BG"]:       w.configure(bg=T["ENTRY_BG"]); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"]); break
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BORDER_GRID"]:    w.configure(bg=T["BORDER_GRID"]); break
            elif cls == "Label":
                bg = w.cget("bg"); fg = w.cget("fg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"]); break
                    if bg == nome_tema["ENTRY_BG"]:       w.configure(bg=T["ENTRY_BG"]); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"]); break
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"]); break
                for nome_tema in TEMAS.values():
                    if fg == nome_tema["ACCENT_VIBRANT"]: w.configure(fg=T["ACCENT_VIBRANT"]); break
                    if fg == nome_tema["FRASE_FG"]:       w.configure(fg=T["FRASE_FG"]); break
                    if fg == nome_tema["FG_TEXT"]:        w.configure(fg=T["FG_TEXT"]); break
            elif cls == "Button":
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["ACCENT_VIBRANT"]: w.configure(bg=T["ACCENT_VIBRANT"], fg="white"); break
                    if bg == nome_tema["ACCENT_SOFT"]:    w.configure(bg=T["ACCENT_SOFT"],    fg="white"); break
                    if bg == nome_tema["BTN_CARDAPIO"]:   w.configure(bg=T["BTN_CARDAPIO"],   fg="white"); break
                    if bg == nome_tema["BTN_VOLTAR"]:     w.configure(bg=T["BTN_VOLTAR"],     fg="white"); break
                    if bg == nome_tema["OBS_BG"]:         w.configure(bg=T["OBS_BG"],  fg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BG_MAIN"]:        w.configure(bg=T["BG_MAIN"], fg=T["ACCENT_VIBRANT"]); break
                    if bg == nome_tema["BG_CARD"]:        w.configure(bg=T["BG_CARD"], fg=T["FRASE_FG"]); break
            elif cls == "Entry":
                w.configure(bg=T["ENTRY_BG"], fg=T["FG_TEXT"], insertbackground=T["ACCENT_VIBRANT"])
            elif cls == "Canvas":
                bg = w.cget("bg")
                for nome_tema in TEMAS.values():
                    if bg == nome_tema["BG_CARD"]: w.configure(bg=T["BG_CARD"]); break
                    if bg == nome_tema["BG_MAIN"]: w.configure(bg=T["BG_MAIN"]); break
            elif cls == "Checkbutton":
                w.configure(bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"],
                            activebackground=T["BG_CARD"], selectcolor=T["ENTRY_BG"])
        except Exception:
            pass
        if w.winfo_children():
            aplicar_tema(w)

def alternar_tema():
    global tema_atual, T
    tema_atual = "escuro" if tema_atual=="claro" else "claro"
    T = TEMAS[tema_atual]; salvar_tema_tk(tema_atual)
    janela.configure(bg=T["BG_MAIN"])
    btn_tema.configure(text="☀️ Tema" if tema_atual=="escuro" else "🌙 Tema",
                       bg=T["ACCENT_VIBRANT"], fg="#b2dfb4")
    aplicar_tema(janela)
    for w in janela.winfo_children():
        if w.winfo_class() == "Toplevel":
            w.configure(bg=T["BG_MAIN"]); aplicar_tema(w)

def iniciar_tkinter(url_publica):
    global janela, T, tema_atual, url_ngrok_global, btn_tema
    url_ngrok_global = url_publica
    tema_atual = carregar_tema()
    T = TEMAS[tema_atual]
    janela = tk.Tk()
    janela.title("EEEP MARWIN — Servidor Admin")
    janela.state("zoomed"); janela.configure(bg=T["BG_MAIN"])
    janela.bind("<F11>", lambda e: janela.attributes("-fullscreen", not janela.attributes("-fullscreen")))

    # ── Barra de topo institucional ──────────────────────────────────────────
    barra_topo = tk.Frame(janela, bg=T["ACCENT_VIBRANT"], height=64)
    barra_topo.pack(fill="x"); barra_topo.pack_propagate(False)
    tk.Label(barra_topo, text="EEEP MARWIN",
             font=("Segoe UI", 15, "bold"), bg=T["ACCENT_VIBRANT"], fg="white").pack(side="left", padx=28, pady=16)
    tk.Label(barra_topo, text="Painel do Servidor Admin",
             font=("Segoe UI", 9), bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="left", padx=(0, 0), pady=22)

    # URL + botão copiar à direita
    url_frame = tk.Frame(barra_topo, bg=T["ACCENT_VIBRANT"]); url_frame.pack(side="right", padx=16)
    tk.Label(url_frame, text=f"● {url_publica}", font=("Segoe UI", 9),
             bg=T["ACCENT_VIBRANT"], fg="#b2dfb4").pack(side="left", padx=(0, 8))
    tk.Button(url_frame, text="Copiar URL", font=("Segoe UI", 8, "bold"),
              bg=T["ACCENT_SOFT"], fg="white", bd=0, padx=10, pady=4, cursor="hand2",
              command=lambda: [janela.clipboard_clear(),
                               janela.clipboard_append(url_publica),
                               messagebox.showinfo("Copiado", "URL copiada para a área de transferência!")]
              ).pack(side="left")

    # Botão tema na barra de topo
    btn_tema = tk.Button(barra_topo, text="🌙 Tema", font=("Segoe UI", 10),
                         bg=T["ACCENT_VIBRANT"], fg="#b2dfb4", bd=0,
                         cursor="hand2", activebackground=T["ACCENT_SOFT"],
                         command=alternar_tema)
    btn_tema.pack(side="right", padx=4)

    # ── Relógio em tempo real na barra de topo ───────────────────────────────
    lbl_relogio = tk.Label(barra_topo, text="", font=("Segoe UI", 10, "bold"),
                           bg=T["ACCENT_VIBRANT"], fg="#b2dfb4")
    lbl_relogio.pack(side="right", padx=(0, 16))

    def _atualizar_relogio():
        agora = _agora_br()
        lbl_relogio.config(text=agora.strftime("%d/%m/%Y  %H:%M:%S"))
        janela.after(1000, _atualizar_relogio)

    _atualizar_relogio()

    # Linha amarela accent
    tk.Frame(janela, bg=T["HIGHLIGHT_YELLOW"], height=4).pack(fill="x")

    # ── Conteúdo central ────────────────────────────────────────────────────
    main = tk.Frame(janela, bg=T["BG_MAIN"]); main.pack(expand=True, fill="both")
    center_wrap = tk.Frame(main, bg=T["BG_MAIN"])
    center_wrap.place(relx=0.5, rely=0.5, anchor="center")

    # Logo / ícone
    try:
        from PIL import Image as _PilImg, ImageTk as _PilImgTk
        _path_marwin = _buscar_logo_png()
        if not _path_marwin:
            _pasta = os.path.dirname(os.path.abspath(__file__))
            _path_marwin = os.path.join(_pasta, "logo_marwin.png")
        if os.path.exists(_path_marwin):
            _img = _PilImg.open(_path_marwin).convert("RGBA")
            _img.thumbnail((520, 520), _PilImg.LANCZOS)
            _logo_tk = _PilImgTk.PhotoImage(_img, master=janela)
            tk.Label(center_wrap, image=_logo_tk, bg=T["BG_MAIN"]).pack(pady=(0, 10))
            center_wrap._logo_ref = _logo_tk   # evitar GC
    except Exception:
        tk.Label(center_wrap, text="🏫", font=("Segoe UI", 44), bg=T["BG_MAIN"]).pack(pady=(0, 8))

    tk.Label(center_wrap, text="Escola Estadual de Educação Profissional",
             font=("Segoe UI", 13), bg=T["BG_MAIN"], fg=T["FRASE_FG"]).pack()
    tk.Label(center_wrap, text="MARWIN",
             font=("Segoe UI", 44, "bold"), bg=T["BG_MAIN"], fg=T["ACCENT_VIBRANT"]).pack(pady=(2, 4))
    tk.Label(center_wrap, text="Painel do Servidor",
             font=("Segoe UI", 11), bg=T["BG_MAIN"], fg=T["FRASE_FG"]).pack()

    tk.Frame(center_wrap, bg=T["BORDER_GRID"], height=1).pack(fill="x", pady=(18, 14))

    # ── Cards de status em tempo real ────────────────────────────────────────
    cards_frame = tk.Frame(center_wrap, bg=T["BG_MAIN"])
    cards_frame.pack(fill="x", pady=(0, 14))

    def _criar_card_status(parent, titulo, col):
        card = tk.Frame(parent, bg=T["BG_CARD"],
                        highlightbackground=T["BORDER_GRID"], highlightthickness=1)
        card.grid(row=0, column=col, padx=6, pady=0, sticky="nsew")
        lbl_titulo = tk.Label(card, text=titulo, font=("Segoe UI", 8),
                              bg=T["BG_CARD"], fg=T["FRASE_FG"])
        lbl_titulo.pack(pady=(8, 2), padx=14)
        lbl_valor = tk.Label(card, text="—", font=("Segoe UI", 22, "bold"),
                             bg=T["BG_CARD"], fg=T["ACCENT_VIBRANT"])
        lbl_valor.pack(pady=(0, 2), padx=14)
        lbl_sub = tk.Label(card, text="", font=("Segoe UI", 8),
                           bg=T["BG_CARD"], fg=T["FRASE_FG"])
        lbl_sub.pack(pady=(0, 8), padx=14)
        return lbl_valor, lbl_sub

    for c in range(3):
        cards_frame.grid_columnconfigure(c, weight=1, uniform="sc")

    lbl_val_ref,  lbl_sub_ref  = _criar_card_status(cards_frame, "Refeições hoje",   0)
    lbl_val_freq, lbl_sub_freq = _criar_card_status(cards_frame, "Presenças hoje",   1)
    lbl_val_db,   lbl_sub_db   = _criar_card_status(cards_frame, "Banco de dados",   2)

    # Indicador de saúde: ponto colorido no card de banco
    _dot_colors = {"ok": "#4caf50", "erro": "#f44336", "verificando": "#fdd835"}

    def _atualizar_cards_status():
        """Atualiza os cards de status a cada 15 segundos.

        A consulta ao banco roda em thread separada (evita travar a UI com
        a latência de rede do Neon), mas TODA atualização de widgets é feita
        de volta na thread principal via janela.after — Tkinter não é
        thread-safe e chamar .config() direto de outra thread pode travar
        ou corromper o loop de eventos da aplicação inteira.
        """
        def _tarefa():
            # Refeitório hoje
            try:
                total_ref = len(_ler_refeitorio_hoje_db())
                ref_result = (str(total_ref), "registros no refeitório")
            except Exception:
                ref_result = ("—", "indisponível")

            # Frequência hoje
            try:
                total_freq = len(_ler_frequencia_hoje_db())
                freq_result = (str(total_freq), "alunos registrados")
            except Exception:
                freq_result = ("—", "indisponível")

            # Saúde do banco
            try:
                conn = get_pg_conn()
                if conn:
                    _release_pg_conn(conn)
                    db_result = ("● Online", _dot_colors["ok"], "PostgreSQL Neon")
                else:
                    db_result = ("● Offline", _dot_colors["erro"], "sem conexão")
            except Exception:
                db_result = ("● Erro", _dot_colors["erro"], "falha na verificação")

            def _aplicar():
                lbl_val_ref.config(text=ref_result[0])
                lbl_sub_ref.config(text=ref_result[1])
                lbl_val_freq.config(text=freq_result[0])
                lbl_sub_freq.config(text=freq_result[1])
                lbl_val_db.config(text=db_result[0], fg=db_result[1])
                lbl_sub_db.config(text=db_result[2])

            janela.after(0, _aplicar)

        threading.Thread(target=_tarefa, daemon=True).start()
        janela.after(15000, _atualizar_cards_status)

    _atualizar_cards_status()

    tk.Button(center_wrap, text="  ABRIR PAINEL ADMIN  ",
              font=("Segoe UI", 17, "bold"),
              bg=T["ACCENT_VIBRANT"], fg="white", bd=0,
              padx=50, pady=20, cursor="hand2",
              activebackground=T["ACCENT_SOFT"], activeforeground="white",
              command=abrir_painel_admin_ctk).pack(fill="x")

    tk.Label(center_wrap,
             text="💚 Servidor ativo — aguardando conexões dos clientes.",
             bg=T["BG_MAIN"], fg=T["FRASE_FG"],
             font=("Segoe UI", 10, "italic")).pack(pady=(24, 0))

    janela.mainloop()

# ==============================================================================
# INICIALIZACAO
# ==============================================================================
if __name__ == "__main__":
    _iniciar_pool_pg()
    try:
        _criar_tabelas_neon()
    except Exception:
        pass
    if _cloud_api_url():
        threading.Thread(target=_sincronizar_tudo_nuvem, daemon=True).start()

    threading.Thread(
        target=lambda: app.run(host="0.0.0.0", port=5000, use_reloader=False, debug=False, threaded=True),
        daemon=True
    ).start()
    time.sleep(1)

    local_ip = _get_local_ip()
    url = f"http://{local_ip}:5000"
    logger.info(f"Servidor MARWIN iniciado em {url}")
    print(f"\n[CLIENTE] Acesse localmente em: {url}")
    print(f"[CLIENTE] Ou use http://localhost:5000 no proprio computador")

    iniciar_tkinter(url)